"""Enterprise Solar Programme -- bulk beneficiary import (rebuild, slice 5).

WHY AN IMPORT IS STAGED AND NOT APPLIED
---------------------------------------
A ministry does not type 4000 schools into a form. It has a spreadsheet, and that
spreadsheet is wrong in a dozen places -- a roof area of 40 where it should be 400, a
region spelled three ways, six schools listed twice because two districts both claimed
them.

An importer that writes straight to the register turns that into a choice between importing
nothing and importing the mess. So NOTHING is written until the operator has seen what will
happen: every row is parsed, mapped, coerced, validated and duplicate-checked into a
staging table first, and the register is only touched on an explicit commit. The raw row is
kept beside the mapped one, so that when somebody asks in six months why a school's roof
area is 40, the answer is still in the database.

WHAT THIS DELIBERATELY DOES NOT DO
----------------------------------
Run in the background. There is no worker and there cannot be one: Render's free tier caps
the account at a single instance (Supervisor correction R1 -- a second service was already
blocked once), so bulk work has to be a durable job table drained by a GitHub-Actions cron.
That machinery lands with slice 7, where the expensive thing (generating and DESIGNING
hundreds of projects) actually lives. Parsing a spreadsheet is not expensive.

So Release 1 parses and commits IN THE REQUEST, under a stated, enforced ceiling
(constants.IMPORT_MAX_ROWS). A cap that is announced is honest engineering; a request that
silently dies at row 8000 is not. The UI says the number.
"""

from __future__ import annotations

import csv
import io
import json

from . import beneficiaries, rbac, txn
from .constants import (
    BENEFICIARY_FIELD_SPEC,
    BENEFICIARY_TYPES,
    IMPORT_MAX_ROWS,
)
from .gates import EnterpriseGateError

_FIELD_KEYS = [f["key"] for f in BENEFICIARY_FIELD_SPEC]
_BENEFICIARY_TYPE_CODES = frozenset(code for code, _ in BENEFICIARY_TYPES)

# The caps that make IMPORT_MAX_ROWS reachable. A row limit checked AFTER parsing is no
# limit at all: a one-row file with a 500 MB cell, or an XLSX that unzips to gigabytes, kills
# the process long before anyone counts the rows. Sized against reality -- a 2000-row
# beneficiary register is a few hundred kilobytes.
MAX_UPLOAD_BYTES = 16 * 1024 * 1024            # 16 MB on the wire
MAX_XLSX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024  # what the zip may claim to expand to
MAX_CELL_CHARS = 4096                          # one cell; the longest real field is an address

# Columns the importer understands beyond the 22 attributes.
_EXTRA_KEYS = ["code", "beneficiary_type"]
_IMPORTABLE_KEYS = _EXTRA_KEYS + _FIELD_KEYS


class ImportError_(EnterpriseGateError):
    """An import rule was broken. Named with a trailing underscore so it cannot shadow the
    Python builtin ImportError, which would be a genuinely confusing thing to catch."""


# --- parsing -----------------------------------------------------------------


def parse_file(filename: str, data: bytes) -> tuple[list[str], list[dict]]:
    """Read a CSV or XLSX into headers plus a list of row dicts.

    Input:  the uploaded filename (only its extension is used) and its bytes.
    Output: (headers, rows) -- rows are {header: cell} with everything as a string.
    Raises: ImportError_ (409) on an unreadable or oversized file.

    Everything comes back as a STRING. The spreadsheet's own idea of a type is not to be
    trusted -- Excel will happily hand back a float for a district code and a datetime for
    something that was never a date -- so coercion happens once, in the validator, against
    the field spec, where the rules live.
    """
    name = (filename or "").lower()

    # THE ROW CAP IS NOT ENOUGH ON ITS OWN (Codex slice-5, MED). It is checked AFTER parsing,
    # so a single-row file with one 500 MB cell, or an XLSX zip bomb, never reaches it -- the
    # process is already dead. The byte cap is what makes the row cap reachable.
    if len(data) > MAX_UPLOAD_BYTES:
        raise ImportError_(
            "IMPORT",
            f"that file is {len(data) // (1024 * 1024)} MB and the limit is "
            f"{MAX_UPLOAD_BYTES // (1024 * 1024)} MB. A {IMPORT_MAX_ROWS}-row spreadsheet "
            "is far smaller than this -- check you uploaded the right file.",
        )

    if name.endswith(".csv") or name.endswith(".txt"):
        headers, rows = _parse_csv(data)
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        _reject_zip_bomb(data)
        headers, rows = _parse_xlsx(data)
    else:
        raise ImportError_(
            "IMPORT", "unsupported file type: upload a .csv or .xlsx spreadsheet"
        )

    if not headers:
        raise ImportError_("IMPORT", "the file has no header row")

    # DUPLICATE HEADERS ARE REFUSED (Codex slice-5, MED). A CSV with two columns both called
    # "School Name" collapses to one in a dict, and the loser's value is gone before the
    # operator ever sees it -- which quietly makes a lie of the promise that the raw row is
    # kept verbatim. There is no safe guess about which column was meant, so we do not make
    # one.
    seen: dict[str, int] = {}
    for header in headers:
        seen[header] = seen.get(header, 0) + 1
    repeated = sorted(h for h, n in seen.items() if n > 1)
    if repeated:
        raise ImportError_(
            "IMPORT",
            "the file has more than one column called "
            + ", ".join(repr(h) for h in repeated)
            + ". Rename them so each column means one thing.",
        )

    if not rows:
        raise ImportError_("IMPORT", "the file has a header row but no data")
    if len(rows) > IMPORT_MAX_ROWS:
        # Refused, not truncated. Importing the first 2000 rows of a 5000-row file and
        # reporting success would be the worst possible outcome: the operator would have no
        # reason to look for the other 3000.
        raise ImportError_(
            "IMPORT",
            f"this file has {len(rows)} rows and the limit is {IMPORT_MAX_ROWS}. "
            "Split it and import in parts.",
        )
    return headers, rows


def _reject_zip_bomb(data: bytes) -> None:
    """Refuse an XLSX that expands to far more than it claims to be.

    Input:  the uploaded bytes.
    Output: none.
    Raises: ImportError_ (409).

    An XLSX is a zip. A few kilobytes of it can declare gigabytes of content, and openpyxl
    will faithfully try to build all of it -- on a 512 MB free-tier instance, that is the
    whole site down, from one authenticated upload. So the archive's own declared sizes are
    inspected BEFORE it is opened: a real spreadsheet compresses perhaps 10:1, and a real
    2000-row register is a few megabytes uncompressed.
    """
    import zipfile

    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            declared = sum(info.file_size for info in archive.infolist())
    except zipfile.BadZipFile as e:
        raise ImportError_("IMPORT", "that file is not a readable .xlsx") from e

    if declared > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise ImportError_(
            "IMPORT",
            "that spreadsheet expands to far more than a beneficiary register should "
            "(refusing to open it)",
        )


def _parse_csv(data: bytes) -> tuple[list[str], list[dict]]:
    """CSV, tolerant about encoding because real spreadsheets are exported from Excel."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - latin-1 decodes any byte string
        raise ImportError_("IMPORT", "could not decode the file as text")

    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip() for h in (reader.fieldnames or []) if h and h.strip()]
    rows = []
    for raw in reader:
        row = {(k or "").strip(): _cell(v)
               for k, v in raw.items() if k and k.strip()}
        if any(row.values()):        # skip the blank lines every spreadsheet ends with
            rows.append(row)
    return headers, rows


def _cell(value) -> str:
    """One cell, as a bounded string.

    Truncated rather than refused: an over-long cell is almost always a stray paragraph in a
    notes column, not an attack, and failing the whole file for it would be obnoxious. The
    bound is what stops a single cell from being the whole of memory. Anything this long was
    never going to fit a beneficiary field anyway -- the validator will reject it on its own
    terms if it matters.
    """
    if value is None:
        return ""
    text = str(value).strip()
    return text[:MAX_CELL_CHARS]


def _parse_xlsx(data: bytes) -> tuple[list[str], list[dict]]:
    """XLSX via openpyxl, read-only so a large file does not build a full object model."""
    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover - openpyxl is installed
        raise ImportError_(
            "IMPORT", "XLSX support is unavailable on this server; upload a CSV instead"
        ) from e

    try:
        book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ImportError_("IMPORT", "that file could not be read as a spreadsheet") from e

    sheet = book.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    # THE COLUMN INDEX TRAVELS WITH THE HEADER (Codex slice-5 round 2, HIGH). Blank header
    # cells are common in a real spreadsheet -- a spacer column, a stray formatted cell past
    # the data. Dropping them from the header list and then reading values back by the
    # POSITION IN THE COMPACTED LIST shifts every column after the blank one left by one:
    # ["Name", "", "Site Code"] would read the empty column's value as the site code, and the
    # preview would look perfectly coherent while the register filled with wrong data. So the
    # original index is kept with each named column, and an unnamed column is simply not read.
    columns = [
        (i, str(h).strip())
        for i, h in enumerate(header_row)
        if h is not None and str(h).strip()
    ]
    headers = [h for _i, h in columns]
    rows = []
    for values in rows_iter:
        row = {}
        for i, header in columns:
            value = values[i] if i < len(values) else None
            row[header] = _cell(value)
        if any(row.values()):
            rows.append(row)
        if len(rows) > IMPORT_MAX_ROWS:
            # Stop READING, not just refuse afterwards: an XLSX can declare a million rows,
            # and materialising them all before saying "too many" is the denial of service
            # the cap was supposed to prevent.
            book.close()
            raise ImportError_(
                "IMPORT",
                f"this file has more than {IMPORT_MAX_ROWS} rows. Split it and import "
                "in parts.",
            )
    book.close()
    return headers, rows


# --- mapping -----------------------------------------------------------------


def auto_map(headers: list[str]) -> dict[str, str]:
    """Guess which spreadsheet column is which beneficiary field.

    Input:  the header row.
    Output: {header: field_key} for the headers it recognised. Unrecognised headers are
            simply absent -- they are shown to the operator to map by hand.

    A GUESS, and treated as one: the mapping is presented for confirmation before anything
    is staged, because silently mapping "Location" to `address` when the ministry meant
    `community` would corrupt 4000 records in a way nobody would notice until a truck
    turned up at the wrong village.
    """
    mapping: dict[str, str] = {}
    for header in headers:
        key = _normalise(header)
        if key in _IMPORTABLE_KEYS:
            mapping[header] = key
            continue
        alias = _ALIASES.get(key)
        if alias:
            mapping[header] = alias
    return mapping


def _normalise(header: str) -> str:
    """'Roof Area (m2)' -> 'roof_area'. Punctuation and units are noise."""
    text = str(header or "").strip().lower()
    for junk in ("(m2)", "(m²)", "(kwh)", "(kwh/month)", "(ghs)", "(kw)", "(usd)"):
        text = text.replace(junk, "")
    out = []
    for ch in text:
        out.append(ch if ch.isalnum() else "_")
    key = "".join(out)
    while "__" in key:
        key = key.replace("__", "_")
    return key.strip("_")


# The names real spreadsheets actually use. Extend freely -- an alias that is wrong costs
# one unmapped column the operator fixes by hand; an alias that is missing costs the same.
_ALIASES: dict[str, str] = {
    "beneficiary": "name",
    "beneficiary_name": "name",
    "school": "name",
    "school_name": "name",
    "facility": "name",
    "facility_name": "name",
    "institution": "name",
    "site_name": "name",
    "id": "code",
    "ref": "code",
    "reference": "code",
    "site_code": "code",
    "school_code": "code",
    "beneficiary_code": "code",
    "type": "beneficiary_type",
    "category": "beneficiary_type",
    "facility_type": "beneficiary_type",
    "site_type": "beneficiary_type",
    "town": "community",
    "village": "community",
    "locality": "community",
    "gps": "gps_coordinates",
    "coordinates": "gps_coordinates",
    "lat_long": "gps_coordinates",
    "latlon": "gps_coordinates",
    "contact": "contact_person",
    "contact_name": "contact_person",
    "phone": "contact_details",
    "telephone": "contact_details",
    "email": "contact_details",
    "contact_phone": "contact_details",
    "students": "occupancy",
    "population": "occupancy",
    "beds": "occupancy",
    "staff": "occupancy",
    "consumption": "electricity_consumption",
    "monthly_consumption": "electricity_consumption",
    "kwh_per_month": "electricity_consumption",
    "monthly_kwh": "electricity_consumption",
    "bill": "electricity_consumption",
    "roof": "roof_area",
    "roof_size": "roof_area",
    "land": "land_availability",
    "land_area": "land_availability",
    "priority": "priority_ranking",
    "rank": "priority_ranking",
    "power_source": "existing_energy_source",
    "energy_source": "existing_energy_source",
    "current_supply": "existing_energy_source",
    "generator": "generator_details",
    "owner": "ownership",
    "owned_by": "ownership",
    "impact": "social_impact_class",
}


def _check_mapping(mapping: dict[str, str] | None) -> None:
    """A column mapping must name real fields, and no two columns may claim the same one.

    Input:  {spreadsheet header: field key}.
    Output: none.
    Raises: ImportError_ (409).

    TWO COLUMNS ONTO ONE FIELD is refused (Codex slice-5, MED) rather than resolved. If both
    "School Name" and "Institution" are mapped to `name`, one of them wins by dict ordering
    -- which is to say, arbitrarily, invisibly, and identically for all 4000 rows. The
    operator made the choice; the operator gets told it does not make sense.
    """
    claimed: dict[str, str] = {}
    for header, key in (mapping or {}).items():
        if not key:
            continue
        if key not in _IMPORTABLE_KEYS:
            raise ImportError_("IMPORT", f"cannot map {header!r} to unknown field {key!r}")
        if key in claimed:
            raise ImportError_(
                "IMPORT",
                f"both {claimed[key]!r} and {header!r} are mapped to {key!r}. "
                "One field, one column.",
            )
        claimed[key] = header


def importable_fields() -> list[dict]:
    """Every field an import may fill, for the mapping UI's dropdowns."""
    labels = {f["key"]: f["key"].replace("_", " ").title()
              for f in BENEFICIARY_FIELD_SPEC}
    labels["code"] = "Code (unique reference)"
    labels["beneficiary_type"] = "Beneficiary Type"
    return [{"value": k, "label": labels[k]} for k in _IMPORTABLE_KEYS]


# --- staging -----------------------------------------------------------------


def stage_import(c, tenant_id: str, user_id: int, programme_id: int, *,
                 filename: str, headers: list[str], rows: list[dict],
                 mapping: dict[str, str], default_type: str = "", audit=None) -> int:
    """Validate every row and write the batch to staging. Touches the register: NEVER.

    Input:  connection, tenant id, acting user, programme id, the filename, the parsed
            headers and rows, the confirmed column mapping, a fallback beneficiary type for
            rows whose file does not carry one, the audit hook.
    Output: the new batch id.
    Raises: EnterprisePermissionError (403), ImportError_ / BeneficiaryError (409 / C13).

    Every row lands in staging with a status: Valid, Error (with the reasons, per field) or
    Duplicate. Nothing is created. commit_batch() is what actually writes.
    """
    beneficiaries._load_programme(c, tenant_id, programme_id)   # C13 FIRST
    rbac.require_permission(c, tenant_id, user_id, "beneficiary.import",
                            programme_id=programme_id)

    if len(rows) > IMPORT_MAX_ROWS:
        raise ImportError_(
            "IMPORT", f"{len(rows)} rows exceeds the {IMPORT_MAX_ROWS}-row import limit"
        )

    # A mapping that points at a field the register does not have would silently drop the
    # column. Refuse it: the operator chose it, so the operator should be told.
    _check_mapping(mapping)

    index = beneficiaries.DuplicateIndex.for_programme(c, tenant_id, programme_id)
    staged = [_stage_row(c, tenant_id, programme_id, i, raw, mapping, default_type, index)
              for i, raw in enumerate(rows, start=1)]

    counts = {"Valid": 0, "Error": 0, "Duplicate": 0}
    for row in staged:
        counts[row["status"]] = counts.get(row["status"], 0) + 1

    audit = audit or txn.audit_on(c)
    with txn.atomic(c):
        cur = c.execute(
            "INSERT INTO enterprise_import_batches "
            "(tenant_id, programme_id, filename, status, column_mapping, default_type, "
            " total_rows, valid_rows, error_rows, duplicate_rows, created_by_user_id) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (tenant_id, programme_id, filename, "Staged", json.dumps(mapping or {}),
             default_type or "", len(staged), counts["Valid"], counts["Error"],
             counts["Duplicate"], user_id),
        )
        batch_id = txn.inserted_id(c, cur)

        for row in staged:
            c.execute(
                "INSERT INTO enterprise_import_rows "
                "(tenant_id, batch_id, row_no, raw_json, mapped_json, status, errors_json) "
                "VALUES (?,?,?,?,?,?,?)",
                (tenant_id, batch_id, row["row_no"], json.dumps(row["raw"]),
                 json.dumps(row["mapped"]), row["status"], json.dumps(row["errors"])),
            )

        _require_audit(
            audit("ENTERPRISE_IMPORT_STAGED", user_id=user_id, tenant_id=tenant_id,
                  details={"programme_id": programme_id, "batch_id": batch_id,
                           "filename": filename, "total": len(staged), **counts}),
            "import staged",
        )
    return batch_id


def _stage_row(c, tenant_id, programme_id, row_no, raw, mapping, default_type,
               index) -> dict:
    """Map, coerce, validate and duplicate-check ONE row. Never raises.

    A raising row-validator would abort the whole file on the first typo, which is the
    behaviour this module exists to avoid. A bad row becomes an Error row and the import
    carries on.
    """
    mapped: dict = {}
    for header, value in (raw or {}).items():
        key = (mapping or {}).get(header)
        if key:
            mapped[key] = value

    errors: list[str] = []

    name = str(mapped.get("name") or "").strip()
    # Canonicalised HERE, not just at write, so the duplicate probe below compares the same
    # form the register stores. 'KP-01', 'kp-01' and 'KP 01' are one school.
    code = beneficiaries.canonical_code(mapped.get("code"))
    beneficiary_type = str(mapped.get("beneficiary_type") or default_type or "").strip()
    beneficiary_type = beneficiary_type.lower().replace(" ", "_")

    if not name:
        errors.append("name is required")
    if beneficiary_type and beneficiary_type not in _BENEFICIARY_TYPE_CODES:
        errors.append(f"beneficiary type {beneficiary_type!r} is not one we recognise")
        beneficiary_type = ""
    if not beneficiary_type:
        errors.append("beneficiary type is required (map a column, or choose a default)")

    # No code in the file? Derive a stable one from the name. Stable MATTERS: it is what
    # makes re-importing the same spreadsheet a no-op instead of a duplicated register, so
    # it must not depend on the row's position in the file (which changes when somebody
    # sorts the sheet).
    if not code and name:
        code = _derive_code(name, mapped.get("community"))

    clean, field_problems = validate_row_fields(mapped)
    errors.extend(field_problems)

    status = "Error" if errors else "Valid"
    if status == "Valid":
        # The index knows the register AND the rows already read from this same file, so a
        # spreadsheet that lists one school twice is caught on the second listing rather than
        # importing it twice (or promising "0 duplicates" and then failing at commit).
        found = index.find(code=code, name=name, community=clean.get("community"))
        if found:
            where, which = found
            status = "Duplicate"
            errors.append(
                f"already in the register as beneficiary #{which}" if where == "register"
                else f"the same site is already on row {which} of this file"
            )
        else:
            index.remember(("row", row_no), code=code, name=name,
                           community=clean.get("community"))

    clean["name"] = name
    return {
        "row_no": row_no,
        "raw": raw,
        "mapped": dict(clean, code=code, beneficiary_type=beneficiary_type),
        "status": status,
        "errors": errors,
    }


def validate_row_fields(mapped: dict) -> tuple[dict, list[str]]:
    """The 22 attributes of one row, through the register's own validator.

    One validator, not two: beneficiaries.validate_fields is what the manual form uses, so
    a value the import accepts is exactly a value the form would have accepted. A separate
    import-side validator is how the two end up disagreeing about what a legal tariff is.
    """
    return beneficiaries.validate_fields(mapped, require_name=False)


def _derive_code(name: str, community) -> str:
    """A stable, readable code for a row that arrived without one.

    'Kpando Senior High' in 'Kpando' -> 'KPANDO-KPANDO-SENIOR-HIGH'.

    Derived from the CONTENT, never from the row number: a spreadsheet that gets sorted
    between two imports must still produce the same codes, or the duplicate check silently
    stops working and the second import doubles the register.

    Built through beneficiaries.canonical_code, which NFKC-normalises -- so a name whose
    accented characters are composed on one machine and decomposed on another still derives
    the same code. Two byte strings that no human can tell apart must not become two schools.
    """
    parts = [p for p in (beneficiaries.canonical_code(community),
                         beneficiaries.canonical_code(name)) if p]
    return beneficiaries.canonical_code("-".join(parts))


# --- reading staged batches --------------------------------------------------


def _load_batch(c, tenant_id: str, batch_id: int) -> dict:
    """Fetch a batch IN THIS TENANT, or raise C13 (which the routes turn into a 404)."""
    row = c.execute(
        "SELECT id, programme_id, filename, status, column_mapping, total_rows, "
        "       valid_rows, error_rows, duplicate_rows, imported_rows, created_at, "
        "       default_type "
        "  FROM enterprise_import_batches WHERE tenant_id=? AND id=?",
        (tenant_id, batch_id),
    ).fetchone()
    if row is None:
        raise ImportError_("C13", "no such import in this organisation")
    return {
        "id": row[0], "programme_id": row[1], "filename": row[2], "status": row[3],
        "column_mapping": _decode(row[4], {}), "total_rows": row[5],
        "valid_rows": row[6], "error_rows": row[7], "duplicate_rows": row[8],
        "imported_rows": row[9], "created_at": row[10], "default_type": row[11],
    }


def get_batch(c, tenant_id: str, batch_id: int, *, limit: int = 500) -> dict:
    """A staged batch with its rows, for the preview screen.

    The row list is CAPPED and the cap is reported (`rows_shown` vs `total_rows`), because
    a 2000-row preview table is not a preview. The COUNTS come from the batch row, which
    the database computed over every row -- so the summary is never a count of the
    truncated page.
    """
    batch = _load_batch(c, tenant_id, batch_id)
    rows = c.execute(
        "SELECT row_no, raw_json, mapped_json, status, errors_json, beneficiary_id "
        "  FROM enterprise_import_rows WHERE tenant_id=? AND batch_id=? "
        " ORDER BY CASE status WHEN 'Error' THEN 0 WHEN 'Duplicate' THEN 1 ELSE 2 END, "
        "          row_no LIMIT ?",
        (tenant_id, batch_id, limit),
    ).fetchall()
    batch["rows"] = [
        {"row_no": r[0], "raw": _decode(r[1], {}), "mapped": _decode(r[2], {}),
         "status": r[3], "errors": _decode(r[4], []), "beneficiary_id": r[5]}
        for r in rows
    ]
    batch["rows_shown"] = len(batch["rows"])
    return batch


def list_batches(c, tenant_id: str, programme_id: int, limit: int = 25) -> list[dict]:
    rows = c.execute(
        "SELECT id, filename, status, total_rows, valid_rows, error_rows, "
        "       duplicate_rows, imported_rows, created_at "
        "  FROM enterprise_import_batches WHERE tenant_id=? AND programme_id=? "
        " ORDER BY id DESC LIMIT ?",
        (tenant_id, programme_id, limit),
    ).fetchall()
    return [
        {"id": r[0], "filename": r[1], "status": r[2], "total_rows": r[3],
         "valid_rows": r[4], "error_rows": r[5], "duplicate_rows": r[6],
         "imported_rows": r[7], "created_at": r[8]}
        for r in rows
    ]


def _decode(raw, default):
    """jsonb comes back decoded from psycopg2; SQLite hands back the TEXT we stored."""
    if isinstance(raw, (dict, list)):
        return raw
    if not raw:
        return default
    try:
        return json.loads(raw)
    except (TypeError, ValueError):
        return default


# --- committing --------------------------------------------------------------


def commit_batch(c, tenant_id: str, user_id: int, batch_id: int, *,
                 include_duplicates: bool = False, audit=None) -> dict:
    """Create a beneficiary for every Valid row. THE ONLY thing that writes the register.

    Input:  connection, tenant id, acting user, batch id, whether to import the rows that
            matched an existing beneficiary, the audit hook.
    Output: {"imported": n, "skipped": n, "failed": n}.
    Raises: EnterprisePermissionError (403), ImportError_ (409 / C13).

    Error rows are never imported -- they failed validation, so there is nothing coherent to
    create. Duplicate rows are skipped BY DEFAULT and only imported if the operator
    explicitly says so, having seen them.

    ONE TRANSACTION for the whole batch, so a failure halfway through leaves no half-imported
    register. A row that fails at the database (a code that raced another import) is
    recorded as an Error row rather than taking the batch down with it.

    Committing twice is a no-op, not a second import: the batch has to be Staged.
    """
    batch = _load_batch(c, tenant_id, batch_id)                 # C13 FIRST
    rbac.require_permission(c, tenant_id, user_id, "beneficiary.import",
                            programme_id=batch["programme_id"])

    if batch["status"] != "Staged":
        raise ImportError_(
            "IMPORT",
            f"this import is already {batch['status']}; it cannot be committed again",
        )

    wanted = ["Valid"] + (["Duplicate"] if include_duplicates else [])
    placeholders = ",".join("?" for _ in wanted)
    rows = c.execute(
        f"SELECT id, row_no, mapped_json FROM enterprise_import_rows "
        f" WHERE tenant_id=? AND batch_id=? AND status IN ({placeholders}) ORDER BY row_no",
        tuple([tenant_id, batch_id] + wanted),
    ).fetchall()

    imported = 0
    failed = 0
    audit = audit or txn.audit_on(c)

    with txn.atomic(c):
        # The batch is claimed FIRST, conditionally on it still being Staged. Two operators
        # pressing Import at the same moment would otherwise both read "Staged", both pass
        # the check above, and both create the entire register.
        cur = c.execute(
            "UPDATE enterprise_import_batches "
            "   SET status='Committed', committed_at=CURRENT_TIMESTAMP "
            " WHERE tenant_id=? AND id=? AND status='Staged'",
            (tenant_id, batch_id),
        )
        if getattr(cur, "rowcount", -1) == 0:
            raise ImportError_(
                "IMPORT", "this import was already committed by somebody else"
            )

        # THE LOOP CALLS write_beneficiary_row, NOT create_beneficiary, AND WRAPS EACH ROW IN
        # ITS OWN SAVEPOINT. Both halves are load-bearing; the second was learned the hard
        # way (Supervisor slice-6.6, HIGH -- the first version of this fix INTRODUCED a bug).
        #
        # WHY NOT create_beneficiary: it re-checked the programme and the permission on every
        # row, and wrote an audit row per row. C13 and `beneficiary.import` are proven ONCE,
        # above, for this programme -- which is the whole batch -- so the re-checks were pure
        # cost. The per-row audit was worse than cost: each audit write takes
        # `pg_advisory_xact_lock` on a CONSTANT key shared by every audit writer in SolarPro.
        # Now the audit row is written ONCE, at the end, which is also what audit.py's caller
        # rule demands ("write the audit row LAST").
        #
        # WHY THE SAVEPOINT: on Postgres a failed statement poisons the whole transaction --
        # psycopg2 puts it in InFailedSqlTransaction and REFUSES every subsequent statement
        # until a rollback. write_beneficiary_row's INSERT is bare, so a duplicate site code
        # aborted the transaction, and the very next line -- the UPDATE that marks the row as
        # an Error -- then raised InFailedSqlTransaction. That is not an EnterpriseGateError,
        # so it escaped this except, escaped the loop, and 500'd the request with the entire
        # batch rolled back: the exact opposite of "one bad row must not cost the other 1999".
        # It was deterministic, not a race -- ticking "import duplicates too" makes row 1 a
        # guaranteed collision -- and NO SQLite test could see it, because SQLite does not
        # poison a transaction on IntegrityError.
        #
        # The savepoint heals it: ROLLBACK TO SAVEPOINT clears the aborted state, and the
        # UPDATE below then runs on a healthy transaction. (txn.atomic takes the savepoint
        # branch here because the batch's own transaction is already open.)
        for row_id, row_no, mapped_json in rows:
            mapped = _decode(mapped_json, {})
            fields = {k: v for k, v in mapped.items() if k in _FIELD_KEYS}
            try:
                with txn.atomic(c):
                    beneficiary_id, _code = beneficiaries.write_beneficiary_row(
                        c, tenant_id, user_id, batch["programme_id"],
                        code=mapped.get("code", ""),
                        name=mapped.get("name", ""),
                        beneficiary_type=mapped.get("beneficiary_type", ""),
                        fields=fields,
                        import_batch_id=batch_id,
                    )
            except EnterpriseGateError as e:
                # A row that the database refused (a code that raced another import, say),
                # or one the register would not validate. Record WHY on the row and carry
                # on: one bad row must not cost the other 1999. It stays in staging as an
                # Error, with the reason, forever.
                #
                # C12 can no longer surface HERE (the audit write moved out of the loop), so
                # the old "C12 is not a row error, re-raise it" special case is gone with the
                # per-row audit that made it necessary. C12 is still enforced -- once, below,
                # for the batch -- and it still aborts the entire import, which is what
                # audit-or-nothing means.
                failed += 1
                c.execute(
                    "UPDATE enterprise_import_rows "
                    "   SET status='Error', errors_json=? "
                    " WHERE tenant_id=? AND id=?",
                    (json.dumps([str(e)]), tenant_id, row_id),
                )
                continue

            imported += 1
            c.execute(
                "UPDATE enterprise_import_rows "
                "   SET status='Imported', beneficiary_id=? "
                " WHERE tenant_id=? AND id=?",
                (beneficiary_id, tenant_id, row_id),
            )

        # COUNTED, NOT INFERRED (Supervisor slice-5, MED). Adding `failed` to error_rows and
        # leaving valid_rows alone made the four numbers stop summing to total_rows: a Valid
        # row that lost a race on the code index stayed counted as Valid AND was counted as an
        # Error. The rows know their own statuses -- ask them.
        tally = dict.fromkeys(("Valid", "Error", "Duplicate", "Imported"), 0)
        for status, n in c.execute(
            "SELECT status, COUNT(*) FROM enterprise_import_rows "
            " WHERE tenant_id=? AND batch_id=? GROUP BY status",
            (tenant_id, batch_id),
        ).fetchall():
            tally[status] = n
        c.execute(
            "UPDATE enterprise_import_batches "
            "   SET imported_rows=?, valid_rows=?, error_rows=?, duplicate_rows=? "
            " WHERE tenant_id=? AND id=?",
            (tally["Imported"], tally["Valid"], tally["Error"], tally["Duplicate"],
             tenant_id, batch_id),
        )

        _require_audit(
            audit("ENTERPRISE_IMPORT_COMMITTED", user_id=user_id, tenant_id=tenant_id,
                  details={"batch_id": batch_id, "programme_id": batch["programme_id"],
                           "imported": imported, "failed": failed,
                           "included_duplicates": include_duplicates}),
            "import commit",
        )

    # `skipped` is the rows the operator DECLINED (the duplicates they did not include), not
    # a catch-all for everything that did not import. Rows that failed validation are
    # `errors`; rows the database refused are `failed`. Reporting "5 skipped" for five rows
    # that were actually broken tells the operator to look in the wrong place.
    return {
        "imported": imported,
        "failed": failed,
        "errors": tally["Error"] - failed,
        "skipped": tally["Duplicate"],
    }


def restage_batch(c, tenant_id: str, user_id: int, batch_id: int, *,
                  mapping: dict[str, str], default_type: str | None = None,
                  audit=None) -> dict:
    """Re-run a STAGED batch through mapping and validation with a corrected mapping.

    Input:  connection, tenant id, acting user, batch id, the new column mapping, the
            fallback beneficiary type (None = keep the one chosen at upload; "" = clear it),
            the audit hook.
    Output: {"Valid": n, "Error": n, "Duplicate": n}.
    Raises: EnterprisePermissionError (403), ImportError_ (409 / C13).

    THIS IS WHY THE RAW ROW IS KEPT. The operator uploads once and then argues with the
    mapping -- "Location" was the community, not the address; that column of numbers is the
    monthly kWh, not the tariff. Re-staging replays the ORIGINAL file bytes, as parsed, back
    through the mapper with the new choices. Without the raw row they would have to re-upload
    a file they may no longer have to hand, and the previous attempt's verdict on each row
    would be lost.

    Only a Staged batch. Once committed, the rows are the provenance of real beneficiaries
    and re-writing them would make the register's history a fiction.
    """
    batch = _load_batch(c, tenant_id, batch_id)                 # C13 FIRST
    rbac.require_permission(c, tenant_id, user_id, "beneficiary.import",
                            programme_id=batch["programme_id"])

    if batch["status"] != "Staged":
        raise ImportError_(
            "IMPORT",
            f"this import is {batch['status']}; its mapping can no longer be changed",
        )

    _check_mapping(mapping)

    # THE UPLOAD-TIME DEFAULT SURVIVES A RE-MAP (Codex slice-5 round 2, MED). The default type
    # is what makes a spreadsheet with no "type" column importable at all. It was passed at
    # upload and never stored, so correcting ONE column silently dropped it -- and every row
    # that relied on it failed with "beneficiary type is required", an error about a thing the
    # operator never touched. None means "keep what was chosen"; "" means "clear it".
    effective_default = (
        batch["default_type"] if default_type is None else str(default_type).strip()
    )

    rows = c.execute(
        "SELECT id, row_no, raw_json FROM enterprise_import_rows "
        " WHERE tenant_id=? AND batch_id=? ORDER BY row_no",
        (tenant_id, batch_id),
    ).fetchall()

    counts = {"Valid": 0, "Error": 0, "Duplicate": 0}
    audit = audit or txn.audit_on(c)
    index = beneficiaries.DuplicateIndex.for_programme(c, tenant_id, batch["programme_id"])

    with txn.atomic(c):
        for row_id, row_no, raw_json in rows:
            staged = _stage_row(c, tenant_id, batch["programme_id"], row_no,
                                _decode(raw_json, {}), mapping, effective_default, index)
            counts[staged["status"]] = counts.get(staged["status"], 0) + 1
            c.execute(
                "UPDATE enterprise_import_rows "
                "   SET mapped_json=?, status=?, errors_json=? "
                " WHERE tenant_id=? AND id=?",
                (json.dumps(staged["mapped"]), staged["status"],
                 json.dumps(staged["errors"]), tenant_id, row_id),
            )

        cur = c.execute(
            "UPDATE enterprise_import_batches "
            "   SET column_mapping=?, default_type=?, valid_rows=?, error_rows=?, "
            "       duplicate_rows=? "
            " WHERE tenant_id=? AND id=? AND status='Staged'",
            (json.dumps(mapping or {}), effective_default, counts["Valid"], counts["Error"],
             counts["Duplicate"], tenant_id, batch_id),
        )
        if getattr(cur, "rowcount", -1) == 0:
            raise ImportError_(
                "IMPORT", "this import is no longer staged; its mapping cannot be changed"
            )

        _require_audit(
            audit("ENTERPRISE_IMPORT_REMAPPED", user_id=user_id, tenant_id=tenant_id,
                  details={"batch_id": batch_id, "programme_id": batch["programme_id"],
                           **counts}),
            "import remap",
        )
    return counts


def cancel_batch(c, tenant_id: str, user_id: int, batch_id: int, *, audit=None) -> None:
    """Throw a staged batch away. The rows stay: what was rejected, and why, is evidence."""
    batch = _load_batch(c, tenant_id, batch_id)                 # C13 first
    rbac.require_permission(c, tenant_id, user_id, "beneficiary.import",
                            programme_id=batch["programme_id"])

    if batch["status"] != "Staged":
        raise ImportError_(
            "IMPORT", f"this import is {batch['status']}; it cannot be cancelled"
        )

    audit = audit or txn.audit_on(c)
    with txn.atomic(c):
        cur = c.execute(
            "UPDATE enterprise_import_batches SET status='Cancelled' "
            " WHERE tenant_id=? AND id=? AND status='Staged'",
            (tenant_id, batch_id),
        )
        if getattr(cur, "rowcount", -1) == 0:
            raise ImportError_("IMPORT", "this import is no longer staged")
        _require_audit(
            audit("ENTERPRISE_IMPORT_CANCELLED", user_id=user_id, tenant_id=tenant_id,
                  details={"batch_id": batch_id,
                           "programme_id": batch["programme_id"]}),
            "import cancel",
        )


def _require_audit(written: bool, what: str) -> None:
    """C12 for this module."""
    from . import gates as gates_mod
    gates_mod.require_audit_written(written, what)
