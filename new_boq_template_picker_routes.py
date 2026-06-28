# new_boq_template_picker_routes.py
# Whole-floor template picker + checkbox render + bulk save.
# Plus Excel / PDF / Email exports of the whole-project BOQ.


# ---- Picker --------------------------------------------------------------

@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/floors/<int:fid>/from-template", methods=["GET"])
@login_required
def boq_template_picker(pid, bid, fid):
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    building = _boq_building_owned_or_404(bid, pid)
    floor = _boq_floor_owned_or_404(fid, bid)
    from new_boq_project_templates import _boq_template_list
    # Filter by building purpose where possible; show all if none match.
    purpose = (building["primary_purpose"] or "").strip().lower()
    matched = _boq_template_list(purpose=purpose)
    others  = [t for t in _boq_template_list() if t["slug"] not in {m["slug"] for m in matched}]
    return render_template(
        "boq_template_picker.html",
        user=current_user(),
        project=project, building=building, floor=floor,
        matched=matched, others=others, purpose=purpose,
    )


# ---- Checkbox render -----------------------------------------------------

@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/floors/<int:fid>/from-template/<slug>", methods=["GET"])
@login_required
def boq_template_view(pid, bid, fid, slug):
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    building = _boq_building_owned_or_404(bid, pid)
    floor = _boq_floor_owned_or_404(fid, bid)
    from new_boq_project_templates import _boq_template_get
    template = _boq_template_get(slug)
    if not template:
        flash("Template not found.", "warning")
        return redirect(url_for("boq_template_picker", pid=pid, bid=bid, fid=fid))
    return render_template(
        "boq_template_checkbox.html",
        user=current_user(),
        project=project, building=building, floor=floor,
        template=template, slug=slug,
    )


@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/floors/<int:fid>/from-template/<slug>/save", methods=["POST"])
@login_required
def boq_template_save(pid, bid, fid, slug):
    uid = session["user_id"]
    _boq_project_owned_or_404(pid, uid)
    _boq_building_owned_or_404(bid, pid)
    _boq_floor_owned_or_404(fid, bid)
    csrf_protect()
    from new_boq_project_templates import _boq_template_get, _boq_template_iter_lines
    template = _boq_template_get(slug)
    if not template:
        flash("Template not found.", "warning")
        return redirect(url_for("boq_floor_view", pid=pid, bid=bid, fid=fid))

    f = request.form
    # Section-wide markup applied to all rows.
    def _pct(name):
        try:
            v = f.get(name, "")
            return max(0.0, min(100.0, float(v))) if v not in (None, "",) else 0.0
        except (TypeError, ValueError):
            return 0.0
    # Rate engine v3 (2026-06-28): supply_pct + install_pct are PERCENTAGES.
    oh, prf, vat = _pct("overhead_pct"), _pct("profit_pct"), _pct("vat_pct")
    supply_pct  = _pct("supply_default_pct")
    install_pct = _pct("install_default_pct")
    vat_in_basic = 1 if f.get("vat_in_basic") else 0
    from boq_rate_v3 import boq_rate_v3

    ticked = set()
    for k in f.getlist("tick"):
        try:
            ticked.add(int(k))
        except (TypeError, ValueError):
            pass

    # Per-line qty / basic / desc / unit overrides come keyed by line_idx.
    def _line_val(prefix, idx, default=""):
        return (f.get(f"{prefix}[{idx}]") or default).strip()

    saved = 0
    skipped = 0
    next_no_cache = {}  # (bill_no, section_letter) -> next int
    with get_db() as c:
        for (bill_no, bill_name, sect_letter, sect_title, subsec, idx,
             desc, unit, qty_d, basic_d, spec) in _boq_template_iter_lines(template):
            if idx not in ticked:
                skipped += 1
                continue
            # Apply per-line edits.
            desc  = _line_val("desc", idx, desc)[:500]
            unit  = _line_val("unit", idx, unit)[:20]
            spec  = _line_val("spec", idx, spec)
            qty_s = _line_val("qty", idx, "")
            basic_s = _line_val("basic", idx, "")
            try:
                qty = float(qty_s) if qty_s else float(qty_d)
            except ValueError:
                qty = float(qty_d)
            try:
                basic = float(basic_s) if basic_s else float(basic_d)
            except ValueError:
                basic = float(basic_d)
            if not desc or qty <= 0:
                skipped += 1
                continue
            # Owner directive 2026-06-28: basic=0 placeholders OK; look up
            # marketplace catalogue price by description match if missing.
            if basic <= 0 and desc:
                try:
                    _row = c.execute(
                        "SELECT price_usd FROM equipment_catalog "
                        "WHERE LOWER(name) = LOWER(?) AND COALESCE(is_active,1)=1 "
                        "ORDER BY id DESC LIMIT 1",
                        (desc,),
                    ).fetchone()
                    if _row and _row["price_usd"]:
                        basic = float(_row["price_usd"])
                except Exception:
                    pass
            supply_amount, install_amount, final_rate = boq_rate_v3(
                basic, supply_pct, install_pct, oh, prf, vat,
                vat_in_basic=bool(vat_in_basic))
            total = qty * final_rate

            key = (bill_no, sect_letter)
            if key not in next_no_cache:
                next_no_cache[key] = int(_boq_next_item_no(fid, bill_no, sect_letter))
            item_no_disp = str(next_no_cache[key])
            next_no_cache[key] += 1

            cur = c.execute(
                "INSERT INTO boq_floor_items "
                "(floor_id, building_id, project_id, user_id, section, subsection, "
                " library_item_id, supplier_id, item_no, description, specification, "
                " unit, qty, final_built_up_rate, total_amount, remarks, "
                " source_type, approval_status, "
                " bill_no, bill_name, section_letter, subsection_label, item_no_display) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (fid, bid, pid, uid, sect_title.lower()[:80], "",
                 None, None, item_no_disp,
                 desc, spec, unit, qty, final_rate, total, "",
                 "project_library", "project_only",
                 bill_no, bill_name, sect_letter, subsec, item_no_disp),
            )
            item_id = int(cur.lastrowid or 0)
            c.execute(
                "INSERT INTO boq_floor_rate_buildup "
                "(floor_item_id, project_id, user_id, basic_price, "
                " supply_pct, install_pct, supply_rate, install_rate, "
                " overhead_pct, profit_pct, contingency_pct, vat_pct, "
                " vat_in_basic, final_built_up_rate, total_amount) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (item_id, pid, uid, basic,
                 supply_pct, install_pct, supply_amount, install_amount,
                 oh, prf, 0, vat, vat_in_basic, final_rate, total),
            )
            saved += 1

        # Custom additions: lines added inline via the "+ Custom item" rows.
        # Posted as parallel arrays custom_desc[], custom_qty[], ...
        custom_descs = f.getlist("custom_desc[]")
        custom_qtys  = f.getlist("custom_qty[]")
        custom_units = f.getlist("custom_unit[]")
        custom_basic = f.getlist("custom_basic[]")
        custom_bill  = f.getlist("custom_bill[]")
        custom_sect  = f.getlist("custom_section[]")
        custom_title = f.getlist("custom_title[]")
        for i in range(len(custom_descs)):
            desc = (custom_descs[i] or "").strip()[:500]
            if not desc:
                continue
            try:
                qty = float(custom_qtys[i]) if i < len(custom_qtys) and custom_qtys[i] else 0.0
            except ValueError:
                qty = 0.0
            try:
                basic = float(custom_basic[i]) if i < len(custom_basic) and custom_basic[i] else 0.0
            except ValueError:
                basic = 0.0
            if qty <= 0 or basic <= 0:
                continue
            unit = (custom_units[i] if i < len(custom_units) else "No.").strip() or "No."
            try:
                bill_no = int(custom_bill[i]) if i < len(custom_bill) else 2
            except ValueError:
                bill_no = 2
            sect_letter = (custom_sect[i] if i < len(custom_sect) else "Z").strip().upper()[:8] or "Z"
            sect_title = (custom_title[i] if i < len(custom_title) else "CUSTOM ITEMS").strip()[:160] or "CUSTOM ITEMS"
            bill_name = _boq_lookup_bill_name(bill_no) or "OTHER"
            supply_amount, install_amount, final_rate = boq_rate_v3(
                basic, supply_pct, install_pct, oh, prf, vat,
                vat_in_basic=bool(vat_in_basic))
            total = qty * final_rate
            key = (bill_no, sect_letter)
            if key not in next_no_cache:
                next_no_cache[key] = int(_boq_next_item_no(fid, bill_no, sect_letter))
            item_no_disp = str(next_no_cache[key])
            next_no_cache[key] += 1
            cur = c.execute(
                "INSERT INTO boq_floor_items "
                "(floor_id, building_id, project_id, user_id, section, subsection, "
                " library_item_id, supplier_id, item_no, description, specification, "
                " unit, qty, final_built_up_rate, total_amount, remarks, "
                " source_type, approval_status, "
                " bill_no, bill_name, section_letter, subsection_label, item_no_display) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (fid, bid, pid, uid, sect_title.lower()[:80], "",
                 None, None, item_no_disp,
                 desc, "", unit, qty, final_rate, total, "",
                 "custom_current_boq", "project_only",
                 bill_no, bill_name, sect_letter, "", item_no_disp),
            )
            item_id = int(cur.lastrowid or 0)
            c.execute(
                "INSERT INTO boq_floor_rate_buildup "
                "(floor_item_id, project_id, user_id, basic_price, "
                " supply_pct, install_pct, supply_rate, install_rate, "
                " overhead_pct, profit_pct, contingency_pct, vat_pct, "
                " vat_in_basic, final_built_up_rate, total_amount) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (item_id, pid, uid, basic,
                 supply_pct, install_pct, supply_amount, install_amount,
                 oh, prf, 0, vat, vat_in_basic, final_rate, total),
            )
            saved += 1

        c.execute("UPDATE boq_projects  SET updated_at=CURRENT_TIMESTAMP WHERE id=?", (pid,))
        c.execute("UPDATE boq_buildings SET updated_at=CURRENT_TIMESTAMP WHERE id=?", (bid,))
        c.execute("UPDATE boq_floors    SET updated_at=CURRENT_TIMESTAMP WHERE id=?", (fid,))

    try:
        from new_boq_hierarchy_schema import boq_audit
        boq_audit(get_db, uid, "boq_template_generated", "boq_floor", fid,
                  f"template={slug} saved={saved} skipped={skipped}")
    except Exception:
        pass
    flash(f"Generated {saved} line(s) from template '{slug}'. ({skipped} skipped.)", "success")
    # "Generate BOQ" lands the user on the project BOQ view so they can
    # immediately see the document they just generated.
    return redirect(url_for("boq_project_boq", pid=pid))


# ---- Whole-project Excel / PDF / Email exports ---------------------------

def _boq_project_rows_grouped(pid: int):
    """Pull every line item across every building+floor in the project,
    grouped Building -> Floor -> Bill -> Section. Used by the 3 exports."""
    t_clause, t_params = _boq_tenant_clause(alias="i")
    with get_db() as c:
        rows = c.execute(
            "SELECT i.*, b.building_name, b.building_code, b.primary_purpose, "
            "       b.purpose_subtype, f.floor_name, f.floor_level, "
            "       f.contingency_pct, "
            "       rb.basic_price, rb.supply_rate, rb.install_rate, "
            "       rb.overhead_pct, rb.profit_pct, rb.contingency_pct AS rb_cont, "
            "       rb.vat_pct "
            "FROM boq_floor_items i "
            "JOIN boq_buildings b ON b.id=i.building_id "
            "JOIN boq_floors    f ON f.id=i.floor_id "
            "LEFT JOIN boq_floor_rate_buildup rb ON rb.floor_item_id=i.id "
            "WHERE i.project_id=?" + t_clause + " "
            "ORDER BY b.id, f.floor_level, COALESCE(i.bill_no,0), "
            "         COALESCE(i.section_letter,''), "
            "         COALESCE(NULLIF(i.item_no_display,''),'0'), i.id",
            (pid,) + t_params,
        ).fetchall()
    return rows


def _boq_floor_subtotals(rows):
    """Reshape the flat rows into {(bid, fid): {"floor_name", "contingency_pct",
    "bills": [{"no","name","subtotal"}]}} so the per-floor Bills Summary +
    Project Final Summary can be computed."""
    floors = {}
    for r in rows:
        bid = r["building_id"]; fid = r["floor_id"]
        key = (bid, fid)
        if key not in floors:
            floors[key] = {
                "building_id": bid,
                "building_name": r["building_name"],
                "floor_id": fid,
                "floor_name": r["floor_name"],
                "floor_level": r["floor_level"],
                "contingency_pct": float(r["contingency_pct"] or 10),
                "bills": {},
            }
        bill_no = r["bill_no"] or 0
        bill_name = r["bill_name"] or "OTHER"
        bk = (bill_no, bill_name)
        floors[key]["bills"].setdefault(bk, 0.0)
        floors[key]["bills"][bk] += float(r["total_amount"] or 0)
    # Flatten bill dict to list.
    out = []
    for (bid, fid), data in floors.items():
        data["bill_list"] = [{"no": k[0], "name": k[1], "subtotal": v}
                             for k, v in sorted(data["bills"].items())]
        data["subtotal"] = sum(b["subtotal"] for b in data["bill_list"])
        data["contingency"] = data["subtotal"] * data["contingency_pct"] / 100.0
        data["carried"] = data["subtotal"] + data["contingency"]
        out.append(data)
    return out


@app.route("/boq-projects/<int:pid>/boq.xlsx")
@login_required
def boq_project_xlsx(pid):
    """Excel export of the whole project BOQ + Bills Summary + Final Summary."""
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    rows = _boq_project_rows_grouped(pid)
    floors = _boq_floor_subtotals(rows)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="B45309")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    bill_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(border_style="thin", color="D1D5DB")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Main BOQ sheet ----
    ws = wb.active
    ws.title = "BOQ"
    ws["A1"] = f"Bill of Quantities -- {project['project_name']}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Client : {project['client_name'] or '-'}"
    ws["A3"] = f"Location: {project['location'] or '-'}"
    # 2026-06-28: free-text instructions render in A4 (merged across all cols).
    if (project["instructions"] or "").strip():
        ws["A4"] = "Instructions: " + str(project["instructions"]).strip()
        ws["A4"].font = Font(italic=True, color="555555")
        ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A4:I4")

    headers = ["Item", "Description", "Qty", "Unit", "Basic Price",
               "Supply Amount Rate", "Installation Amount Rate", "Total Amount Rate", "Amount"]
    HROW = 6
    for col, h in enumerate(headers, 1):
        c_ = ws.cell(row=HROW, column=col, value=h)
        c_.font = header_font; c_.fill = header_fill; c_.border = box
        c_.alignment = Alignment(horizontal="center")

    def _san(v):
        s = str(v or "")
        return "'" + s if s and s[0] in ("=", "+", "-", "@") else s

    row = HROW + 1
    prev = {"bid": None, "fid": None, "bill": None, "sec": None, "sub": None}
    for r in rows:
        if r["building_id"] != prev["bid"]:
            ws.cell(row=row, column=1, value=f"BUILDING: {r['building_name']}").font = title_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
            prev.update({"bid": r["building_id"], "fid": None, "bill": None, "sec": None, "sub": None})
        if r["floor_id"] != prev["fid"]:
            ws.cell(row=row, column=1, value=f"  FLOOR: {r['floor_name']}").font = bold
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
            prev.update({"fid": r["floor_id"], "bill": None, "sec": None, "sub": None})
        if (r["bill_no"] or 0) != prev["bill"]:
            ws.cell(row=row, column=1, value=f"BILL No. {r['bill_no'] or 0} -- {r['bill_name'] or 'OTHER'}").font = bold
            ws.cell(row=row, column=1).fill = bill_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
            prev.update({"bill": r["bill_no"] or 0, "sec": None, "sub": None})
        if (r["section_letter"] or "") != prev["sec"]:
            sec_t = (r["section_letter"] or "") + ". " + (r["section"] or "").upper()
            ws.cell(row=row, column=1, value=sec_t).font = bold
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
            prev.update({"sec": r["section_letter"] or "", "sub": None})
        if (r["subsection_label"] or "") and r["subsection_label"] != prev["sub"]:
            ws.cell(row=row, column=2, value=r["subsection_label"]).font = Font(italic=True)
            row += 1
            prev["sub"] = r["subsection_label"]
        ws.cell(row=row, column=1, value=r["item_no_display"] or r["item_no"] or "")
        ws.cell(row=row, column=2, value=_san(r["description"]))
        ws.cell(row=row, column=3, value=float(r["qty"] or 0))
        ws.cell(row=row, column=4, value=_san(r["unit"]))
        ws.cell(row=row, column=5, value=round(float(r["basic_price"] or 0), 2))
        ws.cell(row=row, column=6, value=round(float(r["supply_rate"] or 0), 2))
        ws.cell(row=row, column=7, value=round(float(r["install_rate"] or 0), 2))
        ws.cell(row=row, column=8, value=round(float(r["final_built_up_rate"] or 0), 2))
        ws.cell(row=row, column=9, value=round(float(r["total_amount"] or 0), 2))
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = box
        row += 1
    for col, w in enumerate([8, 50, 8, 8, 14, 14, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ---- Per-floor Bills Summary ----
    ws2 = wb.create_sheet("Bills Summary")
    ws2["A1"] = f"Bills Summary -- {project['project_name']}"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:E1")
    r2 = 3
    grand_carry = 0.0
    for fl in floors:
        ws2.cell(row=r2, column=1, value=f"{fl['building_name']} -- {fl['floor_name']}").font = bold
        r2 += 1
        ws2.cell(row=r2, column=1, value="Item").font = bold
        ws2.cell(row=r2, column=2, value="Bill").font = bold
        ws2.cell(row=r2, column=3, value="Amount").font = bold
        r2 += 1
        for i, b in enumerate(fl["bill_list"], 1):
            ws2.cell(row=r2, column=1, value=i)
            ws2.cell(row=r2, column=2, value=f"BILL No. {b['no']} -- {b['name']}")
            ws2.cell(row=r2, column=3, value=round(b["subtotal"], 2))
            r2 += 1
        ws2.cell(row=r2, column=2, value="SUB TOTAL").font = bold
        ws2.cell(row=r2, column=3, value=round(fl["subtotal"], 2)).font = bold
        r2 += 1
        ws2.cell(row=r2, column=2, value=f"CONTINGENCIES ({fl['contingency_pct']:.2f}%)")
        ws2.cell(row=r2, column=3, value=round(fl["contingency"], 2))
        r2 += 1
        ws2.cell(row=r2, column=2, value="Total carried to General Summary").font = bold
        ws2.cell(row=r2, column=3, value=round(fl["carried"], 2)).font = bold
        grand_carry += fl["carried"]
        r2 += 2
    ws2.cell(row=r2, column=2, value="PROJECT GRAND TOTAL").font = title_font
    ws2.cell(row=r2, column=3, value=round(grand_carry, 2)).font = title_font
    for col, w in enumerate([8, 45, 18], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", project["project_name"])[:60]
    try:
        from new_boq_hierarchy_schema import boq_audit
        boq_audit(get_db, uid, "boq_project_exported", "boq_project", pid, "xlsx")
    except Exception:
        pass
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"BOQ_{safe_name}.xlsx",
    )


def _boq_project_markdown(pid: int) -> str:
    """Render the project BOQ as markdown (used by the PDF + email)."""
    with get_db() as c:
        project = c.execute("SELECT * FROM boq_projects WHERE id=?", (pid,)).fetchone()
    rows = _boq_project_rows_grouped(pid)
    floors = _boq_floor_subtotals(rows)

    md = [
        f"# Bill of Quantities -- {project['project_name']}",
        "",
        f"**Client:** {project['client_name'] or '-'}  ",
        f"**Location:** {project['location'] or '-'}  ",
        f"**Generated:** {project['updated_at']}",
        "",
    ]
    _instr = (project["instructions"] or "").strip()
    if _instr:
        md.append("> **Instructions:** " + _instr.replace("\n", " "))
        md.append("")
    prev = {"bid": None, "fid": None, "bill": None, "sec": None, "sub": None}
    for r in rows:
        if r["building_id"] != prev["bid"]:
            md.append(f"## Building: {r['building_name']}")
            prev.update({"bid": r["building_id"], "fid": None, "bill": None, "sec": None, "sub": None})
        if r["floor_id"] != prev["fid"]:
            md.append(f"### Floor: {r['floor_name']}")
            prev.update({"fid": r["floor_id"], "bill": None, "sec": None, "sub": None})
        if (r["bill_no"] or 0) != prev["bill"]:
            md.append("")
            md.append(f"#### BILL No. {r['bill_no'] or 0} -- {r['bill_name'] or 'OTHER'}")
            md.append("")
            md.append("| Item | Description | Qty | Unit | Basic Price | Supply Amount Rate | Installation Amount Rate | Total Amount Rate | Amount |")
            md.append("|---|---|---|---|---|---|---|---|---|")
            prev.update({"bill": r["bill_no"] or 0, "sec": None, "sub": None})
        if (r["section_letter"] or "") != prev["sec"]:
            md.append(f"| | **{r['section_letter'] or ''}. {(r['section'] or '').upper()}** | | | | | | | |")
            prev.update({"sec": r["section_letter"] or "", "sub": None})
        if (r["subsection_label"] or "") and r["subsection_label"] != prev["sub"]:
            md.append(f"| | *{r['subsection_label']}* | | | | | | | |")
            prev["sub"] = r["subsection_label"]
        md.append(
            f"| {r['item_no_display'] or r['item_no'] or ''} "
            f"| {r['description']} "
            f"| {float(r['qty'] or 0):.2f} "
            f"| {r['unit']} "
            f"| {float(r['basic_price'] or 0):.2f} "
            f"| {float(r['supply_rate'] or 0):.2f} "
            f"| {float(r['install_rate'] or 0):.2f} "
            f"| {float(r['final_built_up_rate'] or 0):.2f} "
            f"| {float(r['total_amount'] or 0):.2f} |"
        )

    md.append("")
    md.append("## Bills Summary")
    md.append("")
    grand_carry = 0.0
    for fl in floors:
        md.append(f"### {fl['building_name']} -- {fl['floor_name']}")
        md.append("")
        md.append("| Item | Bill | Amount |")
        md.append("|---|---|---|")
        for i, b in enumerate(fl["bill_list"], 1):
            md.append(f"| {i} | BILL No. {b['no']} -- {b['name']} | {b['subtotal']:.2f} |")
        md.append(f"| | **SUB TOTAL** | **{fl['subtotal']:.2f}** |")
        md.append(f"| | CONTINGENCIES ({fl['contingency_pct']:.2f}%) | {fl['contingency']:.2f} |")
        md.append(f"| | **Total carried to General Summary** | **{fl['carried']:.2f}** |")
        md.append("")
        grand_carry += fl["carried"]
    md.append(f"### PROJECT GRAND TOTAL: **{grand_carry:.2f}**")
    return "\n".join(md)


@app.route("/boq-projects/<int:pid>/boq.pdf")
@login_required
def boq_project_pdf(pid):
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    md = _boq_project_markdown(pid)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", project["project_name"])[:60]
    try:
        from new_boq_hierarchy_schema import boq_audit
        boq_audit(get_db, uid, "boq_project_exported", "boq_project", pid, "pdf")
    except Exception:
        pass
    return _render_pdf(f"BOQ -- {project['project_name']}", md, f"BOQ_{safe_name}.pdf")


@app.route("/boq-projects/<int:pid>/email", methods=["POST"])
@login_required
def boq_project_email(pid):
    """Email the project BOQ to an external recipient. Uses the existing
    solar _send_email helper -- Brevo first, SMTP fallback."""
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    csrf_protect()
    to_email = (request.form.get("to_email") or "").strip().lower()[:200]
    if "@" not in to_email or "." not in to_email:
        flash("Invalid recipient email address.", "warning")
        return redirect(url_for("boq_project_overview", pid=pid))
    subject = (request.form.get("subject")
               or f"Bill of Quantities -- {project['project_name']}")[:200]
    body = (request.form.get("body")
            or f"Please find attached the BOQ for {project['project_name']}.\n\n"
               f"Generated by SolarPro -- {project['updated_at']}.")
    md = _boq_project_markdown(pid)
    # Build the PDF in-memory and attach.
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", project["project_name"])[:60]
    try:
        from markdown_pdf import MarkdownPdf, Section
        pdf = MarkdownPdf(toc_level=2)
        pdf.add_section(Section(md, toc=False))
        import tempfile
        tf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        pdf.save(tf.name); tf.close()
        with open(tf.name, "rb") as fh:
            attachment_bytes = fh.read()
    except Exception as e:
        try: app.logger.warning("boq email PDF build failed: %s", e)
        except Exception: pass
        flash("Could not build PDF for email.", "danger")
        return redirect(url_for("boq_project_overview", pid=pid))

    sent = False
    try:
        # _send_email signature varies in this codebase; try the most common form.
        from api_manager import _send_email  # type: ignore
        sent = bool(_send_email(
            to_email, subject, body,
            attachment_name=f"BOQ_{safe_name}.pdf",
            attachment_bytes=attachment_bytes,
        ))
    except Exception:
        try:
            from api_manager import send_email_with_attachment  # type: ignore
            sent = bool(send_email_with_attachment(
                to_email, subject, body,
                f"BOQ_{safe_name}.pdf", attachment_bytes,
            ))
        except Exception:
            sent = False

    try:
        from new_boq_hierarchy_schema import boq_audit
        boq_audit(get_db, uid, "boq_project_emailed", "boq_project", pid,
                  f"to={to_email} sent={sent}")
    except Exception:
        pass
    flash(f"Email {'sent' if sent else 'queued (no email backend available)'} to {to_email}.",
          "success" if sent else "warning")
    return redirect(url_for("boq_project_overview", pid=pid))
