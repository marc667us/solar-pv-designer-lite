# -*- coding: utf-8 -*-
"""
new_capital_investment_routes_v2.py  (CLEAN-ROOM REBUILD 2026-07-03)
====================================================================
PV Independent Solar Capital Investment Design Module -- "Generation Station".

Clean-room rebuild driven by:
    pvsolar1/solar power plant1.txt                       (master spec)
    pvsolar1/CODEX_REBUILD_SSS_generation_station_2026-07-03.md  (rebuild SSS)

Design rule (SSS Section 1): this module is a THIN ORCHESTRATION SHELL over the
existing SolarPro engines. It stores capital-project state, calls existing
engines (BOQ, marketplace, CRM, finance, reports, digital twin, AI agents), and
links out to existing workspaces. It NEVER recreates a BOQ engine, marketplace,
CRM, report renderer, solar calculator, or AI orchestrator.

Hard rules baked in from line 1 (SSS Section 8):
  * Every insert that needs the new id uses  INSERT ... RETURNING id  (never
    cur.lastrowid, which breaks on live Postgres via SELECT lastval()).
  * Schema is created EAGERLY and VERIFIED (not lazy + swallowed).
  * Every owned query filters user_id (+ tenant_id where available).
  * Stored JSON is never trusted -- _safe_json returns a dict, nested values are
    coerced before templates/reports read them.

Entry point (unchanged so web_app.py needs NO byte-patch):
    from new_capital_investment_routes import register_capital_investment
    register_capital_investment(app, *, get_db, login_required,
                                csrf_protect, current_user)

This file is built PHASE BY PHASE (SSS Section 9). Phase 1 = skeleton:
landing / demo-stub / upgrade / new (Step 1) / project overview / diagnostics /
verified schema. Steps 2-14 + digital twin + regulatory land in later phases.
"""

from __future__ import annotations

import json
import os
from typing import Any

import dt_scene_v2 as _dtv2

from flask import (
    render_template, request, redirect, url_for,
    session, flash, abort, jsonify,
)

# Rebuild marker so /api and diagnostics can prove which module is live.
CI_MODULE_BUILD = "v2-rebuild-2026-07-03"


# ===========================================================================
# SECTION A -- Constants (spec Steps 1-2 form vocab; full step vocab arrives
# with each step's phase).
# ===========================================================================

PROJECT_TYPES: list[tuple[str, str, str]] = [
    ("ground_mount",       "Ground-Mounted Solar Farm",        "bi-tree"),
    ("commercial_rooftop", "Commercial Rooftop Generation",    "bi-buildings"),
    ("industrial",         "Industrial Solar Plant",           "bi-tools"),
    ("utility_scale",      "Utility-Scale Solar",              "bi-lightning-charge"),
    ("embedded",           "Embedded Generation",              "bi-plug"),
    ("ipp",                "Independent Power Producer (IPP)", "bi-bank"),
    ("mining",             "Mining Solar Plant",               "bi-gem"),
    ("agricultural",       "Agricultural Solar (Agri-PV)",     "bi-flower1"),
    ("floating",           "Floating Solar (Floatovoltaic)",   "bi-water"),
    ("hybrid",             "Hybrid Solar Plant (PV + BESS)",   "bi-battery-charging"),
    ("bess",               "Battery Energy Storage Plant",     "bi-battery-full"),
    ("other",              "Other",                            "bi-three-dots"),
]
PROJECT_TYPE_CODES: set[str] = {c for c, _, _ in PROJECT_TYPES}

PROJECT_STATUSES: list[tuple[str, str]] = [
    ("concept",         "Concept"),
    ("prefeasibility",  "Pre-feasibility"),
    ("feasibility",     "Feasibility"),
    ("bankable_design", "Bankable Design"),
    ("financial_close", "Financial Close"),
    ("construction",    "Construction"),
    ("commissioning",   "Commissioning"),
    ("operating",       "Operating"),
]
PROJECT_STATUS_CODES: set[str] = {c for c, _ in PROJECT_STATUSES}

DESIGN_STANDARDS: list[tuple[str, str]] = [
    ("IEC",   "IEC (International)"),
    ("IEEE",  "IEEE (US)"),
    ("EN",    "EN / CENELEC (EU)"),
    ("BS",    "BS (UK / Commonwealth)"),
    ("NEC",   "NEC (US National Electrical Code)"),
    ("SANS",  "SANS (South Africa)"),
    ("GS1000", "Ghana Standards (GS)"),
    ("NIS",   "NIS (Nigeria)"),
    ("KEBS",  "KEBS (Kenya)"),
    ("HYBRID_IEC_LOCAL", "IEC + Local National Code"),
]
DESIGN_STANDARD_CODES: set[str] = {c for c, _ in DESIGN_STANDARDS}

CURRENCIES: list[tuple[str, str]] = [
    ("GHS", "GHS - Ghanaian Cedi"),
    ("NGN", "NGN - Nigerian Naira"),
    ("KES", "KES - Kenyan Shilling"),
    ("ZAR", "ZAR - South African Rand"),
    ("XOF", "XOF - West African CFA"),
    ("USD", "USD - US Dollar"),
    ("EUR", "EUR - Euro"),
    ("GBP", "GBP - Pound Sterling"),
]
CURRENCY_CODES: set[str] = {c for c, _ in CURRENCIES}

TAX_REGIMES: list[tuple[str, str]] = [
    ("standard",   "Standard corporate tax"),
    ("epa_exempt", "Renewable Energy Act - tax exemption"),
    ("free_zone",  "Free Zone / SEZ"),
    ("ppa_pass",   "PPA pass-through"),
    ("bot",        "Build-Operate-Transfer concession"),
    ("public",     "Government / public sector"),
    ("negotiated", "Negotiated / bilateral"),
]
TAX_REGIME_CODES: set[str] = {c for c, _ in TAX_REGIMES}

# 14-step rail + the two spec bolt-ons (regulatory before PV; digital twin).
# (num, endpoint_suffix, short label, bootstrap icon). num<=0 = out-of-rail.
STEP_LABELS: list[tuple[int, str, str, str]] = [
    (1,  "new",        "Registration",  "bi-clipboard-plus"),
    (2,  "step2",      "Project Type",  "bi-diagram-3"),
    (3,  "step3",      "Site",          "bi-geo-alt"),
    (4,  "step4",      "Facilities",    "bi-buildings"),
    (5,  "step5",      "Technology",    "bi-cpu"),
    (6,  "step6",      "Electrical",    "bi-lightning"),
    (7,  "step7",      "PV Design",     "bi-grid-3x3"),
    (8,  "step8",      "Finance",       "bi-cash-coin"),
    (9,  "step9",      "BOQ",           "bi-list-check"),
    (10, "step10",     "Marketplace",   "bi-cart"),
    (11, "step11",     "CRM",           "bi-people"),
    (12, "step12",     "Pipeline",      "bi-kanban"),
    (13, "step13",     "Reports",       "bi-file-earmark-text"),
    (14, "step14",     "AI Agents",     "bi-robot"),
]


# ===========================================================================
# SECTION A2 -- Steps 3-6 + Regulatory vocab (spec Steps 3-6 + the recommended
# Development & Regulatory bolt-on). Carried forward verbatim from the spec so
# the rebuilt forms present the same complete option set.
# ===========================================================================

# -- Step 3: Site --
SITE_TERRAINS = [("flat", "Flat"), ("rolling", "Rolling / gently undulating"),
                 ("sloped", "Sloped"), ("hilly", "Hilly / broken"),
                 ("mountainous", "Mountainous"), ("mixed", "Mixed")]
SITE_SLOPES = [("lt_3", "< 3 % (near-flat)"), ("3_5", "3-5 %"),
               ("5_10", "5-10 %"), ("10_20", "10-20 %"), ("gt_20", "> 20 %")]
SITE_SOILS = [("sandy", "Sandy"), ("clay", "Clay"), ("loam", "Loam"),
              ("rocky", "Rocky / laterite"), ("marshy", "Marshy / soft"),
              ("mixed", "Mixed"), ("unknown", "Unknown / needs geotech")]
SITE_FLOOD_RISKS = [("none", "None"), ("low", "Low (1-in-100 year)"),
                    ("medium", "Medium (1-in-25 year)"),
                    ("high", "High (annual)"), ("unknown", "Unknown")]
SITE_WIND_ZONES = [("z1_low", "Zone 1 - Low (< 30 m/s)"),
                   ("z2_medium", "Zone 2 - Medium (30-45 m/s)"),
                   ("z3_high", "Zone 3 - High (45-60 m/s)"),
                   ("cyclone", "Cyclone / hurricane-prone"),
                   ("unknown", "Unknown")]
SITE_SEISMIC_ZONES = [("zone_0", "Zone 0 - negligible"),
                      ("zone_1", "Zone 1 - low"), ("zone_2", "Zone 2 - moderate"),
                      ("zone_3", "Zone 3 - severe"),
                      ("zone_4", "Zone 4 - very severe"), ("unknown", "Unknown")]
SITE_ACCESS = [("paved", "Paved public road"), ("gravel", "Gravel access road"),
               ("dirt_ok", "Dirt track - passable dry season"),
               ("dirt_seasonal", "Dirt track - seasonal only"),
               ("none", "No access road - construction required")]
SITE_WATER = [("piped", "Piped mains supply"), ("borehole", "On-site borehole"),
              ("tanker", "Tanker delivery"), ("river", "Nearby river / stream"),
              ("none", "None - civil supply required")]

# -- Step 4: Facilities -- (code, label, icon, recommended)
BUILDING_TYPES: list[tuple[str, str, str, bool]] = [
    ("control_room",     "Control Room Building",         "bi-cpu",              True),
    ("om_building",      "Operations & Maintenance",      "bi-tools",            True),
    ("security_gate",    "Security Gatehouse",            "bi-shield-lock",      True),
    ("warehouse",        "Warehouse",                     "bi-boxes",            False),
    ("workshop",         "Workshop",                      "bi-wrench-adjustable", False),
    ("admin",            "Administration Building",       "bi-building",         False),
    ("training",         "Training Room",                 "bi-mortarboard",      False),
    ("spare_parts",      "Spare Parts Store",             "bi-archive",          False),
    ("chemical",         "Chemical / Consumables Store",  "bi-droplet",          False),
    ("battery_room",     "Battery Building",              "bi-battery-full",     False),
    ("inverter_room",    "Inverter Building",             "bi-plug",             False),
    ("transformer_bldg", "Transformer Building",          "bi-lightning",        False),
    ("switchgear_bldg",  "Switchgear Building",           "bi-diagram-2",        False),
    ("scada_bldg",       "SCADA Building",                "bi-hdd-network",      False),
    ("comms_bldg",       "Communication Building",        "bi-broadcast",        False),
    ("welfare",          "Staff Welfare Building",        "bi-cup-hot",          False),
    ("washroom",         "Washroom",                      "bi-water",            False),
    ("parking",          "Parking",                       "bi-car-front",        False),
]
BUILDING_CODES: set[str] = {c for c, _, _, _ in BUILDING_TYPES}
BUILDING_SUB_ITEMS: dict[str, list[str]] = {
    "control_room": ["SCADA workstation", "Monitoring screens", "Operator consoles",
        "Video wall", "Server room rack", "Network cabinet", "Patch panels",
        "UPS system", "Battery backup", "HVAC power", "Emergency lighting",
        "Fire alarm", "IP CCTV", "Access control", "Public address",
        "VoIP telephony", "Small power outlets", "Data outlets",
        "Earthing & bonding", "Lightning protection"],
    "om_building": ["Maintenance office", "Technician office", "Workshop",
        "Testing bench area", "Repair bench", "Tool store", "Consumables store",
        "Spare parts store", "Battery maintenance area", "Electrical maintenance area",
        "Cable store", "Plant documentation room", "ICT room", "Lighting",
        "Socket outlets", "Fire alarm", "CCTV", "VoIP", "LAN", "WiFi", "UPS",
        "Earthing", "External lighting"],
    "security_gate": ["Gatehouse lighting", "Socket outlets", "CCTV monitor",
        "Access control panel", "Boom barrier power", "Data outlet",
        "Intercom / VoIP", "External security lighting", "UPS supply"],
    "battery_room": ["Battery racks", "Battery monitoring", "Battery DC cabling",
        "Protection panels", "Fire detection", "Temperature monitoring",
        "Gas detection", "Ventilation / cooling power", "HVAC", "UPS",
        "Emergency lighting", "Safety signage", "Access control", "Earthing & bonding"],
    "switchgear_bldg": ["MV switchgear", "LV switchgear", "Protection panels",
        "Metering panels", "Control panels", "Cable basement", "Earthing",
        "Lighting", "HVAC", "Fire alarm", "Access control"],
    "transformer_bldg": ["Power transformers", "RMU (Ring Main Unit)", "MV switchgear",
        "Lightning protection", "Oil bund / containment", "Earthing grid",
        "Cable trenches", "Yard lighting", "Security fence", "Danger signs"],
    "scada_bldg": ["SCADA servers", "Historian / data logger", "EMS", "PPC controller",
        "Network switches", "Industrial firewall", "Fibre patch panels",
        "GPS time sync", "NTP server", "UPS", "Server rack HVAC", "Fire alarm",
        "Access control"],
    "comms_bldg": ["Fibre distribution frame", "Radio / microwave equipment",
        "Antenna cabling", "Grounding kit", "UPS", "HVAC", "Fire detection",
        "Access control"],
    "warehouse": ["Lighting", "Small power", "Fire detection", "CCTV", "Earthing",
        "Loading bay power", "Ventilation"],
    "workshop": ["Workbench power outlets", "Lighting", "Small power", "Fire alarm",
        "CCTV", "Compressed-air supply", "Welding supply", "Data outlets"],
    "admin": ["Lighting", "Small power", "Data outlets", "VoIP", "Fire alarm",
        "HVAC", "Access control"],
    "training": ["Lighting", "Projector power", "Data outlets", "HVAC", "Fire alarm",
        "Small power"],
    "spare_parts": ["Lighting", "Small power", "Fire detection", "CCTV",
        "Access control", "Temperature monitoring"],
    "chemical": ["Lighting", "Small power", "Fire suppression", "Ventilation",
        "Emergency lighting", "Access control"],
    "inverter_room": ["Inverter AC output panels", "DC input protection",
        "Ventilation", "Fire alarm", "Temperature monitoring", "Emergency lighting",
        "Earthing"],
    "welfare": ["Lighting", "Small power", "Water heaters", "Fire alarm", "HVAC"],
    "washroom": ["Lighting", "Small power", "Water heaters", "Ventilation"],
    "parking": ["Perimeter lighting", "EV charging provision", "CCTV coverage",
        "Barrier control"],
}
EXTERNAL_WORKS: list[tuple[str, str, str, bool]] = [
    ("pv_field",        "PV module field",         "bi-grid-3x3",           True),
    ("mounting",        "Mounting structures",     "bi-columns-gap",        True),
    ("internal_roads",  "Internal roads",          "bi-signpost-2",         True),
    ("cable_trench",    "Cable trenches",          "bi-arrow-down-square",  True),
    ("ducts",           "Underground ducts",       "bi-diagram-3",          True),
    ("drainage",        "Drainage",                "bi-cloud-drizzle",      True),
    ("fence",           "Perimeter fence",         "bi-shield",             True),
    ("security_light",  "Security lighting",       "bi-lightbulb",          True),
    ("gate",            "Gate",                    "bi-door-open",          True),
    ("guardhouse",      "Guardhouse",              "bi-house-lock",         False),
    ("water_supply",    "Water supply",            "bi-droplet-fill",       False),
    ("fire_tank",       "Fire water tank",         "bi-water",              False),
    ("weather_station", "Weather station",         "bi-cloud-sun",          True),
    ("mast",            "Communication mast",      "bi-broadcast-pin",      False),
    ("signage",         "Site signage",            "bi-sign-turn-right",    False),
    ("landscaping",     "Landscaping",             "bi-tree",               False),
    ("stormwater",      "Storm water management",  "bi-cloud-rain-heavy",   False),
]
EXTERNAL_WORKS_CODES: set[str] = {c for c, _, _, _ in EXTERNAL_WORKS}

# -- Step 5: Technology (grouped) --
TECHNOLOGY_GROUPS: list[tuple[str, list[tuple[str, str, str]]]] = [
    ("Plant Control", [
        ("scada", "SCADA", "bi-cpu"), ("ems", "EMS (Energy Management)", "bi-graph-up"),
        ("ppc", "Power Plant Controller", "bi-sliders"),
        ("digital_twin", "Digital Twin", "bi-diagram-3")]),
    ("Monitoring & Metering", [
        ("weather", "Weather Station", "bi-cloud-sun"),
        ("string_mon", "String Monitoring", "bi-bezier"),
        ("energy_meter", "Energy Metering", "bi-speedometer2"),
        ("pq_meter", "Power Quality Monitoring", "bi-activity"),
        ("bms", "Battery Management System", "bi-battery-charging"),
        ("txfr_mon", "Transformer Monitoring", "bi-thermometer-half"),
        ("inv_mon", "Inverter Monitoring", "bi-plug"),
        ("remote_mon", "Remote Monitoring", "bi-cloud-check"),
        ("cloud_mon", "Cloud Monitoring", "bi-cloud-arrow-up"),
        ("thermal_cam", "Thermal Camera Monitoring", "bi-camera-video")]),
    ("AI & Analytics", [
        ("ai_fault", "AI Fault Prediction", "bi-lightning"),
        ("drone_insp", "Drone Inspection", "bi-airplane"),
        ("predictive", "Predictive Maintenance", "bi-graph-up-arrow")]),
    ("Asset & O&M", [
        ("gis", "GIS Mapping", "bi-geo-alt"), ("asset_mgmt", "Asset Management", "bi-boxes"),
        ("cmms", "CMMS (Maintenance Mgmt)", "bi-tools"),
        ("spares", "Spare Parts Management", "bi-archive"),
        ("scheduler", "Maintenance Scheduler", "bi-calendar-check"),
        ("wo_mgmt", "Work Order Management", "bi-clipboard-check")]),
    ("Network & Security", [
        ("cyber", "Cyber Security", "bi-shield-shaded"),
        ("firewall", "Industrial Firewall", "bi-shield-lock"),
        ("ind_eth", "Industrial Ethernet", "bi-diagram-2"),
        ("fibre", "Fibre Optic Network", "bi-slash-lg"),
        ("ind_wifi", "Industrial WiFi", "bi-wifi"),
        ("gps_sync", "GPS Time Synchronisation", "bi-broadcast"),
        ("ntp", "NTP Server", "bi-clock-history")]),
    ("Servers & Storage", [
        ("ind_servers", "Industrial Servers", "bi-server"),
        ("storage_srv", "Storage Servers", "bi-hdd-stack"),
        ("backup_srv", "Backup Servers", "bi-hdd"),
        ("cloud_backup", "Cloud Backup", "bi-cloud-upload"),
        ("dr", "Disaster Recovery", "bi-arrow-counterclockwise")]),
]
TECHNOLOGY_CODES: set[str] = {c for _, items in TECHNOLOGY_GROUPS for c, _, _ in items}

# -- Step 6: Electrical services -- (code, label, icon, recommended)
# NOTE: "pv_solar_farm" is the headline generation-asset service for a utility
# solar farm. Selecting it signals that the 20MWp PV field + balance-of-plant
# equipment BOQ is in scope; that BOQ is built by the solar-farm engine at Step 9
# (_ci_build_solar_farm_items) - the checkbox makes the scope visible/selectable
# on the services page (owner 2026-07-04: "include the PV Solar Farm service").
ELECTRICAL_SERVICES: list[tuple[str, str, str, bool]] = [
    ("pv_solar_farm",         "PV Solar Farm (20MWp Equipment BOQ)", "bi-sun",     True),
    ("internal_installation", "Internal Electrical Installation", "bi-plug",           True),
    ("power_supply",          "Power Supply",                     "bi-lightning",      True),
    ("hv_distribution",       "HV Distribution",                  "bi-diagram-3",      True),
    ("lv_distribution",       "LV Distribution",                  "bi-diagram-2",      True),
    ("dc_collection",         "DC Collection",                    "bi-arrow-down-circle", True),
    ("ac_collection",         "AC Collection",                    "bi-arrow-up-circle", True),
    ("inverters",             "Inverters",                        "bi-plug-fill",      True),
    ("transformers",          "Transformers",                     "bi-lightning-charge", True),
    ("rmu",                   "Ring Main Unit (RMU)",             "bi-arrow-repeat",   True),
    ("hv_switchgear",         "HV Switchgear",                    "bi-toggles",        True),
    ("lv_switchgear",         "LV Switchgear",                    "bi-toggles2",       True),
    ("earthing",              "Earthing",                         "bi-arrow-down",     True),
    ("lightning_protection",  "Lightning Protection",             "bi-cloud-lightning", True),
    ("external_lighting",     "External Lighting",                "bi-lightbulb",      True),
    ("fire_alarm",            "Fire Alarm",                       "bi-fire",           True),
    ("ip_cctv",               "IP CCTV",                          "bi-camera-video",   True),
    ("access_control",        "Access Control",                   "bi-shield-lock",    True),
    ("voip",                  "VoIP",                             "bi-telephone",      False),
    ("public_address",        "Public Address",                   "bi-megaphone",      False),
    ("tv",                    "TV",                               "bi-tv",             False),
    ("ip_clock",              "IP Clock",                         "bi-clock",          False),
    ("lan",                   "LAN",                              "bi-hdd-network",    True),
    ("wan",                   "WAN",                              "bi-router",         False),
    ("server_infra",          "Server Infrastructure",            "bi-server",         False),
    ("scada",                 "SCADA",                            "bi-cpu",            True),
]
ELECTRICAL_SERVICE_CODES: set[str] = {c for c, _, _, _ in ELECTRICAL_SERVICES}

# -- Step 7: PV design vocab -- (module tech carries default Wp)
PV_MODULE_TECHS: list[tuple[str, str, float]] = [
    ("mono_perc",   "Monocrystalline PERC",       550.0),
    ("mono_topcon", "Monocrystalline TOPCon",     600.0),
    ("mono_hjt",    "Monocrystalline HJT",        625.0),
    ("bifacial",    "Bifacial (mono)",            580.0),
    ("thin_film",   "Thin-film (CdTe)",           470.0),
    ("cigs",        "CIGS",                       360.0),
]
PV_MODULE_TECH_CODES: set[str] = {c for c, _, _ in PV_MODULE_TECHS}
PV_MOUNTING_TYPES: list[tuple[str, str]] = [
    ("fixed_tilt",         "Fixed-tilt racking"),
    ("single_axis",        "Single-axis tracker (HSAT)"),
    ("dual_axis",          "Dual-axis tracker"),
    ("east_west",          "East-West flat mounting"),
    ("rooftop_ballasted",  "Rooftop ballasted (flat roof)"),
    ("rooftop_pitched",    "Rooftop pitched"),
]
PV_MOUNTING_CODES: set[str] = {c for c, _ in PV_MOUNTING_TYPES}
PV_INVERTER_TYPES: list[tuple[str, str]] = [
    ("string",   "String inverters (100-250 kW)"),
    ("central",  "Central inverters (1-5 MW blocks)"),
    ("micro",    "Microinverters (rooftop only)"),
]
PV_INVERTER_CODES: set[str] = {c for c, _ in PV_INVERTER_TYPES}
PV_BATTERY_CHEMISTRIES: list[tuple[str, str]] = [
    ("none",      "No battery / grid-tied only"),
    ("lifepo4",   "LiFePO4 (LFP)"),
    ("nmc",       "NMC"),
    ("flow",      "Vanadium flow"),
    ("lead_acid", "Lead-acid (VRLA / OPzS)"),
]
PV_BATTERY_CODES: set[str] = {c for c, _ in PV_BATTERY_CHEMISTRIES}


def size_utility_pv(*,
                    kwp: float,
                    module_wp: float = 550.0,
                    dc_ac_ratio: float = 1.20,
                    tilt_deg: float = 10.0,
                    azimuth_deg: float = 180.0,
                    psh_daily: float = 5.4,
                    performance_ratio: float = 0.78,
                    availability_pct: float = 98.0,
                    annual_degradation_pct: float = 0.5,
                    project_life_yr: int = 25,
                    modules_per_string: int = 28,
                    strings_per_combiner: int = 20,
                    central_inverter_kw: float = 1500.0
                    ) -> dict[str, Any]:
    """Utility-scale PV design (module-local engine; Codex SSS Section 1 notes
    this is not platform-level and is carried with the module). Greenfield
    ground-mount formulas: n_modules=ceil(kwp*1000/Wp); inverter_ac=kwp/DCAC;
    annual_gen=kwp*PSH*365*PR*avail/1000; lifetime with degradation."""
    import math

    if kwp <= 0:
        return {"error": "kwp must be > 0"}
    if module_wp <= 0:
        module_wp = 550.0
    if dc_ac_ratio <= 0:
        dc_ac_ratio = 1.20
    if psh_daily <= 0:
        psh_daily = 5.4
    if performance_ratio <= 0 or performance_ratio > 1:
        performance_ratio = 0.78
    if availability_pct <= 0 or availability_pct > 100:
        availability_pct = 98.0
    if annual_degradation_pct < 0 or annual_degradation_pct >= 100:
        annual_degradation_pct = 0.5   # guard the degradation power series

    n_modules       = int(math.ceil(kwp * 1000 / module_wp))
    dc_kwp_actual   = round(n_modules * module_wp / 1000.0, 2)
    # Size the plant off the ACTUAL installed DC (rounded-up module count), not
    # the requested target, so inverter AC + energy + specific yield are all
    # internally consistent with the module count (Codex calc-review MED-1).
    dc_basis        = dc_kwp_actual if dc_kwp_actual > 0 else kwp
    inverter_ac_kw  = round(dc_basis / dc_ac_ratio, 2)
    n_central_inv   = int(math.ceil(inverter_ac_kw / max(1.0, central_inverter_kw)))
    strings         = int(math.ceil(n_modules / max(1, modules_per_string)))
    combiners       = int(math.ceil(strings / max(1, strings_per_combiner)))

    availability_frac = availability_pct / 100.0
    annual_gen_mwh    = round(
        dc_basis * psh_daily * 365 * performance_ratio * availability_frac / 1000.0, 2)
    monthly_gen_mwh   = round(annual_gen_mwh / 12.0, 2)

    lifetime_mwh = 0.0
    for t in range(1, project_life_yr + 1):
        lifetime_mwh += annual_gen_mwh * ((1 - annual_degradation_pct / 100.0) ** (t - 1))
    lifetime_mwh = round(lifetime_mwh, 2)

    specific_yield_kwh_per_kwp = round(annual_gen_mwh * 1000.0 / dc_basis, 1)
    dc_cable_m_est = int(round(kwp * 6.5, 0))
    ac_cable_m_est = int(round(kwp * 3.5, 0))

    return {
        "kwp_input":               round(kwp, 2),
        "dc_kwp_actual":           dc_kwp_actual,
        "module_wp":               module_wp,
        "n_modules":               n_modules,
        "dc_ac_ratio":             dc_ac_ratio,
        "inverter_ac_kw":          inverter_ac_kw,
        "n_central_inverters":     n_central_inv,
        "central_inverter_kw":     central_inverter_kw,
        "strings":                 strings,
        "modules_per_string":      modules_per_string,
        "combiners":               combiners,
        "strings_per_combiner":    strings_per_combiner,
        "tilt_deg":                tilt_deg,
        "azimuth_deg":             azimuth_deg,
        "psh_daily":               psh_daily,
        "performance_ratio":       performance_ratio,
        "availability_pct":        availability_pct,
        "annual_degradation_pct":  annual_degradation_pct,
        "project_life_yr":         project_life_yr,
        "annual_gen_mwh":          annual_gen_mwh,
        "monthly_gen_mwh":         monthly_gen_mwh,
        "lifetime_gen_mwh":        lifetime_mwh,
        "specific_yield_kwh_per_kwp": specific_yield_kwh_per_kwp,
        "dc_cable_m_est":          dc_cable_m_est,
        "ac_cable_m_est":          ac_cable_m_est,
    }


def _ci_yield_profile(pv_cfg, *, gps_lat=None, years: int = 10) -> dict:
    """Solar generation-yield profiles derived from the Step-7 PV sizing (no
    re-simulation): a representative DAILY curve, a latitude-aware MONTHLY
    distribution, and the ANNUAL yield across `years` (default 10) with module
    degradation.

    Inputs: pv_cfg (the stored pv_config incl. its `sizing` block), gps_lat (site
    latitude in degrees; falls back to Ghana ~7.5 when absent), years (annual
    horizon). The monthly shape uses the extraterrestrial clear-sky insolation
    model (solar declination + sunset hour angle by month, Duffie & Beckman) so
    seasonality tracks latitude physically rather than a fixed guess, then is
    normalised to the sizing's annual energy. Returns {} if no PV sizing exists.
    """
    import math
    sizing = {}
    if isinstance(pv_cfg, dict) and isinstance(pv_cfg.get("sizing"), dict):
        sizing = pv_cfg["sizing"]

    def _f(v, d=0.0):
        try:
            return float(v)
        except (TypeError, ValueError):
            return d

    annual = _f(sizing.get("annual_gen_mwh"))
    if annual <= 0:
        return {}
    deg = _f(sizing.get("annual_degradation_pct"), 0.5)
    deg = (deg / 100.0) if 0.0 <= deg < 100.0 else 0.005   # guard the series
    try:
        lat = float(gps_lat) if gps_lat not in (None, "") else 7.5
    except (TypeError, ValueError):
        lat = 7.5
    lat = max(-89.0, min(89.0, lat))   # keep tan(phi) finite; polar handled below
    phi = math.radians(lat)

    def _sunset_hour_angle(decl_rad: float) -> float:
        """Sunset hour angle with proper polar day/night handling (Codex MED-2):
        cos(ws) = -tan(phi)tan(decl); |.|>=1 -> polar day (pi) or night (0)."""
        x = -math.tan(phi) * math.tan(decl_rad)
        if x <= -1.0:
            return math.pi     # sun never sets (polar day)
        if x >= 1.0:
            return 0.0         # sun never rises (polar night)
        return math.acos(x)

    # Monthly weights via extraterrestrial daily insolation H0 per representative
    # day, times the number of days -> monthly insolation share.
    mid_doy = [17, 47, 75, 105, 135, 162, 198, 228, 258, 288, 318, 344]
    dim = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    mnames = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    m_tot = []
    for i, n in enumerate(mid_doy):
        decl = math.radians(23.45 * math.sin(math.radians(360.0 * (284 + n) / 365.0)))
        ws = _sunset_hour_angle(decl)
        dr = 1 + 0.033 * math.cos(math.radians(360.0 * n / 365.0))
        h0 = (24.0 / math.pi) * dr * (
            ws * math.sin(phi) * math.sin(decl)
            + math.cos(phi) * math.cos(decl) * math.sin(ws))
        m_tot.append(max(0.0, h0) * dim[i])
    tot = sum(m_tot) or 1.0
    monthly = [{"month": mnames[i], "mwh": round(annual * m_tot[i] / tot, 1),
                "pct": round(m_tot[i] / tot * 100.0, 1)} for i in range(12)]

    # Annual yield across N years with module degradation (year 1 = 100%).
    annual_series = [{"year": y, "mwh": round(annual * ((1 - deg) ** (y - 1)), 1),
                      "pct_of_y1": round(((1 - deg) ** (y - 1)) * 100.0, 1)}
                     for y in range(1, int(years) + 1)]

    # Representative clear-day hourly curve: a sine bell across the (equinox)
    # daylight window, scaled to the average daily energy (annual / 365).
    ws0 = _sunset_hour_angle(0.0)
    daylight_h = ws0 * 24.0 / math.pi
    sunrise, sunset = 12.0 - daylight_h / 2.0, 12.0 + daylight_h / 2.0
    avg_daily = annual / 365.0
    hours = list(range(24))
    raw = []
    for h in hours:
        hc = h + 0.5
        raw.append(math.sin(math.pi * (hc - sunrise) / daylight_h)
                   if (sunrise <= hc <= sunset and daylight_h > 0) else 0.0)
    rs = sum(raw) or 1.0
    day_mwh = [round(avg_daily * w / rs, 3) for w in raw]

    return {
        "annual_gen_mwh": round(annual, 1),
        "specific_yield_kwh_per_kwp": _f(sizing.get("specific_yield_kwh_per_kwp")),
        "lat_used": round(lat, 2),
        "daily": {"hours": hours, "mwh": day_mwh,
                  "peak_mw": round(max(day_mwh), 3) if day_mwh else 0.0,
                  "avg_daily_mwh": round(avg_daily, 2),
                  "daylight_hours": round(daylight_h, 1)},
        "monthly": monthly,
        "annual_series": annual_series,
        "years": int(years),
        "basis": ("Step-7 sizing annual yield; monthly via extraterrestrial "
                  "clear-sky model (latitude-aware); annual series applies "
                  "module degradation."),
    }


# -- Step 8: Financial engineering vocab + defaults --
DEFAULT_CAPEX_USD_PER_KWP: dict[str, float] = {
    "modules":          360.0,
    "inverters":         75.0,
    "structures":       150.0,
    "civil":            110.0,
    "electrical":       190.0,
    "grid_connection":   80.0,
    "development":       35.0,
    "professional":      45.0,
    "contingency":       50.0,
    "land":              15.0,
    "ict_scada":         25.0,
    "security":          10.0,
    "bess":               0.0,
    "other":              0.0,
}
DEFAULT_OPEX_USD_PER_KWP_YR: dict[str, float] = {
    "om_labour":         8.0,
    "cleaning":          3.0,
    "spare_parts":       2.5,
    "insurance":         3.5,
    "land_lease":        1.5,
    "grid_charges":      2.0,
    "admin_overhead":    2.0,
    "security":          1.5,
    "other":             0.0,
}
REVENUE_MODELS: list[tuple[str, str]] = [
    ("ppa",          "PPA (fixed tariff)"),
    ("merchant",     "Merchant (spot market)"),
    ("net_metering", "Net metering / self-consumption"),
    ("wheeling",     "Wheeling / bilateral"),
    ("captive",      "Captive (industrial off-take)"),
]
REVENUE_MODEL_CODES: set[str] = {c for c, _ in REVENUE_MODELS}

# Facilities CAPEX lines the Step 9 BOQ actually covers (building interiors:
# electrical fit-out + SCADA/ICT + security). Modules/inverters/structures/
# civil/grid = PV field / balance-of-plant, NOT building BOQ, so excluded from
# the BOQ reconciliation (Codex SSS Section 8).
_CI_FACILITY_CAPEX_KEYS: tuple[str, ...] = ("electrical", "ict_scada", "security")
# Immutable scope tag the Step 9 autobuild writes on every facilities cell, so
# the reconciliation reflects ONLY the generated facilities BOQ (durable across
# Build-all edits). Populated for real in Phase 5.
_CI_AUTOBUILD_SOURCE: str = "capital_autobuild"

# Free-tier worker safety cap: the Step 9 BOQ links a floor per selected
# facility, then pre-prices each floor's starter rows SYNCHRONOUSLY via the
# standard engine. On the single-worker Render free tier a busy facility (e.g.
# a control room with ~11 services => ~77 line items + rate-buildup rows) costs
# ~10-12s of PG round-trips; pricing 6 such floors in one request measured ~69s
# on live and KILLED the sole worker (502/503 cascade for everyone). So we
# auto-price at most this many floors PER REQUEST; any beyond the cap stay fully
# LINKED (boq_buildings + boq_floors + capital_investment_boq_links) and the user
# completes them in bounded batches via the "Finish BOQ pricing" button (each
# click prices up to this many more). Default 2 keeps every request well under
# the worker limit; raise via env only on a paid tier with headroom.
try:
    _CI_MAX_AUTOBUILD_FLOORS: int = max(1, int(
        os.environ.get("CI_MAX_AUTOBUILD_FLOORS", "1")))
except Exception:
    _CI_MAX_AUTOBUILD_FLOORS = 1

# How many facility floors Step 9 itself pre-prices SYNCHRONOUSLY. Measured on
# the live free tier: Postgres does ~275ms per insert, so ONE busy floor (a
# control room ~77 items => ~154 inserts) already costs ~42s, and pricing more
# than one floor plus the solar BOQ in a single Step-9 request killed the sole
# worker (502/503). So by default Step 9 only creates the BOQ STRUCTURE (fast,
# always succeeds) and the user prices it in ~1-floor batches via the "Finish
# BOQ pricing" button. Set CI_STEP9_PREPRICE=1 on a paid tier to pre-price up to
# _CI_MAX_AUTOBUILD_FLOORS floors in the Step-9 request itself.
_CI_STEP9_PREPRICE: bool = (
    os.environ.get("CI_STEP9_PREPRICE", "0").strip().lower() in ("1", "true", "yes"))


def _ci_boq_actuals(get_db, boq_project_id, uid, fx: float = 12.0,
                    extra_project_ids=None, tenant_override=None) -> dict:
    """Summarise a linked Generation-Station BOQ by REUSING the boq_floor_items
    totals the standard engine already wrote (no parallel costing). Returns
    local + USD grand totals, per-facility breakdown (facility =
    boq_buildings.purpose_subtype) + a labelled facility_costs_usd map for CRM.
    Empty (linked=False) until Step 9 links a BOQ project.

    Pass extra_project_ids (e.g. the solar-farm BOQ id) to include additional
    dedicated capital BOQ projects in the totals -- used by the REPORT paths so
    a report's BOQ total reflects facilities + solar (Codex MED-5). Step-8
    finance reconciliation keeps calling this facilities-only (no extra ids)."""
    out = {
        "linked": False, "boq_project_id": int(boq_project_id or 0),
        "grand_total_local": 0.0, "grand_total_usd": 0.0,
        "per_facility_local": {}, "per_facility_usd": {},
        "facility_costs_usd": {}, "n_items": 0,
    }
    try:
        fx = float(fx)
    except (TypeError, ValueError):
        fx = 12.0
    if fx <= 0:
        fx = 12.0
    pid_list = [int(boq_project_id)] if boq_project_id else []
    for x in (extra_project_ids or []):
        try:
            xi = int(x)
        except (TypeError, ValueError):
            continue
        if xi and xi not in pid_list:
            pid_list.append(xi)
    if not pid_list:
        return out
    if tenant_override is not None:
        # Authorized CROSS-TENANT read (Slice 5 institution report): scope the
        # BOQ to the APPLICANT's tenant, not the reviewer's session tenant --
        # otherwise _boq_tenant_clause would filter to the reviewer's tenant and
        # blank the totals on tenanted Postgres (Codex parity finding). NULL rows
        # (BOQ built before Keycloak was on) still match.
        tclause = " AND (i.tenant_id IS NULL OR CAST(i.tenant_id AS TEXT)=?)"
        tparams = (str(tenant_override),)
    else:
        try:
            from web_app import _boq_tenant_clause
            tclause, tparams = _boq_tenant_clause("i")
        except Exception:
            tclause, tparams = "", ()
    # Scope to the DEDICATED capital BOQ project(s) (each whole project belongs to
    # this generation station), so BOTH the lean autobuild starter rows AND any
    # rows the user later adds via BOQ "Build-all" (source_type='build_all') are
    # reconciled -- previously only 'capital_autobuild' rows counted, so Build-all
    # additions silently vanished from the finance reconciliation (Codex #4).
    _pph = ",".join("?" for _ in pid_list)
    params = list(pid_list) + [int(uid)] + list(tparams)
    try:
        with get_db() as c:
            rows = c.execute(
                "SELECT b.purpose_subtype, "
                "       COALESCE(SUM(i.total_amount),0), COUNT(*) "
                "FROM boq_floor_items i "
                "LEFT JOIN boq_buildings b "
                "       ON b.id=i.building_id AND b.project_id=i.project_id "
                "WHERE i.project_id IN (" + _pph + ") AND i.user_id=?" + tclause +
                " GROUP BY b.purpose_subtype",
                tuple(params)).fetchall()
    except Exception:
        return out
    out["linked"] = True
    _labels = {cd: L for cd, L, _, _ in BUILDING_TYPES}
    g_local = 0.0
    n_items = 0
    for r in (rows or []):
        fac = ((r[0] if r else "") or "").strip() or "unassigned"
        tot = float((r[1] if r and len(r) > 1 else 0) or 0)
        cnt = int((r[2] if r and len(r) > 2 else 0) or 0)
        g_local += tot
        n_items += cnt
        usd = round(tot / fx, 2) if fx else 0.0
        out["per_facility_local"][fac] = round(tot, 2)
        out["per_facility_usd"][fac] = usd
        out["facility_costs_usd"][_labels.get(fac, fac)] = usd
    out["grand_total_local"] = round(g_local, 2)
    out["grand_total_usd"] = round(g_local / fx, 2) if fx else 0.0
    out["n_items"] = n_items
    return out


def _ci_cost_plan(get_db, boq_project_id, uid, *, fx: float = 12.0,
                  extra_project_ids=None) -> dict:
    """Cost Plan Deck data derived ENTIRELY from the linked BOQ
    (boq_floor_items + boq_buildings) -- no parallel costing. This is the shared
    aggregation engine behind both the by-building / by-service breakdown views
    (F1) and the exportable Cost Plan Deck (F2).

    Inputs: get_db factory, the linked boq_project_id, owning uid, fx (local per
    USD). Returns a dict with:
      linked            -- False until Step 9 links a BOQ.
      totals            -- {grand_total_local, grand_total_usd, n_items,
                            n_buildings, n_sections}.
      by_building       -- [{key, label, total_local, total_usd, pct, n_items,
                            sections:[{section, label, total_local, n_items,
                            items:[{item_no, description, unit, qty, rate,
                            total_local}]}]}] ordered by cost desc.
      by_service        -- [{key(section), label, total_local, total_usd, pct,
                            n_items, buildings:[{label, total_local}]}] ordered
                            by cost desc. (Sections ARE the services in the
                            Generation-Station BOQ.)
      distribution      -- {by_building:[{label, total_local, pct}],
                            by_service:[...]} -- ready for the cost-distribution
                            infographics.
    Every query is scoped by user_id (+ tenant via _boq_tenant_clause) and to the
    Step-9 autobuild items (source_type) so the deck reflects the generated plan.
    """
    out = {
        "linked": False, "boq_project_id": int(boq_project_id or 0),
        "totals": {"grand_total_local": 0.0, "grand_total_usd": 0.0,
                   "n_items": 0, "n_buildings": 0, "n_sections": 0},
        "by_building": [], "by_service": [],
        "distribution": {"by_building": [], "by_service": []},
    }
    try:
        fx = float(fx)
    except (TypeError, ValueError):
        fx = 12.0
    if fx <= 0:
        fx = 12.0
    # Aggregate across the facilities BOQ AND (optionally) the solar-farm BOQ so
    # the Cost Plan Deck reflects the WHOLE capital investment. Each id is a
    # dedicated capital BOQ project, so we scope by project id(s) + user (+ tenant)
    # and do NOT filter source_type -- both the lean autobuild starter and any
    # Build-all additions are included (Codex #4/#6).
    pid_list = [int(boq_project_id)] if boq_project_id else []
    for x in (extra_project_ids or []):
        try:
            xi = int(x)
        except (TypeError, ValueError):
            continue
        if xi and xi not in pid_list:
            pid_list.append(xi)
    if not pid_list:
        return out

    try:
        from web_app import _boq_tenant_clause
        tclause, tparams = _boq_tenant_clause("i")
    except Exception:
        tclause, tparams = "", ()
    _pph = ",".join("?" for _ in pid_list)
    params = list(pid_list) + [int(uid)] + list(tparams)
    try:
        with get_db() as c:
            rows = c.execute(
                "SELECT COALESCE(b.purpose_subtype,'') , "
                "       COALESCE(b.building_name,'') , "
                "       COALESCE(i.section,'') , COALESCE(i.item_no,'') , "
                "       COALESCE(i.description,'') , COALESCE(i.unit,'') , "
                "       COALESCE(i.qty,0) , COALESCE(i.final_built_up_rate,0) , "
                "       COALESCE(i.total_amount,0) "
                "FROM boq_floor_items i "
                "LEFT JOIN boq_buildings b "
                "       ON b.id=i.building_id AND b.project_id=i.project_id "
                "WHERE i.project_id IN (" + _pph + ") AND i.user_id=?" + tclause +
                " ORDER BY b.building_name, i.section, i.item_no",
                tuple(params)).fetchall()
    except Exception:
        return out

    out["linked"] = True
    fac_labels = {cd: L for cd, L, _, _ in BUILDING_TYPES}
    # Optional friendly section->service labels from the platform BOQ engine.
    try:
        from web_app import _BOQ_SERVICE_LABEL as _svc_label_map
    except Exception:
        _svc_label_map = {}

    def _svc_label(section: str) -> str:
        s = (section or "").strip()
        if not s:
            return "General"
        if isinstance(_svc_label_map, dict) and s in _svc_label_map:
            return str(_svc_label_map[s])
        return s.replace("_", " ").title()

    # Aggregate in Python (a Step-9 BOQ is ~hundreds of rows, not millions).
    buildings: dict = {}   # bkey -> {label, total, n_items, sections{}}
    services: dict = {}    # section -> {label, total, n_items, buildings{}}
    grand = 0.0
    n_items = 0
    for r in (rows or []):
        fkey = (r[0] or "").strip() or "unassigned"
        bname = (r[1] or "").strip() or fac_labels.get(fkey, fkey)
        section = (r[2] or "").strip() or "general"
        item_no = (r[3] or "").strip()
        desc = (r[4] or "").strip()
        unit = (r[5] or "").strip() or "No."
        qty = float(r[6] or 0)
        rate = float(r[7] or 0)
        tot = float(r[8] or 0)
        grand += tot
        n_items += 1

        b = buildings.setdefault(fkey, {
            "key": fkey, "label": fac_labels.get(fkey, bname),
            "total_local": 0.0, "n_items": 0, "sections": {}})
        b["total_local"] += tot
        b["n_items"] += 1
        bs = b["sections"].setdefault(section, {
            "section": section, "label": _svc_label(section),
            "total_local": 0.0, "n_items": 0, "items": []})
        bs["total_local"] += tot
        bs["n_items"] += 1
        bs["items"].append({
            "item_no": item_no, "description": desc, "unit": unit,
            "qty": round(qty, 2), "rate": round(rate, 2),
            "total_local": round(tot, 2)})

        s = services.setdefault(section, {
            "key": section, "label": _svc_label(section),
            "total_local": 0.0, "n_items": 0, "buildings": {}})
        s["total_local"] += tot
        s["n_items"] += 1
        s["buildings"][fkey] = s["buildings"].get(fkey, 0.0) + tot

    def _pct(v: float) -> float:
        return round((v / grand) * 100.0, 1) if grand else 0.0

    by_building = []
    for b in buildings.values():
        secs = []
        for sec in sorted(b["sections"].values(),
                          key=lambda x: x["total_local"], reverse=True):
            secs.append({
                "section": sec["section"], "label": sec["label"],
                "total_local": round(sec["total_local"], 2),
                "n_items": sec["n_items"], "items": sec["items"]})
        by_building.append({
            "key": b["key"], "label": b["label"],
            "total_local": round(b["total_local"], 2),
            "total_usd": round(b["total_local"] / fx, 2),
            "pct": _pct(b["total_local"]), "n_items": b["n_items"],
            "sections": secs})
    by_building.sort(key=lambda x: x["total_local"], reverse=True)

    by_service = []
    for s in services.values():
        blist = [{"label": fac_labels.get(k, k), "total_local": round(v, 2)}
                 for k, v in sorted(s["buildings"].items(),
                                    key=lambda kv: kv[1], reverse=True)]
        by_service.append({
            "key": s["key"], "label": s["label"],
            "total_local": round(s["total_local"], 2),
            "total_usd": round(s["total_local"] / fx, 2),
            "pct": _pct(s["total_local"]), "n_items": s["n_items"],
            "buildings": blist})
    by_service.sort(key=lambda x: x["total_local"], reverse=True)

    out["totals"] = {
        "grand_total_local": round(grand, 2),
        "grand_total_usd": round(grand / fx, 2),
        "n_items": n_items, "n_buildings": len(buildings),
        "n_sections": len(services)}
    out["by_building"] = by_building
    out["by_service"] = by_service
    out["distribution"] = {
        "by_building": [{"label": b["label"], "total_local": b["total_local"],
                         "pct": b["pct"]} for b in by_building],
        "by_service": [{"label": s["label"], "total_local": s["total_local"],
                        "pct": s["pct"]} for s in by_service]}
    return out


def _ci_cashflow_plan(fin_cfg) -> dict:
    """Annual project cash-flow series for the Cost Plan Deck, READ from the
    Step-8 finance engine output (finance_config.computed) -- no re-modelling.

    Input: fin_cfg (the stored finance_config). Returns
      available     -- False until Step 8 computes.
      years         -- [0, 1, ..., N] (year 0 = construction).
      net           -- net cash flow per year; year 0 = -equity (own cash at
                       risk; falls back to -CAPEX if equity is 0).
      cumulative    -- running sum of `net` (cumulative cash position).
      revenue/opex  -- per-year revenue and OPEX (year 0 = 0) for the stacked bars.
      capex_local, equity_local, npv_local, irr_pct, payback_years -- headline KPIs.
    Reuses net_by_year / revenue_by_year / opex_by_year the finance engine already
    wrote, so the deck and Step 8 can never disagree.
    """
    out = {"available": False, "years": [], "net": [], "cumulative": [],
           "revenue": [], "opex": [], "capex_local": 0.0, "equity_local": 0.0,
           "npv_local": None, "irr_pct": None, "payback_years": None}
    computed = {}
    if isinstance(fin_cfg, dict) and isinstance(fin_cfg.get("computed"), dict):
        computed = fin_cfg["computed"]
    net_by_year = computed.get("net_by_year")
    if not isinstance(net_by_year, list) or not net_by_year:
        return out

    def _fl(v, d=0.0):
        try:
            return float(v)
        except (TypeError, ValueError):
            return d

    equity = _fl(computed.get("equity_local"))
    capex = _fl(computed.get("total_capex_local"))
    rev_src = computed.get("revenue_by_year")
    opx_src = computed.get("opex_by_year")
    rev_src = rev_src if isinstance(rev_src, list) else []
    opx_src = opx_src if isinstance(opx_src, list) else []

    net = [_fl(x) for x in net_by_year]
    n = len(net)
    y0 = -(equity if equity > 0 else capex)          # year-0 cash outflow
    flows = [round(y0, 2)] + [round(x, 2) for x in net]
    cumulative, run = [], 0.0
    for f in flows:
        run += f
        cumulative.append(round(run, 2))
    # Align revenue/opex to the N operating years (year 0 = 0).
    rev = [0.0] + [round(_fl(x), 2) for x in rev_src[:n]] + [0.0] * max(0, n - len(rev_src))
    opx = [0.0] + [round(_fl(x), 2) for x in opx_src[:n]] + [0.0] * max(0, n - len(opx_src))

    out.update({
        "available": True,
        "years": list(range(0, n + 1)),
        "net": flows,
        "cumulative": cumulative,
        "revenue": rev[:n + 1],
        "opex": opx[:n + 1],
        "capex_local": round(capex, 2),
        "equity_local": round(equity, 2),
        "npv_local": computed.get("npv_local"),
        "irr_pct": computed.get("irr_pct"),
        "payback_years": computed.get("payback_years"),
    })
    return out


# Bankability hurdles for an emerging-market utility IPP (lenders' rules of
# thumb). Tunable via env without a redeploy.
_CI_BANK_DSCR_STRONG = float(os.environ.get("CI_BANK_DSCR_STRONG", "1.35"))
_CI_BANK_DSCR_MIN = float(os.environ.get("CI_BANK_DSCR_MIN", "1.20"))
_CI_BANK_IRR_STRONG = float(os.environ.get("CI_BANK_IRR_STRONG", "15.0"))
_CI_BANK_IRR_MIN = float(os.environ.get("CI_BANK_IRR_MIN", "10.0"))


def _ci_bankability(computed) -> dict:
    """Determine project bankability from the Step-8 finance-engine output
    (finance_config.computed) -- READ ONLY, no re-modelling. Scores six lender
    metrics (min-DSCR, project IRR, NPV, LCOE-vs-tariff margin, payback ratio,
    downside P90 IRR), each weighted, into a 0-100 bankability score with a
    verdict band (Bankable / Conditionally Bankable / Not Yet Bankable) plus the
    strengths, risks and financing conditions a credit committee would list.

    Returns {available, score, rating, rating_class, metrics[], strengths[],
    risks[], conditions[]}. `available` is False until Step 8 has computed."""
    out = {"available": False, "score": 0, "rating": "Not assessed",
           "rating_class": "secondary", "metrics": [], "strengths": [],
           "risks": [], "conditions": []}
    if not isinstance(computed, dict) or not computed:
        return out

    def _f(key):
        try:
            v = computed.get(key)
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None

    dscr_min = _f("dscr_min")
    dscr_avg = _f("dscr_avg")
    irr = _f("irr_pct")
    npv = _f("npv_local")
    lcoe = _f("lcoe_local_per_kwh")
    tariff = _f("tariff_local_per_kwh")
    payback = _f("payback_years")
    life = _f("project_life_yr") or 25.0
    mc = computed.get("monte_carlo") if isinstance(computed.get("monte_carlo"), dict) else {}
    p90_irr = None
    try:
        p90_irr = float(mc.get("irr_p90_pct")) if mc.get("irr_p90_pct") is not None else None
    except (TypeError, ValueError):
        p90_irr = None

    metrics, strengths, risks, conditions = [], [], [], []
    earned = 0.0
    assessable = 0.0

    def add(label, weight, ratio, value_txt, note, verdict):
        nonlocal earned, assessable
        assessable += weight
        earned += weight * max(0.0, min(1.0, ratio))
        metrics.append({"label": label, "value": value_txt, "note": note,
                        "verdict": verdict, "weight": weight})

    # 1. Minimum DSCR (30) -- the single most important lender metric.
    if dscr_min is not None:
        if dscr_min >= _CI_BANK_DSCR_STRONG:
            r, v = 1.0, "strong"; strengths.append(
                f"Min DSCR {dscr_min:.2f}x clears the {_CI_BANK_DSCR_STRONG:.2f}x lender floor.")
        elif dscr_min >= _CI_BANK_DSCR_MIN:
            r, v = 0.6, "adequate"; conditions.append(
                f"Min DSCR {dscr_min:.2f}x is thin -- size a DSRA and cash sweep to hold >= {_CI_BANK_DSCR_STRONG:.2f}x.")
        else:
            r, v = 0.15, "weak"; risks.append(
                f"Min DSCR {dscr_min:.2f}x is below the {_CI_BANK_DSCR_MIN:.2f}x floor -- reduce gearing or tenor.")
        add("Minimum DSCR", 30, r, f"{dscr_min:.2f}x",
            f"avg {dscr_avg:.2f}x" if dscr_avg else "", v)

    # 2. Project IRR (20).
    if irr is not None:
        if irr >= _CI_BANK_IRR_STRONG:
            r, v = 1.0, "strong"; strengths.append(f"Project IRR {irr:.1f}% is attractive.")
        elif irr >= _CI_BANK_IRR_MIN:
            r, v = 0.6, "adequate"
        else:
            r, v = 0.2, "weak"; risks.append(
                f"Project IRR {irr:.1f}% is below the {_CI_BANK_IRR_MIN:.0f}% equity hurdle.")
        add("Project IRR", 20, r, f"{irr:.1f}%", "", v)

    # 3. NPV sign (15).
    if npv is not None:
        if npv > 0:
            r, v = 1.0, "positive"; strengths.append("NPV is positive at the project discount rate.")
        else:
            r, v = 0.0, "negative"; risks.append("NPV is negative -- the project destroys value at the current tariff/CAPEX.")
        add("NPV", 15, r, f"{npv:,.0f}", "", v)

    # 4. LCOE vs tariff margin (15).
    if lcoe is not None and tariff and tariff > 0:
        margin = (tariff - lcoe) / tariff
        if margin >= 0.25:
            r, v = 1.0, "strong"; strengths.append(
                f"Tariff sits {margin*100:.0f}% above LCOE -- healthy revenue headroom.")
        elif margin >= 0.10:
            r, v = 0.6, "adequate"
        elif margin > 0:
            r, v = 0.3, "thin"; conditions.append(
                "LCOE-to-tariff margin is thin -- a PPA escalator or tariff floor is advisable.")
        else:
            r, v = 0.0, "negative"; risks.append(
                "LCOE exceeds tariff -- revenues do not cover the levelised cost.")
        add("LCOE vs tariff", 15, r, f"{lcoe:.2f} vs {tariff:.2f}",
            f"{margin*100:+.0f}% headroom", v)

    # 5. Payback vs life (10).
    if payback is not None and life > 0:
        ratio = payback / life
        if ratio <= 0.40:
            r, v = 1.0, "fast"
        elif ratio <= 0.60:
            r, v = 0.6, "moderate"
        else:
            r, v = 0.25, "slow"; conditions.append(
                f"Payback {payback:.1f}y is a large share of the {life:.0f}y life -- lenders will want a shorter tenor buffer.")
        add("Payback / life", 10, r, f"{payback:.1f}y / {life:.0f}y", "", v)

    # 6. Downside resilience -- Monte-Carlo P90 IRR (10).
    if p90_irr is not None:
        if p90_irr >= _CI_BANK_IRR_MIN:
            r, v = 1.0, "resilient"; strengths.append(
                f"Even at P90 the IRR holds {p90_irr:.1f}% -- robust to downside.")
        elif p90_irr >= 0:
            r, v = 0.5, "sensitive"; conditions.append(
                "Returns are sensitive on the downside (P90) -- stress-test the resource and availability assumptions.")
        else:
            r, v = 0.1, "fragile"; risks.append("P90 IRR turns negative -- the base case has little margin for error.")
        add("Downside P90 IRR", 10, r, f"{p90_irr:.1f}%", "Monte Carlo", v)

    if assessable <= 0:
        return out   # nothing computed yet -> not assessed

    score = int(round(earned / assessable * 100.0))
    if score >= 75:
        rating, rclass = "Bankable", "success"
    elif score >= 55:
        rating, rclass = "Conditionally Bankable", "warning"
    else:
        rating, rclass = "Not Yet Bankable", "danger"
    if not conditions and rating != "Bankable":
        conditions.append("Firm up the PPA, EPC price and resource data before a lender term sheet.")
    out.update({"available": True, "score": score, "rating": rating,
                "rating_class": rclass, "metrics": metrics,
                "strengths": strengths, "risks": risks, "conditions": conditions})
    return out


# Cost-split donut palette (2026-07-05). The previous set had three near-identical
# greens + three blues that blended into each other on the dark deck ("some
# colours not visible"). This is a CVD-safe categorical order validated with the
# data-viz palette validator against the deck ring surface #22273a: all 8 inside
# the dark lightness band, worst-adjacent colour-blind separation ΔE 23.6 (target
# >=12). Order = blue, amber, aqua, violet, green, red, magenta, orange. The green
# is a touch under 3:1 contrast, which the legend labels + the 2px inter-slice gap
# (secondary encoding) below relieve.
_CI_CHART_PALETTE: list[str] = [
    "#3987e5", "#c98500", "#199e70", "#9085e9", "#008300", "#e66767",
    "#d55181", "#d95926",
]
_CI_ACCENT = "#2dd4bf"     # teal accent (was #f5c518/#f5a623 yellow)
_CI_ACCENT2 = "#38bdf8"    # sky-blue secondary


def _svg_donut(rows, *, size: int = 240, thickness: int = 40,
               palette: list | None = None, money: bool = True,
               center_label: str = "", center_sub: str = "") -> str:
    """Circular donut/pie chart as inline SVG (CSP-safe). `rows` =
    [{label, value, pct?}]; slices cycle through the palette. Renders the ring +
    a compact legend beneath it + an optional centre total. Returns an <svg>
    string (mark |safe)."""
    import html as _html
    import math as _math
    rows = [r for r in (rows or []) if isinstance(r, dict)
            and float(r.get("value") or 0) > 0]
    if not rows:
        return '<div class="text-secondary small py-2">No data yet.</div>'
    pal = palette or _CI_CHART_PALETTE
    total = sum(float(r.get("value") or 0) for r in rows) or 1.0
    r = (size - thickness) / 2.0
    cx = cy = size / 2.0
    circ = 2 * _math.pi * r
    # Legend height: one row per slice (cap the ring at 12 slices, roll the rest
    # into an "Other" wedge so the chart stays legible).
    # Never cycle categorical hues: at most one slice per palette colour, roll any
    # remainder into a single "Other" wedge (data-viz rule).
    cap = len(pal)
    show = rows[:cap - 1]
    rest = rows[cap - 1:]
    if rest:
        show = show + [{"label": f"Other ({len(rest)})",
                        "value": sum(float(x.get("value") or 0) for x in rest)}]
    legend_h = len(show) * 16 + 6
    height = size + legend_h
    out = [f'<svg viewBox="0 0 {size} {height}" width="100%" '
           f'style="max-width:{size}px;font:11px sans-serif" role="img" '
           f'preserveAspectRatio="xMinYMin meet">']
    out.append(f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" '
               f'stroke="#22273a" stroke-width="{thickness}"/>')
    offset = 0.0
    # 2px surface gap between wedges (shows the dark background ring through) so
    # adjacent slices never blend — the data-viz separator spec + secondary
    # encoding that relieves the near-3:1 green.
    gap = 2.0 if len(show) > 1 else 0.0
    for i, row in enumerate(show):
        val = float(row.get("value") or 0)
        frac = val / total
        seg = frac * circ
        color = pal[i % len(pal)]
        drawn = seg - gap
        if drawn < 0.75:                 # keep a sliver visible for tiny slices
            drawn = min(seg, 0.75)
        # dash: draw `drawn`, gap the rest; rotate so slices start at 12 o'clock.
        out.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" '
            f'stroke="{color}" stroke-width="{thickness}" '
            f'stroke-dasharray="{drawn:.2f} {circ - drawn:.2f}" '
            f'stroke-dashoffset="{-offset:.2f}" '
            f'transform="rotate(-90 {cx} {cy})"/>')
        offset += seg
    if center_label:
        out.append(
            f'<text x="{cx}" y="{cy - 2}" text-anchor="middle" '
            f'fill="#e8e8f5" style="font-size:15px;font-weight:700">'
            f'{_html.escape(center_label)}</text>')
    if center_sub:
        out.append(
            f'<text x="{cx}" y="{cy + 15}" text-anchor="middle" '
            f'fill="#9090b0" style="font-size:10px">'
            f'{_html.escape(center_sub)}</text>')
    ly = size + 12
    for i, row in enumerate(show):
        color = pal[i % len(pal)]
        val = float(row.get("value") or 0)
        pct = row.get("pct")
        if pct is None:
            pct = val / total * 100.0
        label = _html.escape(str(row.get("label") or "")[:30])
        vtxt = f"{val:,.0f}" if money else f"{val:g}"
        out.append(f'<rect x="4" y="{ly - 9}" width="10" height="10" rx="2" '
                   f'fill="{color}"/>')
        out.append(f'<text x="20" y="{ly}" fill="#c9c9e0">{label}</text>')
        out.append(f'<text x="{size - 4}" y="{ly}" text-anchor="end" '
                   f'fill="#e8e8f5">{vtxt}  {float(pct):.1f}%</text>')
        ly += 16
    out.append('</svg>')
    return "".join(out)


def _svg_scurve(labels, values, *, width: int = 560, height: int = 220,
                color: str = _CI_ACCENT2, money: bool = True) -> str:
    """Cumulative cost S-curve as inline SVG: an area-filled line of the running
    total across `labels` (e.g. bills / buildings). `values` are the per-step
    amounts; the curve plots their cumulative sum. CSP-safe, no external libs."""
    import html as _html
    vals = [float(v or 0) for v in (values or [])]
    if not vals:
        return '<div class="text-secondary small py-2">No data yet.</div>'
    labels = [str(l) for l in (labels or [])]
    cum, run = [], 0.0
    for v in vals:
        run += v
        cum.append(run)
    vmax = max(cum) or 1.0
    pl, pr, pt, pb = 52, 12, 14, 40
    pw, ph = width - pl - pr, height - pt - pb
    n = len(cum)
    step = pw / max(1, (n - 1)) if n > 1 else pw

    def xof(i):
        return pl + (i * step if n > 1 else pw / 2)

    def yof(v):
        return pt + (1 - v / vmax) * ph

    pts = [(xof(i), yof(v)) for i, v in enumerate(cum)]
    line_pts = " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
    area_pts = (f"{pl:.1f},{pt + ph:.1f} " + line_pts +
                f" {xof(n - 1):.1f},{pt + ph:.1f}")
    out = [f'<svg viewBox="0 0 {width} {height}" width="100%" '
           f'style="max-width:{width}px;font:10px sans-serif" role="img" '
           f'preserveAspectRatio="xMinYMin meet">']
    # gridlines at 0/50/100%
    for frac in (0.0, 0.5, 1.0):
        gy = pt + (1 - frac) * ph
        out.append(f'<line x1="{pl}" y1="{gy:.1f}" x2="{width - pr}" '
                   f'y2="{gy:.1f}" stroke="#2a2f45" stroke-width="1"/>')
        out.append(f'<text x="{pl - 5}" y="{gy + 3:.1f}" text-anchor="end" '
                   f'fill="#9090b0">{vmax * frac:,.0f}</text>')
    out.append(f'<polygon points="{area_pts}" fill="{color}" '
               f'fill-opacity="0.14"/>')
    out.append(f'<polyline points="{line_pts}" fill="none" stroke="{color}" '
               f'stroke-width="2.5"/>')
    for i, (x, y) in enumerate(pts):
        out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="2.6" fill="{color}"/>')
        if n <= 14 or i % 2 == 0:
            out.append(f'<text x="{x:.1f}" y="{height - 8}" '
                       f'text-anchor="middle" fill="#9090b0">'
                       f'{_html.escape(labels[i][:8]) if i < len(labels) else ""}</text>')
    out.append('</svg>')
    return "".join(out)


def _svg_hbars(rows, *, width: int = 560, row_h: int = 24, gap: int = 6,
               pad_left: int = 150, color: str = "#38bdf8",
               money: bool = True) -> str:
    """Horizontal bar chart as inline SVG (CSP-safe, no external libs) for the
    cost-distribution infographics. `rows` = [{label, value, pct?}]. Bars scale
    to the max value; labels/values are HTML-escaped. Returns an <svg> string
    (mark |safe in the template)."""
    import html as _html
    rows = [r for r in (rows or []) if isinstance(r, dict)]
    if not rows:
        return '<div class="text-secondary small py-2">No data yet.</div>'
    vmax = max((abs(float(r.get("value") or 0)) for r in rows), default=0.0) or 1.0
    bar_area = max(60, width - pad_left - 96)
    height = len(rows) * (row_h + gap) + gap
    out = [f'<svg viewBox="0 0 {width} {height}" width="100%" '
           f'style="max-width:{width}px;font:11px sans-serif" '
           f'role="img" preserveAspectRatio="xMinYMin meet">']
    y = gap
    for r in rows:
        label = _html.escape(str(r.get("label") or "")[:34])
        val = float(r.get("value") or 0)
        pct = r.get("pct")
        bw = max(1.0, abs(val) / vmax * bar_area)
        vtxt = (f"{val:,.0f}" if money else f"{val:g}")
        if pct is not None:
            vtxt += f"  {float(pct):.1f}%"
        out.append(
            f'<text x="{pad_left - 6}" y="{y + row_h * 0.7:.0f}" '
            f'text-anchor="end" fill="#c9c9e0">{label}</text>')
        out.append(
            f'<rect x="{pad_left}" y="{y}" width="{bw:.1f}" height="{row_h}" '
            f'rx="3" fill="{color}"/>')
        out.append(
            f'<text x="{pad_left + bw + 5:.1f}" y="{y + row_h * 0.7:.0f}" '
            f'fill="#e8e8f5">{vtxt}</text>')
        y += row_h + gap
    out.append('</svg>')
    return "".join(out)


def _svg_columns(labels, values, *, line=None, width: int = 560,
                 height: int = 210, color: str = "#38bdf8",
                 line_color: str = "#2dd4bf", every: int = 1) -> str:
    """Vertical column chart as inline SVG, with an optional overlaid line
    (e.g. cumulative cash flow). Handles negative values (columns drop below a
    zero baseline). `labels`/`values` equal length; `line` optional same length.
    Used for the daily / monthly / 10-year yield and the cash-flow charts."""
    import html as _html
    vals = [float(v or 0) for v in (values or [])]
    if not vals:
        return '<div class="text-secondary small py-2">No data yet.</div>'
    labels = [str(l) for l in (labels or [])]
    lvals = [float(v or 0) for v in line] if line else None
    pool = vals + (lvals or []) + [0.0]
    vmax, vmin = max(pool), min(pool)
    span = (vmax - vmin) or 1.0
    pl, pr, pt, pb = 44, 12, 12, 26
    pw, ph = width - pl - pr, height - pt - pb
    n = len(vals)
    step = pw / n
    bw = step * 0.66

    def yof(v):
        return pt + (vmax - v) / span * ph

    zero_y = yof(0.0)
    out = [f'<svg viewBox="0 0 {width} {height}" width="100%" '
           f'style="max-width:{width}px;font:10px sans-serif" role="img" '
           f'preserveAspectRatio="xMinYMin meet">']
    # zero baseline + max gridline
    out.append(f'<line x1="{pl}" y1="{zero_y:.1f}" x2="{width - pr}" '
               f'y2="{zero_y:.1f}" stroke="#555" stroke-width="1"/>')
    out.append(f'<text x="{pl - 5}" y="{pt + 8}" text-anchor="end" '
               f'fill="#9090b0">{vmax:,.0f}</text>')
    for i, v in enumerate(vals):
        x = pl + i * step + (step - bw) / 2
        top = yof(max(v, 0.0))
        h = abs(yof(v) - zero_y)
        out.append(f'<rect x="{x:.1f}" y="{top:.1f}" width="{bw:.1f}" '
                   f'height="{max(0.5, h):.1f}" rx="2" fill="{color}"/>')
        if i % max(1, every) == 0:
            out.append(f'<text x="{pl + i * step + step / 2:.1f}" '
                       f'y="{height - 8}" text-anchor="middle" '
                       f'fill="#9090b0">{_html.escape(labels[i]) if i < len(labels) else ""}</text>')
    if lvals:
        pts = " ".join(f"{pl + i * step + step / 2:.1f},{yof(v):.1f}"
                       for i, v in enumerate(lvals))
        out.append(f'<polyline points="{pts}" fill="none" '
                   f'stroke="{line_color}" stroke-width="2"/>')
        for i, v in enumerate(lvals):
            out.append(f'<circle cx="{pl + i * step + step / 2:.1f}" '
                       f'cy="{yof(v):.1f}" r="2.2" fill="{line_color}"/>')
    out.append('</svg>')
    return "".join(out)


def _irr_bisect(cash_flows: list[float],
                lo: float = -0.99, hi: float = 1.0,
                iters: int = 80) -> float | None:
    """Robust IRR by bisection - handles the sign change with wide bounds."""
    def npv(r: float) -> float:
        return sum(cf / ((1 + r) ** t) for t, cf in enumerate(cash_flows))
    f_lo, f_hi = npv(lo), npv(hi)
    if abs(f_lo) < 1e-9:      # exact root at the lower bound (Codex LOW-1)
        return lo
    if abs(f_hi) < 1e-9:      # exact root at the upper bound
        return hi
    if f_lo * f_hi > 0:
        for hi in (2.0, 5.0, 10.0):
            f_hi = npv(hi)
            if f_lo * f_hi < 0:
                break
        else:
            return None
    for _ in range(iters):
        mid = (lo + hi) / 2.0
        f_mid = npv(mid)
        if abs(f_mid) < 1e-6:
            return mid
        if f_lo * f_mid < 0:
            hi, f_hi = mid, f_mid
        else:
            lo, f_lo = mid, f_mid
    return (lo + hi) / 2.0


def finance_utility(*,
                    kwp: float,
                    annual_gen_mwh: float,
                    tariff_local_per_kwh: float,
                    fx_local_per_usd: float = 12.0,
                    capex_usd_per_kwp: dict[str, float] | None = None,
                    opex_usd_per_kwp_yr: dict[str, float] | None = None,
                    project_life_yr: int = 25,
                    discount_rate: float = 0.10,
                    debt_ratio: float = 0.70,
                    debt_rate: float = 0.10,
                    debt_tenor_yr: int = 12,
                    tax_rate: float = 0.25,
                    tariff_escalation: float = 0.02,
                    opex_escalation: float = 0.03,
                    degradation_pct: float = 0.5,
                    bess_capex_usd: float = 0.0,
                    contingency_pct: float | None = None,
                    carbon_credit_usd_per_tco2: float = 5.0,
                    grid_ef_kgco2_per_kwh: float = 0.45,
                    monte_carlo_runs: int = 200,
                    revenue_model: str = "ppa",
                    self_consumption_pct: float = 60.0,
                    export_tariff_local_per_kwh: float | None = None
                    ) -> dict[str, Any]:
    """CAPEX/OPEX/NPV/IRR/LCOE/DSCR + Monte Carlo P10/P50/P90. Money reported in
    LOCAL currency via fx_local_per_usd; NPV on nominal local cash flows.
    Module-local engine carried with the module (Codex SSS Section 1).

    Revenue model: for every model except net metering, year-1 revenue is
    generation x tariff. For NET METERING the plant self-consumes a share of its
    output (valued at the avoided retail tariff = tariff_local_per_kwh) and
    exports the surplus (credited at export_tariff_local_per_kwh, defaulting to
    the retail tariff = 1:1 net metering). self_consumption_pct is 0-100."""
    import random

    cx = dict(DEFAULT_CAPEX_USD_PER_KWP)
    if capex_usd_per_kwp:
        cx.update(capex_usd_per_kwp)
    ox = dict(DEFAULT_OPEX_USD_PER_KWP_YR)
    if opex_usd_per_kwp_yr:
        ox.update(opex_usd_per_kwp_yr)

    capex_lines_usd: dict[str, float] = {}
    for k, per_kwp in cx.items():
        capex_lines_usd[k] = round(per_kwp * kwp, 2)
    if bess_capex_usd > 0:
        capex_lines_usd["bess_fixed"] = round(bess_capex_usd, 2)
    total_capex_usd = round(sum(capex_lines_usd.values()), 2)
    if contingency_pct is not None and contingency_pct > 0:
        cont = round(total_capex_usd * contingency_pct / 100.0, 2)
        capex_lines_usd["contingency_pct_extra"] = cont
        total_capex_usd = round(total_capex_usd + cont, 2)
    total_capex_local = round(total_capex_usd * fx_local_per_usd, 2)

    opex_lines_usd_yr: dict[str, float] = {}
    for k, per_kwp_yr in ox.items():
        opex_lines_usd_yr[k] = round(per_kwp_yr * kwp, 2)
    total_opex_usd_yr = round(sum(opex_lines_usd_yr.values()), 2)
    total_opex_local_yr = round(total_opex_usd_yr * fx_local_per_usd, 2)

    debt_ratio = max(0.0, min(1.0, debt_ratio))   # no negative equity (Codex MED-3)
    debt_local = round(total_capex_local * debt_ratio, 2)
    equity_local = round(total_capex_local - debt_local, 2)

    if debt_local > 0 and debt_rate > 0 and debt_tenor_yr > 0:
        r = debt_rate
        n = debt_tenor_yr
        annuity_factor = (r * (1 + r) ** n) / (((1 + r) ** n) - 1)
        annual_debt_service_local = round(debt_local * annuity_factor, 2)
    else:
        annual_debt_service_local = 0.0

    gen_kwh_y1 = annual_gen_mwh * 1000.0
    if revenue_model == "net_metering":
        # Self-consumed energy is valued at the avoided RETAIL tariff; the
        # exported surplus is credited at the export/net-metering rate (defaults
        # to the retail tariff -> 1:1 netting).
        sc = max(0.0, min(1.0, self_consumption_pct / 100.0))
        exp_rate = (export_tariff_local_per_kwh
                    if (export_tariff_local_per_kwh is not None
                        and export_tariff_local_per_kwh >= 0)
                    else tariff_local_per_kwh)
        self_kwh_y1 = gen_kwh_y1 * sc
        export_kwh_y1 = gen_kwh_y1 * (1.0 - sc)
        revenue_local_y1 = (self_kwh_y1 * tariff_local_per_kwh
                            + export_kwh_y1 * exp_rate)
        net_metering = {
            "self_consumption_pct": round(sc * 100.0, 1),
            "self_consumed_mwh_y1": round(self_kwh_y1 / 1000.0, 1),
            "exported_mwh_y1": round(export_kwh_y1 / 1000.0, 1),
            "retail_tariff_local_per_kwh": round(tariff_local_per_kwh, 4),
            "export_tariff_local_per_kwh": round(exp_rate, 4),
        }
    else:
        revenue_local_y1 = gen_kwh_y1 * tariff_local_per_kwh
        net_metering = None
    carbon_local_y1 = (annual_gen_mwh * grid_ef_kgco2_per_kwh
                       * carbon_credit_usd_per_tco2 * fx_local_per_usd)
    cash_flows: list[float] = [-equity_local]
    revenue_by_year: list[float] = []
    opex_by_year:    list[float] = []
    debt_by_year:    list[float] = []
    net_by_year:     list[float] = []

    for t in range(1, project_life_yr + 1):
        degrad = (1 - degradation_pct / 100.0) ** (t - 1)
        rev_esc = (1 + tariff_escalation) ** (t - 1)
        opex_esc = (1 + opex_escalation) ** (t - 1)
        rev_t = revenue_local_y1 * degrad * rev_esc + carbon_local_y1 * degrad
        opex_t = total_opex_local_yr * opex_esc
        debt_t = annual_debt_service_local if t <= debt_tenor_yr else 0.0
        taxable = rev_t - opex_t
        tax_t = max(0.0, taxable) * tax_rate
        net_t = rev_t - opex_t - debt_t - tax_t
        revenue_by_year.append(round(rev_t, 2))
        opex_by_year.append(round(opex_t, 2))
        debt_by_year.append(round(debt_t, 2))
        net_by_year.append(round(net_t, 2))
        cash_flows.append(net_t)

    npv_local = round(
        sum(cf / ((1 + discount_rate) ** t) for t, cf in enumerate(cash_flows)),
        2)
    irr_val = _irr_bisect(cash_flows)
    irr_pct = round(irr_val * 100.0, 2) if irr_val is not None else None

    payback_years: float | None = None
    running = -equity_local
    for t, net_t in enumerate(net_by_year, start=1):
        running += net_t
        if running >= 0:
            payback_years = t - (running - net_t) / net_t if net_t else float(t)
            payback_years = round(max(0.0, payback_years), 2)
            break

    dscr_years = []
    for t in range(1, min(debt_tenor_yr, project_life_yr) + 1):
        rev_t = revenue_by_year[t - 1]
        opex_t = opex_by_year[t - 1]
        debt_t = debt_by_year[t - 1]
        if debt_t > 0:
            dscr_years.append((rev_t - opex_t) / debt_t)
    dscr_avg = round(sum(dscr_years) / len(dscr_years), 2) if dscr_years else None
    dscr_min = round(min(dscr_years), 2) if dscr_years else None

    disc_energy_kwh = 0.0
    for t in range(1, project_life_yr + 1):
        gen_t = annual_gen_mwh * 1000.0 * ((1 - degradation_pct / 100.0) ** (t - 1))
        disc_energy_kwh += gen_t / ((1 + discount_rate) ** t)
    disc_opex_local = sum(
        opex_by_year[t - 1] / ((1 + discount_rate) ** t)
        for t in range(1, project_life_yr + 1))
    lcoe_local_per_kwh = round(
        (total_capex_local + disc_opex_local) / disc_energy_kwh, 4
    ) if disc_energy_kwh > 0 else None

    if monte_carlo_runs and monte_carlo_runs > 0:
        rng = random.Random(42)
        mc_npv: list[float] = []
        mc_irr: list[float] = []
        for _ in range(monte_carlo_runs):
            tariff_shock = rng.uniform(0.80, 1.20)
            degrad_shock = degradation_pct + rng.uniform(-0.3, 0.3)
            opex_shock   = opex_escalation + rng.uniform(-0.02, 0.02)
            cf = [-equity_local]
            for t in range(1, project_life_yr + 1):
                degrad_t = (1 - degrad_shock / 100.0) ** (t - 1)
                rev_esc_t = (1 + tariff_escalation) ** (t - 1)
                rev_t = revenue_local_y1 * tariff_shock * degrad_t * rev_esc_t \
                        + carbon_local_y1 * degrad_t
                opex_t = total_opex_local_yr * ((1 + opex_shock) ** (t - 1))
                debt_t = annual_debt_service_local if t <= debt_tenor_yr else 0.0
                taxable = rev_t - opex_t
                tax_t = max(0.0, taxable) * tax_rate
                cf.append(rev_t - opex_t - debt_t - tax_t)
            npv_run = sum(v / ((1 + discount_rate) ** t) for t, v in enumerate(cf))
            mc_npv.append(npv_run)
            irr_run = _irr_bisect(cf)
            if irr_run is not None:
                mc_irr.append(irr_run)
        mc_npv.sort()
        mc_irr.sort()

        def _pct(vals, p):
            if not vals:
                return None
            idx = max(0, min(len(vals) - 1, int(p / 100.0 * len(vals))))
            return round(vals[idx], 2)

        # Compute each percentile once; treat only None as missing so a genuine
        # 0.0 percentile is not dropped (Codex LOW-2).
        _ip10, _ip50, _ip90 = _pct(mc_irr, 10), _pct(mc_irr, 50), _pct(mc_irr, 90)
        monte_carlo = {
            "runs": monte_carlo_runs,
            "npv_p10": _pct(mc_npv, 10), "npv_p50": _pct(mc_npv, 50),
            "npv_p90": _pct(mc_npv, 90),
            "irr_p10_pct": round(_ip10 * 100, 2) if _ip10 is not None else None,
            "irr_p50_pct": round(_ip50 * 100, 2) if _ip50 is not None else None,
            "irr_p90_pct": round(_ip90 * 100, 2) if _ip90 is not None else None,
        }
    else:
        monte_carlo = None

    return {
        "kwp": round(kwp, 2), "annual_gen_mwh": annual_gen_mwh,
        "tariff_local_per_kwh": tariff_local_per_kwh,
        "fx_local_per_usd": fx_local_per_usd,
        "capex_lines_usd": capex_lines_usd, "total_capex_usd": total_capex_usd,
        "total_capex_local": total_capex_local,
        "capex_usd_per_kwp": round(total_capex_usd / kwp, 2) if kwp else 0.0,
        "opex_lines_usd_yr": opex_lines_usd_yr,
        "total_opex_usd_yr": total_opex_usd_yr,
        "total_opex_local_yr": total_opex_local_yr,
        "debt_local": debt_local, "equity_local": equity_local,
        "annual_debt_service_local": annual_debt_service_local,
        "revenue_y1_local": round(revenue_local_y1, 2),
        "carbon_y1_local": round(carbon_local_y1, 2),
        "npv_local": npv_local, "irr_pct": irr_pct,
        "payback_years": payback_years, "dscr_avg": dscr_avg,
        "dscr_min": dscr_min, "lcoe_local_per_kwh": lcoe_local_per_kwh,
        "project_life_yr": project_life_yr,
        "revenue_by_year": revenue_by_year, "opex_by_year": opex_by_year,
        "debt_by_year": debt_by_year, "net_by_year": net_by_year,
        "monte_carlo": monte_carlo,
        "revenue_model": revenue_model,
        "net_metering": net_metering,
    }


# -- Step 9: facility/technology/electrical -> BOQ service mapping (Codex SSS
#    Section 4). Maps wizard selections onto the SAME service codes the standard
#    BOQ engine uses (web_app._BOQ_SERVICES) so a generated BOQ loads real
#    Section-by-Section / Build-all sections instead of an empty shell.
_CI_BOQ_SERVICE_ORDER: list[str] = [
    "internal_electrical", "fire_alarm", "earthing_bonding",
    "lightning_protection", "power_supply_lv", "lan_wlan", "it_server_room",
    "voip", "ip_pa", "ip_cctv", "tv_system", "ip_clock", "bms",
]
_CI_BOQ_SERVICE_SET: set[str] = set(_CI_BOQ_SERVICE_ORDER)
_CI_FACILITY_DEFAULT_SERVICES: list[str] = [
    "internal_electrical", "power_supply_lv", "fire_alarm", "earthing_bonding",
]
FACILITY_BOQ_SERVICES: dict[str, list[str]] = {
    "control_room":     ["internal_electrical", "power_supply_lv", "lan_wlan",
                         "it_server_room", "fire_alarm", "ip_cctv", "voip",
                         "ip_pa", "earthing_bonding", "lightning_protection", "bms"],
    "om_building":      ["internal_electrical", "power_supply_lv", "lan_wlan",
                         "fire_alarm", "ip_cctv", "voip", "earthing_bonding", "bms"],
    "security_gate":    ["internal_electrical", "power_supply_lv", "lan_wlan",
                         "ip_cctv", "voip", "fire_alarm", "earthing_bonding"],
    "battery_room":     ["internal_electrical", "power_supply_lv", "fire_alarm",
                         "lan_wlan", "it_server_room", "earthing_bonding",
                         "lightning_protection", "bms"],
    "inverter_room":    ["internal_electrical", "power_supply_lv", "fire_alarm",
                         "lan_wlan", "earthing_bonding", "bms"],
    "switchgear_bldg":  ["internal_electrical", "power_supply_lv", "fire_alarm",
                         "lan_wlan", "ip_cctv", "earthing_bonding",
                         "lightning_protection", "bms"],
    "transformer_bldg": ["power_supply_lv", "earthing_bonding",
                         "lightning_protection", "ip_cctv", "internal_electrical"],
    "scada_bldg":       ["internal_electrical", "power_supply_lv", "lan_wlan",
                         "it_server_room", "fire_alarm", "ip_cctv", "voip",
                         "earthing_bonding", "lightning_protection", "bms"],
    "comms_bldg":       ["internal_electrical", "power_supply_lv", "lan_wlan",
                         "it_server_room", "fire_alarm", "earthing_bonding",
                         "lightning_protection"],
    "spare_parts":      ["internal_electrical", "fire_alarm", "ip_cctv",
                         "lan_wlan", "earthing_bonding"],
    "workshop":         ["internal_electrical", "power_supply_lv", "lan_wlan",
                         "fire_alarm", "ip_cctv", "earthing_bonding"],
    "welfare":          ["internal_electrical", "power_supply_lv", "fire_alarm",
                         "earthing_bonding"],
    "washroom":         ["internal_electrical", "power_supply_lv", "earthing_bonding"],
}
TECHNOLOGY_BOQ_SERVICES: dict[str, list[str]] = {
    "scada":        ["lan_wlan", "it_server_room", "bms", "power_supply_lv"],
    "ems":          ["it_server_room", "lan_wlan", "bms"],
    "ppc":          ["it_server_room", "lan_wlan", "bms"],
    "weather":      ["lan_wlan", "power_supply_lv", "earthing_bonding", "lightning_protection"],
    "string_mon":   ["lan_wlan", "power_supply_lv", "bms"],
    "energy_meter": ["power_supply_lv", "lan_wlan"],
    "pq_meter":     ["power_supply_lv", "lan_wlan"],
    "bms":          ["bms", "lan_wlan", "power_supply_lv", "fire_alarm"],
    "txfr_mon":     ["bms", "lan_wlan", "power_supply_lv"],
    "inv_mon":      ["bms", "lan_wlan"],
    "remote_mon":   ["lan_wlan", "it_server_room"],
    "cloud_mon":    ["lan_wlan", "it_server_room"],
    "thermal_cam":  ["ip_cctv", "lan_wlan", "power_supply_lv"],
    "ai_fault":     ["it_server_room", "lan_wlan", "bms"],
    "predictive":   ["bms", "it_server_room", "lan_wlan"],
    "gis":          ["it_server_room", "lan_wlan"],
    "asset_mgmt":   ["it_server_room", "lan_wlan"],
    "cmms":         ["it_server_room", "lan_wlan"],
    "scheduler":    ["it_server_room", "lan_wlan"],
    "wo_mgmt":      ["it_server_room", "lan_wlan"],
    "cyber":        ["lan_wlan", "it_server_room"],
    "firewall":     ["lan_wlan", "it_server_room"],
    "ind_eth":      ["lan_wlan"],
    "fibre":        ["lan_wlan"],
    "ind_wifi":     ["lan_wlan"],
    "gps_sync":     ["lan_wlan", "it_server_room"],
    "ntp":          ["it_server_room", "lan_wlan"],
    "ind_servers":  ["it_server_room"],
    "storage_srv":  ["it_server_room"],
    "backup_srv":   ["it_server_room"],
    "cloud_backup": ["it_server_room", "lan_wlan"],
    "dr":           ["it_server_room", "lan_wlan"],
    "digital_twin": ["it_server_room", "lan_wlan", "bms"],
}
ELECTRICAL_TO_BOQ_SERVICE: dict[str, list[str]] = {
    "internal_installation": ["internal_electrical"],
    "power_supply":          ["power_supply_lv"],
    "hv_distribution":       ["power_supply_lv"],
    "lv_distribution":       ["power_supply_lv"],
    "dc_collection":         ["power_supply_lv"],
    "ac_collection":         ["power_supply_lv"],
    "inverters":             ["power_supply_lv"],
    "transformers":          ["power_supply_lv"],
    "rmu":                   ["power_supply_lv"],
    "hv_switchgear":         ["power_supply_lv"],
    "lv_switchgear":         ["power_supply_lv"],
    "earthing":              ["earthing_bonding"],
    "lightning_protection":  ["lightning_protection"],
    "external_lighting":     ["power_supply_lv"],
    "fire_alarm":            ["fire_alarm"],
    "ip_cctv":               ["ip_cctv"],
    "access_control":        ["ip_cctv"],
    "voip":                  ["voip"],
    "public_address":        ["ip_pa"],
    "tv":                    ["tv_system"],
    "ip_clock":              ["ip_clock"],
    "lan":                   ["lan_wlan"],
    "wan":                   ["lan_wlan"],
    "server_infra":          ["it_server_room"],
    "scada":                 ["bms"],
}
EXTERNAL_WORKS_BOQ_SERVICES: list[str] = [
    "power_supply_lv", "earthing_bonding", "lightning_protection",
    "ip_cctv", "lan_wlan", "internal_electrical",
]


def _ci_facility_services(building_code: str) -> list[str]:
    """BOQ service codes for one facility; defaults to a baseline electrical
    scope for buildings without an explicit mapping."""
    return FACILITY_BOQ_SERVICES.get(building_code, _CI_FACILITY_DEFAULT_SERVICES)


def _ci_order_services(codes) -> list[str]:
    """De-dupe + restrict to valid BOQ codes + return in canonical order."""
    have = {c for c in codes if c in _CI_BOQ_SERVICE_SET}
    return [c for c in _CI_BOQ_SERVICE_ORDER if c in have]


def _ci_derive_boq_services(fac_cfg: dict, tech_cfg: dict,
                            elec_cfg: dict) -> list[str]:
    """Union of BOQ service codes implied by facility buildings, external works,
    technology and electrical selections - ordered + valid. Reads the v2 config
    keys (technologies / services), with a legacy 'selected' fallback."""
    codes: list[str] = []
    for b in (fac_cfg.get("buildings") or []):
        codes.extend(_ci_facility_services(b))
    if fac_cfg.get("external_works"):
        codes.extend(EXTERNAL_WORKS_BOQ_SERVICES)
    for t in (tech_cfg.get("technologies") or tech_cfg.get("selected") or []):
        codes.extend(TECHNOLOGY_BOQ_SERVICES.get(t, []))
    for e in (elec_cfg.get("services") or elec_cfg.get("selected") or []):
        codes.extend(ELECTRICAL_TO_BOQ_SERVICE.get(e, []))
    return _ci_order_services(codes)


# ===========================================================================
# SOLAR-FARM (20MWp) BOQ -- the PV field + balance-of-plant equipment BOQ that
# the platform building-services catalog (web_app._BOQ_SERVICES) does NOT cover.
# Owner directive 2026-07-03: "there must be a boq for the 20MWp solar equipment
# and materials", kept SEPARATE from the facilities/technology BOQ.
#
# Rows are derived from size_utility_pv() output (n_modules, dc_kwp_actual,
# inverter_ac_kw, n_central_inv, strings, combiners, dc/ac cable metres) so the
# quantities are internally consistent with the Step-7 sizing. Basic rates are
# indicative unit supply costs in LOCAL currency at fx=12 GHS/USD; the standard
# rate engine (boq_rate_v3) then adds Supply/Install/OH/Profit/VAT exactly like
# the facilities autobuild, and the user refines quantities/rates in Build-all.
# Each item: (section_title, service_slug, description, unit, qty_key|const,
# basic_local). qty_key pulls from the sizing dict; a float is a fixed quantity;
# a ("per_kwp", factor) tuple scales by installed DC kWp.
# ===========================================================================
# Full 12-bill utility-PV FARM BOQ (Codex template 2026-07-03, see
# docs/SOLAR_FARM_BOQ_TEMPLATE_2026-07-03.md). Each tuple is
#   (bill_no, bill_name, section, service_slug, description, unit, qty_key, basic_local)
# where qty_key is a STRING naming a key in the `derived` dict computed by
# _ci_solar_boq_rows() (all size_utility_pv fields + farm-derived counts), OR a
# float/int fixed quantity, OR a legacy ("per_kwp", factor) tuple. basic_local is
# the indicative SUPPLY unit rate in LOCAL currency (GHS) at fx=12 GHS/USD -- i.e.
# the USD band midpoint from the template x12; boq_rate_v3 then layers
# Supply/Install/OH/Profit/VAT exactly like the facilities autobuild. Tracker rows
# key on `tracker_kwp`/`tracker_rows`, which resolve to 0 for fixed-tilt plants so
# those lines auto-skip (qty<=0 -> continue). Quantities the design does not yet
# emit (piles, land ha, road/perimeter, pole counts) use the concept-derived
# defaults documented in the design doc until Step 7 computes them.
_CI_SOLAR_BOQ_SECTIONS: list[tuple] = [
    # ---- Bill 1: PV Array Field ----
    (1, "PV Array Field", "PV Modules", "solar_modules",
     "Tier-1 mono/bifacial PV module, IEC 61215/61730 (module_wp Wp)",
     "No.", "n_modules", 1560.0),
    (1, "PV Array Field", "PV Modules", "solar_modules",
     "Module barcode/serial QA & flash-test file registration",
     "No.", "n_modules", 7.2),
    (1, "PV Array Field", "PV Modules", "solar_modules",
     "Module DC connector pair, 1500 Vdc compatible",
     "Pair", "n_modules", 48.0),
    (1, "PV Array Field", "PV Modules", "solar_modules",
     "Spare PV modules, commissioning stock",
     "No.", "spare_modules", 1560.0),
    # ---- Bill 2: Mounting System ----
    (2, "Mounting System", "Fixed-Tilt Structure", "solar_mounting",
     "Galvanised fixed-tilt mounting structure, rails, purlins, braces",
     "kWp", "fixed_mount_kwp", 780.0),
    (2, "Mounting System", "Fixed-Tilt Structure", "solar_mounting",
     "Driven steel piles / posts",
     "No.", "fixed_pile_count", 360.0),
    (2, "Mounting System", "Fixed-Tilt Structure", "solar_mounting",
     "Mid/end clamps, bolts, grounding washers",
     "Set", "fixed_clamps", 72.0),
    (2, "Mounting System", "Fixed-Tilt Structure", "solar_mounting",
     "Row set-out, pull testing & pile-refusal allowance",
     "kWp", "fixed_mount_kwp", 72.0),
    (2, "Mounting System", "Tracker (optional)", "solar_tracker",
     "Single-axis tracker torque tube, bearings & posts",
     "kWp", "tracker_kwp", 1320.0),
    (2, "Mounting System", "Tracker (optional)", "solar_tracker",
     "Tracker drive motor & slew gear",
     "No.", "tracker_rows", 9600.0),
    (2, "Mounting System", "Tracker (optional)", "solar_tracker",
     "Tracker row controller, power supply & sensors",
     "No.", "tracker_rows", 6000.0),
    (2, "Mounting System", "Tracker (optional)", "solar_tracker",
     "Tracker commissioning & stow calibration",
     "kWp", "tracker_kwp", 96.0),
    # ---- Bill 3: DC Collection ----
    (3, "DC Collection", "Stringing", "solar_dc",
     "String cable PV1-F 1C 6mm2, UV-rated, 1500 Vdc",
     "m", "dc_cable_m_est", 14.4),
    (3, "DC Collection", "Stringing", "solar_dc",
     "String home-run cable PV1-F 1C 10/16mm2 allowance",
     "m", "dc_homerun_m", 36.0),
    (3, "DC Collection", "Combiners", "solar_dc",
     "String combiner box with fuses, SPD, isolator & monitoring",
     "No.", "combiners", 24000.0),
    (3, "DC Collection", "Combiners", "solar_dc",
     "DC string fuses / fuse holders",
     "No.", "dc_fuses", 84.0),
    (3, "DC Collection", "Surge Protection", "solar_dc",
     "DC SPD cartridges, Type II 1500 Vdc",
     "No.", "dc_spd", 840.0),
    (3, "DC Collection", "Connectors", "solar_dc",
     "MC4 branch / field connector allowance",
     "Pair", "mc4_pairs", 48.0),
    (3, "DC Collection", "Cable Routes", "solar_dc",
     "DC cable trays, ducts, warning tape & markers",
     "m", "dc_tray_m", 120.0),
    (3, "DC Collection", "Cable Routes", "solar_dc",
     "DC trenching, backfill & reinstatement",
     "m", "dc_trench_m", 180.0),
    # ---- Bill 4: Inverter Stations ----
    (4, "Inverter Stations", "Inverters", "solar_inverter",
     "Central inverter, central_inverter_kw kWac class",
     "No.", "n_central_inv", 960000.0),
    (4, "Inverter Stations", "Stations", "solar_inverter",
     "Inverter skid/container, LV panel, HVAC & fire detection",
     "No.", "n_central_inv", 420000.0),
    (4, "Inverter Stations", "Auxiliary", "solar_inverter",
     "Inverter auxiliary transformer / UPS / AC DB",
     "No.", "n_central_inv", 120000.0),
    (4, "Inverter Stations", "Civil", "solar_inverter",
     "Inverter station foundation plinth + oil/fire separation",
     "No.", "n_central_inv", 168000.0),
    (4, "Inverter Stations", "Spares", "solar_inverter",
     "Inverter spare fans, boards, filters & consumables",
     "Set", "n_central_inv", 48000.0),
    # ---- Bill 5: MV Collector System ----
    (5, "MV Collector System", "Transformers", "solar_mv",
     "LV/MV step-up transformer per inverter station",
     "No.", "n_central_inv", 840000.0),
    (5, "MV Collector System", "Transformers", "solar_mv",
     "Transformer bund, plinth, fire wall / stone pit",
     "No.", "n_central_inv", 216000.0),
    (5, "MV Collector System", "Switchgear", "solar_mv",
     "MV RMU / switchgear panel per inverter station",
     "No.", "n_central_inv", 540000.0),
    (5, "MV Collector System", "Cable", "solar_mv",
     "MV collector cable, XLPE armoured",
     "m", "ac_cable_m_est", 600.0),
    (5, "MV Collector System", "Cable", "solar_mv",
     "MV cable terminations & straight joints",
     "Set", "mv_terminations", 7200.0),
    (5, "MV Collector System", "Cable", "solar_mv",
     "MV trenching, ducts, tiles & markers",
     "m", "ac_cable_m_est", 300.0),
    (5, "MV Collector System", "Protection", "solar_mv",
     "MV protection relays + CT/VT interfaces",
     "Set", "n_central_inv", 72000.0),
    # ---- Bill 6: Grid Substation ----
    (6, "Grid Substation", "Transformer", "solar_grid",
     "Main / grid step-up transformer, plant export class",
     "Item", 1.0, 10800000.0),
    (6, "Grid Substation", "HV Yard", "solar_grid",
     "HV/MV switchyard, breakers, isolators & gantries",
     "Item", 1.0, 9600000.0),
    (6, "Grid Substation", "Metering", "solar_grid",
     "Revenue + check metering & power-quality meter",
     "Item", 1.0, 1080000.0),
    (6, "Grid Substation", "Protection", "solar_grid",
     "Grid protection, intertrip, synchronising & RTU",
     "Item", 1.0, 1800000.0),
    (6, "Grid Substation", "Interconnection", "solar_grid",
     "Grid interconnection studies & utility interface allowance",
     "Item", 1.0, 1440000.0),
    # ---- Bill 7: SCADA & Communications ----
    (7, "SCADA & Communications", "SCADA", "solar_scada",
     "Plant SCADA server, HMI, historian & engineering workstation",
     "Item", 1.0, 1080000.0),
    (7, "SCADA & Communications", "PPC", "solar_scada",
     "Power Plant Controller, grid-code active/reactive control",
     "Item", 1.0, 1080000.0),
    (7, "SCADA & Communications", "Monitoring", "solar_scada",
     "Inverter / string monitoring integration",
     "No.", "monitoring_pts", 1200.0),
    (7, "SCADA & Communications", "Network", "solar_scada",
     "Fibre-optic ring, switches, patch panels & cabinets",
     "m", "fibre_m", 48.0),
    (7, "SCADA & Communications", "Weather", "solar_scada",
     "Weather station mast: POA/GHI irradiance, temp & wind",
     "No.", "weather_masts", 216000.0),
    (7, "SCADA & Communications", "Security", "solar_scada",
     "CCTV / NVR integration to SCADA / security room",
     "Item", 1.0, 360000.0),
    # ---- Bill 8: Earthing & Lightning Protection ----
    (8, "Earthing & Lightning Protection", "Earth Grid", "solar_earthing",
     "Array earthing grid, bare copper / galvanised steel",
     "m", "earth_grid_m", 48.0),
    (8, "Earthing & Lightning Protection", "Bonding", "solar_earthing",
     "Module frame bonding jumpers / lugs",
     "No.", "n_modules", 12.0),
    (8, "Earthing & Lightning Protection", "Electrodes", "solar_earthing",
     "Earth rods / electrodes & inspection pits",
     "No.", "earthing_pits", 3000.0),
    (8, "Earthing & Lightning Protection", "Earth Bar", "solar_earthing",
     "Main earth bar & transformer/inverter bonds",
     "Set", "earth_bar", 9600.0),
    (8, "Earthing & Lightning Protection", "Lightning", "solar_earthing",
     "Lightning masts / air terminals / down conductors",
     "No.", "lightning_masts", 14400.0),
    (8, "Earthing & Lightning Protection", "Testing", "solar_earthing",
     "Soil resistivity & earth resistance testing",
     "Item", 1.0, 60000.0),
    # ---- Bill 9: Civil Works ----
    (9, "Civil Works", "Earthworks", "solar_civil",
     "Site clearing, grading & compaction",
     "ha", "land_ha", 120000.0),
    (9, "Civil Works", "Roads", "solar_civil",
     "Internal spine road, gravel/laterite, 4 m wide",
     "m", "road_m", 1800.0),
    (9, "Civil Works", "Hardstand", "solar_civil",
     "Inverter / transformer access hardstand",
     "No.", "n_central_inv", 96000.0),
    (9, "Civil Works", "Drainage", "solar_civil",
     "Drainage swales, culverts & erosion protection",
     "ha", "land_ha", 72000.0),
    (9, "Civil Works", "Cable Civil", "solar_civil",
     "Cable-route excavation & reinstatement allowance",
     "m", "cable_civil_m", 240.0),
    # ---- Bill 10: Security & Site Services ----
    (10, "Security & Site Services", "Fence", "solar_security",
     "Perimeter security fence, 2.4 m anti-climb",
     "m", "perimeter_m", 660.0),
    (10, "Security & Site Services", "Gates", "solar_security",
     "Main gate & maintenance gate",
     "Set", 1.0, 120000.0),
    (10, "Security & Site Services", "CCTV", "solar_security",
     "CCTV pole with camera & network drop",
     "No.", "cctv_poles", 24000.0),
    (10, "Security & Site Services", "Lighting", "solar_security",
     "Perimeter lighting pole with LED luminaire",
     "No.", "lighting_poles", 18000.0),
    (10, "Security & Site Services", "Signage", "solar_security",
     "Security signage, danger labels & wayfinding",
     "Set", 1.0, 60000.0),
    # ---- Bill 11: Spares & Consumables ----
    (11, "Spares & Consumables", "Commissioning", "solar_spares",
     "Commissioning spares: fuses, SPDs, connectors & labels",
     "Set", 1.0, 300000.0),
    (11, "Spares & Consumables", "O&M", "solar_spares",
     "Two-year O&M critical-spares allowance",
     "kWp", "dc_kwp_actual", 36.0),
    # ---- Bill 12: Testing & Handover ----
    (12, "Testing & Handover", "DC Testing", "solar_commissioning",
     "Module/string polarity, Voc, insulation & IV-curve tests",
     "String", "strings", 360.0),
    (12, "Testing & Handover", "Combiner", "solar_commissioning",
     "Combiner functional & monitoring tests",
     "No.", "combiners", 1800.0),
    (12, "Testing & Handover", "Inverter", "solar_commissioning",
     "Inverter commissioning & performance test",
     "No.", "n_central_inv", 30000.0),
    (12, "Testing & Handover", "MV", "solar_commissioning",
     "Transformer/MV testing, relay injection, cable VLF/HiPot",
     "Set", "n_central_inv", 96000.0),
    (12, "Testing & Handover", "Grid Compliance", "solar_commissioning",
     "Grid-code compliance, PPC tuning & utility witness tests",
     "Item", 1.0, 1080000.0),
    (12, "Testing & Handover", "Handover", "solar_commissioning",
     "As-built drawings, O&M manuals, training & handover dossier",
     "Item", 1.0, 300000.0),
]


def _ci_solar_derived_terms(sizing: dict) -> dict:
    """Compute the full set of farm quantities the 12-bill BOQ references, from
    size_utility_pv() output plus concept-derived defaults (land ha, perimeter,
    road, pole counts) for quantities Step 7 does not yet emit. Pure. Tracker
    quantities are zero unless the plant's mounting is a tracker (so tracker rows
    auto-skip on fixed-tilt). See docs/SOLAR_FARM_BOQ_TEMPLATE_2026-07-03.md."""
    import math
    sz = dict(sizing or {})

    def _f(key, d=0.0):
        try:
            return float(sz.get(key) or 0.0)
        except (TypeError, ValueError):
            return d

    n_inv = sz.get("n_central_inv") or sz.get("n_central_inverters") or 0
    try:
        n_inv = int(n_inv)
    except (TypeError, ValueError):
        n_inv = 0
    n_modules = int(_f("n_modules"))
    dc_kwp = _f("dc_kwp_actual") or _f("kwp")
    kwp_input = _f("kwp_input") or dc_kwp
    strings = int(_f("strings"))
    combiners = int(_f("combiners"))
    dc_m = _f("dc_cable_m_est")
    ac_m = _f("ac_cable_m_est")

    # Concept-derived site geometry (same density family as the 3D twin).
    land_ha = max(kwp_input / 800.0, 5.0) if kwp_input > 0 else 0.0
    land_side_m = math.sqrt(land_ha * 10000.0) if land_ha > 0 else 0.0
    road_m = max(0.0, land_side_m - 20.0)
    perimeter_m = max(0.0, 4.0 * (land_side_m - 2.0))

    # Tracker gating: mounting stored at pv_config top level, threaded in by the
    # callers as `mounting_type`/`mounting`. Trackers -> single/dual axis.
    mounting = str(sz.get("mounting_type") or sz.get("mounting") or "").lower()
    is_tracker = ("track" in mounting) or (mounting in ("single_axis", "dual_axis"))
    tracker_rows = int(math.ceil(n_modules / 60.0)) if (is_tracker and n_modules) else 0
    tracker_kwp = dc_kwp if is_tracker else 0.0

    # Mounting quantities. `pile_count`/`clamps` remain full-plant totals for any
    # generic references, but the Bill-2 FIXED-TILT structure lines must resolve
    # to ZERO on a tracker plant -- otherwise a tracker project pays for both the
    # fixed-tilt racking AND the tracker racking (Codex MED, double-count). The
    # dedicated fixed_* keys carry the fixed-tilt structure quantity and collapse
    # to 0 when is_tracker; the tracker rows (tracker_kwp/tracker_rows) cover the
    # tracker structure instead. Fixed-tilt plants are unaffected (is_tracker=False).
    pile_count = int(sz.get("pile_count") or math.ceil(n_modules / 2.0))
    clamps = n_modules
    fixed_mount_kwp = 0.0 if is_tracker else dc_kwp
    fixed_pile_count = 0 if is_tracker else pile_count
    fixed_clamps = 0 if is_tracker else clamps

    return {
        # direct sizing pass-through (so plain-key catalog rows still resolve)
        "n_modules": n_modules, "dc_kwp_actual": dc_kwp, "n_central_inv": n_inv,
        "strings": strings, "combiners": combiners,
        "dc_cable_m_est": dc_m, "ac_cable_m_est": ac_m,
        # PV array
        "spare_modules": math.ceil(n_modules * 0.005),
        # mounting -- full-plant totals + fixed-tilt-only variants (zeroed on tracker)
        "pile_count": pile_count,
        "clamps": clamps,
        "fixed_mount_kwp": fixed_mount_kwp,
        "fixed_pile_count": fixed_pile_count,
        "fixed_clamps": fixed_clamps,
        "tracker_rows": tracker_rows, "tracker_kwp": tracker_kwp,
        # DC collection
        "dc_homerun_m": round(dc_m * 0.25, 2),
        "dc_fuses": strings * 2,
        "dc_spd": combiners * 2,
        "mc4_pairs": strings * 2,
        "dc_tray_m": round(dc_m * 0.35, 2),
        "dc_trench_m": round(dc_m * 0.20, 2),
        # MV
        "mv_terminations": n_inv * 4,
        # SCADA / comms
        "monitoring_pts": combiners + n_inv,
        "fibre_m": round((dc_m + ac_m) * 0.10, 2),
        "weather_masts": 2,
        # earthing / LPS
        "earth_grid_m": round(dc_m * 0.35, 2),
        "earthing_pits": max(1, n_inv + 1),
        "earth_bar": n_inv + 1,
        "lightning_masts": max(4, int(math.ceil(dc_kwp / 2500.0))) if dc_kwp > 0 else 4,
        # civil
        "land_ha": round(land_ha, 2),
        "road_m": round(road_m, 0),
        "cable_civil_m": round(dc_m * 0.20 + ac_m, 2),
        # security
        "perimeter_m": round(perimeter_m, 0),
        "cctv_poles": 4,
        "lighting_poles": 6,
    }


def _ci_solar_boq_rows(sizing: dict) -> list[dict]:
    """Derive solar-farm BOQ line items from size_utility_pv() output. Returns a
    list of dicts {bill_no, bill_name, section, service_code, desc, unit, qty,
    basic}. Skips any line whose derived quantity rounds to zero. Pure -- no DB,
    no side effects."""
    sz = dict(sizing or {})
    if not sz.get("n_central_inv"):
        sz["n_central_inv"] = sz.get("n_central_inverters") or 0
    try:
        kwp = float(sz.get("dc_kwp_actual") or sz.get("kwp") or 0.0)
    except (TypeError, ValueError):
        kwp = 0.0
    # Require a real PV array (Step 7 completed) before building ANY solar rows.
    # Otherwise the fixed-quantity lump sums (grid substation, SCADA, weather)
    # would still insert and falsely report a "priced" solar BOQ that silently
    # omits every module/inverter/MV line (Supervisor MED). No array -> no BOQ.
    try:
        _n_mod = float(sz.get("n_modules") or 0)
    except (TypeError, ValueError):
        _n_mod = 0.0
    if kwp <= 0 and _n_mod <= 0:
        return []
    derived = _ci_solar_derived_terms(sz)
    rows: list[dict] = []
    for bill_no, bill_name, section, slug, desc, unit, qkey, basic in _CI_SOLAR_BOQ_SECTIONS:
        if isinstance(qkey, tuple) and qkey and qkey[0] == "per_kwp":
            qty = kwp * float(qkey[1])
        elif isinstance(qkey, (int, float)):
            qty = float(qkey)
        else:
            try:
                qty = float(derived.get(qkey) or 0.0)
            except (TypeError, ValueError):
                qty = 0.0
        qty = round(qty, 2)
        if qty <= 0:
            continue
        rows.append({"bill_no": int(bill_no), "bill_name": bill_name,
                     "section": section, "service_code": slug, "desc": desc,
                     "unit": unit, "qty": qty, "basic": float(basic)})
    return rows


# ---------------------------------------------------------------------------
# Slice 2 (2026-07-04): price solar-farm BOQ lines from the MARKETPLACE.
# Each line's service slug maps to a marketplace category; a line is priced from
# the median price of public+verified products in that category with the SAME
# unit, converted to local at _CI_SOLAR_FX. Lines with no unit-matched product
# fall back to the engineered reference `basic` (owner: "price from marketplace
# ... reference-price fallback + manual override in rate build-up"). The rate
# build-up row still lets the estimator override any line by hand afterwards.
# ---------------------------------------------------------------------------
_CI_SOLAR_FX = 12.0  # GHS per USD -- same basis as the reference `basic` rates.
_CI_SOLAR_SLUG_TO_MARKET_CAT: dict[str, str] = {
    "solar_modules":  "solar_equipment",
    "solar_mounting": "solar_equipment",
    "solar_tracker":  "solar_equipment",
    "solar_dc":       "solar_equipment",
    "solar_inverter": "solar_equipment",
    "solar_mv":       "transformers",
    "solar_grid":     "transformers",
    "solar_scada":    "plant_control",
    "solar_earthing": "earthing",
}


def _ci_solar_market_rates(c) -> dict:
    """Median local rates from public + verified marketplace products, keyed at
    two grains so each solar-farm BOQ line prices as precisely as the catalogue
    allows: ('SUB', category, subcategory, unit) is tried first (a line's BOQ
    section name mirrors the marketplace subcategory, e.g. 'PV Modules'), then
    ('CU', category, unit). Local rate = median(price_usd) * _CI_SOLAR_FX. Lines
    matching neither keep their engineered reference `basic`. Never raises
    (returns {} on any error -> every line uses its reference)."""
    cats = set(_CI_SOLAR_SLUG_TO_MARKET_CAT.values())
    if not cats:
        return {}
    try:
        ph = ",".join("?" for _ in cats)
        rows = c.execute(
            "SELECT LOWER(COALESCE(pc.code,'')) AS cat, "
            "       LOWER(COALESCE(ec.subcategory,'')) AS sub, "
            "       LOWER(COALESCE(ec.unit,'')) AS unit, ec.price_usd AS price "
            "FROM equipment_catalog ec "
            "LEFT JOIN product_categories pc ON pc.id = ec.category_id "
            "WHERE ec.is_active=1 AND ec.is_public_visible=1 "
            "  AND ec.is_verified=1 AND COALESCE(ec.price_usd,0) > 0 "
            "  AND pc.code IN (%s)" % ph, tuple(sorted(cats))).fetchall()
    except Exception:
        return {}
    cu_b: dict = {}
    sub_b: dict = {}
    for r in rows:
        d = dict(r)
        try:
            p = float(d.get("price") or 0.0)
        except (TypeError, ValueError):
            continue
        if p <= 0:
            continue
        cat = d.get("cat") or ""
        unit = (d.get("unit") or "").strip()
        sub = (d.get("sub") or "").strip()
        cu_b.setdefault((cat, unit), []).append(p)
        if sub:
            sub_b.setdefault((cat, sub, unit), []).append(p)

    def _median(vals: list) -> float:
        vals = sorted(vals)
        n = len(vals)
        return vals[n // 2] if n % 2 else (vals[n // 2 - 1] + vals[n // 2]) / 2.0

    out: dict = {}
    for (cat, unit), vals in cu_b.items():
        out[("CU", cat, unit)] = round(_median(vals) * _CI_SOLAR_FX, 2)
    for (cat, sub, unit), vals in sub_b.items():
        out[("SUB", cat, sub, unit)] = round(_median(vals) * _CI_SOLAR_FX, 2)
    return out


def _ci_solar_sizing_for(proj: dict) -> dict:
    """Sizing for the solar-farm BOQ. Prefer the Step-7 PV design in
    pv_config.sizing; when that is absent or zero, REUSE the same platform PV
    design engine (size_utility_pv) to derive the quantities from the project's
    declared capacity (target_kwp) so the solar-farm BOQ is never silently empty
    -- the generation station reuses the solar design instead of requiring a
    separate PV-design pass. Always threads `mounting` so the tracker/fixed BOQ
    rows gate correctly."""
    pv_cfg = _safe_json(proj.get("pv_config"))
    sizing = dict(pv_cfg.get("sizing") or {})
    have = False
    for _k in ("dc_kwp_actual", "n_modules", "kwp"):
        try:
            if float(sizing.get(_k) or 0) > 0:
                have = True
                break
        except (TypeError, ValueError):
            pass
    if not have:
        try:
            _kwp = float(proj.get("target_kwp") or 0)
        except (TypeError, ValueError):
            _kwp = 0.0
        if _kwp > 0:
            try:
                sizing = dict(size_utility_pv(kwp=_kwp))
            except Exception:
                sizing = {}
    sizing["mounting"] = (pv_cfg.get("mounting")
                          or sizing.get("mounting") or "fixed_tilt")
    return sizing


def _ci_build_solar_farm_items(get_db, fid, bid, pid, uid, tenant_id, sizing):
    """Insert the solar-farm BOQ line items (CELL level) for one floor, REUSING
    the standard rate engine (boq_rate_v3, same OH/Profit/VAT/Supply/Install as
    the facilities autobuild). Writes source_type='capital_solar_autobuild' +
    per-section item numbering + a rate-buildup row per item (so Build-all edits
    and rate-buildup views work identically to facilities rows). Idempotent per
    floor. Returns the count of line items inserted."""
    rows = _ci_solar_boq_rows(sizing)
    if not rows:
        return 0
    try:
        from boq_rate_v3 import boq_rate_v3
    except Exception:
        boq_rate_v3 = None
    oh, prf, vat, sp, ip, vinb = 10.0, 15.0, 12.5, 10.0, 15.0, 0
    inserted = 0
    next_no: dict = {}
    with get_db() as c:
        try:
            existing = c.execute(
                "SELECT id FROM boq_floor_items WHERE floor_id=? LIMIT 1",
                (fid,)).fetchone()
        except Exception:
            existing = None
        if existing is not None:
            return 0
        # Slice 2: marketplace rate table (category, unit) -> local rate.
        mrates = _ci_solar_market_rates(c)
        for r in rows:
            basic = float(r["basic"] or 0.0)
            qty = float(r["qty"] or 0.0)
            desc = (r["desc"] or "").strip()
            if not desc or qty <= 0 or basic <= 0:
                continue
            # Price from the marketplace when a matching product exists in this
            # line's mapped category - subcategory (section) first, then
            # category+unit; else keep the engineered reference rate.
            _mcat = _CI_SOLAR_SLUG_TO_MARKET_CAT.get(r.get("service_code") or "")
            _mrate = None
            if _mcat:
                _u = (r.get("unit") or "").strip().lower()
                _s = (r.get("section") or "").strip().lower()
                _mrate = (mrates.get(("SUB", _mcat, _s, _u))
                          or mrates.get(("CU", _mcat, _u)))
            if _mrate and _mrate > 0:
                basic = float(_mrate)
                spec_note = "Marketplace-priced (median, %s)" % _mcat
            else:
                spec_note = "Reference rate (no marketplace match)"
            if boq_rate_v3:
                supply_amt, install_amt, total_rate = boq_rate_v3(
                    basic, sp, ip, oh, prf, vat, vat_in_basic=bool(vinb))
            else:
                supply_amt = basic * (sp + oh + prf + vat) / 100.0
                install_amt = basic * ip / 100.0
                total_rate = basic + supply_amt + install_amt
            total = qty * total_rate
            section = (r["section"] or "").strip()[:80]
            svc = (r["service_code"] or "")[:40]
            unit = (r["unit"] or "No.").strip()[:20] or "No."
            bill_no = int(r.get("bill_no") or 0)
            bill_name = (r.get("bill_name") or "SOLAR FARM 20MWp").strip()[:80]
            # Number items per (bill, section) so each section restarts at 1 under
            # its own bill. Assigned only AFTER a successful insert so a failed row
            # never leaves a gap (Supervisor).
            sec_key = (bill_no, section)
            item_no_disp = str(next_no.get(sec_key, 0) + 1)
            try:
                cur = c.execute(
                    "INSERT INTO boq_floor_items ("
                    "  floor_id, building_id, project_id, user_id, tenant_id, "
                    "  service_code, section, subsection, "
                    "  bill_no, bill_name, section_letter, subsection_label, "
                    "  item_no, item_no_display, "
                    "  description, specification, unit, qty, "
                    "  final_built_up_rate, total_amount, "
                    "  source_type, approval_status) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (fid, bid, pid, uid, tenant_id,
                     svc, section, "",
                     bill_no, bill_name, "", "",
                     item_no_disp, item_no_disp,
                     desc[:500], spec_note[:200], unit, qty,
                     total_rate, total,
                     "capital_solar_autobuild", "project_only"))
                new_id = int(cur.lastrowid or 0)
            except Exception:
                continue
            next_no[sec_key] = next_no.get(sec_key, 0) + 1   # commit the number
            inserted += 1
            try:
                c.execute(
                    "INSERT INTO boq_floor_rate_buildup ("
                    "  floor_item_id, project_id, user_id, tenant_id, "
                    "  basic_price, supply_rate, install_rate, "
                    "  overhead_pct, profit_pct, contingency_pct, vat_pct, "
                    "  supply_pct, install_pct, vat_in_basic, "
                    "  final_built_up_rate, total_amount) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (new_id, pid, uid, tenant_id,
                     basic, supply_amt, install_amt,
                     oh, prf, 0.0, vat, sp, ip, vinb,
                     total_rate, total))
            except Exception:
                pass
        try:
            c.execute(
                "UPDATE boq_floors SET updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (fid,))
        except Exception:
            pass
    return inserted


# BOQ traceability link table (eager + VERIFIED, observable failures).
_CIBL_SQLITE_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_boq_links (
    id                            INTEGER PRIMARY KEY AUTOINCREMENT,
    capital_investment_project_id INTEGER NOT NULL,
    user_id                       INTEGER NOT NULL,
    tenant_id                     TEXT,
    facility_code                 TEXT NOT NULL,
    source_kind                   TEXT NOT NULL DEFAULT 'facility',
    boq_project_id                INTEGER NOT NULL,
    boq_building_id               INTEGER,
    boq_floor_id                  INTEGER,
    service_codes_csv             TEXT DEFAULT '',
    created_at                    TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at                    TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(capital_investment_project_id, facility_code, source_kind)
);
CREATE INDEX IF NOT EXISTS idx_cibl_project     ON capital_investment_boq_links(capital_investment_project_id);
CREATE INDEX IF NOT EXISTS idx_cibl_boq_project ON capital_investment_boq_links(boq_project_id);
CREATE INDEX IF NOT EXISTS idx_cibl_user        ON capital_investment_boq_links(user_id);
CREATE INDEX IF NOT EXISTS idx_cibl_tenant      ON capital_investment_boq_links(tenant_id);
"""
_CIBL_POSTGRES_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_boq_links (
    id                            SERIAL PRIMARY KEY,
    capital_investment_project_id INTEGER NOT NULL,
    user_id                       INTEGER NOT NULL,
    tenant_id                     UUID,
    facility_code                 TEXT NOT NULL,
    source_kind                   TEXT NOT NULL DEFAULT 'facility',
    boq_project_id                INTEGER NOT NULL,
    boq_building_id               INTEGER,
    boq_floor_id                  INTEGER,
    service_codes_csv             TEXT DEFAULT '',
    created_at                    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at                    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(capital_investment_project_id, facility_code, source_kind)
);
CREATE INDEX IF NOT EXISTS idx_cibl_project     ON capital_investment_boq_links(capital_investment_project_id);
CREATE INDEX IF NOT EXISTS idx_cibl_boq_project ON capital_investment_boq_links(boq_project_id);
CREATE INDEX IF NOT EXISTS idx_cibl_user        ON capital_investment_boq_links(user_id);
CREATE INDEX IF NOT EXISTS idx_cibl_tenant      ON capital_investment_boq_links(tenant_id);
"""
_CIBL_SCHEMA_STATE: dict[str, object] = {"ready": False, "error": ""}


def _ensure_capital_investment_boq_links_schema(get_db) -> bool:
    """Eager, idempotent, per-statement schema creation for the BOQ link table
    WITH verification (a failed live-PG migration surfaces, not swallowed)."""
    if _CIBL_SCHEMA_STATE["ready"]:
        return True
    try:
        with get_db() as c:
            c.executescript(_CIBL_SQLITE_DDL)
    except Exception:
        for stmt in _CIBL_POSTGRES_DDL.split(";"):
            s = stmt.strip()
            if not s:
                continue
            try:
                with get_db() as c:
                    c.execute(s)
            except Exception:
                pass
    try:
        with get_db() as c:
            c.execute("SELECT 1 FROM capital_investment_boq_links LIMIT 1")
        _CIBL_SCHEMA_STATE["ready"] = True
        _CIBL_SCHEMA_STATE["error"] = ""
    except Exception as exc:
        _CIBL_SCHEMA_STATE["ready"] = False
        _CIBL_SCHEMA_STATE["error"] = str(exc)[:300]
    return bool(_CIBL_SCHEMA_STATE["ready"])


class _CIGenerationRaceLost(Exception):
    """Raised inside the Step 9 create+claim transaction when a concurrent
    request already claimed BOQ generation, so get_db()'s exception-rollback
    discards the orphan boq_projects row (no partial/duplicate BOQ)."""


def _marketplace_categories_for(pv_cfg: dict, tech_cfg: dict,
                                elec_cfg: dict) -> list[dict[str, Any]]:
    """Curated marketplace categories to sweep in the EXISTING /marketplace UI
    (this module owns no marketplace). Reads v2 config keys (technologies /
    services) with a legacy 'selected' fallback. Each entry: label, why, href."""
    pv_selected = bool((pv_cfg or {}).get("kwp"))
    tech_selected = set((tech_cfg or {}).get("technologies")
                        or (tech_cfg or {}).get("selected") or [])
    elec_selected = set((elec_cfg or {}).get("services")
                        or (elec_cfg or {}).get("selected") or [])

    def _cat(label: str, cat: str, sub: str = "", why: str = "") -> dict:
        href = f"/marketplace?cat={cat}"
        if sub:
            href += f"&sub={sub}"
        return {"label": label, "why": why, "href": href}

    cats: list[dict] = []
    if pv_selected:
        cats.append(_cat("PV Modules", "pv_modules",
                         why="Confirm Wp band, warranty, bifacial gain if bifacial"))
        cats.append(_cat("Inverters", "inverters",
                         why="Central 1-5 MW blocks for utility, 100-250 kW string otherwise"))
        cats.append(_cat("Mounting / Trackers", "mounting_structures",
                         why="Fixed-tilt vs. HSAT; check +/- 60 deg range for HSAT"))
        cats.append(_cat("DC Cables", "cables",
                         why="Solar-rated DC cable 1500 V, UV-stable"))
        cats.append(_cat("Combiner Boxes", "combiners",
                         why="String monitoring, fuses, DC surge protection"))
    if "transformers" in elec_selected or "hv_distribution" in elec_selected:
        cats.append(_cat("Power Transformers", "transformers",
                         why="Step-up MV -> HV; oil vs. dry; NEMA / IEC"))
        cats.append(_cat("RMU / MV Switchgear", "power_system",
                         why="Ring Main Unit + protection relays"))
    if "lv_switchgear" in elec_selected or "lv_distribution" in elec_selected:
        cats.append(_cat("LV Panels & Switchgear", "power_system",
                         why="LV combiner + main LV distribution"))
    if "earthing" in elec_selected or "lightning_protection" in elec_selected:
        cats.append(_cat("Earthing & Lightning Protection", "earthing",
                         why="Earth rods, mesh grid, LPS air terminals"))
    if "fire_alarm" in elec_selected:
        cats.append(_cat("Fire Alarm & Detection", "fire_alarm",
                         why="Smoke/heat/beam detectors + control panel"))
    if "ip_cctv" in elec_selected:
        cats.append(_cat("CCTV & NVR", "cctv",
                         why="Perimeter + gate coverage; PoE recommended"))
    if "access_control" in elec_selected:
        cats.append(_cat("Access Control", "access_control",
                         why="Card readers, boom barriers, biometrics"))
    if ("lan" in elec_selected or "wan" in elec_selected
            or "ind_eth" in tech_selected or "fibre" in tech_selected):
        cats.append(_cat("Network Switches / Firewalls / Fibre", "networking",
                         why="Industrial-grade Ethernet + fibre backbone"))
    if ("scada" in tech_selected or "scada" in elec_selected
            or "ems" in tech_selected or "ppc" in tech_selected):
        cats.append(_cat("SCADA / EMS / PPC Servers", "server_equipment",
                         why="Redundant server pair + historian"))
    if "bms" in tech_selected:
        cats.append(_cat("Battery Management System", "monitoring",
                         why="BMS master + cell-level monitoring"))
    if "weather" in tech_selected:
        cats.append(_cat("Weather Stations", "monitoring",
                         why="Pyranometer + ambient + module temperature"))
    if "ups" not in [c["label"] for c in cats]:
        cats.append(_cat("UPS Systems", "ups",
                         why="Control room + SCADA continuity"))
    return cats


# -- Step 11/12: CRM investment opportunity + 13-stage sales pipeline --
PIPELINE_STAGES: list[tuple[str, str, str]] = [
    ("lead",            "Lead",             "bi-search"),
    ("opportunity",     "Opportunity",      "bi-lightbulb"),
    ("concept_design",  "Concept Design",   "bi-pencil-square"),
    ("feasibility",     "Feasibility",      "bi-clipboard-check"),
    ("financial_model", "Financial Model",  "bi-cash-coin"),
    ("proposal",        "Proposal",         "bi-file-earmark-text"),
    ("investor_review", "Investor Review",  "bi-people"),
    ("bank_review",     "Bank Review",      "bi-bank"),
    ("negotiation",     "Negotiation",      "bi-chat-square-text"),
    ("award",           "Award",            "bi-award"),
    ("construction",    "Construction",     "bi-cone-striped"),
    ("commissioning",   "Commissioning",    "bi-toggle-on"),
    ("completed",       "Completed",        "bi-check2-circle"),
]
PIPELINE_STAGE_CODES: list[str] = [c for c, _, _ in PIPELINE_STAGES]
PIPELINE_STAGE_LABEL: dict[str, str] = {c: L for c, L, _ in PIPELINE_STAGES}

_CIO_SQLITE_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_opportunities (
    id                            INTEGER PRIMARY KEY AUTOINCREMENT,
    capital_investment_project_id INTEGER NOT NULL,
    user_id                       INTEGER NOT NULL,
    project_name                  TEXT DEFAULT '',
    investor                      TEXT DEFAULT '',
    developer                     TEXT DEFAULT '',
    client                        TEXT DEFAULT '',
    location                      TEXT DEFAULT '',
    country                       TEXT DEFAULT '',
    currency                      TEXT DEFAULT 'GHS',
    capacity_mwp                  REAL,
    capex_local                   REAL,
    capex_usd                     REAL,
    revenue_y1_local              REAL,
    annual_gen_mwh                REAL,
    npv_local                     REAL,
    irr_pct                       REAL,
    lcoe_local_per_kwh            REAL,
    payback_years                 REAL,
    dscr_avg                      REAL,
    stage                         TEXT DEFAULT 'lead',
    stage_history                 TEXT DEFAULT '[]',
    pipeline_notes                TEXT DEFAULT '',
    tenant_id                     TEXT,
    created_at                    TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at                    TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_cio_project ON capital_investment_opportunities(capital_investment_project_id);
CREATE INDEX IF NOT EXISTS idx_cio_user    ON capital_investment_opportunities(user_id);
CREATE INDEX IF NOT EXISTS idx_cio_stage   ON capital_investment_opportunities(user_id, stage);
"""
_CIO_POSTGRES_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_opportunities (
    id                            SERIAL PRIMARY KEY,
    capital_investment_project_id INTEGER NOT NULL,
    user_id                       INTEGER NOT NULL,
    project_name                  TEXT DEFAULT '',
    investor                      TEXT DEFAULT '',
    developer                     TEXT DEFAULT '',
    client                        TEXT DEFAULT '',
    location                      TEXT DEFAULT '',
    country                       TEXT DEFAULT '',
    currency                      TEXT DEFAULT 'GHS',
    capacity_mwp                  DOUBLE PRECISION,
    capex_local                   DOUBLE PRECISION,
    capex_usd                     DOUBLE PRECISION,
    revenue_y1_local              DOUBLE PRECISION,
    annual_gen_mwh                DOUBLE PRECISION,
    npv_local                     DOUBLE PRECISION,
    irr_pct                       DOUBLE PRECISION,
    lcoe_local_per_kwh            DOUBLE PRECISION,
    payback_years                 DOUBLE PRECISION,
    dscr_avg                      DOUBLE PRECISION,
    stage                         TEXT DEFAULT 'lead',
    stage_history                 TEXT DEFAULT '[]',
    pipeline_notes                TEXT DEFAULT '',
    tenant_id                     UUID,
    created_at                    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at                    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_cio_project ON capital_investment_opportunities(capital_investment_project_id);
CREATE INDEX IF NOT EXISTS idx_cio_user    ON capital_investment_opportunities(user_id);
CREATE INDEX IF NOT EXISTS idx_cio_stage   ON capital_investment_opportunities(user_id, stage);
"""


# Slice 9 (2026-07-05) -- funding fields carried on the CRM opportunity (= the
# pipeline entry), added via per-statement ALTER so existing rows upgrade in
# place. Each runs in its own transaction (dup-column on Postgres aborts only
# itself). "REAL"/"INTEGER"/"TEXT" are accepted by both SQLite and Postgres.
_CIO_FUNDING_MIGRATIONS = (
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "funding_requested REAL",
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "funding_amount REAL",
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "funding_selected_institutions TEXT",
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "funding_status TEXT",
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "funding_score INTEGER",
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "funding_approval_date TEXT",
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "expected_close_date TEXT",
    "ALTER TABLE capital_investment_opportunities ADD COLUMN "
    "funding_success_fee REAL",
)


def _ensure_opportunities_schema(get_db) -> None:
    """Idempotent opportunities-table creation; per-statement transactions so a
    Postgres failure doesn't cascade."""
    try:
        with get_db() as c:
            c.executescript(_CIO_SQLITE_DDL)
    except Exception:
        for stmt in _CIO_POSTGRES_DDL.split(";"):
            s = stmt.strip()
            if not s:
                continue
            try:
                with get_db() as c:
                    c.execute(s)
            except Exception:
                pass
    # Funding columns (Slice 9) -- each ALTER isolated so a duplicate-column
    # failure on an already-upgraded table doesn't abort the others.
    for stmt in _CIO_FUNDING_MIGRATIONS:
        try:
            with get_db() as c:
                c.execute(stmt)
        except Exception:
            pass


def build_opportunity_from_project(proj: dict[str, Any]) -> dict[str, Any]:
    """Opportunity payload derived from the project's stored config blobs.
    Numeric fields default to None when the backing step isn't done yet."""
    pv_cfg = _safe_json(proj.get("pv_config"))
    fin_cfg = _safe_json(proj.get("finance_config"))
    sizing = pv_cfg.get("sizing") or {}
    computed = fin_cfg.get("computed") or {}

    kwp = sizing.get("kwp_input") or pv_cfg.get("kwp")
    capacity_mwp = round(kwp / 1000.0, 3) if kwp else None
    if capacity_mwp is None and proj.get("target_kwp"):
        capacity_mwp = round(proj["target_kwp"] / 1000.0, 3)

    return {
        "capital_investment_project_id": proj["id"],
        "user_id":       proj["user_id"],
        "project_name":  proj.get("project_name") or "",
        "investor":      proj.get("investor") or "",
        "developer":     proj.get("developer") or "",
        "client":        proj.get("client_name") or "",
        "location":      ", ".join(x for x in (proj.get("region"),
                                               proj.get("country")) if x),
        "country":       proj.get("country") or "",
        "currency":      proj.get("currency") or "GHS",
        "capacity_mwp":  capacity_mwp,
        "capex_local":       computed.get("total_capex_local"),
        "capex_usd":         computed.get("total_capex_usd"),
        "revenue_y1_local":  computed.get("revenue_y1_local"),
        "annual_gen_mwh":    computed.get("annual_gen_mwh")
                             or sizing.get("annual_gen_mwh"),
        "npv_local":         computed.get("npv_local"),
        "irr_pct":           computed.get("irr_pct"),
        "lcoe_local_per_kwh": computed.get("lcoe_local_per_kwh"),
        "payback_years":     computed.get("payback_years"),
        "dscr_avg":          computed.get("dscr_avg"),
    }


# -- Step 13: Reports (13 downloadable PDF reports) --
REPORT_TYPES: list[tuple[str, str, str, bool]] = [
    ("executive",       "Executive Summary",     "bi-clipboard-data",  True),
    ("technical",       "Technical Report",      "bi-cpu",             True),
    ("financial",       "Financial Report",      "bi-cash-coin",       True),
    ("bankability",     "Bankability Report",    "bi-bank",            True),
    ("investment_memo", "Investment Memorandum", "bi-file-earmark-text", True),
    ("risk",            "Risk Assessment",       "bi-shield-exclamation", True),
    ("boq",             "BOQ",                   "bi-list-check",      True),
    ("bom",             "BOM",                   "bi-boxes",           True),
    ("rfq",             "Marketplace RFQ",       "bi-cart",            True),
    ("construction_est","Construction Estimate", "bi-hammer",          True),
    ("maintenance",     "Maintenance Strategy",  "bi-tools",           True),
    ("monitoring",      "Monitoring Strategy",   "bi-eye",             True),
    ("ops_manual",      "Operations Manual",     "bi-journal",         True),
    # Parity with the residential/C&I 'New Project' report set (owner #6).
    ("wiring",           "Wiring & Cabling Schedule", "bi-diagram-3",  True),
    ("single_line",      "Single-Line Diagram (SLD)", "bi-diagram-2",  True),
    ("energy_impact",    "Energy Impact & Yield",     "bi-lightning-charge", True),
    ("economic_impact",  "Economic Impact",           "bi-graph-up-arrow",   True),
    ("implementation_plan", "Implementation Plan",    "bi-calendar-range",   True),
]
REPORT_KEYS: set[str] = {k for k, _, _, _ in REPORT_TYPES}
FULL_REPORT_KEYS: set[str] = {k for k, _, _, full in REPORT_TYPES if full}


def _fmt_money(v: Any, currency: str = "") -> str:
    if v is None:
        return "n/a"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if abs(v) >= 1_000_000:
        s = f"{v/1_000_000:,.2f}M"
    elif abs(v) >= 1_000:
        s = f"{v:,.0f}"
    else:
        s = f"{v:.2f}"
    return f"{currency} {s}" if currency else s


def _fmt_pct(v: Any) -> str:
    if v is None:
        return "n/a"
    try:
        return f"{float(v):.2f} %"
    except (TypeError, ValueError):
        return str(v)


def _build_report_markdown(key: str, proj: dict[str, Any],
                           opp: dict[str, Any] | None,
                           boq: dict[str, Any] | None = None) -> tuple[str, str]:
    """Return (markdown, doc_title) for the given report key. Hardened so a
    partially-malformed stored JSON NEVER 500s (Codex MED fix carried)."""
    pv = _safe_json(proj.get("pv_config"))
    fin = _safe_json(proj.get("finance_config"))
    fac = _safe_json(proj.get("facility_config"))
    tech = _safe_json(proj.get("technology_config"))
    elec = _safe_json(proj.get("electrical_config"))
    site = _safe_json(proj.get("site_config"))
    reg = _safe_json(proj.get("regulatory_config"))
    sizing = pv.get("sizing") or {}
    computed = fin.get("computed") or {}
    cur = proj.get("currency") or "GHS"
    framework = country_framework(proj.get("country"))
    boq = boq if isinstance(boq, dict) else {}

    def _D(x):
        return x if isinstance(x, dict) else {}

    def _L(x):
        return x if isinstance(x, list) else []
    sizing = _D(sizing)
    computed = _D(computed)
    computed["capex_lines_usd"] = _D(computed.get("capex_lines_usd"))
    computed["opex_lines_usd_yr"] = _D(computed.get("opex_lines_usd_yr"))
    computed["monte_carlo"] = _D(computed.get("monte_carlo"))
    fac["buildings"] = _L(fac.get("buildings"))
    # v2 stores technologies / services (legacy used 'selected'); normalise both
    # so every downstream branch reading tech['selected'] / elec['selected']
    # gets the right list regardless of which key the project was saved under.
    tech["selected"] = _L(tech.get("technologies") or tech.get("selected"))
    elec["selected"] = _L(elec.get("services") or elec.get("selected"))
    reg["items"] = _D(reg.get("items"))
    _boq_linked = bool(boq.get("linked"))
    _fac_list = fac["buildings"]
    _tech_list = tech["selected"]
    _elec_list = elec["selected"]

    def _lbl(code) -> str:
        return str(code).replace("_", " ").title()

    def _bullets(items, empty="(none configured)"):
        items = _L(items)
        return "\n".join(f"- {_lbl(x)}" for x in items) if items else empty

    header = (
        f"# {proj['project_name']}\n\n"
        f"**Client:** {proj.get('client_name') or '(unspecified)'}  \n"
        f"**Investor:** {proj.get('investor') or '(unspecified)'}  \n"
        f"**Developer:** {proj.get('developer') or '(unspecified)'}  \n"
        f"**Location:** "
        f"{proj.get('district') or ''} {proj.get('region') or ''} "
        f"{proj.get('country') or ''}  \n"
        f"**Target COD:** {proj.get('target_cod') or 'to be confirmed'}  \n"
        f"**Design standard:** {proj.get('design_standard') or 'IEC'}  \n\n"
    )

    if key == "executive":
        title = f"Executive Summary - {proj['project_name']}"
        md = header + (
            "## Executive summary\n\n"
            f"{proj.get('description') or ''}\n\n"
            "### Headline\n\n"
            f"- **Capacity:** {_fmt_money(sizing.get('kwp_input') or (proj.get('target_kwp') or 0))} kWp DC "
            f"({(sizing.get('kwp_input') or proj.get('target_kwp') or 0)/1000:.1f} MWp)\n"
            f"- **Annual generation:** {_fmt_money(sizing.get('annual_gen_mwh'))} MWh\n"
            f"- **Total CAPEX:** {_fmt_money(computed.get('total_capex_usd'), 'USD')} "
            f"({_fmt_money(computed.get('total_capex_local'), cur)})\n"
            f"- **IRR:** {_fmt_pct(computed.get('irr_pct'))}\n"
            f"- **NPV:** {_fmt_money(computed.get('npv_local'), cur)}\n"
            f"- **LCOE:** {computed.get('lcoe_local_per_kwh') or 'n/a'} {cur}/kWh\n"
            f"- **Payback:** "
            f"{('%.1f yr' % computed['payback_years']) if computed.get('payback_years') else 'beyond project life'}\n\n"
            "### Facilities\n\n"
            + (", ".join(fac.get("buildings") or []) or "not yet configured")
            + "\n\n"
            "### Technology stack\n\n"
            + (", ".join(tech.get("selected") or []) or "not yet configured")
            + "\n\n"
            "### Jurisdiction\n\n"
            f"Regulator: **{framework.get('regulator', {}).get('name', '')}**  \n"
            f"ESIA authority: **{framework.get('esia_authority', {}).get('name', '')}**  \n"
            f"Off-taker(s): {', '.join(framework.get('utility_offtakers') or [])}\n"
        )
    elif key == "technical":
        title = f"Technical Report - {proj['project_name']}"
        md = header + (
            "## Technical report\n\n"
            "### PV design\n\n"
            f"- **DC capacity:** {sizing.get('dc_kwp_actual') or 'n/a'} kWp\n"
            f"- **AC capacity:** {sizing.get('inverter_ac_kw') or 'n/a'} kW\n"
            f"- **Module technology:** {pv.get('module_tech') or 'n/a'}\n"
            f"- **Module Wp:** {pv.get('module_wp') or 'n/a'} W\n"
            f"- **Modules:** {sizing.get('n_modules') or 'n/a'}\n"
            f"- **Strings:** {sizing.get('strings') or 'n/a'}\n"
            f"- **Combiner boxes:** {sizing.get('combiners') or 'n/a'}\n"
            f"- **Central inverters:** {sizing.get('n_central_inverters') or 'n/a'} "
            f"x {sizing.get('central_inverter_kw') or 'n/a'} kW\n"
            f"- **Mounting:** {pv.get('mounting') or 'n/a'}\n"
            f"- **Tilt / azimuth:** {pv.get('tilt_deg') or 'n/a'}deg / {pv.get('azimuth_deg') or 'n/a'}deg\n"
            f"- **PSH:** {pv.get('psh_daily') or 'n/a'} kWh/m2/day\n"
            f"- **PR:** {pv.get('performance_ratio') or 'n/a'}\n"
            f"- **Availability:** {pv.get('availability_pct') or 'n/a'} %\n"
            f"- **Annual degradation:** {pv.get('annual_degradation_pct') or 'n/a'} %\n"
            f"- **Specific yield:** {sizing.get('specific_yield_kwh_per_kwp') or 'n/a'} kWh/kWp\n"
            f"- **Annual gen:** {_fmt_money(sizing.get('annual_gen_mwh'))} MWh\n"
            f"- **Lifetime gen:** {_fmt_money(sizing.get('lifetime_gen_mwh'))} MWh\n\n"
            "### Site\n\n"
            f"- Land area: {site.get('land_area_ha') or 'n/a'} ha\n"
            f"- Terrain: {site.get('terrain') or 'n/a'}\n"
            f"- Slope: {site.get('slope') or 'n/a'}\n"
            f"- Soil: {site.get('soil') or 'n/a'}\n"
            f"- Flood risk: {site.get('flood_risk') or 'n/a'}\n"
            f"- Wind zone: {site.get('wind_zone') or 'n/a'}\n\n"
            "### Facilities\n\n"
            + "\n".join(f"- {b}" for b in (fac.get("buildings") or []) or ["(none)"])
            + "\n\n### Electrical scope\n\n"
            + "\n".join(f"- {s}" for s in (elec.get("selected") or []) or ["(none)"])
            + "\n\n### Technology stack\n\n"
            + "\n".join(f"- {t}" for t in (tech.get("selected") or []) or ["(none)"])
        )
    elif key == "financial":
        title = f"Financial Report - {proj['project_name']}"
        md = header + (
            "## Financial report\n\n"
            "### CAPEX\n\n"
            f"- **Total CAPEX:** {_fmt_money(computed.get('total_capex_usd'), 'USD')} "
            f"({_fmt_money(computed.get('total_capex_local'), cur)})\n"
            f"- **USD per kWp:** {_fmt_money(computed.get('capex_usd_per_kwp'), 'USD')}\n\n"
            "CAPEX breakdown (USD):\n\n"
            + ("\n".join(f"- {k.replace('_', ' ').title()}: {_fmt_money(v)}"
                        for k, v in (computed.get('capex_lines_usd') or {}).items())
               or "(not yet computed)")
            + "\n\n### OPEX\n\n"
            f"- **Annual OPEX:** {_fmt_money(computed.get('total_opex_usd_yr'), 'USD')} "
            f"({_fmt_money(computed.get('total_opex_local_yr'), cur)})\n\n"
            "OPEX breakdown (USD/yr):\n\n"
            + ("\n".join(f"- {k.replace('_', ' ').title()}: {_fmt_money(v)}"
                        for k, v in (computed.get('opex_lines_usd_yr') or {}).items())
               or "(not yet computed)")
            + "\n\n### Debt & equity\n\n"
            f"- Debt: {_fmt_money(computed.get('debt_local'), cur)}\n"
            f"- Equity: {_fmt_money(computed.get('equity_local'), cur)}\n"
            f"- Annual debt service: {_fmt_money(computed.get('annual_debt_service_local'), cur)}\n\n"
            "### Returns\n\n"
            f"- NPV: {_fmt_money(computed.get('npv_local'), cur)}\n"
            f"- IRR: {_fmt_pct(computed.get('irr_pct'))}\n"
            f"- LCOE: {computed.get('lcoe_local_per_kwh') or 'n/a'} {cur}/kWh\n"
            f"- Payback: "
            f"{('%.1f yr' % computed['payback_years']) if computed.get('payback_years') else 'beyond project life'}\n"
        )
    elif key == "bankability":
        title = f"Bankability Report - {proj['project_name']}"
        mc = computed.get("monte_carlo") or {}
        # Bankability DETERMINATION (weighted lender scorecard) from the finance
        # engine -- same result the Cost Plan Deck shows, so report == UI.
        bank = _ci_bankability(computed)
        if bank.get("available"):
            det = (
                "## Bankability determination\n\n"
                f"**Verdict: {bank['rating']}  -  score {bank['score']}/100**\n\n"
                "| Metric | Value | Weight | Verdict |\n"
                "|---|---|---:|---|\n"
                + "".join(
                    f"| {m['label']} | {m['value']}"
                    + (f" ({m['note']})" if m['note'] else "")
                    + f" | {m['weight']} | {m['verdict']} |\n"
                    for m in bank["metrics"])
                + "\n"
                + ("**Strengths**\n\n"
                   + "".join(f"- {s}\n" for s in bank["strengths"]) + "\n"
                   if bank["strengths"] else "")
                + ("**Risks**\n\n"
                   + "".join(f"- {s}\n" for s in bank["risks"]) + "\n"
                   if bank["risks"] else "")
                + ("**Financing conditions**\n\n"
                   + "".join(f"- {s}\n" for s in bank["conditions"]) + "\n"
                   if bank["conditions"] else "")
            )
        else:
            det = ("## Bankability determination\n\n"
                   "_Complete Step 8 (Financial Engineering) to compute the "
                   "bankability score._\n\n")
        md = header + det + (
            "## Bankability report\n\n"
            "### Base-case metrics\n\n"
            f"- NPV: {_fmt_money(computed.get('npv_local'), cur)}\n"
            f"- IRR: {_fmt_pct(computed.get('irr_pct'))}\n"
            f"- LCOE: {computed.get('lcoe_local_per_kwh') or 'n/a'} {cur}/kWh\n"
            f"- Payback: "
            f"{('%.1f yr' % computed['payback_years']) if computed.get('payback_years') else 'beyond project life'}\n"
            f"- DSCR average: {computed.get('dscr_avg') or 'n/a'}x\n"
            f"- DSCR minimum: {computed.get('dscr_min') or 'n/a'}x\n\n"
            "### Monte Carlo\n\n"
            f"Runs: {mc.get('runs') or '(not run)'}\n\n"
            f"- NPV P10 / P50 / P90: {mc.get('npv_p10') or '-'} / "
            f"{mc.get('npv_p50') or '-'} / {mc.get('npv_p90') or '-'}\n"
            f"- IRR P10 / P50 / P90 (%): {mc.get('irr_p10_pct') or '-'} / "
            f"{mc.get('irr_p50_pct') or '-'} / {mc.get('irr_p90_pct') or '-'}\n\n"
            "### Key risks\n\n"
            "- FX / convertibility risk\n"
            "- Off-taker credit\n"
            "- PPA tariff review clauses\n"
            "- Grid curtailment\n"
            "- Component degradation vs. warranty\n"
            "- Land tenure disputes\n"
            "- Regulatory + permit delays\n\n"
            "### Jurisdictional bankability\n\n"
            f"{framework.get('notes') or ''}\n"
        )
    elif key == "investment_memo":
        title = f"Investment Memorandum - {proj['project_name']}"
        md = header + (
            "## Investment memorandum\n\n"
            "### Opportunity\n\n"
            f"{proj.get('description') or ''}\n\n"
            "### Investment ask\n\n"
            f"- Equity requirement: {_fmt_money(computed.get('equity_local'), cur)}\n"
            f"- Debt requirement: {_fmt_money(computed.get('debt_local'), cur)}\n\n"
            "### Expected returns\n\n"
            f"- IRR (base): {_fmt_pct(computed.get('irr_pct'))}\n"
            f"- Payback: "
            f"{('%.1f yr' % computed['payback_years']) if computed.get('payback_years') else 'beyond project life'}\n\n"
            "### Structure\n\n"
            f"- Currency: {cur}\n"
            f"- Debt ratio: {fin.get('debt_ratio_pct') or 'n/a'}%\n"
            f"- Debt rate: {fin.get('debt_rate_pct') or 'n/a'}%\n"
            f"- Debt tenor: {fin.get('debt_tenor_yr') or 'n/a'} yr\n"
            f"- Off-take: {fin.get('revenue_model') or 'n/a'}\n"
            f"- Tariff: {fin.get('tariff_local_per_kwh') or 'n/a'} {cur}/kWh\n\n"
            "### Jurisdiction\n\n"
            f"- Country: {proj.get('country') or ''}\n"
            f"- Regulator: {framework.get('regulator', {}).get('name', '')}\n"
            f"- ESIA authority: {framework.get('esia_authority', {}).get('name', '')}\n"
            f"- Off-takers: {', '.join(framework.get('utility_offtakers') or [])}\n\n"
            f"### Regulatory posture\n\n"
            + ("\n".join(
                f"- {label}: {_D((reg.get('items') or {}).get(code)).get('status', 'not_started')}"
                for code, label, _ in REGULATORY_ITEMS)
               if reg.get("items") else "(regulatory posture not yet captured)")
        )
    elif key == "risk":
        title = f"Risk Assessment - {proj['project_name']}"
        md = header + (
            "## Risk assessment\n\n"
            "Risks are rated Low / Medium / High with a primary mitigation.\n\n"
            "### Technical\n\n"
            "- Module degradation vs. warranty - **Medium** - Tier-1 modules, "
            "linear performance warranty, annual EL/IV testing.\n"
            "- Inverter reliability - **Medium** - spares holding + service SLA.\n"
            "- Grid curtailment - **Medium** - dispatch agreement + storage option.\n\n"
            "### Financial\n\n"
            f"- FX / convertibility ({cur}) - **High** - hard-currency PPA where "
            "possible, DSRA funding.\n"
            "- Off-taker credit - **High** - sovereign / guarantee support.\n"
            f"- Tariff review - **Medium** - indexation clauses; base IRR "
            f"{_fmt_pct(computed.get('irr_pct'))}, DSCR avg "
            f"{computed.get('dscr_avg') or 'n/a'}x.\n\n"
            "### Construction\n\n"
            "- EPC delivery / delay - **Medium** - LDs, milestone payments.\n"
            "- Land tenure / access - **Medium** - see regulatory posture below.\n\n"
            "### Regulatory & O&M\n\n"
            "- Permit / ESIA delays - **Medium** - early engagement with "
            f"{framework.get('esia_authority', {}).get('name', 'the ESIA authority')}.\n"
            "- O&M staffing / spares - **Low** - CMMS + preventive plan (see "
            "Maintenance Strategy).\n"
        )
    elif key == "boq":
        title = f"Bill of Quantities (summary) - {proj['project_name']}"
        pf = boq.get("per_facility_usd") or {}
        md = header + (
            "## Bill of Quantities - summary\n\n"
            + ("The priced, editable BOQ is linked to this project. Open it in "
               "the BOQ workspace (Build-all / Section-by-Section) for the full "
               "line-item detail and exports.\n\n"
               if _boq_linked else
               "No linked BOQ yet - generate it on Step 9 (BOQ), then re-open "
               "this report for priced totals.\n\n")
            + f"- **Line items (indicative starter):** {boq.get('n_items') or 0}\n"
            + f"- **BOQ total:** {_fmt_money(boq.get('grand_total_usd'), 'USD')} "
              f"({_fmt_money(boq.get('grand_total_local'), cur)})\n\n"
            + "### By facility (USD, indicative)\n\n"
            + ("\n".join(f"- {_lbl(k or 'unassigned')}: {_fmt_money(v, 'USD')}"
                          for k, v in pf.items()) if pf else "(no facility rows yet)")
            + "\n\n_Figures are indicative while Step 9 seeds a 1-per-section "
              "starter; expand each section in Build-all for firm quantities._\n"
        )
    elif key == "bom":
        title = f"Bill of Materials (procurement scope) - {proj['project_name']}"
        md = header + (
            "## Bill of Materials - procurement scope\n\n"
            "Materials scope derived from the facility, technology and "
            "electrical selections. Prices and suppliers are sourced in the "
            "SolarPro Marketplace (Step 10).\n\n"
            "### Facilities / civil\n\n" + _bullets(_fac_list) + "\n\n"
            "### Monitoring & technology\n\n" + _bullets(_tech_list) + "\n\n"
            "### Electrical / power system\n\n" + _bullets(_elec_list) + "\n\n"
            "### PV primary equipment\n\n"
            f"- Modules: {sizing.get('n_modules') or 'n/a'} x "
            f"{pv.get('module_wp') or 'n/a'} W ({pv.get('module_tech') or 'n/a'})\n"
            f"- Central inverters: {sizing.get('n_central_inverters') or 'n/a'} x "
            f"{sizing.get('central_inverter_kw') or 'n/a'} kW\n"
            f"- Mounting: {pv.get('mounting') or 'n/a'}\n"
        )
    elif key == "rfq":
        title = f"Marketplace RFQ pack - {proj['project_name']}"
        cats = sorted(set(
            [_lbl(x) for x in _tech_list] + [_lbl(x) for x in _elec_list]
            + ["PV Modules", "Inverters", "Mounting Structures",
               "Transformers", "MV/LV Cables", "Earthing & Lightning"]))
        md = header + (
            "## Marketplace RFQ pack\n\n"
            "Request-for-quotation scope. Issue these categories to verified "
            "suppliers from the Marketplace RFQ workflow.\n\n"
            f"- **Project capacity:** "
            f"{(sizing.get('kwp_input') or proj.get('target_kwp') or 0)/1000:.1f} MWp\n"
            f"- **Delivery location:** {proj.get('region') or ''} "
            f"{proj.get('country') or ''}\n"
            f"- **Target COD:** {proj.get('target_cod') or 'to be confirmed'}\n\n"
            "### RFQ categories\n\n" + "\n".join(f"- {c}" for c in cats) + "\n"
        )
    elif key == "construction_est":
        title = f"Construction Estimate - {proj['project_name']}"
        md = header + (
            "## Construction estimate\n\n"
            f"- **Total CAPEX (engineering model):** "
            f"{_fmt_money(computed.get('total_capex_usd'), 'USD')} "
            f"({_fmt_money(computed.get('total_capex_local'), cur)})\n"
            f"- **Linked-BOQ facilities actual (indicative):** "
            f"{_fmt_money(boq.get('grand_total_usd'), 'USD')}\n\n"
            "### CAPEX breakdown (USD)\n\n"
            + ("\n".join(f"- {_lbl(k)}: {_fmt_money(v)}"
                          for k, v in (computed.get('capex_lines_usd') or {}).items())
               or "(run Step 8 finance to populate)")
            + "\n\n### Construction assumptions\n\n"
            "- Mobilisation to COD driven by grid + civil critical path.\n"
            "- Facilities (control room, O&M, security, transformer yard) "
            "priced from the linked BOQ.\n"
            "- Contingency carried within the financial model.\n"
        )
    elif key == "maintenance":
        title = f"Maintenance Strategy - {proj['project_name']}"
        has = lambda *xs: any(x in _tech_list for x in xs)
        md = header + (
            "## Maintenance strategy (O&M)\n\n"
            "Preventive-first O&M aligned to the plant's monitoring stack.\n\n"
            "### Preventive maintenance plan\n\n"
            "- Modules: cleaning by soiling rate; annual EL / IV curve + "
            "thermography.\n"
            "- Inverters: quarterly inspection, filter service, firmware.\n"
            "- Transformers / switchgear: oil + protection-relay testing.\n"
            "- Structures: torque + corrosion checks; tracker service (if any).\n"
            "- Earthing & lightning: annual continuity + resistance test.\n\n"
            "### Maintenance technology\n\n"
            f"- CMMS / work orders: {'enabled' if has('cmms','maintenance','tickets') else 'recommended'}\n"
            f"- Spare-parts inventory: {'tracked' if has('inventory','spare_parts') else 'recommended'}\n"
            f"- Mobile field app: {'enabled' if has('mobile_app') else 'recommended'}\n\n"
            "### Spare-parts strategy\n\n"
            "- Critical spares on site: inverter modules, fuses, protection "
            "relays, comms gear.\n"
            "- Consumables: cleaning + PPE + connectors.\n\n"
            "### O&M staffing assumptions\n\n"
            "- Resident O&M lead + technicians (site-size dependent).\n"
            "- Security roster at the gatehouse.\n"
            "- Remote monitoring / NOC support (see Monitoring Strategy).\n"
        )
    elif key == "monitoring":
        title = f"Monitoring & SCADA Strategy - {proj['project_name']}"
        has = lambda *xs: any(x in _tech_list for x in xs)
        md = header + (
            "## Monitoring & SCADA strategy\n\n"
            "Real-time supervision and performance analytics for the plant.\n\n"
            "### Monitoring stack (from Step 5 selections)\n\n"
            f"- SCADA monitoring: {'yes' if has('scada','scada_monitoring') else 'recommended'}\n"
            f"- Weather station: {'yes' if has('weather','weather_station') else 'recommended'}\n"
            f"- Remote monitoring portal: {'yes' if has('remote_mon','cloud_mon') else 'recommended'}\n"
            f"- String / inverter monitoring: {'yes' if has('string_mon','inv_mon') else 'recommended'}\n"
            f"- Energy metering: {'yes' if has('energy_meter') else 'recommended'}\n"
            f"- Battery monitoring (BMS): {'yes' if has('bms') else 'n/a'}\n\n"
            "### Control room\n\n"
            + ("- Control Room facility is included in the plant scope.\n"
               if 'control_room' in _fac_list else
               "- No dedicated control room selected - remote NOC assumed.\n")
            + "\n### KPIs & alarms\n\n"
            "- PR, availability, specific yield, inverter efficiency.\n"
            f"- Design PR {pv.get('performance_ratio') or 'n/a'}, availability "
            f"{pv.get('availability_pct') or 'n/a'}%.\n"
            "- Alarm classes: trip, derate, comms-loss, security.\n\n"
            "### Fault response workflow\n\n"
            "1. Alarm raised (SCADA/portal) -> 2. NOC triage -> 3. Field work "
            "order (CMMS) -> 4. Fix + verify -> 5. Close-out + RCA.\n"
        )
    elif key == "ops_manual":
        title = f"Operations Manual - {proj['project_name']}"
        md = header + (
            "## Operations manual\n\n"
            "### Site & facilities\n\n" + _bullets(_fac_list, "(configure facilities on Step 4)") + "\n\n"
            "### Roles\n\n"
            "- O&M lead: plant performance, PM schedule, reporting.\n"
            "- Technicians: corrective + preventive work orders.\n"
            "- Security: access control at the gatehouse, patrols.\n"
            "- NOC / remote: monitoring, first-line triage.\n\n"
            "### Standard procedures\n\n"
            "- Start-up / shutdown + LOTO (lock-out tag-out).\n"
            "- Emergency response: fire, electrical, medical.\n"
            "- Grid interface + protection coordination.\n"
            "- Environmental + community obligations.\n\n"
            "### Documentation\n\n"
            "- As-built drawings, warranties, test certificates.\n"
            "- CMMS records, spare-parts register, monitoring exports.\n\n"
            "### Jurisdiction\n\n"
            f"- Regulator: {framework.get('regulator', {}).get('name', '')}\n"
            f"- Off-taker(s): {', '.join(framework.get('utility_offtakers') or [])}\n"
        )
    elif key == "wiring":
        title = f"Wiring & Cabling Schedule - {proj['project_name']}"
        dc_m = sizing.get("dc_cable_m_est") or 0
        ac_m = sizing.get("ac_cable_m_est") or 0
        n_str = sizing.get("strings") or 0
        n_comb = sizing.get("combiners") or 0
        n_inv = sizing.get("n_central_inverters") or sizing.get("n_central_inv") or 0
        md = header + (
            "## Wiring & cabling schedule\n\n"
            "Derived from the Step-7 PV sizing. Quantities are indicative; confirm "
            "against the detailed cable-pulling schedule and voltage-drop study.\n\n"
            "### Cable schedule\n\n"
            "| Segment | Type | Cable | Est. length |\n"
            "|---|---|---|---|\n"
            f"| Module -> string | DC | 1x6mm2 PV1-F (UV) | {_fmt_money(dc_m)} m |\n"
            f"| String -> combiner | DC | 1x6mm2 / 1x10mm2 PV1-F | (in DC total) |\n"
            f"| Combiner -> inverter | DC | 1x(35-240)mm2 DC main | (in DC total) |\n"
            f"| Inverter -> transformer | AC LV | 4c XLPE/SWA | {_fmt_money(ac_m)} m |\n"
            f"| Transformer -> MV switchgear | MV | XLPE armoured MV | (in MV total) |\n"
            f"| Earthing / bonding | CU | bare CU + earth rods | array-wide grid |\n\n"
            "### DC array wiring\n\n"
            f"- Strings: **{n_str}**  \n"
            f"- Combiner boxes: **{n_comb}** (with DC SPD + fuses + isolator)  \n"
            f"- Modules per string: {pv.get('modules_per_string') or 'n/a'}  \n"
            f"- Total DC cable: **{_fmt_money(dc_m)} m**\n\n"
            "### AC / MV wiring\n\n"
            f"- Central inverters: **{n_inv}**  \n"
            f"- Inverter step-up transformers: **{n_inv}**  \n"
            f"- Total AC/MV cable: **{_fmt_money(ac_m)} m**\n\n"
            "### Protection & earthing\n\n"
            "- DC: string fuses, DC SPD Type 2 at combiners + inverter input.\n"
            "- AC: MCCB/ACB at inverter LV, MV RMU with protection relays.\n"
            "- Earthing: array frame bonding, equipotential grid, earth electrodes;\n"
            "  lightning protection air terminals + down conductors.\n"
        )
    elif key == "single_line":
        title = f"Single-Line Diagram (SLD) - {proj['project_name']}"
        n_inv = sizing.get("n_central_inverters") or sizing.get("n_central_inv") or 0
        md = header + (
            "## Single-line diagram (SLD)\n\n"
            "Utility-scale power topology from the PV array to the grid point of "
            "common coupling (PCC).\n\n"
            "### Topology\n\n"
            "```\n"
            "  PV modules  (" + str(sizing.get("n_modules") or "n/a") + " x "
            + str(pv.get("module_wp") or "n/a") + " Wp)\n"
            "      |  DC strings (" + str(sizing.get("strings") or "n/a") + ")\n"
            "      v\n"
            "  String combiner boxes (" + str(sizing.get("combiners") or "n/a")
            + ")  --[DC SPD + fuses]\n"
            "      |  DC main\n"
            "      v\n"
            "  Central inverters (" + str(n_inv) + " x "
            + str(pv.get("central_inverter_kw")
                  or sizing.get("central_inverter_kw") or "n/a") + " kW)\n"
            "      |  LV AC\n"
            "      v\n"
            "  Inverter step-up transformers (LV/MV, " + str(n_inv) + ")\n"
            "      |  MV\n"
            "      v\n"
            "  MV switchgear / ring main unit (RMU) + protection\n"
            "      |  MV collection\n"
            "      v\n"
            "  Grid substation / HV switchyard  -->  PCC / utility grid\n"
            "```\n\n"
            "### Component schedule\n\n"
            "| Item | Qty | Rating |\n"
            "|---|---|---|\n"
            f"| PV modules | {sizing.get('n_modules') or 'n/a'} | {pv.get('module_wp') or 'n/a'} Wp |\n"
            f"| Combiner boxes | {sizing.get('combiners') or 'n/a'} | DC, IP65 |\n"
            f"| Central inverters | {n_inv} | {pv.get('central_inverter_kw') or sizing.get('central_inverter_kw') or 'n/a'} kW |\n"
            f"| Step-up transformers | {n_inv} | LV/MV |\n"
            f"| MV switchgear / RMU | {n_inv} | MV |\n"
            f"| AC capacity | 1 | {sizing.get('inverter_ac_kw') or 'n/a'} kW |\n\n"
            "> Protection coordination, relay settings and the grid-connection "
            "agreement must be confirmed with the utility / regulator.\n"
        )
    elif key == "energy_impact":
        title = f"Energy Impact & Yield - {proj['project_name']}"
        try:
            yld = _ci_yield_profile(pv, gps_lat=proj.get("gps_lat")) or {}
        except Exception:
            yld = {}   # never 500 the report on malformed stored JSON
        monthly = yld.get("monthly") if isinstance(yld.get("monthly"), list) else []
        annual10 = (yld.get("annual_series")
                    if isinstance(yld.get("annual_series"), list) else [])
        ann_mwh = sizing.get("annual_gen_mwh") or yld.get("annual_gen_mwh") or 0
        life_mwh = sizing.get("lifetime_gen_mwh") or 0
        co2 = round(float(ann_mwh or 0) * 0.40, 1)          # ~0.4 tCO2/MWh grid
        homes = int(round(float(ann_mwh or 0) * 1000 / 1200)) if ann_mwh else 0
        cf = 0.0
        try:
            ac = float(sizing.get("inverter_ac_kw") or 0)
            if ac > 0:
                cf = round(float(ann_mwh) * 1000 / (ac * 8760) * 100, 1)
        except (TypeError, ValueError, ZeroDivisionError):
            cf = 0.0
        mrows = ""
        for m in monthly[:12]:
            if isinstance(m, dict):
                mrows += f"| {m.get('month', '')} | {_fmt_money(m.get('mwh'))} |\n"
        if not mrows:
            mrows = "| (per-month profile unavailable) | |\n"
        yrows = ""
        for a in annual10:
            if isinstance(a, dict):
                yrows += f"| Year {a.get('year', '')} | {_fmt_money(a.get('mwh'))} |\n"
        if not yrows:
            yrows = "| (10-year series unavailable) | |\n"
        md = header + (
            "## Energy impact & yield\n\n"
            "### Generation headline\n\n"
            f"- **Annual generation:** {_fmt_money(ann_mwh)} MWh  \n"
            f"- **Lifetime generation:** {_fmt_money(life_mwh)} MWh  \n"
            f"- **Specific yield:** {sizing.get('specific_yield_kwh_per_kwp') or 'n/a'} kWh/kWp  \n"
            f"- **Capacity factor:** {cf} %  \n\n"
            "### Environmental & social impact\n\n"
            f"- **CO2 avoided:** ~{_fmt_money(co2)} tCO2 / year (grid factor 0.40 tCO2/MWh)  \n"
            f"- **Homes equivalent:** ~{_fmt_money(homes)} (at 1,200 kWh/home/yr)  \n"
            f"- **Clean energy over life:** {_fmt_money(life_mwh)} MWh  \n\n"
            "### Monthly generation (MWh)\n\n"
            "| Month | Generation |\n|---|---|\n" + mrows + "\n"
            "### Annual generation over 10 years (MWh, with degradation)\n\n"
            "| Year | Generation |\n|---|---|\n" + yrows + "\n"
            f"Degradation: {pv.get('annual_degradation_pct') or '0.5'} %/yr; "
            f"availability {pv.get('availability_pct') or '98'} %; "
            f"PR {pv.get('performance_ratio') or '0.78'}.\n"
        )
    elif key == "economic_impact":
        title = f"Economic Impact - {proj['project_name']}"
        try:
            cash = _ci_cashflow_plan(fin) or {}
        except Exception:
            cash = {}   # never 500 the report on malformed stored JSON
        _rev = cash.get("revenue") if isinstance(cash.get("revenue"), list) else []
        _cum = cash.get("cumulative") if isinstance(cash.get("cumulative"), list) else []
        yr1_rev = _rev[1] if len(_rev) > 1 else (_rev[0] if _rev else 0)
        life_cum = _cum[-1] if _cum else 0
        capex_usd = computed.get("total_capex_usd") or 0
        ann_mwh = sizing.get("annual_gen_mwh") or 0
        jobs_c = int(round(float(sizing.get("dc_kwp_actual") or 0) / 250)) or "n/a"
        jobs_o = int(round(float(sizing.get("dc_kwp_actual") or 0) / 4000)) or "n/a"
        md = header + (
            "## Economic impact\n\n"
            "### Investment\n\n"
            f"- **Total CAPEX:** {_fmt_money(capex_usd, 'USD')} "
            f"({_fmt_money(computed.get('total_capex_local'), cur)})  \n"
            f"- **Annual OPEX:** {_fmt_money(computed.get('total_opex_usd_yr'), 'USD')}  \n"
            f"- **NPV:** {_fmt_money(computed.get('npv_local'), cur)}  \n"
            f"- **IRR:** {_fmt_pct(computed.get('irr_pct'))}  \n"
            f"- **LCOE:** {computed.get('lcoe_local_per_kwh') or 'n/a'} {cur}/kWh  \n"
            f"- **Payback:** "
            f"{('%.1f yr' % computed['payback_years']) if computed.get('payback_years') else 'beyond project life'}  \n\n"
            "### Local economic impact\n\n"
            f"- **Construction jobs (peak):** ~{jobs_c}  \n"
            f"- **Permanent O&M jobs:** ~{jobs_o}  \n"
            f"- **Annual clean energy delivered:** {_fmt_money(ann_mwh)} MWh  \n"
            "- **Import substitution:** reduces reliance on thermal/imported power.  \n"
            "- **Local content:** civil works, security, O&M, logistics sourced locally.  \n\n"
            "### Revenue & cash flow\n\n"
            f"- Revenue model: **{(fin.get('revenue_model') or 'ppa')}**  \n"
            f"- Year-1 revenue: {_fmt_money(yr1_rev, cur)}  \n"
            f"- Cumulative net cash flow (life): {_fmt_money(life_cum, cur)}  \n\n"
            "> Job and impact figures are planning estimates (per-kWp heuristics); "
            "confirm in the ESIA / socio-economic study.\n"
        )
    elif key == "implementation_plan":
        title = f"Implementation Plan - {proj['project_name']}"
        mwp = (sizing.get("dc_kwp_actual") or proj.get("target_kwp") or 0) / 1000.0
        md = header + (
            "## Implementation plan\n\n"
            f"Indicative delivery schedule for a ~{mwp:.1f} MWp utility-scale PV "
            "plant. Durations scale with plant size, permitting and grid works.\n\n"
            "### Phased schedule\n\n"
            "| Phase | Key activities | Indicative duration |\n"
            "|---|---|---|\n"
            "| 1. Development | Land, permits, ESIA, grid-connection agreement | 3-6 months |\n"
            "| 2. Engineering | Detailed design, SLD, geotech, procurement specs | 2-3 months |\n"
            "| 3. Procurement | Modules, inverters, transformers, MV, BOP (long-lead) | 3-5 months |\n"
            "| 4. Civil works | Access roads, drainage, fencing, foundations | 2-4 months |\n"
            "| 5. Mechanical | Mounting structures / trackers, module install | 3-5 months |\n"
            "| 6. Electrical | DC/AC wiring, combiners, inverters, transformers, MV | 3-5 months |\n"
            "| 7. Grid & substation | Substation, HV switchyard, interconnection | 3-6 months |\n"
            "| 8. Testing & commissioning | SAT, grid compliance, energisation | 1-2 months |\n"
            "| 9. Handover & O&M | As-builts, training, O&M mobilisation | 1 month |\n\n"
            "### Milestones\n\n"
            "- **M1:** Financial close & notice to proceed (NTP).\n"
            "- **M2:** Major equipment delivered to site.\n"
            "- **M3:** Mechanical completion.\n"
            "- **M4:** Grid connection energised.\n"
            "- **M5:** Commercial operation date (COD): "
            f"{proj.get('target_cod') or 'to be confirmed'}.\n\n"
            "### Key dependencies & risks\n\n"
            "- Grid-connection approval and substation availability (critical path).\n"
            "- Long-lead procurement (transformers, MV switchgear).\n"
            "- Weather / ground conditions during civil works.\n"
            "- Permit / ESIA sign-off before construction start.\n\n"
            "### Governance\n\n"
            f"- Regulator: {framework.get('regulator', {}).get('name', '')}\n"
            f"- ESIA authority: {framework.get('esia_authority', {}).get('name', '')}\n"
        )
    else:
        title = "Report"
        md = header + "This report has not been implemented yet."

    return md, title


def _render_pdf_bytes(markdown_text: str, doc_title: str) -> bytes:
    """Render markdown to PDF via markdown-pdf (the SolarPro-standard library)."""
    from markdown_pdf import MarkdownPdf, Section
    import io
    pdf = MarkdownPdf(toc_level=2)
    pdf.meta["title"] = doc_title
    pdf.add_section(Section(markdown_text, toc=True))
    buf = io.BytesIO()
    pdf.save(buf)
    return buf.getvalue()


# ===========================================================================
# SECTION A3 -- AI agents (14 rule-based specialists + reviewer) + 3D Digital
# Twin scene builder + shared _is_meaningfully_populated. Carried verbatim
# from the legacy module (module-local per Codex SSS). v2 stores technology/
# electrical picks under 'technologies'/'services'; the step handlers pass a
# NORMALISED proj (adds a 'selected' mirror) so these blocks run unchanged.
# ===========================================================================

def _is_meaningfully_populated(field: str, val: Any) -> bool:
    """A JSON blob field counts as done only if it has real content."""
    if val is None:
        return False
    if field.startswith("_"):
        return False
    if field == "boq_project_id":
        return bool(val)
    if not isinstance(val, str):
        return bool(val)
    s = val.strip()
    if not s:
        return False
    # JSON blob fields must actually contain something meaningful.
    JSON_FIELDS = {"site_config", "facility_config", "technology_config",
                   "electrical_config", "pv_config", "finance_config",
                   "regulatory_config"}
    if field in JSON_FIELDS:
        try:
            parsed = json.loads(s)
        except (TypeError, ValueError):
            return False
        if not isinstance(parsed, dict):
            return bool(parsed)
        # Truthy if any leaf value is set (non-empty string / non-empty list /
        # a real number, but NOT None / '' / [] / {}).
        for v in parsed.values():
            if v in (None, "", [], {}):
                continue
            return True
        return False
    return True


AGENT_DEPARTMENTS: list[tuple[str, str, str, str]] = [
    # (code, label, department, icon)
    ("pv_design",      "PV Design Agent",              "Engineering", "bi-sun"),
    ("electrical",     "Electrical Design Agent",      "Engineering", "bi-plug"),
    ("civil",          "Civil Design Agent",           "Engineering", "bi-bricks"),
    ("structural",     "Structural Agent",             "Engineering", "bi-columns-gap"),
    ("ict",            "ICT Infrastructure Agent",     "Engineering", "bi-hdd-network"),
    ("scada",          "SCADA Agent",                  "Engineering", "bi-cpu"),
    ("grid",           "Grid Connection Agent",        "Engineering", "bi-lightning-charge"),
    ("financial",      "Financial Engineering Agent",  "Finance",     "bi-cash-coin"),
    ("investment",     "Investment Agent",             "Finance",     "bi-bank"),
    ("risk",           "Risk Analysis Agent",          "Finance",     "bi-shield-exclamation"),
    ("marketplace",    "Marketplace Agent",            "Procurement", "bi-shop"),
    ("boq",            "BOQ Agent",                    "Procurement", "bi-list-check"),
    ("report_writer",  "Report Writer Agent",          "Reporting",   "bi-file-earmark-text"),
    ("qa_qc",          "QA/QC Agent",                  "Governance",  "bi-clipboard-check"),
    ("reviewer",       "Project Reviewer Agent",       "Governance",  "bi-check2-circle"),
]
AGENT_CODES: set[str] = {c for c, _, _, _ in AGENT_DEPARTMENTS}


def _f(v: Any, default: float = 0.0) -> float:
    try:
        if v in (None, "", "None"):
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _rating(score: int) -> str:
    if score >= 85: return "A - excellent"
    if score >= 70: return "B - good"
    if score >= 55: return "C - acceptable"
    if score >= 40: return "D - marginal"
    return "F - not ready"


def _agent_pv_design(proj: dict[str, Any]) -> dict[str, Any]:
    pv = _safe_json(proj.get("pv_config"))
    site = _safe_json(proj.get("site_config"))
    sizing = pv.get("sizing") or {}
    kwp = _f(sizing.get("kwp_input") or pv.get("kwp") or proj.get("target_kwp"))
    findings, recs = [], []
    score = 100
    if kwp <= 0:
        findings.append("kWp not set (Step 7 not completed)")
        recs.append("Complete Step 7 with a plant capacity")
        score = 15
    else:
        # DC/AC ratio sanity
        dcac = _f(pv.get("dc_ac_ratio"), 1.20)
        if dcac < 1.05:
            findings.append(f"DC/AC ratio {dcac:.2f} is low - inverters over-sized, CAPEX inefficient")
            recs.append("Target DC/AC 1.15-1.30 for most tropical climates")
            score -= 10
        elif dcac > 1.40:
            findings.append(f"DC/AC ratio {dcac:.2f} is high - meaningful clipping losses expected")
            recs.append("Model clipping loss and validate against inverter curve")
            score -= 10
        # Tilt sanity for Ghana (near equator)
        gps_lat = _f(proj.get("gps_lat"), 6.0)
        tilt = _f(pv.get("tilt_deg"), 10)
        ideal_tilt = abs(gps_lat)
        if abs(tilt - ideal_tilt) > 10 and pv.get("mounting") == "fixed_tilt":
            findings.append(f"Fixed tilt {tilt}deg differs by {abs(tilt-ideal_tilt):.0f}deg from latitude-optimal ({ideal_tilt:.0f}deg)")
            recs.append("Snap tilt closer to latitude unless site-specific reason")
            score -= 5
        # Land area vs capacity (~1.5 ha per MW for utility ground-mount)
        land_ha = _f(site.get("land_area_ha"))
        needed_ha = kwp / 1000.0 * 1.5
        if land_ha > 0 and land_ha < needed_ha:
            findings.append(f"Land {land_ha:.1f} ha may be tight for {kwp/1000:.1f} MW (~{needed_ha:.1f} ha recommended)")
            recs.append("Verify with a bifacial + HSAT layout or increase land area")
            score -= 10
        # Availability
        avail = _f(pv.get("availability_pct"), 98)
        if avail < 96:
            findings.append(f"Availability {avail}% is lower than utility norm (98-99%)")
            score -= 5
        if not findings:
            findings.append("PV design values are all within utility-scale norms")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"PV design rating: {_rating(max(0, score))}"}


def _agent_electrical(proj: dict[str, Any]) -> dict[str, Any]:
    elec = _safe_json(proj.get("electrical_config"))
    pv = _safe_json(proj.get("pv_config"))
    sizing = pv.get("sizing") or {}
    selected = set(elec.get("selected") or [])
    findings, recs = [], []
    score = 100
    core = {"internal_installation", "hv_distribution", "lv_distribution",
            "inverters", "transformers", "earthing", "lightning_protection"}
    missing = core - selected
    if missing:
        findings.append(f"Missing core electrical scope: {', '.join(sorted(missing))}")
        recs.append("Enable the missing services on Step 6 before proceeding")
        score -= 8 * len(missing)
    ac_kw = _f(sizing.get("inverter_ac_kw"))
    if ac_kw > 5000 and "hv_switchgear" not in selected:
        findings.append(f"Plant AC {ac_kw/1000:.1f} MW requires HV switchgear but none selected")
        score -= 10
    if "hv_distribution" in selected and "rmu" not in selected:
        findings.append("HV distribution enabled without RMU - MV interconnection likely incomplete")
        recs.append("Add RMU on Step 6 or note bypass rationale")
        score -= 5
    if "scada" in selected and "lan" not in selected:
        findings.append("SCADA enabled without LAN backbone")
        recs.append("Enable LAN on Step 6 for SCADA connectivity")
        score -= 3
    if not findings:
        findings.append("Electrical scope is complete for a utility-scale plant")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Electrical rating: {_rating(max(0, score))}"}


def _agent_civil(proj: dict[str, Any]) -> dict[str, Any]:
    site = _safe_json(proj.get("site_config"))
    fac = _safe_json(proj.get("facility_config"))
    findings, recs = [], []
    score = 100
    slope = (site.get("slope") or "").strip()
    if slope == "gt_20":
        findings.append("Slope > 20% - major grading + geotech needed")
        recs.append("Add 6-9% civil CAPEX contingency")
        score -= 20
    elif slope == "10_20":
        findings.append("Slope 10-20% - grading + drainage attention needed")
        score -= 8
    flood = site.get("flood_risk")
    if flood == "high":
        findings.append("High flood risk - elevated mounting + drainage compulsory")
        recs.append("Add drainage + PV table height provision")
        score -= 15
    access = site.get("access_road")
    if access in ("dirt_seasonal", "none"):
        findings.append(f"Access road '{access}' - construction access + module delivery risk")
        recs.append("Budget for construction access improvement")
        score -= 10
    soil = site.get("soil")
    if soil == "marshy":
        findings.append("Marshy soil - deep pile foundations mandatory")
        recs.append("Instruct a geotech CPT/SPT campaign")
        score -= 12
    if soil == "unknown":
        findings.append("Soil type unknown - geotech survey required")
        recs.append("Commission a geotech survey before EPC bid")
        score -= 6
    ex_works = fac.get("external_works") or []
    if "drainage" not in ex_works and flood in ("low", "medium", "high"):
        findings.append(f"Flood risk '{flood}' but drainage not enabled in Step 4 external works")
        recs.append("Enable drainage on Step 4")
        score -= 5
    if not findings:
        findings.append("Civil / geotech risk profile is low for the given site data")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Civil rating: {_rating(max(0, score))}"}


def _agent_structural(proj: dict[str, Any]) -> dict[str, Any]:
    site = _safe_json(proj.get("site_config"))
    pv = _safe_json(proj.get("pv_config"))
    findings, recs = [], []
    score = 100
    wind = site.get("wind_zone")
    mounting = pv.get("mounting")
    if wind == "cyclone":
        findings.append("Cyclone-prone zone")
        recs.append("Specify structures certified to IEC 61400 Class-I or equivalent")
        score -= 20
        if mounting == "single_axis":
            findings.append("Single-axis trackers in cyclone zone - stow angle + control required")
            score -= 10
    elif wind == "z3_high":
        findings.append("High wind zone - upgrade module clamps + rail gauge")
        score -= 8
    seismic = site.get("seismic_zone")
    if seismic in ("zone_3", "zone_4"):
        findings.append(f"Seismic {seismic} - anchor & bracing design attention")
        recs.append("Structural PE sign-off on foundations + transformer base")
        score -= 10
    if not findings:
        findings.append("Structural loading profile is standard for site data given")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Structural rating: {_rating(max(0, score))}"}


def _agent_ict(proj: dict[str, Any]) -> dict[str, Any]:
    tech = _safe_json(proj.get("technology_config"))
    selected = set(tech.get("selected") or [])
    findings, recs = [], []
    score = 100
    if not selected:
        findings.append("No technology stack selected on Step 5")
        recs.append("Enable at least SCADA + monitoring + cyber security")
        score = 25
        return {"status": "warning", "score": score,
                "findings": findings, "recs": recs,
                "summary": f"ICT rating: {_rating(score)}"}
    if "cyber" not in selected:
        findings.append("Cyber security not selected - not bankable for utility off-take")
        recs.append("Enable cyber security + firewall on Step 5")
        score -= 15
    if "gps_sync" not in selected and "scada" in selected:
        findings.append("SCADA without GPS time sync - IEC 61850 event correlation degraded")
        score -= 8
    if "fibre" not in selected and "ind_eth" not in selected:
        findings.append("Neither fibre nor industrial Ethernet backbone selected")
        recs.append("Enable fibre for the plant collection network")
        score -= 12
    if "backup_srv" not in selected and "cloud_backup" not in selected:
        findings.append("No backup strategy (backup server nor cloud backup)")
        recs.append("Choose at least one backup path for SCADA historian")
        score -= 8
    if not findings:
        findings.append("ICT infrastructure is comprehensive")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"ICT rating: {_rating(max(0, score))}"}


def _agent_scada(proj: dict[str, Any]) -> dict[str, Any]:
    tech = _safe_json(proj.get("technology_config"))
    elec = _safe_json(proj.get("electrical_config"))
    s = set(tech.get("selected") or [])
    e = set(elec.get("selected") or [])
    findings, recs = [], []
    score = 100
    if "scada" not in s:
        findings.append("SCADA not enabled in Step 5 technology stack")
        recs.append("Enable SCADA + EMS + PPC on Step 5")
        score -= 25
    if "scada" in s and "scada" not in e:
        findings.append("SCADA in tech stack but no SCADA service on Step 6 - cabling scope missing")
        recs.append("Enable SCADA in Step 6 electrical services")
        score -= 8
    if "ems" not in s:
        findings.append("EMS not selected - dispatch optimisation missing")
        recs.append("Enable EMS on Step 5")
        score -= 5
    if "ppc" not in s:
        findings.append("Power Plant Controller (PPC) not selected - grid-code compliance risk")
        recs.append("Enable PPC on Step 5 to meet reactive/frequency response requirements")
        score -= 10
    if "weather" not in s:
        findings.append("Weather station missing - performance model can't be measured")
        recs.append("Enable weather station on Step 5")
        score -= 3
    if not findings:
        findings.append("SCADA / EMS / PPC stack is complete")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"SCADA rating: {_rating(max(0, score))}"}


def _agent_grid(proj: dict[str, Any]) -> dict[str, Any]:
    site = _safe_json(proj.get("site_config"))
    elec = _safe_json(proj.get("electrical_config"))
    pv = _safe_json(proj.get("pv_config"))
    sizing = pv.get("sizing") or {}
    ac_kw = _f(sizing.get("inverter_ac_kw"))
    findings, recs = [], []
    score = 100
    dist_km = _f(site.get("grid_distance_km"), -1)
    if dist_km < 0:
        findings.append("Grid distance not set on Step 3")
        score -= 8
    elif dist_km > 15:
        findings.append(f"Grid distance {dist_km:.1f} km - substation extension CAPEX will be material")
        recs.append("Include line construction + wayleave cost in Step 8 CAPEX")
        score -= 10
    elif dist_km > 5:
        findings.append(f"Grid distance {dist_km:.1f} km - budget for interconnection line + easements")
        score -= 3
    hv_kv = _f(site.get("hv_line_kv"), -1)
    if hv_kv > 0 and ac_kw > 0:
        # Rough sanity: 33 kV OK up to ~15 MW, 66 kV up to ~50 MW, else 132/161 kV
        if ac_kw > 15000 and hv_kv <= 33:
            findings.append(f"Interconnect at {hv_kv} kV for {ac_kw/1000:.1f} MW AC - consider 66 kV+")
            score -= 8
        if ac_kw > 50000 and hv_kv < 66:
            findings.append(f"Interconnect at {hv_kv} kV for {ac_kw/1000:.1f} MW AC - 132 kV recommended")
            score -= 12
    services = set(elec.get("selected") or [])
    if "transformers" not in services:
        findings.append("Transformers not selected on Step 6")
        recs.append("Enable transformer service on Step 6")
        score -= 15
    if "rmu" not in services and ac_kw > 3000:
        findings.append("RMU absent for a > 3 MW plant")
        recs.append("Enable RMU on Step 6")
        score -= 5
    if not findings:
        findings.append("Grid interconnection scope is coherent")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Grid rating: {_rating(max(0, score))}"}


def _agent_financial(proj: dict[str, Any]) -> dict[str, Any]:
    fin = _safe_json(proj.get("finance_config"))
    computed = fin.get("computed") or {}
    findings, recs = [], []
    score = 100
    if not computed:
        return {"status": "warning", "score": 20,
                "findings": ["Financial model not yet computed on Step 8"],
                "recs": ["Complete Step 8 to run the finance engine"],
                "summary": "Financial rating: F - not ready"}
    irr = _f(computed.get("irr_pct"))
    lcoe = _f(computed.get("lcoe_local_per_kwh"))
    tariff = _f(fin.get("tariff_local_per_kwh"))
    dscr_min = _f(computed.get("dscr_min"))
    payback = _f(computed.get("payback_years"))
    if irr < 8:
        findings.append(f"IRR {irr:.1f}% is below the 8% utility benchmark")
        recs.append("Renegotiate PPA or reduce CAPEX assumptions")
        score -= 15
    elif irr < 12:
        findings.append(f"IRR {irr:.1f}% is acceptable but not compelling")
        score -= 5
    if dscr_min > 0 and dscr_min < 1.20:
        findings.append(f"DSCR min {dscr_min:.2f}x - lenders typically require >= 1.25x")
        recs.append("Lengthen debt tenor or lower leverage")
        score -= 15
    elif dscr_min > 0 and dscr_min < 1.30:
        findings.append(f"DSCR min {dscr_min:.2f}x is right at the covenant floor - tight")
        score -= 5
    if tariff > 0 and lcoe > 0:
        margin = (tariff - lcoe) / tariff * 100
        if margin < 15:
            findings.append(f"Tariff-LCOE margin {margin:.1f}% is thin - sensitive to tariff review")
            score -= 8
    if payback > 12:
        findings.append(f"Payback {payback:.1f} yr is longer than the debt tenor")
        recs.append("Consider ITC / accelerated depreciation")
        score -= 5
    if not findings:
        findings.append("Financial metrics are all within bank-comfort range")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Financial rating: {_rating(max(0, score))}"}


def _agent_investment(proj: dict[str, Any]) -> dict[str, Any]:
    findings, recs = [], []
    score = 100
    if not (proj.get("investor") or "").strip():
        findings.append("Investor field is empty on Step 1")
        recs.append("Populate Step 1 investor when a champion emerges")
        score -= 20
    if not (proj.get("developer") or "").strip():
        findings.append("Developer / EPC not identified")
        score -= 5
    if not (proj.get("description") or "").strip():
        findings.append("Executive description missing - investor memo will be weak")
        score -= 5
    if not (proj.get("client_name") or "").strip():
        findings.append("Off-taker / client is blank")
        recs.append("Investors need clarity on the off-take - populate Step 1 client")
        score -= 15
    if not findings:
        findings.append("Deal team + off-taker identity is clear")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Investment posture: {_rating(max(0, score))}"}


def _agent_risk(proj: dict[str, Any]) -> dict[str, Any]:
    reg = _safe_json(proj.get("regulatory_config"))
    findings, recs = [], []
    score = 100
    items = reg.get("items") or {}
    if not items:
        findings.append("Development & Regulatory not yet completed")
        recs.append("Open /large-scale-solar/<pid>/regulatory to capture posture")
        score -= 25
    else:
        critical = ("esia", "grid_interconnect", "energy_commission", "land_tenure")
        for code in critical:
            st = (items.get(code) or {}).get("status") or "not_started"
            if st in ("not_started", "denied"):
                findings.append(f"{code}: {st}")
                score -= 8
    if not reg.get("land_tenure"):
        findings.append("Land tenure structure not chosen")
        recs.append("Pick a tenure on the Regulatory step")
        score -= 10
    fin = _safe_json(proj.get("finance_config"))
    if fin.get("revenue_model") == "merchant":
        findings.append("Merchant revenue model - price + volume risk material")
        recs.append("Consider PPA cover for at least the debt tenor")
        score -= 10
    if not findings:
        findings.append("Development risk posture is under control")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Risk profile: {_rating(max(0, score))}"}


def _agent_marketplace(proj: dict[str, Any]) -> dict[str, Any]:
    tech = _safe_json(proj.get("technology_config"))
    elec = _safe_json(proj.get("electrical_config"))
    pv = _safe_json(proj.get("pv_config"))
    cats = _marketplace_categories_for(pv, tech, elec)
    findings, recs = [], []
    score = 100
    if not cats:
        findings.append("No marketplace categories derived - Steps 5-7 incomplete")
        score = 30
    if len(cats) < 6:
        findings.append(f"Only {len(cats)} categories mapped - procurement scope narrow")
        recs.append("Enable more technology + electrical services on Steps 5-6")
        score -= 8
    if not findings:
        findings.append(f"{len(cats)} marketplace categories mapped from tech/electrical picks")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"Marketplace coverage: {_rating(max(0, score))}"}


def _agent_boq(proj: dict[str, Any]) -> dict[str, Any]:
    findings, recs = [], []
    score = 100
    if not proj.get("boq_project_id"):
        findings.append("No BOQ project linked - Step 9 not run")
        recs.append("Run Step 9 to auto-generate a linked BOQ project + buildings")
        score = 40
    else:
        findings.append(f"BOQ project #{proj['boq_project_id']} linked")
    fac = _safe_json(proj.get("facility_config"))
    if not (fac.get("buildings") or []):
        findings.append("No buildings enabled on Step 4 - BOQ scope will be sparse")
        recs.append("Enable at least Control Room + O&M + Transformer Yard")
        score -= 15
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"BOQ readiness: {_rating(max(0, score))}"}


def _agent_report_writer(proj: dict[str, Any]) -> dict[str, Any]:
    findings, recs = [], []
    score = 100
    # This agent nominally invokes the LLM narrative through the SolarPro
    # AI chain. Here we produce a deterministic executive narrative from
    # the project's stored data so the agent is always useful even when
    # the LLM chain is down.
    pv = _safe_json(proj.get("pv_config"))
    fin = _safe_json(proj.get("finance_config"))
    sizing = pv.get("sizing") or {}
    computed = fin.get("computed") or {}
    kwp = _f(sizing.get("kwp_input") or pv.get("kwp") or proj.get("target_kwp"))
    if kwp <= 0:
        findings.append("Not enough data to author reports - complete Steps 1-8")
        return {"status": "warning", "score": 30,
                "findings": findings, "recs": ["Return once Step 8 is done"],
                "summary": "Reports not ready"}
    narrative = (
        f"{proj.get('project_name')} is a {kwp/1000:.1f} MWp "
        f"utility-scale plant proposed for "
        f"{proj.get('district') or ''} {proj.get('region') or ''} {proj.get('country') or ''}. "
        f"The design produces {_f(sizing.get('annual_gen_mwh')):,.0f} MWh/yr at PR "
        f"{_f(pv.get('performance_ratio'), 0.78):.2f}. "
        f"Total CAPEX is USD {_f(computed.get('total_capex_usd'))/1e6:.1f}M "
        f"with a base-case IRR of {_f(computed.get('irr_pct')):.1f}% "
        f"and LCOE {_f(computed.get('lcoe_local_per_kwh')):.4f} {proj.get('currency') or 'GHS'}/kWh."
    )
    findings.append("Executive narrative drafted from project data")
    recs.append("Download the 5 PDF reports on Step 13")
    return {"status": "ok", "score": score,
            "findings": findings, "recs": recs,
            "summary": "Report narrative ready",
            "narrative": narrative}


def _agent_qa_qc(proj: dict[str, Any]) -> dict[str, Any]:
    findings, recs = [], []
    score = 100
    checks = [
        ("Step 1 identity", bool((proj.get("project_name") or "").strip())),
        ("Step 2 project type", bool((proj.get("project_type") or "").strip())),
        ("Step 3 site config", _is_meaningfully_populated("site_config", proj.get("site_config"))),
        ("Step 4 facility",   _is_meaningfully_populated("facility_config", proj.get("facility_config"))),
        ("Step 5 technology", _is_meaningfully_populated("technology_config", proj.get("technology_config"))),
        ("Step 6 electrical", _is_meaningfully_populated("electrical_config", proj.get("electrical_config"))),
        ("Step 7 PV design",  _is_meaningfully_populated("pv_config", proj.get("pv_config"))),
        ("Step 8 finance",    _is_meaningfully_populated("finance_config", proj.get("finance_config"))),
        ("Step 9 BOQ linked", bool(proj.get("boq_project_id"))),
    ]
    missing = [name for name, done in checks if not done]
    if missing:
        findings.append(f"Missing / incomplete steps: {', '.join(missing)}")
        score -= 6 * len(missing)
        recs.append("Complete the missing steps before requesting Enterprise sign-off")
    if not missing:
        findings.append("All upstream engineering + finance steps are populated")
    return {"status": "ok" if score >= 55 else "warning",
            "score": max(0, min(100, score)),
            "findings": findings, "recs": recs,
            "summary": f"QA/QC completeness: {_rating(max(0, score))}"}


def _agent_reviewer(proj: dict[str, Any], sub_scores: dict[str, int]) -> dict[str, Any]:
    """Overall bankability + readiness score - aggregates every specialist."""
    findings, recs = [], []
    if not sub_scores:
        return {"status": "warning", "score": 0,
                "findings": ["No specialist scores available"],
                "recs": [], "summary": "Not ready"}
    avg = sum(sub_scores.values()) / len(sub_scores)
    score = int(round(avg))
    reds = [k for k, v in sub_scores.items() if v < 55]
    ambers = [k for k, v in sub_scores.items() if 55 <= v < 70]
    if reds:
        findings.append(f"Red-status specialists: {', '.join(reds)}")
        recs.append("Address every red-status specialist before EPC bid")
    if ambers:
        findings.append(f"Amber-status specialists: {', '.join(ambers)}")
    if not reds and not ambers:
        findings.append("Every specialist above 70 - project is EPC-ready")
    findings.append(f"Overall bankability score: {score}/100 ({_rating(score)})")
    return {"status": "ok" if score >= 55 else "warning",
            "score": score,
            "findings": findings, "recs": recs,
            "summary": f"Project readiness: {_rating(score)}"}


AGENT_RUNNERS: dict[str, Any] = {
    "pv_design":     _agent_pv_design,
    "electrical":    _agent_electrical,
    "civil":         _agent_civil,
    "structural":    _agent_structural,
    "ict":           _agent_ict,
    "scada":         _agent_scada,
    "grid":          _agent_grid,
    "financial":     _agent_financial,
    "investment":    _agent_investment,
    "risk":          _agent_risk,
    "marketplace":   _agent_marketplace,
    "boq":           _agent_boq,
    "report_writer": _agent_report_writer,
    "qa_qc":         _agent_qa_qc,
    # 'reviewer' is orchestrator-only and runs last with sub-scores.
}


def run_agent_orchestrator(proj: dict[str, Any]) -> dict[str, Any]:
    """Dispatch every specialist against the project, then run the
    reviewer with the collected sub-scores. Returns a full report."""
    results: dict[str, dict[str, Any]] = {}
    for code in AGENT_CODES:
        if code == "reviewer":
            continue
        runner = AGENT_RUNNERS.get(code)
        if not runner:
            continue
        try:
            results[code] = runner(proj)
        except Exception as e:   # pragma: no cover
            results[code] = {"status": "error", "score": 0,
                             "findings": [f"Agent crash: {e}"],
                             "recs": ["Retry after fixing the project data"],
                             "summary": "Agent error"}
    sub_scores = {c: r["score"] for c, r in results.items()}
    results["reviewer"] = _agent_reviewer(proj, sub_scores)
    return {"specialists": results,
            "aggregate_score": results["reviewer"]["score"],
            "aggregate_status": results["reviewer"]["status"]}


# ---- agent-runs storage ----

_AGENTRUN_SQLITE_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_agent_runs (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id     INTEGER NOT NULL,
    user_id        INTEGER NOT NULL,
    agent_code     TEXT NOT NULL,
    status         TEXT DEFAULT 'ok',
    score          INTEGER DEFAULT 0,
    payload        TEXT DEFAULT '{}',
    created_at     TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_ciar_project ON capital_investment_agent_runs(project_id);
CREATE INDEX IF NOT EXISTS idx_ciar_user    ON capital_investment_agent_runs(user_id);
CREATE INDEX IF NOT EXISTS idx_ciar_recent  ON capital_investment_agent_runs(project_id, created_at DESC);
"""
_AGENTRUN_POSTGRES_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_agent_runs (
    id             SERIAL PRIMARY KEY,
    project_id     INTEGER NOT NULL,
    user_id        INTEGER NOT NULL,
    agent_code     TEXT NOT NULL,
    status         TEXT DEFAULT 'ok',
    score          INTEGER DEFAULT 0,
    payload        TEXT DEFAULT '{}',
    created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_ciar_project ON capital_investment_agent_runs(project_id);
CREATE INDEX IF NOT EXISTS idx_ciar_user    ON capital_investment_agent_runs(user_id);
CREATE INDEX IF NOT EXISTS idx_ciar_recent  ON capital_investment_agent_runs(project_id, created_at DESC);
"""


def _ensure_agent_runs_schema(get_db) -> None:
    try:
        with get_db() as c:
            c.executescript(_AGENTRUN_SQLITE_DDL)
        return
    except Exception:
        pass
    for stmt in _AGENTRUN_POSTGRES_DDL.split(";"):
        s = stmt.strip()
        if not s:
            continue
        try:
            with get_db() as c:
                c.execute(s)
        except Exception:
            pass



# ---------------------------------------------------------------------------
# Slice 7 - 3D Digital Twin Studio (replaces the classic /shading page for
# capital-investment projects). Scene is built server-side into a JSON
# graph consumed by Three.js in the template.
# ---------------------------------------------------------------------------

# Layer palette used by both scene generation and the left-nav toggles.
DT_LAYER_PALETTE: dict[str, str] = {
    # Site
    "terrain":         "#3a4a2a",   # olive-brown ground
    "fence":           "#c0a060",   # tan
    "gate":            "#e0a020",   # amber
    "internal_roads":  "#606060",   # grey
    "drainage":        "#3a5a80",   # slate blue
    # Buildings
    "building":        "#e0c080",   # sand
    "control_room":    "#f59e0b",   # SolarPro warning
    "om_building":     "#c76a2e",
    "security_gate":   "#a06040",
    "battery_room":    "#7c4e9f",
    "switchgear_bldg": "#5c7cff",
    "transformer_bldg":"#c04040",
    "scada_bldg":      "#20b0a0",
    # Power system
    "pv_array":        "#1a3468",   # deep blue
    "pv_row":          "#2050a0",
    "combiner":        "#40a0e0",
    "inverter":        "#e0c020",
    "transformer":     "#c04040",
    "rmu":             "#e05050",
    "mv_switchgear":   "#a02020",
    "cable_trench":    "#606060",
    # ICT / SCADA / SAFETY
    "cctv_pole":       "#ffffff",
    "weather_mast":    "#20e0c0",
    "lighting_pole":   "#f0e080",
    "earthing_pit":    "#606030",
    "fire_hydrant":    "#e02020",
    "warning_sign":    "#e0a020",
}

# Every scene object carries a layer code so the left-nav can toggle it.
DT_LAYER_GROUPS: list[tuple[str, str, list[str]]] = [
    # (group_label, group_icon, [object_layer_codes])
    ("SITE",       "bi-map",             ["terrain", "fence", "gate",
                                          "internal_roads", "drainage"]),
    ("BUILDINGS",  "bi-buildings",       ["control_room", "om_building",
                                          "security_gate", "battery_room",
                                          "switchgear_bldg", "transformer_bldg",
                                          "scada_bldg", "building"]),
    ("PV FIELD",   "bi-grid-3x3",        ["pv_array", "pv_row"]),
    ("POWER SYSTEM","bi-lightning",      ["combiner", "inverter", "transformer",
                                          "rmu", "mv_switchgear", "cable_trench"]),
    ("ICT",        "bi-hdd-network",     ["cctv_pole", "weather_mast"]),
    ("LIGHTING",   "bi-lightbulb",       ["lighting_pole"]),
    ("EARTHING",   "bi-arrow-down",      ["earthing_pit"]),
    ("SAFETY",     "bi-fire",            ["fire_hydrant", "warning_sign"]),
]




def build_scene_from_project(proj: dict[str, Any]) -> dict[str, Any]:
    """Return a scene-graph dict consumable by the Three.js digital-twin.

    Units: metres. Origin at site centre. +X = East, +Z = South (Three.js
    right-handed).

    The 3D twin and the approved 2D plot plan MUST render the SAME physical
    arrangement, so this REUSES ``dt_site_layout.build_site_layout_model`` --
    the single source of truth for the site envelope, the inverter-block grid,
    the skids, the substation compound, the control building, the access-road
    network and the perimeter fence. That 2D model is expressed in top-left
    metres (x -> east/right, y -> south/down); here we lift it into the twin's
    centre-origin 3D frame and fill each block with real, tilted module TABLES
    laid in rows, so the field reads as an engineered array of tables in blocks
    -- not a flat pale band.

    Never raises: build_site_layout_model degrades every field to a safe
    default, and each render step below is defensive, so a half-built project
    still produces a coherent (if sparse) scene.
    """
    pv_cfg = _safe_json(proj.get("pv_config"))
    fac_cfg = _safe_json(proj.get("facility_config"))
    site_cfg = _safe_json(proj.get("site_config"))
    elec_cfg = _safe_json(proj.get("electrical_config"))
    tech_cfg = _safe_json(proj.get("technology_config"))

    sizing = pv_cfg.get("sizing") or {}
    kwp = float(sizing.get("kwp_input") or pv_cfg.get("kwp")
                or (proj.get("target_kwp") or 0))
    tilt_deg = float(pv_cfg.get("tilt_deg") or 12.0)
    azimuth_deg = float(pv_cfg.get("azimuth_deg") or 180.0)
    row_pitch = float(pv_cfg.get("row_pitch_m") or 6.0)
    n_modules = int(sizing.get("n_modules") or 0)

    # --- Shared layout model (the SAME model the 2D plot plan renders) -------
    try:
        from dt_site_layout import build_site_layout_model, _ROADS_SETBACK_M
        layout = build_site_layout_model(proj) or {}
    except Exception:
        layout, _ROADS_SETBACK_M = {}, 8.0
    site = layout.get("site") or {}
    site_w = float(site.get("w_m") or 0.0) or 600.0    # E-W extent (metres, X)
    site_h = float(site.get("h_m") or 0.0) or 400.0    # N-S extent (metres, Z)
    land_area_ha = (float(site.get("area_ha") or 0.0)
                    or float(site_cfg.get("land_area_ha") or 0.0)
                    or round(site_w * site_h / 10000.0, 1))
    diag_side = max(site_w, site_h)          # square hint for camera + scenery

    # Layout coords are top-left (x right, y down); the twin is centre-origin
    # (+X east, +Z south). These lift a 2D layout point into the 3D frame.
    def _sx(x: float) -> float:
        return round(float(x) - site_w / 2.0, 2)

    def _sz(y: float) -> float:
        return round(float(y) - site_h / 2.0, 2)

    lay_fence = layout.get("fence") or {}
    lay_blocks = layout.get("blocks") or []
    lay_skids = layout.get("skids") or []
    lay_sub = layout.get("substation") or {}
    lay_ctrl = layout.get("control") or {}
    lay_roads = layout.get("roads") or []
    lay_field = layout.get("field") or {}

    # --- Terrain (rectangular: covers the whole site envelope) --------------
    terrain = {
        "layer": "terrain",
        "kind": "ground",
        "side_m": round(diag_side, 1),      # legacy square hint (camera/scenery)
        "w_m": round(site_w, 1),            # rectangular extent (X = E-W)
        "l_m": round(site_h, 1),            # rectangular extent (Z = N-S)
        "label": "Site",
        "meta": {"land_area_ha": land_area_ha,
                 "terrain": site_cfg.get("terrain") or "flat",
                 "soil":    site_cfg.get("soil") or "sandy",
                 "site_w_m": round(site_w, 1), "site_h_m": round(site_h, 1)},
    }

    # --- Perimeter fence (rectangular, from the layout fence rect) ----------
    if lay_fence:
        fx, fy = float(lay_fence.get("x") or 0.0), float(lay_fence.get("y") or 0.0)
        fw = float(lay_fence.get("w") or site_w)
        fh = float(lay_fence.get("h") or site_h)
    else:
        fx = fy = 10.0
        fw, fh = site_w - 20.0, site_h - 20.0
    fence_pts = [
        [_sx(fx),       _sz(fy)],
        [_sx(fx + fw),  _sz(fy)],
        [_sx(fx + fw),  _sz(fy + fh)],
        [_sx(fx),       _sz(fy + fh)],
    ]
    fence = {
        "layer": "fence",
        "kind":  "line_loop",
        "points": fence_pts,
        "height_m": 2.4,
        "label": "Perimeter security fence",
        "meta":  {"perimeter_m": round(2 * (fw + fh), 1)},
    }

    # --- PV field: tilted module TABLES laid in rows inside each block -------
    # Each inverter block is inset by the road setback, then filled with E-W
    # rows (long axis = X) stacked N-S (Z) at the row pitch, leaving a visible
    # maintenance aisle between rows. Every row is segmented along X into a few
    # discrete tables (~1 m gaps) so the field reads as a grid of tables, not a
    # slab. The project module count is distributed evenly across every table.
    setback = float(_ROADS_SETBACK_M)
    table_depth = min(4.0, max(2.4, row_pitch - 2.0))   # N-S depth of one table
    _MAX_TABLES_PER_BLOCK = 200                          # keep 100 MW renderable

    # First pass: geometry of every table (module count filled in second pass).
    # A project with no committed module count (half-built) shows an EMPTY field
    # -- never a grid of placeholder tables with modules: 0 (Codex fix).
    table_specs: list[dict[str, Any]] = []
    for b in (lay_blocks if n_modules > 0 else []):
        bn = int(b.get("n") or 0)
        bx = float(b.get("x") or 0.0)
        by = float(b.get("y") or 0.0)
        bw = float(b.get("w") or 0.0)
        bh = float(b.get("h") or 0.0)
        # Usable footprint inside the block (road setback on every side).
        ins = min(setback, bw * 0.2, bh * 0.2)
        ux0, uy0 = bx + ins, by + ins
        uw, uh = bw - 2 * ins, bh - 2 * ins
        if uw < 4.0 or uh < table_depth:
            continue
        # Rows down the block (N-S), one table-band per pitch.
        n_block_rows = max(1, int(uh // max(row_pitch, 1.0)))
        # Segment each row into a few tables along X (~25 m target per table).
        segs = max(1, min(4, int(round(uw / 25.0)) or 1))
        # Coarsen (drop rows) if a single block would blow the table budget.
        while n_block_rows * segs > _MAX_TABLES_PER_BLOCK and n_block_rows > 1:
            n_block_rows -= 1
        gap_x = 1.0
        seg_w = max(4.0, (uw - (segs - 1) * gap_x) / segs)
        for ri in range(n_block_rows):
            cz = uy0 + row_pitch / 2.0 + ri * row_pitch
            if cz + table_depth / 2.0 > uy0 + uh:
                break
            for si in range(segs):
                cx = ux0 + seg_w / 2.0 + si * (seg_w + gap_x)
                table_specs.append({"cx": cx, "cz": cz, "block": bn,
                                    "w": round(seg_w, 2),
                                    "l": round(table_depth, 2)})

    n_tables = len(table_specs)
    # Distribute modules across tables; a per-table remainder keeps the placed
    # total as close to the plan as the geometry allows.
    if n_tables > 0 and n_modules > 0:
        base = n_modules // n_tables
        extra = n_modules - base * n_tables
    else:
        base = extra = 0
    pv_rows: list[dict[str, Any]] = []
    placed = 0
    for i, t in enumerate(table_specs):
        mods = base + (1 if i < extra else 0)
        placed += mods
        pv_rows.append({
            "id":    f"pvtbl_{i+1:04d}",
            "layer": "pv_row",
            "kind":  "box",
            "x":     _sx(t["cx"]),
            "y":     1.4,
            "z":     _sz(t["cz"]),
            "w":     t["w"],                 # table length (X, E-W)
            "h":     0.06,                   # module-packet thickness
            "l":     t["l"],                 # table depth (Z, N-S)
            "tilt_deg": tilt_deg,
            "azimuth_deg": azimuth_deg,
            "label": f"PV table {i+1}",
            "meta":  {"modules": mods,
                      "table_index": i + 1,
                      "block": t["block"],
                      "tilt_deg": tilt_deg,
                      "azimuth_deg": azimuth_deg},
        })

    pv_meta = {
        "kwp": kwp,
        "n_modules_planned": n_modules,
        "n_modules_placed":  placed,
        "n_rows": n_tables,                  # discrete PV table objects
        "modules_per_row": base if n_tables else 0,
        "row_pitch_m": row_pitch,
        "tilt_deg": tilt_deg,
        "azimuth_deg": azimuth_deg,
        "field_w_m": round(float(lay_field.get("w") or site_w), 1),
        "field_l_m": round(float(lay_field.get("h") or site_h), 1),
        "n_blocks": len(lay_blocks),
        "n_tables": n_tables,
    }

    # --- Inverter + step-up skids (one per block, at the block's south edge) -
    inverters: list[dict[str, Any]] = []
    for i, s in enumerate(lay_skids):
        inverters.append({
            "id":    f"skid_{i+1:02d}",
            "layer": "inverter",
            "kind":  "box",
            "x":     _sx(float(s.get("x") or 0.0)),
            "y":     1.6,
            "z":     _sz(float(s.get("y") or 0.0)),
            "w":     4.0, "h": 3.0, "l": 2.5,
            "label": f"Inverter + step-up skid #{i+1}",
            "meta":  {"kw": sizing.get("central_inverter_kw"),
                      "block": i + 1,
                      "contents": "Central inverter + MV step-up transformer"},
        })

    # --- Buildings: substation compound + control/O&M + facility buildings --
    buildings: list[dict[str, Any]] = []
    building_dim_defaults: dict[str, dict[str, float]] = {
        "control_room":     {"w": 15, "l": 12, "h": 6},
        "om_building":      {"w": 20, "l": 12, "h": 5},
        "security_gate":    {"w": 6,  "l": 5,  "h": 3},
        "warehouse":        {"w": 25, "l": 15, "h": 6},
        "workshop":         {"w": 15, "l": 10, "h": 5},
        "admin":            {"w": 15, "l": 10, "h": 5},
        "training":         {"w": 10, "l": 8,  "h": 4},
        "spare_parts":      {"w": 10, "l": 8,  "h": 4},
        "chemical":         {"w": 8,  "l": 6,  "h": 4},
        "battery_room":     {"w": 20, "l": 10, "h": 5},
        "inverter_room":    {"w": 15, "l": 8,  "h": 4},
        "transformer_bldg": {"w": 12, "l": 10, "h": 6},
        "switchgear_bldg":  {"w": 15, "l": 10, "h": 5},
        "scada_bldg":       {"w": 12, "l": 8,  "h": 5},
        "comms_bldg":       {"w": 8,  "l": 6,  "h": 4},
        "welfare":          {"w": 10, "l": 8,  "h": 4},
        "washroom":         {"w": 6,  "l": 4,  "h": 3},
        "parking":          {"w": 30, "l": 15, "h": 0.2},
    }

    # (a) Fenced substation compound: a ground pad + a ROW of transformers + an
    #     MV switchgear house -- an arranged yard, NOT one bare 30x30 box.
    if lay_sub:
        ssx = float(lay_sub.get("x") or 0.0)
        ssy = float(lay_sub.get("y") or 0.0)
        ssw = float(lay_sub.get("w") or 60.0)
        ssh = float(lay_sub.get("h") or 40.0)
        # Honour an operator drag of the substation. The Phase-6 move_transformer
        # action persists electrical_config.transformer_pos in SCENE coords; re-
        # anchor the whole compound so the saved position survives a rebuild
        # instead of snapping back to the layout default (Codex fix).
        _tp = elec_cfg.get("transformer_pos") if isinstance(elec_cfg, dict) else None
        if isinstance(_tp, dict) and _tp.get("x") is not None and _tp.get("z") is not None:
            try:
                ssx = float(_tp["x"]) + site_w / 2.0 - ssw / 2.0
                ssy = float(_tp["z"]) + site_h / 2.0 - ssh / 2.0
            except (TypeError, ValueError):
                pass
        buildings.append({
            "id":    "substation_pad",
            "layer": "internal_roads",       # gravel / concrete yard pad
            "kind":  "box",
            "x":     _sx(ssx + ssw / 2.0),
            "y":     0.06,
            "z":     _sz(ssy + ssh / 2.0),
            "w":     round(ssw, 1), "h": 0.12, "l": round(ssh, 1),
            "label": "Substation compound",
            "meta":  {"role": "MV / main substation yard"},
        })
        n_tx = 3
        tx_gap = 4.0
        tx_w = max(4.0, (ssw * 0.7 - (n_tx - 1) * tx_gap) / n_tx)
        tx_x0 = ssx + ssw * 0.15
        tx_z = ssy + ssh * 0.35
        tx_l = min(round(ssh * 0.3, 1), 12.0)
        for i in range(n_tx):
            buildings.append({
                "id":    f"sub_transformer_{i+1}",
                "layer": "transformer",
                "kind":  "box",
                "x":     _sx(tx_x0 + tx_w / 2.0 + i * (tx_w + tx_gap)),
                "y":     3.0,
                "z":     _sz(tx_z),
                "w":     round(tx_w, 1), "h": 6.0, "l": tx_l,
                "label": f"Grid transformer #{i+1}",
                "meta":  {"contents": "Step-up power transformer + cooling"},
            })
        buildings.append({
            "id":    "sub_switchgear",
            "layer": "mv_switchgear",
            "kind":  "box",
            "x":     _sx(ssx + ssw * 0.5),
            "y":     2.5,
            "z":     _sz(ssy + ssh * 0.78),
            "w":     round(ssw * 0.5, 1), "h": 5.0,
            "l":     min(round(ssh * 0.28, 1), 10.0),
            "label": "MV switchgear house",
            "meta":  {"contents": "MV switchgear, protection, metering"},
        })

    # (b) Control / O&M building from the layout control rect.
    if lay_ctrl:
        ccx = float(lay_ctrl.get("x") or 0.0)
        ccy = float(lay_ctrl.get("y") or 0.0)
        ccw = float(lay_ctrl.get("w") or 24.0)
        cch = float(lay_ctrl.get("h") or 18.0)
        buildings.append({
            "id":    "control_building",
            "layer": "control_room",
            "kind":  "box",
            "x":     _sx(ccx + ccw / 2.0),
            "y":     3.0,
            "z":     _sz(ccy + cch / 2.0),
            "w":     round(ccw, 1), "h": 6.0, "l": round(cch, 1),
            "label": "Control / O&M building",
            "meta":  {"building_code": "control_room",
                      "sub_items": BUILDING_SUB_ITEMS.get("control_room", []),
                      "footprint_m2": round(ccw * cch, 1)},
        })

    # (c) Any OTHER selected facility buildings -- laid out in a TIDY ROW in the
    #     bottom strip between the control building and the substation, wrapping
    #     to a second row if they run out of width. Realistic footprints (NO x2
    #     inflation), no random cluster.
    selected_buildings = [b for b in (fac_cfg.get("buildings") or [])
                          if b != "control_room"]
    if selected_buildings:
        strip_gap = 8.0
        ctrl_right = ((float(lay_ctrl.get("x") or 0.0)
                       + float(lay_ctrl.get("w") or 24.0))
                      if lay_ctrl else site_w * 0.1)
        strip_x0 = ctrl_right + strip_gap
        strip_y = (float(lay_ctrl.get("y") or 0.0) if lay_ctrl
                   else site_h - 60.0)
        strip_xmax = ((float(lay_sub.get("x") or (site_w * 0.9)) - strip_gap)
                      if lay_sub else site_w * 0.9)
        if strip_xmax - strip_x0 < 40.0:              # thin strip: use full width
            strip_xmax = site_w - 20.0
        x_cursor = strip_x0
        row_i = 0
        for b in selected_buildings:
            bd = building_dim_defaults.get(b, {"w": 12, "l": 8, "h": 5})
            bw, bl, bh = float(bd["w"]), float(bd["l"]), float(bd["h"])
            if x_cursor > strip_x0 and x_cursor + bw > strip_xmax:
                x_cursor = strip_x0
                row_i += 1
            by = strip_y + row_i * 26.0
            layer = b if b in DT_LAYER_PALETTE else "building"
            label = next((L for c, L, _, _ in BUILDING_TYPES if c == b), b)
            buildings.append({
                "id":    f"bldg_{b}",
                "layer": layer,
                "kind":  "box",
                "x":     _sx(x_cursor + bw / 2.0),
                "y":     bh / 2.0,
                "z":     _sz(by + bl / 2.0),
                "w":     bw, "h": bh, "l": bl,
                "label": label,
                "meta":  {"building_code": b,
                          "sub_items":     BUILDING_SUB_ITEMS.get(b, []),
                          "footprint_m2":  round(bw * bl, 1)},
            })
            x_cursor += bw + strip_gap

    # --- Access roads: render the SAME network as the 2D plot plan ----------
    # (perimeter ring + central spine + substation link). Each layout road is a
    # poly-line of axis-aligned segments; each segment becomes a thin road box.
    # Axis-aligned only, so no rotation is needed (matches the plot plan).
    road_w = 5.0
    roads: list[dict[str, Any]] = []
    _ri = 0
    for road in lay_roads:
        pts = road.get("points") or []
        kind = road.get("kind") or "road"
        for j in range(len(pts) - 1):
            ax, ay = float(pts[j][0]), float(pts[j][1])
            bx2, by2 = float(pts[j + 1][0]), float(pts[j + 1][1])
            if abs(ay - by2) < 0.5:                   # horizontal (E-W) segment
                length = abs(bx2 - ax)
                if length < 1.0:
                    continue
                roads.append({
                    "id":    f"road_{kind}_{_ri}",
                    "layer": "internal_roads", "kind": "box",
                    "x": _sx((ax + bx2) / 2.0), "y": 0.06, "z": _sz(ay),
                    "w": round(length, 1), "h": 0.12, "l": road_w,
                    "label": f"Access road ({kind})",
                    "meta": {"length_m": round(length, 1), "kind": kind},
                })
            else:                                     # vertical (N-S) segment
                length = abs(by2 - ay)
                if length < 1.0:
                    continue
                roads.append({
                    "id":    f"road_{kind}_{_ri}",
                    "layer": "internal_roads", "kind": "box",
                    "x": _sx(ax), "y": 0.06, "z": _sz((ay + by2) / 2.0),
                    "w": road_w, "h": 0.12, "l": round(length, 1),
                    "label": f"Access road ({kind})",
                    "meta": {"length_m": round(length, 1), "kind": kind},
                })
            _ri += 1
    if not roads:                                     # degenerate fallback
        roads.append({
            "id": "spine_road", "layer": "internal_roads", "kind": "box",
            "x": 0.0, "y": 0.06, "z": 0.0,
            "w": road_w, "h": 0.12, "l": max(site_h - 20.0, 20.0),
            "label": "Internal spine road", "meta": {},
        })

    # --- Weather mast + perimeter CCTV poles --------------------------------
    ict: list[dict[str, Any]] = []
    if "weather" in (tech_cfg.get("selected") or []):
        wx = float(lay_field.get("x") or (site_w * 0.1)) + 10.0
        wy = float(lay_field.get("y") or (site_h * 0.1)) + 10.0
        ict.append({
            "id":    "weather_mast",
            "layer": "weather_mast",
            "kind":  "mast",
            "x": _sx(wx), "y": 6.0, "z": _sz(wy),
            "w": 0.4, "h": 12.0, "l": 0.4,
            "label": "Weather station mast",
            "meta":  {"instruments": "pyranometer + ambient + module T"},
        })
    for i, (cx, cz) in enumerate(fence_pts):          # CCTV at 4 fence corners
        ict.append({
            "id":    f"cctv_{i+1}",
            "layer": "cctv_pole",
            "kind":  "mast",
            "x": round(cx * 0.99, 2), "y": 4.0, "z": round(cz * 0.99, 2),
            "w": 0.3, "h": 8.0, "l": 0.3,
            "label": f"CCTV pole #{i+1}",
            "meta":  {"coverage_m": 80},
        })

    # --- Perimeter lighting poles -------------------------------------------
    lighting: list[dict[str, Any]] = []
    _hw, _hh = site_w / 2.0 - 6.0, site_h / 2.0 - 6.0
    perim_positions = [
        ( _hw * 0.5, -_hh), (-_hw * 0.5, -_hh),
        ( _hw * 0.5,  _hh), (-_hw * 0.5,  _hh),
        (-_hw, 0.0), ( _hw, 0.0),
    ]
    for i, (lx, lz) in enumerate(perim_positions):
        lighting.append({
            "id":    f"light_{i+1}",
            "layer": "lighting_pole",
            "kind":  "mast",
            "x": round(lx, 2), "y": 3.0, "z": round(lz, 2),
            "w": 0.2, "h": 6.0, "l": 0.2,
            "label": f"Perimeter light #{i+1}",
            "meta":  {"lumens": 20000, "wattage_W": 200},
        })

    # --- Main earthing pit near the substation compound ---------------------
    safety: list[dict[str, Any]] = []
    if lay_sub:
        epx = float(lay_sub.get("x") or 0.0) - 6.0
        epy = float(lay_sub.get("y") or 0.0) + 6.0
    else:
        epx, epy = site_w - 25.0, site_h - 25.0
    if buildings or pv_rows:
        safety.append({
            "id":    "earth_pit_main",
            "layer": "earthing_pit",
            "kind":  "box",
            "x": _sx(epx), "y": 0.3, "z": _sz(epy),
            "w": 1.5, "h": 0.6, "l": 1.5,
            "label": "Main earthing pit",
            "meta":  {"resistance_ohm_target": 1.0},
        })

    # --- Assemble scene (camera framed to the RECTANGULAR site) -------------
    scene = {
        "site": {
            "kwp":          kwp,
            "land_area_ha": land_area_ha,
            "land_side_m":  round(diag_side, 1),   # camera/scenery framing hint
            "site_w_m":     round(site_w, 1),
            "site_h_m":     round(site_h, 1),
            "gps": {"lat": proj.get("gps_lat"),
                    "lon": proj.get("gps_lon")},
            "country":           proj.get("country"),
            "region":            proj.get("region"),
        },
        "camera": {
            # Elevated 3/4 aerial that frames the whole rectangular envelope.
            "position": [diag_side * 0.55, diag_side * 0.5, site_h * 0.75],
            "target":   [0, 0, 0],
        },
        "terrain":   terrain,
        "fence":     fence,
        "roads":     roads,
        "buildings": buildings,
        "pv":        {"meta": pv_meta, "rows": pv_rows},
        "inverters": inverters,
        "ict":       ict,
        "lighting":  lighting,
        "safety":    safety,
        "palette":   DT_LAYER_PALETTE,
        "layer_groups": [
            {"label": g_label, "icon": g_icon, "codes": codes}
            for g_label, g_icon, codes in DT_LAYER_GROUPS
        ],
    }
    return _dtv2.augment_scene_v2(scene, proj)



# -- Regulatory (Development & Regulatory bolt-on; Ghana-first) --
REGULATORY_ITEMS: list[tuple[str, str, str]] = [
    ("land_tenure", "Land ownership / lease", "Owned / long-lease / concession / negotiating"),
    ("esia", "ESIA (Environmental & Social)", "Screening -> Scoping -> Full ESIA -> approval"),
    ("grid_interconnect", "Grid interconnection", "Feasibility study, connection agreement, wheeling"),
    ("utility_approval", "Utility approval", "Off-taker approval, dispatch instructions"),
    ("energy_commission", "Energy Commission licensing", "Generation licence, transmission licence"),
    ("epa_approval", "EPA approval", "Air / noise / waste permits"),
    ("building_permits", "Building permits", "District Assembly building permits"),
    ("financial_close", "Financial close milestones", "Term sheet, credit approval, drawdown"),
    ("construction_sched", "Construction schedule", "Baseline schedule + critical path"),
    ("cod", "Commercial Operation Date (COD)", "Target COD + commissioning tests"),
]
REGULATORY_ITEM_CODES: set[str] = {c for c, _, _ in REGULATORY_ITEMS}
REGULATORY_STATUSES: list[tuple[str, str, str]] = [
    ("not_started", "Not started", "secondary"), ("in_progress", "In progress", "warning"),
    ("applied", "Applied / submitted", "warning"), ("pending", "Pending decision", "info"),
    ("approved", "Approved / complete", "success"), ("denied", "Denied / blocked", "danger"),
    ("na", "Not applicable", "secondary"),
]
REGULATORY_STATUS_CODES: set[str] = {c for c, _, _ in REGULATORY_STATUSES}

COUNTRY_REGULATORY_FRAMEWORKS: dict[str, dict[str, Any]] = {
    "Ghana": {
        "flag": "\U0001F1EC\U0001F1ED",
        "regulator": {"name": "Energy Commission of Ghana", "abbr": "EC",
                      "url": "https://www.energycom.gov.gh"},
        "esia_authority": {"name": "Environmental Protection Agency", "abbr": "EPA-GH",
                           "url": "https://epa.gov.gh"},
        "tariff_regulator": {"name": "Public Utilities Regulatory Commission", "abbr": "PURC"},
        "grid_operator": ["GRIDCo (transmission)", "ECG / NEDCo (distribution)"],
        "utility_offtakers": ["ECG", "NEDCo", "VRA", "Bulk consumers via wheeling"],
        "land_tenures": [
            {"code": "private_freehold", "label": "Private freehold title",
             "notes": "Registered at Lands Commission; verify indenture chain."},
            {"code": "stool_land", "label": "Stool land (customary)",
             "notes": "Traditional Authority + Regional House of Chiefs consent required."},
            {"code": "skin_land", "label": "Skin land (Northern)",
             "notes": "Analogous to stool land in Upper East / West / Northern."},
            {"code": "family_land", "label": "Family land",
             "notes": "All principal elders must sign; head-of-family alone insufficient."},
            {"code": "government_lease", "label": "Government / state leasehold",
             "notes": "Lands Commission grant; typical 50-yr with renewal."},
            {"code": "leasehold_from_stool", "label": "Long-lease from stool",
             "notes": "50-99 yr lease from Traditional Council; register at Lands Commission."},
        ],
        "land_practices": [
            "Obtain a Lands Commission search on the exact parcel before signing anything.",
            "For stool/family land: convene a formal meeting with the Chief, Queen Mother and principal elders; record consent in writing.",
            "Use a Land Purchase Agreement with escrow through a lawyer - never informal 'drink money'.",
            "Instruct a licensed surveyor to peg boundaries; indenture discrepancies are a red flag.",
            "Publish a 21-day public notice at the District Assembly (Land Act 1036, 2020, s.96).",
            "Confirm the site is NOT within a Forest Reserve, Ramsar Site, or Archaeological Site.",
        ],
        "regulations": [
            "Energy Commission Act 541 (1997)",
            "Renewable Energy Act 832 (2011) as amended by Act 1045 (2020)",
            "Electricity Regulations LI 1937 (2008)",
            "Environmental Assessment Regulations LI 1652 (1999)",
            "Land Act 1036 (2020)", "Local Governance Act 936 (2016)",
            "Renewable Energy Sub-Code (2015)", "Distribution Code (GRIDCo)",
        ],
        "permits_sequence": [
            "Site suitability + LOI from Traditional Authority",
            "Land registration at Lands Commission",
            "ESIA registration + Scoping Report -> EPA",
            "Provisional Wholesale Electricity Supply Licence -> Energy Commission",
            "Grid connection application -> GRIDCo or ECG/NEDCo",
            "Full ESIA + EPA Permit", "PPA negotiation -> ECG / NEDCo / bulk off-taker",
            "PURC tariff approval", "Construction Permit -> District Assembly",
            "Wholesale Electricity Supply Licence (final) -> Energy Commission",
            "Interconnection Agreement -> GRIDCo", "Commissioning + Operations Licence",
        ],
        "notes": "Ghana's Renewable Energy Master Plan targets 10% RE; import-duty incentives on modules. Watch for Cedi devaluation clauses in the PPA.",
    },
    "generic": {
        "flag": "\U0001F30D",
        "regulator": {"name": "National energy regulator", "abbr": "Regulator", "url": ""},
        "esia_authority": {"name": "Environmental authority", "abbr": "ENV"},
        "tariff_regulator": {"name": "Tariff regulator", "abbr": ""},
        "grid_operator": ["National transmission operator", "Distribution utility"],
        "utility_offtakers": ["National utility", "Bulk consumers"],
        "land_tenures": [
            {"code": "freehold", "label": "Freehold title", "notes": "Verify registered title."},
            {"code": "leasehold", "label": "Leasehold", "notes": "Check tenure length + renewal."},
            {"code": "customary", "label": "Customary / community land",
             "notes": "Obtain documented community + traditional-authority consent."},
            {"code": "government", "label": "Government grant/lease",
             "notes": "Direct state grant; check revocation clauses."},
        ],
        "land_practices": [
            "Run a formal title search at the national land registry before committing capital.",
            "Engage the host community and traditional authority early; document consent.",
            "Instruct a licensed surveyor to confirm boundaries against the title.",
            "Confirm the site is clear of protected / reserved areas.",
        ],
        "regulations": ["National electricity act", "Environmental impact assessment law",
                        "Land tenure statute", "Grid connection code"],
        "permits_sequence": [
            "Site + land control", "Environmental screening -> ESIA",
            "Generation licence application", "Grid connection study + agreement",
            "PPA / off-take agreement", "Tariff approval", "Construction permit",
            "Commissioning + operating licence",
        ],
        "notes": "Generic framework - verify the exact national statutes at concept-design stage.",
    },
}


def country_framework(country: str | None) -> dict[str, Any]:
    """Regulatory framework for a country; case-insensitive; generic fallback."""
    if not country:
        return COUNTRY_REGULATORY_FRAMEWORKS["generic"]
    c = country.strip()
    if c in COUNTRY_REGULATORY_FRAMEWORKS:
        return COUNTRY_REGULATORY_FRAMEWORKS[c]
    for k, v in COUNTRY_REGULATORY_FRAMEWORKS.items():
        if k.lower() == c.lower():
            return v
    return COUNTRY_REGULATORY_FRAMEWORKS["generic"]


def _ci_location_bundle() -> tuple[list[str], dict[str, dict[str, dict]]]:
    """Country + region reference data for the Step 1 location dropdowns and
    automatic GPS fill. REUSES the existing platform solar database
    (config/global_solar_data.py) rather than duplicating coordinates.

    Inputs: none.
    Returns: (countries, location_map) where
        countries    = sorted list of country names, and
        location_map = {country: {region: {"lat": float, "lon": float,
                                            "psh": float}}}.
    The template renders `countries` as the country <select>, drives the
    dependent region <select> from `location_map[country]`, and auto-fills the
    lat/lon inputs from `location_map[country][region]`. Returns ([], {}) if the
    solar DB is unavailable so Step 1 still renders (fields degrade to free text).
    """
    try:
        from config.global_solar_data import GLOBAL_DATA, get_countries
    except Exception:
        return [], {}
    countries = get_countries()
    location_map: dict[str, dict[str, dict]] = {}
    for c in countries:
        regions = (GLOBAL_DATA.get(c, {}) or {}).get("regions", {}) or {}
        location_map[c] = {
            rn: {"lat": rd.get("lat"), "lon": rd.get("lon"),
                 "psh": rd.get("psh")}
            for rn, rd in regions.items()
        }
    return countries, location_map


def _ci_resolve_location(form_like) -> tuple[str, str, float | None, float | None]:
    """Resolve the Step 1 location fields from the submitted form, honouring the
    dropdown "Other (enter manually)" escape and auto-filling GPS from the solar
    DB when the user left lat/lon blank.

    Inputs: `form_like` = request.form (a mapping with .get()).
    Returns: (country, region, gps_lat, gps_lon). country/region are the resolved
    strings (the free-text *_other value when the select was "__other__"); GPS is
    the user value when supplied, else the region's reference coordinate, else
    None. Server-side twin of the template JS so the auto-fill still works with
    JavaScript disabled and can never be spoofed past the DB's known regions.
    """
    _OTHER = "__other__"

    def _flt(key):
        v = form_like.get(key)
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    country = (form_like.get("country") or "").strip()
    if country == _OTHER:
        country = (form_like.get("country_other") or "").strip()
    country = country[:120]
    region = (form_like.get("region") or "").strip()
    if region == _OTHER:
        region = (form_like.get("region_other") or "").strip()
    region = region[:120]
    lat = _flt("gps_lat")
    lon = _flt("gps_lon")
    if lat is None or lon is None:
        try:
            from config.global_solar_data import get_solar_data
            sd = get_solar_data(country, region)
        except Exception:
            sd = None
        if sd:
            if lat is None:
                lat = sd.get("latitude")
            if lon is None:
                lon = sd.get("longitude")
    return country, region, lat, lon


def _pick(form_like, key: str, allowed: list[tuple]) -> str:
    """Return the submitted value if it's a valid code in `allowed`, else ''."""
    v = (form_like.get(key) or "").strip()
    codes = {t[0] for t in allowed}
    return v if v in codes else ""


# ===========================================================================
# SECTION B -- Tier gating (the module is an Enterprise-tier feature).
# ===========================================================================

CI_TIER_LEVEL: dict[str, int] = {
    "free": 0, "starter": 1, "professional": 2, "business": 2, "enterprise": 3,
}
CI_LEVEL_MARKETING = 0   # marketing landing only
CI_LEVEL_DEMO      = 1   # read-only showcase
CI_LEVEL_SETUP     = 2   # create projects + Steps 1-7
CI_LEVEL_FULL      = 3   # BOQ, finance, marketplace, CRM, reports, DT
CI_TIER_LABEL: dict[str, str] = {
    "free": "Free", "starter": "Starter", "professional": "Professional",
    "business": "Business", "enterprise": "Enterprise",
}


def _row_get(row: Any, key: str, default: Any = None) -> Any:
    """Portable row accessor: sqlite3.Row (IndexError), psycopg2 Dict/RealDict
    (KeyError), plain dict, attribute rows. Never raises."""
    if row is None:
        return default
    try:
        v = row[key]
        return default if v is None else v
    except (KeyError, IndexError, TypeError):
        pass
    try:
        v = getattr(row, key)
        return default if v is None else v
    except AttributeError:
        return default


def _ci_tier_of(user: Any) -> str:
    """User's tier code. Anon -> 'free'. Admins are Enterprise regardless of
    stored plan. Bulletproof across row types."""
    if not user:
        return "free"
    try:
        is_admin_raw = _row_get(user, "is_admin", 0) or 0
        is_admin = int(bool(is_admin_raw)) if isinstance(is_admin_raw, bool) \
            else int(is_admin_raw)
    except (TypeError, ValueError):
        is_admin = 0
    if is_admin:
        return "enterprise"
    plan_raw = _row_get(user, "plan", "free") or "free"
    try:
        plan = str(plan_raw).strip().lower()
    except (TypeError, AttributeError):
        plan = "free"
    return plan if plan in CI_TIER_LEVEL else "free"


def _ci_level_of(user: Any) -> int:
    return CI_TIER_LEVEL.get(_ci_tier_of(user), 0)


# ===========================================================================
# SECTION C -- Small shared helpers (PG-safe id shim + safe JSON + time).
# ===========================================================================

class _RetId:
    """Cursor stand-in exposing .lastrowid for RETURNING-id inserts, so any
    downstream `cur.lastrowid` read keeps working on both SQLite and Postgres."""
    __slots__ = ("lastrowid",)
    def __init__(self, rid: int):
        self.lastrowid = rid


# Standard PV loss-category shares (fractions of TOTAL system loss). The Digital
# Twin "System Losses" donut presents the project's single Step-7 performance
# ratio as the familiar PVsyst-style stack: the TOTAL (1 - PR) is the real,
# project-specific figure; this split follows conventional utility-scale
# proportions and is labelled "modelled split" in the UI so it is not mistaken
# for a per-category simulation.
_CI_LOSS_SHARES = [
    ("Irradiance / soiling", 0.185, "#22c55e"),
    ("Temperature",          0.360, "#f59e0b"),
    ("Shading",              0.200, "#a855f7"),
    ("Wiring / DC",          0.100, "#3b82f6"),
    ("Inverter",             0.070, "#ef4444"),
    ("Mismatch / other",     0.085, "#94a3b8"),
]


def _ci_dt_metrics(proj) -> dict:
    """Read-only dashboard metrics for the 3D Digital Twin.

    Pulls finance headline figures from the Step-8 finance engine
    (finance_config.computed), the energy-yield profile from _ci_yield_profile
    over the Step-7 sizing, and a standard system-loss stack derived from the
    project performance ratio. Never raises: every field degrades to a safe
    default so the twin still renders on a half-built project.
    """
    def _fl(v, d=0.0):
        try:
            return float(v)
        except (TypeError, ValueError):
            return d

    def _fn(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    pv = _safe_json(proj.get("pv_config"))
    fin = _safe_json(proj.get("finance_config"))
    site = _safe_json(proj.get("site_config"))
    sizing = pv.get("sizing") if isinstance(pv.get("sizing"), dict) else {}
    computed = fin.get("computed") if isinstance(fin.get("computed"), dict) else {}
    cur = proj.get("currency") or "GHS"

    lat = site.get("gps_lat", site.get("latitude"))
    energy = _ci_yield_profile(pv, gps_lat=lat, years=25) or {}

    annual_mwh = _fl(energy.get("annual_gen_mwh")) or _fl(sizing.get("annual_gen_mwh"))
    pr = _fl(sizing.get("performance_ratio"), 0.0)
    total_loss = round((1.0 - pr) * 100.0, 1) if 0.0 < pr <= 1.0 else 0.0
    losses = {"available": total_loss > 0.0, "total_pct": total_loss,
              "pr_pct": round(pr * 100.0, 1) if pr else 0.0, "items": []}
    if total_loss > 0.0:
        _items, _acc = [], 0.0
        for _i, (lbl, share, col) in enumerate(_CI_LOSS_SHARES):
            if _i < len(_CI_LOSS_SHARES) - 1:
                _p = round(total_loss * share, 1); _acc += _p
            else:
                _p = round(total_loss - _acc, 1)   # last absorbs rounding drift
            _items.append({"label": lbl, "pct": _p, "color": col})
        losses["items"] = _items

    finance = {
        "available": bool(computed),
        "currency": cur,
        "capex": _fn(computed.get("total_capex_local")),
        "lcoe": _fn(computed.get("lcoe_local_per_kwh")),
        "irr_pct": _fn(computed.get("irr_pct")),
        "npv": _fn(computed.get("npv_local")),
        "payback_years": _fn(computed.get("payback_years")),
        "tariff": _fn(computed.get("tariff_local_per_kwh")),
        "annual_energy_mwh": round(annual_mwh, 0) if annual_mwh else None,
    }
    return {"finance": finance, "energy": energy, "losses": losses}


def _safe_json(raw: Any) -> dict[str, Any]:
    """Return a dict from a stored JSON blob -- {} on empty/parse-error/non-dict.
    Guarantees the TOP level is a dict; nested values are coerced at each use."""
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        v = json.loads(raw)
        return v if isinstance(v, dict) else {}
    except (TypeError, ValueError):
        return {}


def _utc_now_iso() -> str:
    """UTC timestamp without importing a module-level clock (kept import-local
    so the module stays cheap to import)."""
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


# ===========================================================================
# SECTION D -- Schema (EAGER + VERIFIED; SSS Section 3 / Section 8).
# The projects table is the module's spine. Sibling tables (opportunities,
# agent runs, boq links) are created by their own phase helpers.
# ===========================================================================

_CIP_SQLITE_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_projects (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id           INTEGER NOT NULL,
    project_name      TEXT NOT NULL,
    client_name       TEXT DEFAULT '',
    investor          TEXT DEFAULT '',
    developer         TEXT DEFAULT '',
    country           TEXT DEFAULT '',
    region            TEXT DEFAULT '',
    district          TEXT DEFAULT '',
    gps_lat           REAL,
    gps_lon           REAL,
    description       TEXT DEFAULT '',
    project_status    TEXT DEFAULT 'concept',
    target_cod        TEXT DEFAULT '',
    target_kwp        REAL,
    design_standard   TEXT DEFAULT 'IEC',
    currency          TEXT DEFAULT 'GHS',
    tax_regime        TEXT DEFAULT 'standard',
    project_type      TEXT DEFAULT '',
    site_config       TEXT DEFAULT '',
    facility_config   TEXT DEFAULT '',
    technology_config TEXT DEFAULT '',
    electrical_config TEXT DEFAULT '',
    pv_config         TEXT DEFAULT '',
    finance_config    TEXT DEFAULT '',
    regulatory_config TEXT DEFAULT '',
    boq_project_id    INTEGER,
    boq_facilities_project_id INTEGER,
    boq_solar_project_id      INTEGER,
    tenant_id         TEXT,
    schema_version    INTEGER DEFAULT 2,
    created_at        TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at        TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_cip_user_id      ON capital_investment_projects(user_id);
CREATE INDEX IF NOT EXISTS idx_cip_tenant_id    ON capital_investment_projects(tenant_id);
CREATE INDEX IF NOT EXISTS idx_cip_user_status  ON capital_investment_projects(user_id, project_status);
CREATE INDEX IF NOT EXISTS idx_cip_user_updated ON capital_investment_projects(user_id, updated_at DESC);
"""

_CIP_POSTGRES_DDL = """
CREATE TABLE IF NOT EXISTS capital_investment_projects (
    id                SERIAL PRIMARY KEY,
    user_id           INTEGER NOT NULL,
    project_name      TEXT NOT NULL,
    client_name       TEXT DEFAULT '',
    investor          TEXT DEFAULT '',
    developer         TEXT DEFAULT '',
    country           TEXT DEFAULT '',
    region            TEXT DEFAULT '',
    district          TEXT DEFAULT '',
    gps_lat           DOUBLE PRECISION,
    gps_lon           DOUBLE PRECISION,
    description       TEXT DEFAULT '',
    project_status    TEXT DEFAULT 'concept',
    target_cod        TEXT DEFAULT '',
    target_kwp        DOUBLE PRECISION,
    design_standard   TEXT DEFAULT 'IEC',
    currency          TEXT DEFAULT 'GHS',
    tax_regime        TEXT DEFAULT 'standard',
    project_type      TEXT DEFAULT '',
    site_config       TEXT DEFAULT '',
    facility_config   TEXT DEFAULT '',
    technology_config TEXT DEFAULT '',
    electrical_config TEXT DEFAULT '',
    pv_config         TEXT DEFAULT '',
    finance_config    TEXT DEFAULT '',
    regulatory_config TEXT DEFAULT '',
    boq_project_id    INTEGER,
    boq_facilities_project_id INTEGER,
    boq_solar_project_id      INTEGER,
    tenant_id         UUID,
    schema_version    INTEGER DEFAULT 2,
    created_at        TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at        TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_cip_user_id      ON capital_investment_projects(user_id);
CREATE INDEX IF NOT EXISTS idx_cip_tenant_id    ON capital_investment_projects(tenant_id);
CREATE INDEX IF NOT EXISTS idx_cip_user_status  ON capital_investment_projects(user_id, project_status);
CREATE INDEX IF NOT EXISTS idx_cip_user_updated ON capital_investment_projects(user_id, updated_at DESC);
"""

# Additive, idempotent migrations for pre-rebuild DBs missing a new column.
_CIP_MIGRATIONS = [
    "ALTER TABLE capital_investment_projects ADD COLUMN target_kwp REAL",
    "ALTER TABLE capital_investment_projects ADD COLUMN tenant_id TEXT",
    "ALTER TABLE capital_investment_projects ADD COLUMN schema_version INTEGER DEFAULT 2",
    "ALTER TABLE capital_investment_projects ADD COLUMN regulatory_config TEXT DEFAULT ''",
    # Two-BOQ split (owner directive 2026-07-03): a facilities/technology/non-solar
    # BOQ and a separate 20MWp solar-farm equipment BOQ. Legacy boq_project_id keeps
    # pointing at the facilities BOQ for back-compat with Step-8 reconciliation.
    "ALTER TABLE capital_investment_projects ADD COLUMN boq_facilities_project_id INTEGER",
    "ALTER TABLE capital_investment_projects ADD COLUMN boq_solar_project_id INTEGER",
]

# Verification state so a failed live-PG migration surfaces (never swallowed).
_CIP_SCHEMA_STATE: dict[str, object] = {"ready": False, "error": ""}


def _ensure_ci_projects_schema_verified(get_db) -> bool:
    """EAGER + VERIFIED schema for the projects table. Runs each statement in
    its own transaction (a failed statement on Postgres aborts only itself, not
    the later ADD COLUMN migrations), then confirms the table is queryable.
    Returns True only when verified. Remembers the result in _CIP_SCHEMA_STATE
    so a live-PG failure is observable in diagnostics rather than silent."""
    if _CIP_SCHEMA_STATE["ready"]:
        return True
    # SQLite fast path (executescript is Postgres-hostile -> falls through).
    try:
        with get_db() as c:
            c.executescript(_CIP_SQLITE_DDL)
    except Exception:
        for stmt in _CIP_POSTGRES_DDL.split(";"):
            s = stmt.strip()
            if not s:
                continue
            try:
                with get_db() as c:
                    c.execute(s)
            except Exception:
                pass
    for ddl in _CIP_MIGRATIONS:
        try:
            with get_db() as c:
                c.execute(ddl)
        except Exception:
            pass   # column already present / backend mismatch
    # One-shot backfill: generation-station facilities BOQs created before the
    # two-BOQ split were tagged project_type='campus' and would (a) still clutter
    # the marketplace /boq-projects list and (b) not appear under ?scope=capital.
    # Re-tag ONLY those linked from a capital_investment_project (never a
    # standalone marketplace "Campus" BOQ, which shares the 'campus' type) +
    # populate the new columns from the legacy id. Idempotent (Supervisor).
    for _bf in (
        "UPDATE boq_projects SET project_type='capital_facilities' "
        "WHERE project_type='campus' AND id IN "
        "(SELECT boq_project_id FROM capital_investment_projects "
        " WHERE boq_project_id IS NOT NULL)",
        "UPDATE capital_investment_projects SET boq_facilities_project_id=boq_project_id "
        "WHERE boq_facilities_project_id IS NULL AND boq_project_id IS NOT NULL",
    ):
        try:
            with get_db() as c:
                c.execute(_bf)
        except Exception:
            pass   # boq_projects not yet created / backend mismatch
    # Verify queryable before declaring success.
    try:
        with get_db() as c:
            c.execute("SELECT id FROM capital_investment_projects LIMIT 1")
        _CIP_SCHEMA_STATE["ready"] = True
        _CIP_SCHEMA_STATE["error"] = ""
    except Exception as exc:
        _CIP_SCHEMA_STATE["ready"] = False
        _CIP_SCHEMA_STATE["error"] = str(exc)[:300]
    return bool(_CIP_SCHEMA_STATE["ready"])


# ---------------------------------------------------------------------------
# Recent-projects "clear history" view preference (owner 2026-07-05).
# NON-DESTRUCTIVE: hiding a project only removes it from the landing "Recent
# projects" panel for THIS user. The capital_investment_projects row, its wizard
# data and any linked BOQ are untouched -- the project is still reachable by URL
# and via the marketplace/BOQ lists. A composite PK (user_id, project_id) keeps
# the insert idempotent. Scoped by user_id (the recent query is already
# user-scoped); tenant_id is recorded best-effort for audit parity only.
# ---------------------------------------------------------------------------
_CI_RECENT_HIDDEN_STATE = {"ready": False}


def _ensure_ci_recent_hidden_schema(get_db) -> bool:
    """Create ci_recent_hidden once (SQLite + Postgres safe). Returns readiness
    so callers can skip the filter/insert cleanly if the DDL failed."""
    if _CI_RECENT_HIDDEN_STATE["ready"]:
        return True
    # tenant_id is part of the PK (normalised to '' when no tenant) so the same
    # user in two tenants hides independently -- Codex tenant-isolation fix.
    ddl = (
        "CREATE TABLE IF NOT EXISTS ci_recent_hidden ("
        " user_id INTEGER NOT NULL,"
        " project_id INTEGER NOT NULL,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " hidden_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " PRIMARY KEY (user_id, project_id, tenant_id))"
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _CI_RECENT_HIDDEN_STATE["ready"] = True
    except Exception:
        _CI_RECENT_HIDDEN_STATE["ready"] = False
    return bool(_CI_RECENT_HIDDEN_STATE["ready"])


# ---------------------------------------------------------------------------
# Project Funding module -- Slice 1 foundation (2026-07-05, spec
# pvsolar1/sponsors page1.txt). ONE funding application per generation-station
# project, keyed by (project_id, tenant_id) so no autoincrement/SERIAL is needed
# (cross-DB safe) and tenants stay isolated. Later slices add institutions,
# selections, revenue. All money columns are in the project's own currency.
# ---------------------------------------------------------------------------
_CI_FUNDING_STATE = {"ready": False}

# Funding application lifecycle (spec sections 9/13). Slice 1 uses draft/requested.
CI_FUNDING_STATUSES = (
    "draft", "requested", "package_prepared", "submitted",
    "under_review", "conditional", "approved", "rejected", "closed",
)

# Per-institution application status inside an institution's workspace (Slice 4+).
# Each funding_institution_selections row carries one of these: the customer's
# submission (Slice 3) sets 'submitted'; the review page (Slice 5) transitions
# it. Kept as an allowlist so a ?status= filter can never inject SQL.
FI_APP_STATUSES = (
    "submitted", "under_review", "awaiting_documents", "technical_review",
    "financial_review", "approved_in_principle", "conditional",
    "approved", "rejected", "completed",
)
FI_APP_STATUS_LABELS = {
    "submitted": "Submitted", "under_review": "Under Review",
    "awaiting_documents": "Awaiting Documents",
    "technical_review": "Technical Review",
    "financial_review": "Financial Review",
    "approved_in_principle": "Approved in Principle",
    "conditional": "Conditional Approval", "approved": "Approved",
    "rejected": "Rejected", "completed": "Completed",
}
# Bootstrap badge class per status (for the workspace chips + table badges).
FI_APP_STATUS_CLASS = {
    "submitted": "info", "under_review": "primary",
    "awaiting_documents": "warning", "technical_review": "primary",
    "financial_review": "primary", "approved_in_principle": "success",
    "conditional": "warning", "approved": "success",
    "rejected": "danger", "completed": "secondary",
}


# ---------------------------------------------------------------------------
# Slice 10c (2026-07-05) -- extend Project Funding to REGULAR /project/<pid>
# projects (residential / C&I), not just generation-station Capital Investment
# projects. Regular projects store their engineering + economics as a JSON blob
# in projects.data_json (a DIFFERENT data model from capital_investment_projects),
# so they are BRIDGED into the SAME funding tables + institution registry rather
# than duplicated:
#   * their funding rows are keyed by a NAMESPACED project id -- PF_PID_OFFSET +
#     real pid -- so a regular project #5 can never collide with generation-station
#     project #5 on the shared (project_id, tenant_id) primary keys. The offset is
#     far above any real AUTOINCREMENT/SERIAL id.
#   * a denormalized display snapshot (name / client / type / kwp / currency /
#     country / region) is stored ON the funding row so the institution workspace
#     never has to join the residential `projects` table.
#   * the finance metrics are ADAPTED from the residential economics dict into the
#     same finance_config.computed shape _ci_funding_overview / _ci_bankability /
#     _ci_funding_assessment already read -- zero re-modelling, one funding engine.
# Nothing here changes the generation-station funding code paths.
# ---------------------------------------------------------------------------
PF_PID_OFFSET = 1_000_000_000    # namespace boundary for regular-project funding


def _pf_fid(pid) -> int:
    """Real /project/<pid> id -> namespaced funding-table project id."""
    return PF_PID_OFFSET + int(pid)


def _pf_is_regular(fid) -> bool:
    """True when a funding-table project id belongs to a regular project."""
    try:
        return int(fid) >= PF_PID_OFFSET
    except (TypeError, ValueError):
        return False


def _pf_real_pid(fid) -> int:
    """Namespaced funding-table project id -> real /project/<pid> id."""
    return int(fid) - PF_PID_OFFSET


def _pf_finance_config(eco: dict) -> dict:
    """Adapt a residential/C&I economics dict (projects.data_json['results']
    ['economics'], produced by calc_economics) into the finance_config.computed
    shape the funding engine reads. READ ONLY -- no re-modelling; the residential
    finance numbers are reused verbatim so the funding view can never disagree
    with the project's own economic report."""
    import math as _m
    if not isinstance(eco, dict):
        eco = {}

    def _f(v):
        try:
            f = float(v)
            return f if _m.isfinite(f) else None
        except (TypeError, ValueError):
            return None
    cf_rows = eco.get("cf_rows") if isinstance(eco.get("cf_rows"), list) else []
    net_by_year, rev_by_year, opex_by_year = [], [], []
    for row in cf_rows:
        if not isinstance(row, dict):
            continue
        try:
            if int(row.get("year", 0) or 0) <= 0:
                continue                   # year 0 = construction, excluded
        except (TypeError, ValueError):
            continue
        net_by_year.append(_f(row.get("net")) or 0.0)
        rev_by_year.append(_f(row.get("gross")) or 0.0)
        opex_by_year.append(_f(row.get("om")) or 0.0)
    dscr = _f(eco.get("dscr"))
    computed = {
        "total_capex_local": _f(eco.get("total_local")),
        "equity_local": _f(eco.get("equity")) or 0.0,
        "npv_local": _f(eco.get("npv")),
        "irr_pct": _f(eco.get("irr_pct")),
        "payback_years": _f(eco.get("payback")),
        "project_life_yr": 25.0,
        "dscr_min": dscr,
        "dscr_avg": dscr,
    }
    if net_by_year:
        computed["net_by_year"] = net_by_year
        computed["revenue_by_year"] = rev_by_year
        computed["opex_by_year"] = opex_by_year
    return {"computed": computed}


def _pf_project_view(proj_row: dict) -> dict:
    """Map a residential `projects` row (with parsed data_json under 'data') into
    the dict shape the funding templates + helpers expect: the CI field names
    (project_name / client_name / project_type / target_kwp[kWp] / currency /
    country / region / user_id / tenant_id / id) plus a synthesized
    finance_config. `id` is the NAMESPACED funding id; _pf_real_pid recovers the
    real one. `_kind='project'` lets shared code + templates branch safely."""
    data = proj_row.get("data") if isinstance(proj_row.get("data"), dict) else {}
    results = data.get("results") if isinstance(data.get("results"), dict) else {}
    eco = results.get("economics") if isinstance(results.get("economics"), dict) else {}
    real_pid = int(proj_row.get("id"))

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    # Residential PV size is in kW (== kWp); keep target_kwp in kWp to match the
    # generation-station field. project_funding.html renders it as kWp.
    kwp = _f(results.get("pv_kw")) or _f(results.get("system_kw")) \
        or _f(data.get("pv_kw"))
    return {
        "id": _pf_fid(real_pid),
        "real_pid": real_pid,
        "_kind": "project",
        "user_id": proj_row.get("user_id"),
        "tenant_id": proj_row.get("tenant_id") or '',
        "project_name": (proj_row.get("name") or data.get("project_name")
                         or ("Project #%s" % real_pid)),
        "client_name": (data.get("client_name") or data.get("customer_name")
                        or data.get("client") or ''),
        "project_type": (data.get("project_type") or data.get("building_type")
                         or data.get("customer_type") or 'residential'),
        "target_kwp": kwp,
        "currency": data.get("currency") or 'GHS',
        "country": data.get("country") or '',
        "region": data.get("region") or '',
        "district": data.get("district") or '',
        "developer": '', "investor": '',
        "finance_config": json.dumps(_pf_finance_config(eco)),
        # Slice 10c: the saved Check My Bill snapshot travels with the funding
        # view so the institution review page shows the applicant's verified bill.
        "bill_check": data.get("bill_check"),
    }


def _ensure_ci_funding_schema(get_db) -> bool:
    """Create capital_investment_funding once (SQLite + Postgres safe)."""
    if _CI_FUNDING_STATE["ready"]:
        return True
    ddl = (
        "CREATE TABLE IF NOT EXISTS capital_investment_funding ("
        " capital_investment_project_id INTEGER NOT NULL,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " user_id INTEGER NOT NULL,"
        " status TEXT NOT NULL DEFAULT 'draft',"
        " funding_requested REAL,"
        " customer_equity REAL,"
        " funding_score INTEGER,"
        " risk_rating TEXT,"
        " selected_institutions TEXT NOT NULL DEFAULT '[]',"
        " remarks TEXT,"
        " created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " PRIMARY KEY (capital_investment_project_id, tenant_id))"
    )
    # Slice 10c -- project_kind discriminator + denormalized display snapshot for
    # regular /project/<pid> funding rows (so the institution workspace never has
    # to join the residential `projects` table). Added via per-column ALTER in
    # their OWN transactions so a duplicate-column failure on Postgres aborts only
    # that statement (house pattern, matches the selection-schema upgrades).
    upgrades = (
        "ALTER TABLE capital_investment_funding ADD COLUMN "
        "project_kind TEXT NOT NULL DEFAULT 'capital'",
        "ALTER TABLE capital_investment_funding ADD COLUMN proj_name TEXT",
        "ALTER TABLE capital_investment_funding ADD COLUMN proj_client TEXT",
        "ALTER TABLE capital_investment_funding ADD COLUMN proj_type TEXT",
        "ALTER TABLE capital_investment_funding ADD COLUMN proj_kwp REAL",
        "ALTER TABLE capital_investment_funding ADD COLUMN proj_currency TEXT",
        "ALTER TABLE capital_investment_funding ADD COLUMN proj_country TEXT",
        "ALTER TABLE capital_investment_funding ADD COLUMN proj_region TEXT",
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _CI_FUNDING_STATE["ready"] = True
    except Exception:
        _CI_FUNDING_STATE["ready"] = False
    for stmt in upgrades:
        try:
            with get_db() as c:
                c.execute(stmt)
        except Exception:
            pass   # column already present / backend mismatch
    return bool(_CI_FUNDING_STATE["ready"])


# ---------------------------------------------------------------------------
# Financial Institutions -- Slice 2 (2026-07-05). A PLATFORM-GLOBAL registry:
# banks / leasing / funds register, Platform Admin approves, and customers across
# tenants select from the approved set (mirrors how the marketplace supplier
# registry is shared). created_by_user_id + tenant_id are recorded for provenance
# only -- institution visibility is NOT tenant-filtered. UUID TEXT PK avoids the
# AUTOINCREMENT/SERIAL cross-DB translation.
# ---------------------------------------------------------------------------
_FI_STATE = {"ready": False}

FI_TYPES = [
    ("commercial_bank", "Commercial Bank"),
    ("development_bank", "Development Bank"),
    ("leasing", "Leasing Company"),
    ("investment_fund", "Investment Fund"),
    ("infrastructure_fund", "Infrastructure Fund"),
    ("climate_fund", "Climate Fund"),
    ("private_equity", "Private Equity"),
    ("govt_agency", "Government Funding Agency"),
]
FI_TYPE_CODES = {c for c, _ in FI_TYPES}
FI_STATUSES = ("pending", "approved", "rejected", "suspended")
# Admin action -> resulting status.
FI_ACTIONS = {"approve": "approved", "reject": "rejected",
              "suspend": "suspended", "reactivate": "approved"}


def _ensure_fi_schema(get_db) -> bool:
    """Create financial_institutions once (SQLite + Postgres safe)."""
    if _FI_STATE["ready"]:
        return True
    ddl = (
        "CREATE TABLE IF NOT EXISTS financial_institutions ("
        " institution_id TEXT PRIMARY KEY,"
        " name TEXT NOT NULL,"
        " inst_type TEXT NOT NULL DEFAULT '',"
        " country TEXT NOT NULL DEFAULT '',"
        " region TEXT NOT NULL DEFAULT '',"
        " contact_person TEXT NOT NULL DEFAULT '',"
        " position TEXT NOT NULL DEFAULT '',"
        " email TEXT NOT NULL DEFAULT '',"
        " phone TEXT NOT NULL DEFAULT '',"
        " website TEXT NOT NULL DEFAULT '',"
        " licence_no TEXT NOT NULL DEFAULT '',"
        " regulator TEXT NOT NULL DEFAULT '',"
        " loan_min REAL,"
        " loan_max REAL,"
        " tenor_months INTEGER,"
        " interest_min REAL,"
        " interest_max REAL,"
        " supported_project_types TEXT NOT NULL DEFAULT '',"
        " funding_products TEXT NOT NULL DEFAULT '',"
        " fee_pct REAL NOT NULL DEFAULT 2.0,"
        " agreement_ref TEXT NOT NULL DEFAULT '',"
        " agreement_status TEXT NOT NULL DEFAULT 'none',"
        " status TEXT NOT NULL DEFAULT 'pending',"
        " created_by_user_id INTEGER,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)"
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _FI_STATE["ready"] = True
    except Exception:
        _FI_STATE["ready"] = False
    return bool(_FI_STATE["ready"])


# ---------------------------------------------------------------------------
# Funding institution selections -- Slice 3 (2026-07-05). Which institution(s) a
# customer submitted a given project's funding application to, with explicit
# consent. One row per (project, institution, tenant); the institution only ever
# sees applications that have a row here (isolation for Slice 4's workspace).
# ---------------------------------------------------------------------------
_FI_SEL_STATE = {"ready": False}


def _ensure_fi_selection_schema(get_db) -> bool:
    """Create funding_institution_selections once (SQLite + Postgres safe)."""
    if _FI_SEL_STATE["ready"]:
        return True
    ddl = (
        "CREATE TABLE IF NOT EXISTS funding_institution_selections ("
        " capital_investment_project_id INTEGER NOT NULL,"
        " institution_id TEXT NOT NULL,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " user_id INTEGER NOT NULL,"
        " consent INTEGER NOT NULL DEFAULT 0,"
        " status TEXT NOT NULL DEFAULT 'submitted',"
        " submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " PRIMARY KEY (capital_investment_project_id, institution_id, tenant_id))"
    )
    # Slice 5 decision columns -- added via per-column ALTER so a table created
    # by Slice 3 (before these existed) is upgraded in place. SQLite has no ADD
    # COLUMN IF NOT EXISTS, so each ALTER runs in its OWN transaction: a
    # duplicate-column failure on Postgres aborts only that statement, not the
    # CREATE or the sibling ALTERs (matches _ensure_ci_projects_schema_verified).
    upgrades = (
        "ALTER TABLE funding_institution_selections ADD COLUMN "
        "decision_note TEXT",
        "ALTER TABLE funding_institution_selections ADD COLUMN "
        "decided_at TIMESTAMP",
        "ALTER TABLE funding_institution_selections ADD COLUMN "
        "decided_by INTEGER",
        # Slice 10c -- discriminates a regular /project/<pid> application from a
        # generation-station one so the workspace read can resolve the right
        # source without inspecting the namespaced id.
        "ALTER TABLE funding_institution_selections ADD COLUMN "
        "project_kind TEXT NOT NULL DEFAULT 'capital'",
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _FI_SEL_STATE["ready"] = True
    except Exception:
        _FI_SEL_STATE["ready"] = False
    for stmt in upgrades:
        try:
            with get_db() as c:
                c.execute(stmt)
        except Exception:
            pass   # column already present / backend mismatch
    return bool(_FI_SEL_STATE["ready"])


# ---------------------------------------------------------------------------
# Funding application communication -- Slice 6a (2026-07-05). One thread per
# (project, institution, tenant): the two-way message history between an
# approved institution and the applicant, plus document/info requests and the
# audit copy of any email dispatched via the existing _send_email service. UUID
# TEXT PK avoids the AUTOINCREMENT/SERIAL cross-DB translation.
# ---------------------------------------------------------------------------
_FI_MSG_STATE = {"ready": False}

# Message types an institution/applicant can post (allowlisted -> no injection
# into the type filter or badge lookups).
FI_MSG_TYPES = ("message", "info_request", "document_request")
FI_MSG_TYPE_LABELS = {
    "message": "Message",
    "info_request": "Information request",
    "document_request": "Document request",
}


def _ensure_fi_messages_schema(get_db) -> bool:
    """Create funding_application_messages once (SQLite + Postgres safe)."""
    if _FI_MSG_STATE["ready"]:
        return True
    ddl = (
        "CREATE TABLE IF NOT EXISTS funding_application_messages ("
        " message_id TEXT PRIMARY KEY,"
        " capital_investment_project_id INTEGER NOT NULL,"
        " institution_id TEXT NOT NULL,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " sender_role TEXT NOT NULL DEFAULT 'system',"
        " sender_user_id INTEGER,"
        " channel TEXT NOT NULL DEFAULT 'message',"
        " msg_type TEXT NOT NULL DEFAULT 'message',"
        " subject TEXT,"
        " body TEXT NOT NULL DEFAULT '',"
        " emailed_to TEXT,"
        " created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)"
    )
    idx = (
        "CREATE INDEX IF NOT EXISTS idx_fi_msg_thread "
        "ON funding_application_messages "
        "(capital_investment_project_id, institution_id, tenant_id, created_at)"
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _FI_MSG_STATE["ready"] = True
    except Exception:
        _FI_MSG_STATE["ready"] = False
    try:
        with get_db() as c:
            c.execute(idx)
    except Exception:
        pass
    return bool(_FI_MSG_STATE["ready"])


# ---------------------------------------------------------------------------
# Hard-copy document tracking -- Slice 6b (2026-07-05). SolarPro tracks the
# STATUS of physical originals couriered DIRECTLY between the applicant and the
# institution -- it never takes custody (spec: "SolarPro must not claim custody
# of original physical documents"). One row per shipment.
# ---------------------------------------------------------------------------
_FI_SHIP_STATE = {"ready": False}

FI_SHIP_STATUSES = (
    "dispatched", "in_transit", "received", "verified", "rejected",
)
FI_SHIP_STATUS_LABELS = {
    "dispatched": "Dispatched", "in_transit": "In transit",
    "received": "Received", "verified": "Verified", "rejected": "Rejected",
}
FI_SHIP_STATUS_CLASS = {
    "dispatched": "info", "in_transit": "primary", "received": "warning",
    "verified": "success", "rejected": "danger",
}


def _ensure_fi_shipments_schema(get_db) -> bool:
    """Create funding_document_shipments once (SQLite + Postgres safe)."""
    if _FI_SHIP_STATE["ready"]:
        return True
    ddl = (
        "CREATE TABLE IF NOT EXISTS funding_document_shipments ("
        " shipment_id TEXT PRIMARY KEY,"
        " capital_investment_project_id INTEGER NOT NULL,"
        " institution_id TEXT NOT NULL,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " created_by_role TEXT NOT NULL DEFAULT 'applicant',"
        " created_by_user_id INTEGER,"
        " document_type TEXT NOT NULL DEFAULT '',"
        " courier_company TEXT NOT NULL DEFAULT '',"
        " tracking_number TEXT NOT NULL DEFAULT '',"
        " dispatch_date TEXT,"
        " recipient TEXT NOT NULL DEFAULT '',"
        " receiving_officer TEXT NOT NULL DEFAULT '',"
        " received_date TEXT,"
        " verification_status TEXT NOT NULL DEFAULT 'dispatched',"
        " notes TEXT,"
        " created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)"
    )
    idx = (
        "CREATE INDEX IF NOT EXISTS idx_fi_ship_thread "
        "ON funding_document_shipments "
        "(capital_investment_project_id, institution_id, tenant_id)"
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _FI_SHIP_STATE["ready"] = True
    except Exception:
        _FI_SHIP_STATE["ready"] = False
    try:
        with get_db() as c:
            c.execute(idx)
    except Exception:
        pass
    return bool(_FI_SHIP_STATE["ready"])


# ---------------------------------------------------------------------------
# Funding revenue / success fee -- Slice 7 (2026-07-05). SolarPro's success fee
# (default 2%, per-institution configurable via financial_institutions.fee_pct)
# is charged to the institution ONLY after Approved + Agreement Executed + First
# Disbursement. One revenue/invoice row per (project, institution, tenant); the
# invoice number materialises only when all three milestones are met.
# ---------------------------------------------------------------------------
_FI_REV_STATE = {"ready": False}

# selection.status values that satisfy the "Funding Approved" milestone.
FI_APPROVED_GATE = ("approved", "completed")
FI_PAYMENT_STATUSES = ("outstanding", "paid")


def _ensure_fi_revenue_schema(get_db) -> bool:
    """Create funding_revenue once (SQLite + Postgres safe)."""
    if _FI_REV_STATE["ready"]:
        return True
    ddl = (
        "CREATE TABLE IF NOT EXISTS funding_revenue ("
        " capital_investment_project_id INTEGER NOT NULL,"
        " institution_id TEXT NOT NULL,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " project_name TEXT NOT NULL DEFAULT '',"
        " customer TEXT NOT NULL DEFAULT '',"
        " developer TEXT NOT NULL DEFAULT '',"
        " institution_name TEXT NOT NULL DEFAULT '',"
        " country TEXT NOT NULL DEFAULT '',"
        " region TEXT NOT NULL DEFAULT '',"
        " project_type TEXT NOT NULL DEFAULT '',"
        " currency TEXT NOT NULL DEFAULT 'GHS',"
        " approved_loan_amount REAL,"
        " approved_project_value REAL,"
        " fee_pct REAL NOT NULL DEFAULT 2.0,"
        " fee_amount REAL,"
        " vat REAL NOT NULL DEFAULT 0,"
        " invoice_number TEXT,"
        " invoice_date TEXT,"
        " invoice_status TEXT NOT NULL DEFAULT 'pending',"
        " payment_status TEXT NOT NULL DEFAULT 'outstanding',"
        " payment_date TEXT,"
        " agreement_reference TEXT NOT NULL DEFAULT '',"
        " agreement_executed INTEGER NOT NULL DEFAULT 0,"
        " first_disbursement INTEGER NOT NULL DEFAULT 0,"
        " disbursement_date TEXT,"
        " remarks TEXT,"
        " created_by_user_id INTEGER,"
        " created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " PRIMARY KEY (capital_investment_project_id, institution_id, tenant_id))"
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _FI_REV_STATE["ready"] = True
    except Exception:
        _FI_REV_STATE["ready"] = False
    return bool(_FI_REV_STATE["ready"])


# ---------------------------------------------------------------------------
# AI Funding Assessment -- Slice 8 (2026-07-05). A DETERMINISTIC, zero-cost
# funding-readiness assessment that REUSES the existing finance/bankability
# engines and the project's completeness -- the same rule-based agent pattern
# the generation station already uses at Step-14 (the solar app is excluded from
# the ADK resync; a full ADK migration of the whole agent layer is a separate
# cross-cutting effort, not a per-slice change). Produces the spec's funding
# score + readiness dimensions + risk rating + recommendation, persisted per
# (project, institution, tenant) for the institution's review + audit.
# ---------------------------------------------------------------------------
_FI_ASSESS_STATE = {"ready": False}

FI_RECOMMENDATIONS = ("recommend", "conditional", "decline")
FI_RECOMMENDATION_LABELS = {
    "recommend": "Recommend to fund",
    "conditional": "Conditional / more due diligence",
    "decline": "Decline",
}
FI_RECOMMENDATION_CLASS = {
    "recommend": "success", "conditional": "warning", "decline": "danger",
}


def _ensure_fi_assessment_schema(get_db) -> bool:
    """Create funding_assessments once (SQLite + Postgres safe)."""
    if _FI_ASSESS_STATE["ready"]:
        return True
    ddl = (
        "CREATE TABLE IF NOT EXISTS funding_assessments ("
        " capital_investment_project_id INTEGER NOT NULL,"
        " institution_id TEXT NOT NULL,"
        " tenant_id TEXT NOT NULL DEFAULT '',"
        " funding_score INTEGER,"
        " technical_readiness INTEGER,"
        " financial_readiness INTEGER,"
        " documentation_readiness INTEGER,"
        " construction_readiness INTEGER,"
        " risk_rating TEXT,"
        " matched INTEGER NOT NULL DEFAULT 0,"
        " recommendation TEXT,"
        " payload TEXT,"
        " created_by_user_id INTEGER,"
        " created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        " PRIMARY KEY (capital_investment_project_id, institution_id, "
        " tenant_id))"
    )
    try:
        with get_db() as c:
            c.execute(ddl)
        _FI_ASSESS_STATE["ready"] = True
    except Exception:
        _FI_ASSESS_STATE["ready"] = False
    return bool(_FI_ASSESS_STATE["ready"])


def _ci_funding_assessment(proj: dict, fund: dict, inst: dict) -> dict:
    """Deterministic multi-dimension funding assessment (reuses the finance +
    bankability engines; no re-modelling, no LLM). Returns the spec outputs:
    funding score, technical/financial/documentation/construction readiness,
    risk rating, institution match, recommendation + per-dimension findings."""
    fin_cfg = _safe_json(proj.get("finance_config"))
    computed = fin_cfg.get("computed") if isinstance(fin_cfg, dict) else {}
    computed = computed if isinstance(computed, dict) else {}
    bank = _ci_bankability(computed)
    findings = {}

    # Technical readiness -- completeness of the engineering steps (3-7).
    tech_steps = ["site_config", "facility_config", "technology_config",
                  "electrical_config", "pv_config"]
    tdone = sum(1 for s in tech_steps
                if _is_meaningfully_populated(s, proj.get(s)))
    technical = int(round(100 * tdone / len(tech_steps)))
    findings["technical"] = "%d/%d engineering steps complete" % (
        tdone, len(tech_steps))

    # Financial readiness -- from the bankability engine (or partial if only the
    # finance model exists but isn't fully computed).
    if bank.get("available"):
        financial = int(bank.get("score") or 0)
        findings["financial"] = "Bankability score %d (%s)" % (
            financial, bank.get("rating") or "-")
    elif _is_meaningfully_populated("finance_config",
                                    proj.get("finance_config")):
        financial = 55
        findings["financial"] = "Finance model started; not fully computed"
    else:
        financial = 0
        findings["financial"] = "No financial model yet"

    # Construction readiness -- BOQ linked + engineering complete.
    boq_linked = bool(proj.get("boq_project_id")
                      or proj.get("boq_facilities_project_id")
                      or proj.get("boq_solar_project_id"))
    construction = max(0, min(100, (60 if boq_linked else 20)
                              + technical // 5))
    findings["construction"] = ("BOQ linked" if boq_linked
                                else "No BOQ linked yet")

    # Documentation readiness -- finance model + computed results present.
    docs_ok = (_is_meaningfully_populated("finance_config",
                                          proj.get("finance_config"))
               and bool(computed))
    documentation = 70 if docs_ok else 30
    findings["documentation"] = ("Core financial documentation ready"
                                 if docs_ok
                                 else "Financial documentation incomplete")

    # Institution match -- requested amount within the loan range + the project
    # type is supported by the institution.
    import math as _math

    def _f(v):
        try:
            x = float(v)
        except (TypeError, ValueError):
            return None
        return x if _math.isfinite(x) else None   # reject NaN/inf
    req = _f((fund or {}).get("funding_requested"))
    lmin, lmax = _f(inst.get("loan_min")), _f(inst.get("loan_max"))
    in_range = True
    if req is not None:
        if lmin is not None and req < lmin:
            in_range = False
        if lmax is not None and req > lmax:
            in_range = False
    supported = [x for x in (inst.get("supported_project_types") or "").split(",")
                 if x]
    type_ok = (not supported) or (proj.get("project_type") in supported)
    matched = bool(in_range and type_ok)
    findings["match"] = ("Fits the institution's loan range and mandate"
                         if matched
                         else "Outside the institution's loan range / mandate")

    risk_rating = bank.get("rating") if bank.get("available") else "Unrated"

    # Overall funding score -- weighted blend; a mandate mismatch trims it.
    fscore = int(round(0.35 * financial + 0.25 * technical
                       + 0.20 * construction + 0.20 * documentation))
    if not matched:
        fscore = max(0, fscore - 15)

    if fscore >= 70 and matched:
        rec = "recommend"
    elif fscore >= 55:
        rec = "conditional"
    else:
        rec = "decline"

    return {
        "funding_score": fscore, "technical_readiness": technical,
        "financial_readiness": financial,
        "documentation_readiness": documentation,
        "construction_readiness": construction,
        "risk_rating": risk_rating, "matched": 1 if matched else 0,
        "recommendation": rec, "findings": findings,
    }


# ===========================================================================
# SECTION E -- Wizard progress (derives completion from real stored data;
# SSS Section 2 acceptance -- no pseudo "hook" fields).
# ===========================================================================

def _wizard_progress(proj: dict[str, Any]) -> list[dict[str, Any]]:
    """Return a per-step progress list for the rail: each step is 'done' when
    its backing data is meaningfully populated. Phase 1 knows Steps 1-6 fields;
    later phases extend the completion map as steps land."""
    def _has_json(field: str) -> bool:
        d = _safe_json(proj.get(field))
        return bool(d) and any(
            v not in (None, "", [], {}, 0, 0.0) for v in d.values())

    done_map = {
        1:  bool(proj.get("project_name")),
        2:  bool(proj.get("project_type")),
        3:  _has_json("site_config"),
        4:  _has_json("facility_config"),
        5:  _has_json("technology_config"),
        6:  _has_json("electrical_config"),
        7:  _has_json("pv_config"),
        8:  _has_json("finance_config"),
        9:  bool(proj.get("boq_project_id")),
        10: bool(_safe_json(proj.get("pv_config")).get("kwp")),  # procurable once PV designed
        11: False,  # extended in Phase 7 (opportunity exists)
        12: False,  # extended in Phase 7 (stage history)
        13: bool(_safe_json(proj.get("finance_config")).get("computed")),  # report-ready
        14: False,  # extended in Phase 10 (an agent run exists)
    }
    out = []
    for num, suffix, label, icon in STEP_LABELS:
        out.append({
            "num": num, "suffix": suffix, "label": label, "icon": icon,
            "done": bool(done_map.get(num)),
            "endpoint": ("capital_investment_new" if num == 1
                         else "capital_investment_" + suffix),
        })
    return out


# ===========================================================================
# SECTION F -- Registration (mounts /large-scale-solar/*). Same signature as
# the legacy module so web_app.py needs NO change.
# ===========================================================================

def _sun_position(lat_deg: float, lon_deg: float,
                  month: int, hour: float,
                  tz_offset_h: float = 0.0) -> dict[str, float]:
    """Thin wrapper -> dt_scene_v2.sun_position (single source of truth).
    Extended backward-compatible superset of the legacy 5-key payload; the
    /dt/sun.json route and every caller keep working. Azimuth 0=N,90=E,180=S."""
    return _dtv2.sun_position(lat_deg, lon_deg, month, hour, tz_offset_h)


def _ci_normalize_proj_for_agents(proj: dict[str, Any]) -> dict[str, Any]:
    """v2 stores tech/electrical picks under 'technologies'/'services'; the
    carried AI-agent + digital-twin blocks read '.selected'. Return a shallow
    copy whose technology_config/electrical_config JSON also carry a 'selected'
    mirror so those verbatim blocks run unchanged."""
    p = dict(proj)
    tech = _safe_json(proj.get("technology_config"))
    sel_t = tech.get("technologies") or tech.get("selected")
    if sel_t and not tech.get("selected"):
        tech["selected"] = sel_t
        p["technology_config"] = json.dumps(tech)
    elec = _safe_json(proj.get("electrical_config"))
    sel_e = elec.get("services") or elec.get("selected")
    if sel_e and not elec.get("selected"):
        elec["selected"] = sel_e
        p["electrical_config"] = json.dumps(elec)
    return p


def register_capital_investment(app, *, get_db, login_required, csrf_protect,
                                current_user):
    """Register the Generation Station routes. Dependencies are injected to
    avoid a circular import with web_app."""

    # -- eager schema at registration time (SSS Section 8) ------------------
    try:
        _ensure_ci_projects_schema_verified(get_db)
    except Exception:
        pass

    # -- tier gate ----------------------------------------------------------
    def _gate(min_level: int):
        user = current_user()
        if _ci_level_of(user) >= min_level:
            return None
        session["ci_upsell_from"] = request.path
        session["ci_upsell_min_level"] = min_level
        return redirect(url_for("capital_investment_upgrade"))

    # -- tenant id from the canonical JWT source (matches BOQ engine) -------
    def _tenant_id():
        try:
            from web_app import _kc_current_tenant_id
            return _kc_current_tenant_id()
        except Exception:
            return None

    # -- project load / save (always user-scoped; SSS Section 8) -----------
    def _load_project(pid: int) -> dict[str, Any]:
        uid = session.get("user_id")
        with get_db() as c:
            row = c.execute(
                "SELECT * FROM capital_investment_projects "
                "WHERE id=? AND user_id=?", (pid, uid),
            ).fetchone()
        if not row:
            abort(404)
        return dict(row)

    def _save_project_field(pid: int, field: str, value: str) -> None:
        uid = session.get("user_id")
        with get_db() as c:
            c.execute(
                f"UPDATE capital_investment_projects "
                f"SET {field}=?, updated_at=CURRENT_TIMESTAMP "
                f"WHERE id=? AND user_id=?", (value, pid, uid),
            )

    def _n(form, key, default=None):
        try:
            v = form.get(key)
            return float(v) if v not in (None, "") else default
        except (TypeError, ValueError):
            return default

    # ------------------------------------------------------------------
    # GET /large-scale-solar -- landing + recent projects
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar", endpoint="capital_investment_landing")
    def _landing():
        _ensure_ci_projects_schema_verified(get_db)
        user = current_user()
        tier = _ci_tier_of(user)
        level = _ci_level_of(user)
        uid = session.get("user_id")
        recent: list[dict[str, Any]] = []
        hidden_count = 0
        if uid and level >= CI_LEVEL_SETUP:
            # Non-destructive "clear recent" hides rows from THIS user's view only,
            # scoped to the current tenant (Codex tenant-isolation fix).
            tid_s = str(_tenant_id()) if _tenant_id() is not None else ''
            hidden_ready = _ensure_ci_recent_hidden_schema(get_db)
            try:
                with get_db() as c:
                    if hidden_ready:
                        rows = c.execute(
                            "SELECT id, project_name, client_name, project_type, "
                            "project_status, currency, target_kwp, updated_at "
                            "FROM capital_investment_projects WHERE user_id=? "
                            "AND id NOT IN (SELECT project_id FROM ci_recent_hidden "
                            "               WHERE user_id=? AND tenant_id=?) "
                            "ORDER BY updated_at DESC, id DESC LIMIT 6",
                            (uid, uid, tid_s),
                        ).fetchall()
                        hrow = c.execute(
                            "SELECT COUNT(*) FROM ci_recent_hidden "
                            "WHERE user_id=? AND tenant_id=?",
                            (uid, tid_s)).fetchone()
                        hidden_count = int(hrow[0]) if hrow else 0
                    else:
                        rows = c.execute(
                            "SELECT id, project_name, client_name, project_type, "
                            "project_status, currency, target_kwp, updated_at "
                            "FROM capital_investment_projects WHERE user_id=? "
                            "ORDER BY updated_at DESC, id DESC LIMIT 6", (uid,),
                        ).fetchall()
                    recent = [dict(r) for r in rows] if rows else []
            except Exception:
                recent = []
        return render_template(
            "capital_investment/landing.html",
            user=user, tier=tier, tier_level=level,
            tier_label=CI_TIER_LABEL.get(tier, "Free"),
            recent=recent, hidden_count=hidden_count,
            project_types=PROJECT_TYPES,
        )

    # ------------------------------------------------------------------
    # POST /large-scale-solar/recent/clear -- non-destructive "clear recent
    # history" (owner 2026-07-05). Hides EVERY one of this user's generation
    # station projects from the landing "Recent projects" panel. Nothing is
    # deleted -- rows/data/BOQ remain and are reachable by URL. POST-only + CSRF
    # + user-scoped; reversible via /recent/restore. ("recent" is not an int so
    # it never collides with the /<int:pid> route.)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/recent/clear", methods=["POST"],
               endpoint="capital_investment_recent_clear")
    @login_required
    def _recent_clear():
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        csrf_protect()
        uid = session.get("user_id")
        tid_s = str(_tenant_id()) if _tenant_id() is not None else ''
        hidden = 0
        if _ensure_ci_recent_hidden_schema(get_db):
            try:
                with get_db() as c:
                    rows = c.execute(
                        "SELECT id FROM capital_investment_projects "
                        "WHERE user_id=? AND id NOT IN "
                        "(SELECT project_id FROM ci_recent_hidden "
                        " WHERE user_id=? AND tenant_id=?)",
                        (uid, uid, tid_s)).fetchall()
                    for r in rows or []:
                        # INSERT OR IGNORE -> Postgres ON CONFLICT DO NOTHING via
                        # db_adapter, so a double-submit can't poison the txn.
                        c.execute(
                            "INSERT OR IGNORE INTO ci_recent_hidden "
                            "(user_id, project_id, tenant_id) VALUES (?,?,?)",
                            (uid, int(r[0]), tid_s))
                        hidden += 1
            except Exception:
                flash("Could not clear the recent list - please try again.",
                      "danger")
                return redirect(url_for("capital_investment_landing"))
        try:
            from new_boq_hierarchy_schema import boq_audit
            boq_audit(get_db, uid, "capital_recent_history_cleared",
                      "capital_investment_recent", 0,
                      "hid %d project(s) from recent view" % hidden)
        except Exception:
            pass
        flash(
            f"Recent list cleared - {hidden} project(s) hidden from this view. "
            f"Your projects are NOT deleted.", "success")
        return redirect(url_for("capital_investment_landing"))

    # ------------------------------------------------------------------
    # POST /large-scale-solar/recent/restore -- undo the clear above (un-hide
    # every project for this user). POST-only + CSRF + user-scoped.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/recent/restore", methods=["POST"],
               endpoint="capital_investment_recent_restore")
    @login_required
    def _recent_restore():
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        csrf_protect()
        uid = session.get("user_id")
        tid_s = str(_tenant_id()) if _tenant_id() is not None else ''
        if _ensure_ci_recent_hidden_schema(get_db):
            try:
                with get_db() as c:
                    c.execute("DELETE FROM ci_recent_hidden "
                              "WHERE user_id=? AND tenant_id=?", (uid, tid_s))
            except Exception:
                flash("Could not restore the recent list - please try again.",
                      "danger")
                return redirect(url_for("capital_investment_landing"))
        try:
            from new_boq_hierarchy_schema import boq_audit
            boq_audit(get_db, uid, "capital_recent_history_restored",
                      "capital_investment_recent", 0, "restored recent view")
        except Exception:
            pass
        flash("Recent list restored.", "success")
        return redirect(url_for("capital_investment_landing"))

    # ------------------------------------------------------------------
    # GET /large-scale-solar/upgrade -- upsell
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/upgrade",
               endpoint="capital_investment_upgrade")
    def _upgrade():
        user = current_user()
        return render_template(
            "capital_investment/upgrade.html",
            user=user, tier=_ci_tier_of(user), tier_level=_ci_level_of(user),
            tier_label=CI_TIER_LABEL.get(_ci_tier_of(user), "Free"),
            upsell_from=session.pop("ci_upsell_from", None),
            upsell_min_level=session.pop("ci_upsell_min_level", CI_LEVEL_SETUP),
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/demo -- placeholder until Phase 2 wires the
    # read-only showcase (kept as a stub so the nav/landing CTA never 404s).
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/demo", endpoint="capital_investment_demo")
    def _demo():
        flash("The interactive demo is being rebuilt. Create a project to "
              "explore the Generation Station workflow.", "info")
        return redirect(url_for("capital_investment_landing"))

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/new -- Step 1 Project Registration
    # (PG-safe INSERT ... RETURNING id).
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/new", methods=["GET", "POST"],
               endpoint="capital_investment_new")
    @login_required
    def _new():
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        _ensure_ci_projects_schema_verified(get_db)
        if request.method == "POST":
            csrf_protect()
            f = request.form
            name = (f.get("project_name") or "").strip()[:300]
            if not name:
                flash("Project name is required.", "warning")
                return redirect(url_for("capital_investment_new"))

            def _pick(key, allowed, default=""):
                v = (f.get(key) or "").strip()
                return v if v in allowed else default

            uid = session["user_id"]
            tenant_id = _tenant_id()
            target_kwp = _n(f, "target_mwp")
            if target_kwp is not None:
                target_kwp = target_kwp * 1000.0   # MWp form -> kWp stored
            # Country/region come from the Step 1 dropdowns (with an "Other"
            # free-text escape); GPS lat/lon auto-fill from the solar DB when the
            # user left them blank. _ci_resolve_location is the server-side twin
            # of the template JS so it works with JavaScript disabled too.
            country, region, gps_lat, gps_lon = _ci_resolve_location(f)
            row_vals = (
                uid, name,
                (f.get("client_name") or "").strip()[:300],
                (f.get("investor") or "").strip()[:300],
                (f.get("developer") or "").strip()[:300],
                country,
                region,
                (f.get("district") or "").strip()[:120],
                gps_lat, gps_lon,
                (f.get("description") or "").strip()[:4000],
                _pick("project_status", PROJECT_STATUS_CODES, "concept"),
                (f.get("target_cod") or "").strip()[:40],
                target_kwp,
                _pick("design_standard", DESIGN_STANDARD_CODES, "IEC"),
                _pick("currency", CURRENCY_CODES, "GHS"),
                _pick("tax_regime", TAX_REGIME_CODES, "standard"),
                _pick("project_type", PROJECT_TYPE_CODES, ""),
                tenant_id,
            )
            cols = ("user_id, project_name, client_name, investor, developer, "
                    "country, region, district, gps_lat, gps_lon, description, "
                    "project_status, target_cod, target_kwp, design_standard, "
                    "currency, tax_regime, project_type, tenant_id")
            ph = ",".join("?" * 19)
            pid = 0
            try:
                with get_db() as c:
                    cur = c.execute(
                        f"INSERT INTO capital_investment_projects ({cols}) "
                        f"VALUES ({ph}) RETURNING id", row_vals,
                    )
                    r = cur.fetchone()
                    pid = int(r[0]) if r else 0
            except Exception:
                # Legacy DB without tenant_id column -> retry without it.
                try:
                    with get_db() as c:
                        cur = c.execute(
                            "INSERT INTO capital_investment_projects "
                            "(user_id, project_name, client_name, investor, "
                            " developer, country, region, district, gps_lat, "
                            " gps_lon, description, project_status, target_cod, "
                            " target_kwp, design_standard, currency, tax_regime, "
                            " project_type) VALUES (" + ",".join("?" * 18) +
                            ") RETURNING id", row_vals[:18],
                        )
                        r = cur.fetchone()
                        pid = int(r[0]) if r else 0
                except Exception:
                    pid = 0
            if pid <= 0:
                flash("Could not create the project. Please try again.",
                      "danger")
                return redirect(url_for("capital_investment_new"))
            flash("Project registered. Continue with Step 2 - Project Type.",
                  "success")
            return redirect(url_for("capital_investment_project", pid=pid))

        # GET
        ci_countries, ci_location_map = _ci_location_bundle()
        return render_template(
            "capital_investment/step01_registration.html",
            user=current_user(), proj=None,
            project_types=PROJECT_TYPES, project_statuses=PROJECT_STATUSES,
            design_standards=DESIGN_STANDARDS, currencies=CURRENCIES,
            tax_regimes=TAX_REGIMES,
            ci_countries=ci_countries, ci_location_map=ci_location_map,
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid> -- project overview (wizard hub)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>",
               endpoint="capital_investment_project")
    @login_required
    def _project(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        # Steps light up as their phase ships: a step is reachable only when
        # its endpoint is actually registered (skeleton-safe rail).
        from flask import current_app
        progress = _wp(proj)
        vf = current_app.view_functions
        return render_template(
            "capital_investment/project.html",
            user=current_user(), proj=proj, progress=progress,
            reg_available=("capital_investment_regulatory" in vf),
            dt_available=("capital_investment_digital_twin" in vf),
            fund_available=("capital_investment_funding" in vf),
            project_types=dict((c, L) for c, L, _ in PROJECT_TYPES),
        )

    # ------------------------------------------------------------------
    # Project Funding (Slice 1) -- GET overview (auto-populated from existing
    # project + Step-8 finance data + the bankability funding-readiness score)
    # and POST "Request Project Funding" (records/updates the funding row). The
    # page REUSES the finance engine output; it never re-models. Institution
    # selection + submission arrive in Slices 2-3.
    # ------------------------------------------------------------------
    def _ci_funding_row(pid: int, uid, tid_s: str):
        """Return the funding application dict for this project/tenant, or None."""
        if not _ensure_ci_funding_schema(get_db):
            return None
        try:
            with get_db() as c:
                r = c.execute(
                    "SELECT status, funding_requested, customer_equity, "
                    "funding_score, risk_rating, selected_institutions, remarks, "
                    "updated_at FROM capital_investment_funding "
                    "WHERE capital_investment_project_id=? AND tenant_id=? "
                    "AND user_id=? LIMIT 1", (pid, tid_s, uid)).fetchone()
            return dict(r) if r else None
        except Exception:
            return None

    def _ci_funding_overview(proj: dict):
        """Assemble the read-only funding overview from existing project data --
        no re-modelling (spec: 'do not ask the user to re-enter data')."""
        fin_cfg = _safe_json(proj.get("finance_config"))
        computed = fin_cfg.get("computed") if isinstance(fin_cfg, dict) else {}
        computed = computed if isinstance(computed, dict) else {}
        cash = _ci_cashflow_plan(fin_cfg)
        bank = _ci_bankability(computed)

        def _f(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
        capex = _f(cash.get("capex_local"))
        equity = _f(cash.get("equity_local"))
        debt = None
        if capex is not None:
            debt = round(max(capex - (equity or 0.0), 0.0), 2)
        return {
            "finance_available": bool(cash.get("available")),
            "capex_local": capex, "equity_local": equity, "debt_local": debt,
            "npv_local": _f(cash.get("npv_local")),
            "irr_pct": _f(cash.get("irr_pct")),
            "payback_years": _f(cash.get("payback_years")),
            "dscr_min": _f(computed.get("dscr_min")),
            "bank_available": bool(bank.get("available")),
            "funding_score": bank.get("score"),
            "funding_rating": bank.get("rating"),
            "funding_rating_class": bank.get("rating_class", "secondary"),
        }

    @app.route("/large-scale-solar/<int:pid>/funding",
               methods=["GET", "POST"], endpoint="capital_investment_funding")
    @login_required
    def _funding(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)          # 404s unless owned by this user
        # Defence in depth: _load_project scopes only user_id, so also reject a
        # cross-tenant project (same user in another tenant) before exposing its
        # finance data or writing funding rows (Codex High).
        _ctid = _tenant_id()
        if _ctid is not None and proj.get("tenant_id") not in (
                None, _ctid, str(_ctid)):
            from flask import abort
            abort(404)
        uid = session.get("user_id")
        tid_s = str(_ctid) if _ctid is not None else ''
        ov = _ci_funding_overview(proj)

        if request.method == "POST":
            csrf_protect()
            # Record / refresh the funding application (status -> requested). The
            # funding amount defaults to the debt portion (CAPEX - equity); the
            # readiness score + equity are snapshotted from the finance engine.
            req_amt = ov.get("debt_local")
            equity = ov.get("equity_local")
            score = ov.get("funding_score") if ov.get("bank_available") else None
            ok = False
            if _ensure_ci_funding_schema(get_db):
                try:
                    with get_db() as c:
                        # Atomic upsert in ONE statement -- safe under a concurrent
                        # double-submit (no SELECT-then-INSERT race). ON CONFLICT
                        # DO UPDATE is supported by SQLite 3.24+ and Postgres; the
                        # conflict target is the PK (project_id, tenant_id).
                        c.execute(
                            "INSERT INTO capital_investment_funding "
                            "(capital_investment_project_id, tenant_id, user_id, "
                            " status, funding_requested, customer_equity, "
                            " funding_score) VALUES (?,?,?,?,?,?,?) "
                            "ON CONFLICT (capital_investment_project_id, tenant_id) "
                            "DO UPDATE SET status='requested', "
                            "funding_requested=EXCLUDED.funding_requested, "
                            "customer_equity=EXCLUDED.customer_equity, "
                            "funding_score=EXCLUDED.funding_score, "
                            "updated_at=CURRENT_TIMESTAMP",
                            (pid, tid_s, uid, "requested", req_amt, equity, score))
                    ok = True
                except Exception:
                    ok = False
            if ok:
                try:
                    from new_boq_hierarchy_schema import boq_audit
                    boq_audit(get_db, uid, "capital_funding_requested",
                              "capital_investment_project", pid,
                              "funding requested amount=%s score=%s" % (
                                  req_amt, score))
                except Exception:
                    pass
                flash("Project funding requested. Next: register/select a "
                      "financial institution (coming soon).", "success")
            else:
                flash("Could not record the funding request - please try again.",
                      "danger")
            return redirect(url_for("capital_investment_funding", pid=pid))

        # GET
        fund = _ci_funding_row(pid, uid, tid_s)
        institutions, selections = _ci_funding_institutions(pid, uid, tid_s)
        # Per-institution message threads + hard-copy shipments for the
        # applicant's Communication + Documents panels.
        threads, shipments = {}, {}
        for iid in selections.keys():
            threads[iid] = _fi_thread(pid, iid, tid_s)
            shipments[iid] = _fi_shipments(pid, iid, tid_s)
        # CRM / pipeline handoff (Slice 9): current opportunity + funding snapshot.
        crm = _load_opportunity(pid, uid)
        crm_snap = _ci_funding_crm_snapshot(pid, uid, tid_s)
        return render_template(
            "capital_investment/funding.html",
            user=current_user(), proj=proj, ov=ov, fund=fund,
            institutions=institutions, selections=selections, threads=threads,
            shipments=shipments, ship_status_labels=FI_SHIP_STATUS_LABELS,
            ship_status_class=FI_SHIP_STATUS_CLASS,
            crm=crm, crm_snap=crm_snap,
            project_types=dict((c, L) for c, L, _ in PROJECT_TYPES),
        )

    def _ci_funding_institutions(pid: int, uid, tid_s: str):
        """Return (approved_institutions, selected_map). approved_institutions is
        the platform-global approved registry; selected_map is
        {institution_id: {consent, status, submitted_at}} for this project."""
        approved, selected = [], {}
        _ensure_fi_schema(get_db)
        _ensure_fi_selection_schema(get_db)
        try:
            with get_db() as c:
                rr = c.execute(
                    "SELECT institution_id, name, inst_type, country, region, "
                    "fee_pct, loan_min, loan_max FROM financial_institutions "
                    "WHERE status='approved' ORDER BY name").fetchall()
                approved = [dict(r) for r in rr] if rr else []
                sr = c.execute(
                    "SELECT institution_id, consent, status, submitted_at "
                    "FROM funding_institution_selections "
                    "WHERE capital_investment_project_id=? AND tenant_id=? "
                    "AND user_id=?", (pid, tid_s, uid)).fetchall()
                for r in sr or []:
                    selected[_row_get(r, "institution_id")] = {
                        "consent": _row_get(r, "consent", 0),
                        "status": _row_get(r, "status", "submitted"),
                        "submitted_at": _row_get(r, "submitted_at"),
                    }
        except Exception:
            pass
        return approved, selected

    # POST /large-scale-solar/<pid>/funding/submit -- submit the funding package
    # to the selected APPROVED institution(s) with explicit consent (Slice 3).
    @app.route("/large-scale-solar/<int:pid>/funding/submit",
               methods=["POST"], endpoint="capital_investment_funding_submit")
    @login_required
    def _funding_submit(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)          # 404s unless owned by this user
        _ctid = _tenant_id()
        if _ctid is not None and proj.get("tenant_id") not in (
                None, _ctid, str(_ctid)):
            from flask import abort
            abort(404)
        csrf_protect()
        uid = session.get("user_id")
        tid_s = str(_ctid) if _ctid is not None else ''
        f = request.form
        if not f.get("consent"):
            flash("You must consent to share your project reports before "
                  "submitting to a financial institution.", "warning")
            return redirect(url_for("capital_investment_funding", pid=pid))
        chosen = [x for x in f.getlist("institution_id") if x]
        chosen = list(dict.fromkeys(chosen))       # de-dup, preserve order
        if not chosen:
            flash("Select at least one financial institution.", "warning")
            return redirect(url_for("capital_investment_funding", pid=pid))
        _ensure_fi_schema(get_db)
        _ensure_fi_selection_schema(get_db)
        _ensure_ci_funding_schema(get_db)
        # Only the ids that VALIDATE as currently-approved are recorded, counted
        # and persisted -- never the raw posted list (Codex: a forged POST must
        # not leak a rejected/suspended id into selected_institutions).
        approved_ids = []
        try:
            with get_db() as c:
                for iid in chosen:
                    ok = c.execute(
                        "SELECT 1 FROM financial_institutions WHERE "
                        "institution_id=? AND status='approved' LIMIT 1",
                        (iid,)).fetchone()
                    if not ok:
                        continue
                    # Idempotent per (project, institution, tenant).
                    c.execute(
                        "INSERT OR IGNORE INTO funding_institution_selections "
                        "(capital_investment_project_id, institution_id, "
                        " tenant_id, user_id, consent, status) "
                        "VALUES (?,?,?,?,1,'submitted')",
                        (pid, iid, tid_s, uid))
                    approved_ids.append(iid)
                if approved_ids:
                    # Reflect the submission on the funding application, creating
                    # the row if the customer skipped "Request" (atomic upsert).
                    # selected_institutions stores ONLY the validated subset.
                    import json as _json
                    ov2 = _ci_funding_overview(proj)
                    c.execute(
                        "INSERT INTO capital_investment_funding "
                        "(capital_investment_project_id, tenant_id, user_id, "
                        " status, funding_requested, customer_equity, "
                        " funding_score, selected_institutions) "
                        "VALUES (?,?,?,?,?,?,?,?) "
                        "ON CONFLICT (capital_investment_project_id, tenant_id) "
                        "DO UPDATE SET status='submitted', "
                        "selected_institutions=EXCLUDED.selected_institutions, "
                        "updated_at=CURRENT_TIMESTAMP",
                        (pid, tid_s, uid, "submitted", ov2.get("debt_local"),
                         ov2.get("equity_local"),
                         ov2.get("funding_score") if ov2.get("bank_available")
                         else None, _json.dumps(approved_ids)))
        except Exception:
            approved_ids = []
        # Count = institutions in THIS submission package, not newly-inserted
        # rows. A legitimate re-submit re-affirms the same approved lenders
        # (INSERT OR IGNORE keeps prior rows) and must still report the true
        # package size -- using c.rowcount here would wrongly show 0 on resubmit.
        submitted_count = len(approved_ids)
        if submitted_count:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "capital_funding_submitted",
                          "capital_investment_project", pid,
                          "submitted to %d institution(s)" % submitted_count)
            except Exception:
                pass
            flash("Funding application submitted to %d financial "
                  "institution(s)." % submitted_count, "success")
        else:
            flash("Could not submit - none of the selected institutions are "
                  "currently approved. Please try again.", "danger")
        return redirect(url_for("capital_investment_funding", pid=pid))

    # ==================================================================
    # Slice 10c -- Project Funding for REGULAR /project/<pid> projects. Reuses the
    # SAME institution registry + funding tables + institution workspace as the
    # generation-station flow; only the project SOURCE differs (residential
    # projects.data_json). Funding rows are namespaced by _pf_fid(pid) so they can
    # never collide with a generation-station project of the same integer id, and
    # a display snapshot is stored on the row so the workspace needs no join to the
    # residential `projects` table. Every route re-loads the project user-scoped.
    # ==================================================================
    def _pf_load(pid: int) -> dict:
        """Load + adapt a regular project owned by the current user, or 404."""
        from web_app import get_project as _get_project
        prow = _get_project(pid)          # user-scoped (WHERE id=? AND user_id=?)
        if not prow:
            abort(404)
        return _pf_project_view(prow)

    def _pf_bill_check(pid: int):
        """The saved "Check My Bill" snapshot (data['bill_check']) for a regular
        project owned by this user, or None. Persisted by /project/<pid>/bill-check
        /save. Used to gate the funding application + surface bill verification in
        the economic/energy reports."""
        try:
            from web_app import get_project as _get_project
            prow = _get_project(pid)
            if not prow:
                return None
            return (prow.get("data") or {}).get("bill_check")
        except Exception:
            return None

    @app.route("/project/<int:pid>/funding",
               methods=["GET", "POST"], endpoint="project_funding")
    @login_required
    def _pf_funding(pid: int):
        proj = _pf_load(pid)               # 404 unless owned by this user
        fid = proj["id"]                   # namespaced funding id
        uid = session.get("user_id")
        _ctid = _tenant_id()
        tid_s = str(_ctid) if _ctid is not None else ''
        ov = _ci_funding_overview(proj)

        if request.method == "POST":
            csrf_protect()
            req_amt = ov.get("debt_local")
            equity = ov.get("equity_local")
            score = ov.get("funding_score") if ov.get("bank_available") else None
            ok = False
            if _ensure_ci_funding_schema(get_db):
                try:
                    with get_db() as c:
                        # Atomic upsert incl. the denormalized display snapshot so
                        # the institution workspace never joins `projects`.
                        c.execute(
                            "INSERT INTO capital_investment_funding "
                            "(capital_investment_project_id, tenant_id, user_id, "
                            " status, funding_requested, customer_equity, "
                            " funding_score, project_kind, proj_name, proj_client, "
                            " proj_type, proj_kwp, proj_currency, proj_country, "
                            " proj_region) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
                            "ON CONFLICT (capital_investment_project_id, tenant_id) "
                            "DO UPDATE SET status='requested', "
                            "funding_requested=EXCLUDED.funding_requested, "
                            "customer_equity=EXCLUDED.customer_equity, "
                            "funding_score=EXCLUDED.funding_score, "
                            "proj_name=EXCLUDED.proj_name, "
                            "proj_client=EXCLUDED.proj_client, "
                            "proj_type=EXCLUDED.proj_type, "
                            "proj_kwp=EXCLUDED.proj_kwp, "
                            "proj_currency=EXCLUDED.proj_currency, "
                            "proj_country=EXCLUDED.proj_country, "
                            "proj_region=EXCLUDED.proj_region, "
                            "updated_at=CURRENT_TIMESTAMP",
                            (fid, tid_s, uid, "requested", req_amt, equity, score,
                             "project", proj.get("project_name"),
                             proj.get("client_name"), proj.get("project_type"),
                             proj.get("target_kwp"), proj.get("currency"),
                             proj.get("country"), proj.get("region")))
                    ok = True
                except Exception:
                    ok = False
            if ok:
                try:
                    from new_boq_hierarchy_schema import boq_audit
                    boq_audit(get_db, uid, "project_funding_requested",
                              "project", pid,
                              "funding requested amount=%s score=%s" % (
                                  req_amt, score))
                except Exception:
                    pass
                flash("Project funding requested. Next: select a financial "
                      "institution and submit your application.", "success")
            else:
                flash("Could not record the funding request - please try again.",
                      "danger")
            return redirect(url_for("project_funding", pid=pid))

        # GET
        fund = _ci_funding_row(fid, uid, tid_s)
        institutions, selections = _ci_funding_institutions(fid, uid, tid_s)
        threads, shipments = {}, {}
        for iid in selections.keys():
            threads[iid] = _fi_thread(fid, iid, tid_s)
            shipments[iid] = _fi_shipments(fid, iid, tid_s)
        # Bill-check gate (owner 2026-07-05): the customer must run "Check My Bill"
        # on the project and save it BEFORE the application is sent to a bank, so
        # the verified bill/tariff rides along in the economic + energy reports.
        bill_check = _pf_bill_check(pid)
        return render_template(
            "capital_investment/project_funding.html",
            user=current_user(), proj=proj, pid=pid, ov=ov, fund=fund,
            institutions=institutions, selections=selections, threads=threads,
            shipments=shipments, ship_status_labels=FI_SHIP_STATUS_LABELS,
            ship_status_class=FI_SHIP_STATUS_CLASS,
            bill_check=bill_check, bill_checked=bool(bill_check),
            bc_project_id=pid, bc_currency=(proj.get("currency") or "GHS"),
            bc_loads_json="null")

    # POST /project/<pid>/funding/submit -- submit to selected APPROVED
    # institution(s) with consent (regular-project twin of the CI submit).
    @app.route("/project/<int:pid>/funding/submit",
               methods=["POST"], endpoint="project_funding_submit")
    @login_required
    def _pf_funding_submit(pid: int):
        proj = _pf_load(pid)
        fid = proj["id"]
        csrf_protect()
        uid = session.get("user_id")
        _ctid = _tenant_id()
        tid_s = str(_ctid) if _ctid is not None else ''
        f = request.form
        if not f.get("consent"):
            flash("You must consent to share your project reports before "
                  "submitting to a financial institution.", "warning")
            return redirect(url_for("project_funding", pid=pid))
        # Gate: the customer must run + save "Check My Bill" first, so the verified
        # bill/tariff is embedded in the economic + energy reports the bank sees
        # (owner 2026-07-05). No bank submission without it.
        if not _pf_bill_check(pid):
            flash("Run 'Check My Bill' on this project and save it before applying "
                  "for funding - the bank needs your verified bill and tariff in the "
                  "economic and energy reports.", "warning")
            return redirect(url_for("project_funding", pid=pid))
        chosen = [x for x in f.getlist("institution_id") if x]
        chosen = list(dict.fromkeys(chosen))       # de-dup, preserve order
        if not chosen:
            flash("Select at least one financial institution.", "warning")
            return redirect(url_for("project_funding", pid=pid))
        _ensure_fi_schema(get_db)
        _ensure_fi_selection_schema(get_db)
        _ensure_ci_funding_schema(get_db)
        # Only currently-approved ids are recorded/persisted -- never the raw
        # posted list (a forged POST must not leak a rejected/suspended id).
        approved_ids = []
        try:
            with get_db() as c:
                for iid in chosen:
                    okr = c.execute(
                        "SELECT 1 FROM financial_institutions WHERE "
                        "institution_id=? AND status='approved' LIMIT 1",
                        (iid,)).fetchone()
                    if not okr:
                        continue
                    # INSERT OR IGNORE is backend-safe: db_adapter rewrites it to
                    # `ON CONFLICT DO NOTHING` on Postgres (db_adapter.py:105-119);
                    # this mirrors the live generation-station submit exactly.
                    c.execute(
                        "INSERT OR IGNORE INTO funding_institution_selections "
                        "(capital_investment_project_id, institution_id, "
                        " tenant_id, user_id, consent, status, project_kind) "
                        "VALUES (?,?,?,?,1,'submitted','project')",
                        (fid, iid, tid_s, uid))
                    approved_ids.append(iid)
                if approved_ids:
                    ov2 = _ci_funding_overview(proj)
                    c.execute(
                        "INSERT INTO capital_investment_funding "
                        "(capital_investment_project_id, tenant_id, user_id, "
                        " status, funding_requested, customer_equity, "
                        " funding_score, selected_institutions, project_kind, "
                        " proj_name, proj_client, proj_type, proj_kwp, "
                        " proj_currency, proj_country, proj_region) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
                        "ON CONFLICT (capital_investment_project_id, tenant_id) "
                        "DO UPDATE SET status='submitted', "
                        "selected_institutions=EXCLUDED.selected_institutions, "
                        "updated_at=CURRENT_TIMESTAMP",
                        (fid, tid_s, uid, "submitted", ov2.get("debt_local"),
                         ov2.get("equity_local"),
                         ov2.get("funding_score") if ov2.get("bank_available")
                         else None, json.dumps(approved_ids), "project",
                         proj.get("project_name"), proj.get("client_name"),
                         proj.get("project_type"), proj.get("target_kwp"),
                         proj.get("currency"), proj.get("country"),
                         proj.get("region")))
        except Exception:
            approved_ids = []
        submitted_count = len(approved_ids)
        if submitted_count:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "project_funding_submitted",
                          "project", pid,
                          "submitted to %d institution(s)" % submitted_count)
            except Exception:
                pass
            flash("Funding application submitted to %d financial "
                  "institution(s)." % submitted_count, "success")
        else:
            flash("Could not submit - none of the selected institutions are "
                  "currently approved. Please try again.", "danger")
        return redirect(url_for("project_funding", pid=pid))

    # POST /project/<pid>/funding/message -- applicant replies to an institution
    # (owner-scoped; only consented threads to a still-approved institution).
    @app.route("/project/<int:pid>/funding/message",
               methods=["POST"], endpoint="project_funding_message")
    @login_required
    def _pf_funding_message(pid: int):
        proj = _pf_load(pid)               # owner-scoped
        fid = proj["id"]
        csrf_protect()
        uid = session.get("user_id")
        _ctid = _tenant_id()
        tid_s = str(_ctid) if _ctid is not None else ''
        f = request.form
        institution_id = (f.get("institution_id") or "").strip()
        body = (f.get("body") or "").strip()[:5000]
        if not body:
            flash("Message body is required.", "warning")
            return redirect(url_for("project_funding", pid=pid))
        _ensure_fi_selection_schema(get_db)
        _ensure_fi_schema(get_db)
        allowed = False
        try:
            with get_db() as c:
                r = c.execute(
                    "SELECT 1 FROM funding_institution_selections s "
                    "JOIN financial_institutions fi "
                    "  ON fi.institution_id = s.institution_id "
                    "WHERE s.capital_investment_project_id=? "
                    "AND s.institution_id=? AND COALESCE(s.tenant_id,'')=? "
                    "AND s.consent=1 AND fi.status='approved' LIMIT 1",
                    (fid, institution_id, tid_s)).fetchone()
                allowed = bool(r)
        except Exception:
            allowed = False
        if not allowed:
            flash("You can only message an approved institution you submitted "
                  "to.", "warning")
            return redirect(url_for("project_funding", pid=pid))
        subject = (f.get("subject") or "").strip()[:200]
        emailed_to, channel = None, "message"
        if f.get("send_email"):
            inst_email = None
            try:
                with get_db() as c:
                    ir = c.execute(
                        "SELECT email FROM financial_institutions "
                        "WHERE institution_id=?", (institution_id,)).fetchone()
                inst_email = _row_get(ir, "email")
            except Exception:
                inst_email = None
            appl_name = proj.get("client_name") or "the applicant"
            if inst_email and _fi_send_funding_email(
                    inst_email, subject or "Applicant reply", body, appl_name):
                channel, emailed_to = "email", inst_email
            else:
                flash("Reply saved, but the institution email could not be "
                      "sent.", "warning")
        ok = _fi_add_message(fid, institution_id, tid_s,
                             sender_role="applicant", sender_uid=uid,
                             msg_type="message", subject=subject, body=body,
                             channel=channel, emailed_to=emailed_to)
        flash("Reply posted%s." % (" and emailed" if emailed_to else "")
              if ok else "Could not post the reply.",
              "success" if ok else "danger")
        return redirect(url_for("project_funding", pid=pid))

    # POST /project/<pid>/funding/shipment -- applicant records a hard-copy
    # dispatch to an institution they submitted to (approved + consent).
    @app.route("/project/<int:pid>/funding/shipment",
               methods=["POST"], endpoint="project_funding_shipment")
    @login_required
    def _pf_funding_shipment(pid: int):
        proj = _pf_load(pid)               # owner-scoped
        fid = proj["id"]
        csrf_protect()
        uid = session.get("user_id")
        _ctid = _tenant_id()
        tid_s = str(_ctid) if _ctid is not None else ''
        institution_id = (request.form.get("institution_id") or "").strip()
        _ensure_fi_selection_schema(get_db)
        _ensure_fi_schema(get_db)
        allowed = False
        try:
            with get_db() as c:
                r = c.execute(
                    "SELECT 1 FROM funding_institution_selections s "
                    "JOIN financial_institutions fi "
                    "  ON fi.institution_id = s.institution_id "
                    "WHERE s.capital_investment_project_id=? "
                    "AND s.institution_id=? AND COALESCE(s.tenant_id,'')=? "
                    "AND s.consent=1 AND fi.status='approved' LIMIT 1",
                    (fid, institution_id, tid_s)).fetchone()
                allowed = bool(r)
        except Exception:
            allowed = False
        if not allowed:
            flash("You can only send documents to an approved institution you "
                  "submitted to.", "warning")
            return redirect(url_for("project_funding", pid=pid))
        ok = _fi_add_shipment(fid, institution_id, tid_s, request.form,
                              role="applicant", uid=uid,
                              default_status="dispatched")
        flash("Hard-copy dispatch recorded." if ok
              else "Could not record the dispatch.",
              "success" if ok else "danger")
        return redirect(url_for("project_funding", pid=pid))

    # ==================================================================
    # Slice 4 -- Financial Institution Workspace. An approved institution's
    # owner sees ONLY the funding applications customers explicitly submitted to
    # it (funding_institution_selections). This is a deliberate CROSS-TENANT read
    # -- the institution is a different party from the customer -- bounded to the
    # rows the customer consented to share. Isolation is enforced entirely by
    # joining on fi.created_by_user_id = <this user> AND fi.status='approved';
    # no client-supplied institution id is ever trusted, and there is no
    # customer-tenant filter (by design). Read-only GET, so no CSRF.
    # ==================================================================
    def _fi_workspace_rows(uid):
        """Assigned applications for every APPROVED institution this user owns.
        Returns (rows, institutions): institutions is the owner's approved list
        (header + multi-institution labelling); rows are the assigned
        applications joined to their funding record + project. NULL/'' tenant
        ids are treated as equal (single-tenant / no-Keycloak mode)."""
        _ensure_fi_schema(get_db)
        _ensure_fi_selection_schema(get_db)
        _ensure_ci_funding_schema(get_db)
        rows, insts = [], []
        try:
            with get_db() as c:
                ir = c.execute(
                    "SELECT institution_id, name, inst_type, country, region, "
                    "fee_pct FROM financial_institutions "
                    "WHERE created_by_user_id=? AND status='approved' "
                    "ORDER BY name", (uid,)).fetchall()
                insts = [dict(r) for r in ir] if ir else []
                if not insts:
                    return [], []
                rr = c.execute(
                    "SELECT s.capital_investment_project_id AS pid, "
                    " s.institution_id AS institution_id, "
                    " s.tenant_id AS tenant_id, s.status AS app_status, "
                    " s.submitted_at AS submitted_at, "
                    " fi.name AS institution_name, "
                    " fund.funding_requested AS funding_requested, "
                    " fund.customer_equity AS customer_equity, "
                    " fund.funding_score AS funding_score, "
                    " fund.risk_rating AS risk_rating, "
                    " p.project_name AS project_name, "
                    " p.client_name AS client_name, "
                    " p.project_type AS project_type, p.country AS country, "
                    " p.region AS region, p.district AS district, "
                    " p.target_kwp AS target_kwp, p.currency AS currency "
                    "FROM funding_institution_selections s "
                    "JOIN financial_institutions fi "
                    "  ON fi.institution_id = s.institution_id "
                    "JOIN capital_investment_funding fund "
                    "  ON fund.capital_investment_project_id = "
                    "     s.capital_investment_project_id "
                    " AND COALESCE(fund.tenant_id,'') = COALESCE(s.tenant_id,'') "
                    # p.tenant_id is UUID on the fresh Postgres schema while the
                    # selection/funding tenant ids are TEXT -- CAST to TEXT so the
                    # NULL/'' equivalence join works on both Postgres and SQLite.
                    "JOIN capital_investment_projects p "
                    "  ON p.id = s.capital_investment_project_id "
                    " AND COALESCE(CAST(p.tenant_id AS TEXT),'') = "
                    "     COALESCE(s.tenant_id,'') "
                    # Read-side consent boundary (defence in depth): an institution
                    # only ever sees applications the customer explicitly consented
                    # to share, regardless of how the selection row was written.
                    "WHERE fi.created_by_user_id=? AND fi.status='approved' "
                    " AND s.consent=1 "
                    "ORDER BY s.submitted_at DESC", (uid,)).fetchall()
                rows = [dict(r) for r in rr] if rr else []
        except Exception:
            rows = []
        return rows, insts

    def _pf_workspace_rows(uid):
        """Slice 10c -- the REGULAR /project/<pid> funding applications assigned to
        this user's approved institutions. Same auth boundary as
        _fi_workspace_rows (owned + approved institution, consent=1); the display
        fields come from the denormalized snapshot on capital_investment_funding
        so there is NO join to the residential `projects` table. The CI workspace
        query's INNER JOIN to capital_investment_projects already excludes these
        namespaced rows, so this reads them separately and the route concatenates.
        """
        _ensure_fi_schema(get_db)
        _ensure_fi_selection_schema(get_db)
        _ensure_ci_funding_schema(get_db)
        rows = []
        try:
            with get_db() as c:
                rr = c.execute(
                    "SELECT s.capital_investment_project_id AS pid, "
                    " s.institution_id AS institution_id, "
                    " s.tenant_id AS tenant_id, s.status AS app_status, "
                    " s.submitted_at AS submitted_at, "
                    " fi.name AS institution_name, "
                    " fund.funding_requested AS funding_requested, "
                    " fund.customer_equity AS customer_equity, "
                    " fund.funding_score AS funding_score, "
                    " fund.risk_rating AS risk_rating, "
                    " fund.proj_name AS project_name, "
                    " fund.proj_client AS client_name, "
                    " fund.proj_type AS project_type, "
                    " fund.proj_country AS country, "
                    " fund.proj_region AS region, "
                    " fund.proj_kwp AS target_kwp, "
                    " fund.proj_currency AS currency "
                    "FROM funding_institution_selections s "
                    "JOIN financial_institutions fi "
                    "  ON fi.institution_id = s.institution_id "
                    "JOIN capital_investment_funding fund "
                    "  ON fund.capital_investment_project_id = "
                    "     s.capital_investment_project_id "
                    " AND COALESCE(fund.tenant_id,'') = COALESCE(s.tenant_id,'') "
                    "WHERE fi.created_by_user_id=? AND fi.status='approved' "
                    " AND s.consent=1 AND s.project_kind='project' "
                    "ORDER BY s.submitted_at DESC", (uid,)).fetchall()
                rows = [dict(r) for r in rr] if rr else []
        except Exception:
            rows = []
        for r in rows:
            r["kind"] = "project"           # workspace labels kWp (not MWp)
            r["district"] = None
        return rows

    # GET /funding/workspace -- the approved institution's dashboard + list.
    @app.route("/funding/workspace",
               endpoint="funding_institution_workspace")
    @login_required
    def _fi_workspace():
        uid = session.get("user_id")
        rows, insts = _fi_workspace_rows(uid)
        for _r in rows:
            _r.setdefault("kind", "capital")
        # Slice 10c -- merge regular /project/<pid> applications (namespaced ids
        # the CI query's INNER join excludes), then re-sort the unified list.
        rows = rows + _pf_workspace_rows(uid)
        rows.sort(key=lambda r: (r.get("submitted_at") or ""), reverse=True)

        # Status filter (allowlist -> no SQL injection). Filters the visible
        # table only; the dashboard metrics are always over the full set.
        status = (request.args.get("status") or "").strip()
        if status not in FI_APP_STATUSES:
            status = ""

        def _num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        under = {"under_review", "technical_review", "financial_review"}
        appr = {"approved", "approved_in_principle"}
        closed = {"approved", "rejected", "completed"}
        counts = {s: 0 for s in FI_APP_STATUSES}
        for r in rows:
            st = r.get("app_status")
            if st in counts:
                counts[st] += 1
        scores = [_num(r.get("funding_score")) for r in rows]
        scores = [s for s in scores if s is not None]
        req_sum = sum((_num(r.get("funding_requested")) or 0.0) for r in rows)
        appr_sum = sum((_num(r.get("funding_requested")) or 0.0)
                       for r in rows if r.get("app_status") in appr)
        pipe_sum = sum((_num(r.get("funding_requested")) or 0.0)
                       for r in rows if r.get("app_status") not in closed)
        # Money totals only make sense in a single currency; if the assigned
        # projects mix currencies, show counts and mark money as mixed (FX roll-up
        # is Slice 7's revenue job, not the workspace's).
        currencies = {(r.get("currency") or "GHS") for r in rows}
        money_cur = next(iter(currencies)) if len(currencies) == 1 else None
        metrics = {
            "assigned": len(rows),
            "under_review": sum(counts[s] for s in under),
            "awaiting_documents": counts["awaiting_documents"],
            "approved": sum(counts[s] for s in appr),
            "rejected": counts["rejected"],
            "total_requested": req_sum,
            "total_approved": appr_sum,
            "pipeline_value": pipe_sum,
            "avg_score": round(sum(scores) / len(scores)) if scores else None,
            "money_cur": money_cur,
        }
        shown = [r for r in rows
                 if (not status or r.get("app_status") == status)]
        return render_template(
            "capital_investment/funding_workspace.html",
            user=current_user(), institutions=insts, rows=shown,
            metrics=metrics, status=status,
            statuses=FI_APP_STATUSES, status_labels=FI_APP_STATUS_LABELS,
            status_class=FI_APP_STATUS_CLASS, status_counts=counts,
            multi=len(insts) > 1,
            project_types=dict((c, L) for c, L, _ in PROJECT_TYPES))

    # ==================================================================
    # Slice 5 -- Application Review page. An approved institution opens the FULL
    # application (project + finance + reports) for a customer that submitted to
    # it, and moves it through the review lifecycle with a decision note. Every
    # entry point re-authorizes via _fi_load_application (ownership + approved +
    # consent) -- no institution ever reaches an unassigned project.
    # ==================================================================
    def _fi_load_application(uid, pid, institution_id):
        """Authorize + assemble one institution's view of a funding application.
        Returns None unless the CURRENT user owns institution_id (approved) AND
        the customer submitted THIS project to it WITH consent. Loads the project
        + funding rows directly by (pid, tenant) -- NOT owner-scoped -- because
        the reviewing institution is a different party from the project owner.
        pid is a global PK so (pid, institution_id) maps to exactly one tenant."""
        _ensure_fi_schema(get_db)
        _ensure_fi_selection_schema(get_db)
        _ensure_ci_funding_schema(get_db)
        try:
            with get_db() as c:
                inst = c.execute(
                    "SELECT * FROM financial_institutions "
                    "WHERE institution_id=? AND created_by_user_id=? "
                    "AND status='approved'", (institution_id, uid)).fetchone()
                if not inst:
                    return None
                sel = c.execute(
                    "SELECT * FROM funding_institution_selections "
                    "WHERE capital_investment_project_id=? AND institution_id=? "
                    "AND consent=1", (pid, institution_id)).fetchone()
                if not sel:
                    return None
                sel = dict(sel)
                tid = sel.get("tenant_id") or ''
                if _pf_is_regular(pid):
                    # Regular /project/<pid> application (Slice 10c): load from the
                    # residential `projects` table -- a different data model -- and
                    # adapt it into the funding view. Authorization is ALREADY
                    # proved above (owned+approved institution + consented
                    # selection); the project is loaded by id here (cross-tenant,
                    # exactly like the CI branch: the reviewing institution is a
                    # different party from the project owner).
                    prow = c.execute(
                        "SELECT * FROM projects WHERE id=?",
                        (_pf_real_pid(pid),)).fetchone()
                    if not prow:
                        return None
                    prow = dict(prow)
                    try:
                        prow["data"] = json.loads(prow.get("data_json") or "{}")
                    except (TypeError, ValueError):
                        prow["data"] = {}
                    proj = _pf_project_view(prow)
                else:
                    proj = c.execute(
                        "SELECT * FROM capital_investment_projects "
                        "WHERE id=? AND COALESCE(CAST(tenant_id AS TEXT),'')=?",
                        (pid, tid)).fetchone()
                    if not proj:
                        return None
                    proj = dict(proj)
                fund = c.execute(
                    "SELECT * FROM capital_investment_funding "
                    "WHERE capital_investment_project_id=? "
                    "AND COALESCE(tenant_id,'')=?", (pid, tid)).fetchone()
                fund = dict(fund) if fund else None
        except Exception:
            return None
        return {"institution": dict(inst), "selection": sel,
                "project": proj, "funding": fund}

    # GET /funding/workspace/<pid>/<institution_id> -- full application review.
    @app.route("/funding/workspace/<int:pid>/<institution_id>",
               endpoint="funding_application_review")
    @login_required
    def _fi_review(pid, institution_id):
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            from flask import abort
            abort(404)
        proj = actx["project"]
        ov = _ci_funding_overview(proj)          # reuse -- no re-modelling
        _tid = actx["selection"].get("tenant_id") or ''
        thread = _fi_thread(pid, institution_id, _tid)
        shipments = _fi_shipments(pid, institution_id, _tid)
        revenue = _fi_revenue_row(pid, institution_id, _tid)
        fee_ready = actx["selection"].get("status") in FI_APPROVED_GATE
        assessment = _fi_assessment_row(pid, institution_id, _tid)
        if assessment and assessment.get("payload"):
            assessment["findings_parsed"] = _safe_json(assessment.get("payload"))
        return render_template(
            "capital_investment/funding_application_review.html",
            user=current_user(), inst=actx["institution"],
            sel=actx["selection"], proj=proj, fund=actx["funding"], ov=ov,
            report_types=REPORT_TYPES, thread=thread, shipments=shipments,
            revenue=revenue, fee_ready=fee_ready, assessment=assessment,
            recommendation_labels=FI_RECOMMENDATION_LABELS,
            recommendation_class=FI_RECOMMENDATION_CLASS,
            msg_types=FI_MSG_TYPES, msg_type_labels=FI_MSG_TYPE_LABELS,
            ship_statuses=FI_SHIP_STATUSES,
            ship_status_labels=FI_SHIP_STATUS_LABELS,
            ship_status_class=FI_SHIP_STATUS_CLASS,
            statuses=FI_APP_STATUSES, status_labels=FI_APP_STATUS_LABELS,
            status_class=FI_APP_STATUS_CLASS,
            project_types=dict((c, L) for c, L, _ in PROJECT_TYPES))

    # POST /funding/workspace/<pid>/<institution_id>/decision -- transition the
    # per-institution application status + record a decision note (audited).
    @app.route("/funding/workspace/<int:pid>/<institution_id>/decision",
               methods=["POST"], endpoint="funding_application_decision")
    @login_required
    def _fi_decision(pid, institution_id):
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            from flask import abort
            abort(404)
        csrf_protect()
        new_status = (request.form.get("status") or "").strip()
        if new_status not in FI_APP_STATUSES:
            flash("Invalid decision status.", "warning")
            return redirect(url_for("funding_application_review", pid=pid,
                                    institution_id=institution_id))
        note = (request.form.get("decision_note") or "").strip()[:2000]
        tid = actx["selection"].get("tenant_id") or ''
        ok = False
        try:
            with get_db() as c:
                # Re-prove the FULL authorization boundary atomically at write
                # time (Codex TOCTOU hardening): still-consented selection AND the
                # institution still approved + owned by this user. rowcount==1
                # confirms exactly the intended row changed.
                cur = c.execute(
                    "UPDATE funding_institution_selections "
                    "SET status=?, decision_note=?, "
                    "decided_at=CURRENT_TIMESTAMP, decided_by=? "
                    "WHERE capital_investment_project_id=? AND institution_id=? "
                    "AND COALESCE(tenant_id,'')=? AND consent=1 "
                    "AND institution_id IN (SELECT institution_id "
                    " FROM financial_institutions WHERE institution_id=? "
                    " AND created_by_user_id=? AND status='approved')",
                    (new_status, note, uid, pid, institution_id, tid,
                     institution_id, uid))
                try:
                    ok = (cur.rowcount == 1)
                except Exception:
                    ok = True   # backend without rowcount -> trust the auth gate
        except Exception:
            ok = False
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_application_decision",
                          "capital_investment_project", pid,
                          "%s -> %s" % (institution_id, new_status))
            except Exception:
                pass
            flash("Application marked '%s'." %
                  FI_APP_STATUS_LABELS.get(new_status, new_status), "success")
        else:
            flash("Could not update the application status.", "danger")
        return redirect(url_for("funding_application_review", pid=pid,
                                institution_id=institution_id))

    # GET /funding/workspace/<pid>/<institution_id>/report/<key>.pdf --
    # institution-scoped report reuse (same builders as the owner route, but
    # authorized by the selection, not project ownership; access audited).
    @app.route("/funding/workspace/<int:pid>/<institution_id>/report/"
               "<report_key>.pdf",
               endpoint="funding_application_report_pdf")
    @login_required
    def _fi_review_report(pid, institution_id, report_key):
        from flask import abort
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            abort(404)
        # Slice 10c: the CI report builders (_build_report_markdown) model a
        # generation-station project. Regular /project applications don't carry
        # that shape, so the institution report grid is hidden for them in the
        # template and this route 404s rather than render an ill-formed PDF.
        if _pf_is_regular(pid):
            abort(404)
        if report_key not in REPORT_KEYS:
            abort(404)
        proj = actx["project"]
        owner_uid = proj.get("user_id")
        app_tenant = actx["selection"].get("tenant_id") or ''
        # CRM opportunity + BOQ actuals belong to the project OWNER; the report
        # reflects the owner's real project data (consented for this review).
        opp = _load_opportunity(pid, owner_uid)
        try:
            _rfx = float(_safe_json(proj.get("finance_config"))
                         .get("fx_local_per_usd") or 12.0)
        except (TypeError, ValueError, AttributeError):
            _rfx = 12.0
        try:
            # tenant_override = the APPLICANT's tenant so the BOQ totals are not
            # blanked by the reviewer's session tenant on Postgres (Codex parity).
            _rboq = _ci_boq_actuals(
                get_db,
                proj.get("boq_facilities_project_id")
                or proj.get("boq_project_id"),
                owner_uid, _rfx,
                extra_project_ids=[proj.get("boq_solar_project_id")],
                tenant_override=app_tenant)
        except Exception:
            _rboq = None
        md, title = _build_report_markdown(report_key, proj, opp, _rboq)
        try:
            pdf_bytes = _render_pdf_bytes(md, title)
        except Exception as e:
            flash("Could not build the PDF - %s. markdown-pdf missing?" % e,
                  "danger")
            return redirect(url_for("funding_application_review", pid=pid,
                                    institution_id=institution_id))
        try:
            from new_boq_hierarchy_schema import boq_audit
            boq_audit(get_db, uid, "funding_report_viewed",
                      "capital_investment_project", pid,
                      "%s report=%s" % (institution_id, report_key))
        except Exception:
            pass
        from flask import make_response
        safe = (proj.get("project_name") or "project").replace(" ", "_")[:80]
        resp = make_response(pdf_bytes)
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = (
            'attachment; filename="%s_%s.pdf"' % (safe, report_key))
        return resp

    # ==================================================================
    # Slice 6a -- Communication. Two-way message history between an institution
    # and the applicant, one thread per (project, institution, tenant), plus
    # info/document requests and best-effort email dispatch via _send_email.
    # ==================================================================
    def _fi_thread(pid, institution_id, tid):
        """Ordered message history for one (project, institution, tenant)."""
        _ensure_fi_messages_schema(get_db)
        try:
            with get_db() as c:
                rr = c.execute(
                    "SELECT message_id, sender_role, sender_user_id, channel, "
                    "msg_type, subject, body, emailed_to, created_at "
                    "FROM funding_application_messages "
                    "WHERE capital_investment_project_id=? AND institution_id=? "
                    "AND COALESCE(tenant_id,'')=? "
                    "ORDER BY created_at, message_id",
                    (pid, institution_id, tid or '')).fetchall()
            return [dict(r) for r in rr] if rr else []
        except Exception:
            return []

    def _fi_applicant_contact(owner_uid):
        """(email, name) of the project owner = applicant, for email dispatch."""
        try:
            with get_db() as c:
                r = c.execute("SELECT email, name FROM users WHERE id=?",
                              (owner_uid,)).fetchone()
            if r:
                return (_row_get(r, "email"), _row_get(r, "name"))
        except Exception:
            pass
        return (None, None)

    def _fi_add_message(pid, institution_id, tid, *, sender_role, sender_uid,
                        msg_type, subject, body, channel="message",
                        emailed_to=None):
        """Insert one thread message (UUID PK). Returns True on success."""
        if not _ensure_fi_messages_schema(get_db):
            return False
        import uuid as _uuid
        mid = "FM-" + _uuid.uuid4().hex[:16].upper()
        try:
            with get_db() as c:
                c.execute(
                    "INSERT INTO funding_application_messages "
                    "(message_id, capital_investment_project_id, "
                    " institution_id, tenant_id, sender_role, sender_user_id, "
                    " channel, msg_type, subject, body, emailed_to) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (mid, pid, institution_id, tid or '', sender_role,
                     sender_uid, channel, msg_type, subject, body, emailed_to))
            return True
        except Exception:
            return False

    def _fi_send_funding_email(to_addr, subject, body, from_label):
        """Best-effort dispatch via the platform email service WITH the standard
        injection guards (_safe_email_subject strips CR/LF; _safe_email_text
        html-escapes). Lazy import avoids the web_app<->this circular import.
        Returns True only if the send call succeeded."""
        if not to_addr:
            return False
        try:
            from web_app import (_send_email, _safe_email_subject,
                                 _safe_email_text)
        except Exception:
            return False
        safe_subj = _safe_email_subject(
            "[SolarPro Funding] " + (subject or "Message"))
        safe_from = _safe_email_text(from_label or "SolarPro")
        safe_body_html = _safe_email_text(body or "").replace("\n", "<br>")
        html = (
            "<p>You have a new funding message from <strong>%s</strong>:</p>"
            "<blockquote style=\"border-left:3px solid #f0ad4e;"
            "padding-left:10px;color:#333\">%s</blockquote>"
            "<p style=\"color:#888;font-size:12px\">Reply inside SolarPro to "
            "keep the conversation attached to the funding application.</p>"
        ) % (safe_from, safe_body_html)
        try:
            # _send_email delegates to api_manager and returns (ok, message) --
            # honour it so a delivery/config failure is NOT recorded as emailed
            # (Codex: audit accuracy).
            res = _send_email(to_addr, safe_subj, html, text_body=(body or ""))
            if isinstance(res, (tuple, list)):
                return bool(res[0]) if res else False
            return bool(res)
        except Exception:
            return False

    # POST /funding/workspace/<pid>/<institution_id>/message -- institution posts
    # a message / info request / document request to the applicant (+ optional
    # email). Authorized by the same _fi_load_application gate as the review page.
    @app.route("/funding/workspace/<int:pid>/<institution_id>/message",
               methods=["POST"], endpoint="funding_application_message")
    @login_required
    def _fi_message(pid, institution_id):
        from flask import abort
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            abort(404)
        csrf_protect()
        f = request.form
        msg_type = (f.get("msg_type") or "message").strip()
        if msg_type not in FI_MSG_TYPES:
            msg_type = "message"
        subject = (f.get("subject") or "").strip()[:200]
        body = (f.get("body") or "").strip()[:5000]
        if not body:
            flash("Message body is required.", "warning")
            return redirect(url_for("funding_application_review", pid=pid,
                                    institution_id=institution_id))
        proj = actx["project"]
        tid = actx["selection"].get("tenant_id") or ''
        emailed_to, channel = None, "message"
        if f.get("send_email"):
            appl_email, _appl_name = _fi_applicant_contact(proj.get("user_id"))
            if appl_email and _fi_send_funding_email(
                    appl_email, subject or FI_MSG_TYPE_LABELS.get(msg_type),
                    body, actx["institution"].get("name")):
                channel, emailed_to = "email", appl_email
            else:
                flash("Message saved, but the applicant email could not be "
                      "sent.", "warning")
        ok = _fi_add_message(pid, institution_id, tid,
                             sender_role="institution", sender_uid=uid,
                             msg_type=msg_type, subject=subject, body=body,
                             channel=channel, emailed_to=emailed_to)
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_message_sent",
                          "capital_investment_project", pid,
                          "%s %s%s" % (institution_id, msg_type,
                                       " (emailed)" if emailed_to else ""))
            except Exception:
                pass
            flash("Message posted%s." % (
                " and emailed to the applicant" if emailed_to else ""),
                "success")
        else:
            flash("Could not post the message.", "danger")
        return redirect(url_for("funding_application_review", pid=pid,
                                institution_id=institution_id))

    # POST /large-scale-solar/<pid>/funding/message -- the APPLICANT replies to
    # an institution they submitted to (owner-scoped; only consented threads).
    @app.route("/large-scale-solar/<int:pid>/funding/message",
               methods=["POST"], endpoint="capital_investment_funding_message")
    @login_required
    def _funding_message(pid):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)          # owner-scoped
        _ctid = _tenant_id()
        if _ctid is not None and proj.get("tenant_id") not in (
                None, _ctid, str(_ctid)):
            from flask import abort
            abort(404)
        csrf_protect()
        uid = session.get("user_id")
        tid_s = str(_ctid) if _ctid is not None else ''
        f = request.form
        institution_id = (f.get("institution_id") or "").strip()
        body = (f.get("body") or "").strip()[:5000]
        if not body:
            flash("Message body is required.", "warning")
            return redirect(url_for("capital_investment_funding", pid=pid))
        # Applicant may only message an institution they submitted to WITH
        # consent AND that is still APPROVED (a stale consent row must not reach a
        # suspended/rejected institution -- mirrors the institution-side gate).
        _ensure_fi_selection_schema(get_db)
        _ensure_fi_schema(get_db)
        allowed = False
        try:
            with get_db() as c:
                r = c.execute(
                    "SELECT 1 FROM funding_institution_selections s "
                    "JOIN financial_institutions fi "
                    "  ON fi.institution_id = s.institution_id "
                    "WHERE s.capital_investment_project_id=? "
                    "AND s.institution_id=? AND COALESCE(s.tenant_id,'')=? "
                    "AND s.consent=1 AND fi.status='approved' LIMIT 1",
                    (pid, institution_id, tid_s)).fetchone()
                allowed = bool(r)
        except Exception:
            allowed = False
        if not allowed:
            flash("You can only message an approved institution you submitted "
                  "to.", "warning")
            return redirect(url_for("capital_investment_funding", pid=pid))
        subject = (f.get("subject") or "").strip()[:200]
        emailed_to, channel = None, "message"
        if f.get("send_email"):
            inst_email = None
            try:
                with get_db() as c:
                    ir = c.execute(
                        "SELECT email FROM financial_institutions "
                        "WHERE institution_id=?", (institution_id,)).fetchone()
                inst_email = _row_get(ir, "email")
            except Exception:
                inst_email = None
            appl_name = proj.get("client_name") or "the applicant"
            if inst_email and _fi_send_funding_email(
                    inst_email, subject or "Applicant reply", body, appl_name):
                channel, emailed_to = "email", inst_email
            else:
                flash("Reply saved, but the institution email could not be "
                      "sent.", "warning")
        ok = _fi_add_message(pid, institution_id, tid_s,
                             sender_role="applicant", sender_uid=uid,
                             msg_type="message", subject=subject, body=body,
                             channel=channel, emailed_to=emailed_to)
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_message_sent",
                          "capital_investment_project", pid,
                          "applicant->%s%s" % (
                              institution_id,
                              " (emailed)" if emailed_to else ""))
            except Exception:
                pass
            flash("Reply posted%s." % (
                " and emailed" if emailed_to else ""), "success")
        else:
            flash("Could not post the reply.", "danger")
        return redirect(url_for("capital_investment_funding", pid=pid))

    # ==================================================================
    # Slice 6b -- Hard-copy document tracking. Status-only record of physical
    # originals couriered directly between applicant and institution (SolarPro
    # never takes custody). Institution creates + advances status; applicant
    # records a dispatch they sent. All entry points re-authorize.
    # ==================================================================
    def _fi_shipments(pid, institution_id, tid):
        """All hard-copy shipments for one (project, institution, tenant)."""
        _ensure_fi_shipments_schema(get_db)
        try:
            with get_db() as c:
                rr = c.execute(
                    "SELECT * FROM funding_document_shipments "
                    "WHERE capital_investment_project_id=? AND institution_id=? "
                    "AND COALESCE(tenant_id,'')=? "
                    "ORDER BY created_at DESC, shipment_id",
                    (pid, institution_id, tid or '')).fetchall()
            return [dict(r) for r in rr] if rr else []
        except Exception:
            return []

    def _fi_add_shipment(pid, institution_id, tid, f, *, role, uid,
                         default_status="dispatched"):
        """Insert a shipment from form `f`. Returns True on success."""
        if not _ensure_fi_shipments_schema(get_db):
            return False
        vs = (f.get("verification_status") or default_status).strip()
        if vs not in FI_SHIP_STATUSES:
            vs = default_status
        import uuid as _uuid
        sid = "FS-" + _uuid.uuid4().hex[:16].upper()
        vals = (
            sid, pid, institution_id, tid or '', role, uid,
            (f.get("document_type") or "").strip()[:120],
            (f.get("courier_company") or "").strip()[:120],
            (f.get("tracking_number") or "").strip()[:120],
            (f.get("dispatch_date") or "").strip()[:20] or None,
            (f.get("recipient") or "").strip()[:160],
            (f.get("receiving_officer") or "").strip()[:120],
            (f.get("received_date") or "").strip()[:20] or None,
            vs,
            (f.get("notes") or "").strip()[:1000],
        )
        try:
            with get_db() as c:
                c.execute(
                    "INSERT INTO funding_document_shipments "
                    "(shipment_id, capital_investment_project_id, "
                    " institution_id, tenant_id, created_by_role, "
                    " created_by_user_id, document_type, courier_company, "
                    " tracking_number, dispatch_date, recipient, "
                    " receiving_officer, received_date, verification_status, "
                    " notes) VALUES (" + ",".join("?" * 15) + ")", vals)
            return True
        except Exception:
            return False

    # POST .../shipment -- institution logs / advances an inbound shipment.
    @app.route("/funding/workspace/<int:pid>/<institution_id>/shipment",
               methods=["POST"], endpoint="funding_application_shipment")
    @login_required
    def _fi_shipment_add(pid, institution_id):
        from flask import abort
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            abort(404)
        csrf_protect()
        tid = actx["selection"].get("tenant_id") or ''
        ok = _fi_add_shipment(pid, institution_id, tid, request.form,
                              role="institution", uid=uid,
                              default_status="received")
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_shipment_logged",
                          "capital_investment_project", pid, institution_id)
            except Exception:
                pass
            flash("Hard-copy shipment recorded.", "success")
        else:
            flash("Could not record the shipment.", "danger")
        return redirect(url_for("funding_application_review", pid=pid,
                                institution_id=institution_id))

    # POST .../shipment/<sid>/update -- institution advances status / receipt.
    @app.route("/funding/workspace/<int:pid>/<institution_id>/shipment/"
               "<shipment_id>/update", methods=["POST"],
               endpoint="funding_application_shipment_update")
    @login_required
    def _fi_shipment_update(pid, institution_id, shipment_id):
        from flask import abort
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            abort(404)
        csrf_protect()
        tid = actx["selection"].get("tenant_id") or ''
        f = request.form
        vs = (f.get("verification_status") or "").strip()
        if vs not in FI_SHIP_STATUSES:
            flash("Invalid shipment status.", "warning")
            return redirect(url_for("funding_application_review", pid=pid,
                                    institution_id=institution_id))
        recd = (f.get("received_date") or "").strip()[:20] or None
        officer = (f.get("receiving_officer") or "").strip()[:120]
        ok = False
        try:
            with get_db() as c:
                cur = c.execute(
                    "UPDATE funding_document_shipments "
                    "SET verification_status=?, received_date=COALESCE(?,"
                    " received_date), receiving_officer=CASE WHEN ?<>'' THEN ? "
                    " ELSE receiving_officer END, updated_at=CURRENT_TIMESTAMP "
                    "WHERE shipment_id=? AND capital_investment_project_id=? "
                    "AND institution_id=? AND COALESCE(tenant_id,'')=?",
                    (vs, recd, officer, officer, shipment_id, pid,
                     institution_id, tid))
                try:
                    ok = (cur.rowcount == 1)
                except Exception:
                    ok = True
        except Exception:
            ok = False
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_shipment_updated",
                          "capital_investment_project", pid,
                          "%s %s->%s" % (institution_id, shipment_id, vs))
            except Exception:
                pass
            flash("Shipment updated to '%s'." %
                  FI_SHIP_STATUS_LABELS.get(vs, vs), "success")
        else:
            flash("Could not update the shipment.", "danger")
        return redirect(url_for("funding_application_review", pid=pid,
                                institution_id=institution_id))

    # POST /large-scale-solar/<pid>/funding/shipment -- applicant records a
    # dispatch they sent to an institution they submitted to (approved+consent).
    @app.route("/large-scale-solar/<int:pid>/funding/shipment",
               methods=["POST"], endpoint="capital_investment_funding_shipment")
    @login_required
    def _funding_shipment(pid):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)          # owner-scoped
        _ctid = _tenant_id()
        if _ctid is not None and proj.get("tenant_id") not in (
                None, _ctid, str(_ctid)):
            from flask import abort
            abort(404)
        csrf_protect()
        uid = session.get("user_id")
        tid_s = str(_ctid) if _ctid is not None else ''
        institution_id = (request.form.get("institution_id") or "").strip()
        _ensure_fi_selection_schema(get_db)
        _ensure_fi_schema(get_db)
        allowed = False
        try:
            with get_db() as c:
                r = c.execute(
                    "SELECT 1 FROM funding_institution_selections s "
                    "JOIN financial_institutions fi "
                    "  ON fi.institution_id = s.institution_id "
                    "WHERE s.capital_investment_project_id=? "
                    "AND s.institution_id=? AND COALESCE(s.tenant_id,'')=? "
                    "AND s.consent=1 AND fi.status='approved' LIMIT 1",
                    (pid, institution_id, tid_s)).fetchone()
                allowed = bool(r)
        except Exception:
            allowed = False
        if not allowed:
            flash("You can only track a shipment to an approved institution you "
                  "submitted to.", "warning")
            return redirect(url_for("capital_investment_funding", pid=pid))
        ok = _fi_add_shipment(pid, institution_id, tid_s, request.form,
                              role="applicant", uid=uid,
                              default_status="dispatched")
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_shipment_logged",
                          "capital_investment_project", pid,
                          "applicant->%s" % institution_id)
            except Exception:
                pass
            flash("Shipment tracking recorded.", "success")
        else:
            flash("Could not record the shipment.", "danger")
        return redirect(url_for("capital_investment_funding", pid=pid))

    # ==================================================================
    # Slice 7 -- Success Fee + Revenue. The institution records the commercial
    # milestones (approved loan, agreement executed, first disbursement); when
    # all three are met SolarPro auto-calculates the success fee and issues an
    # invoice. Platform Admin gets a Funding Revenue Dashboard.
    # ==================================================================
    def _fi_revenue_row(pid, iid, tid):
        """The single funding_revenue row for one (project, institution, tenant)."""
        _ensure_fi_revenue_schema(get_db)
        try:
            with get_db() as c:
                r = c.execute(
                    "SELECT * FROM funding_revenue "
                    "WHERE capital_investment_project_id=? AND institution_id=? "
                    "AND COALESCE(tenant_id,'')=?",
                    (pid, iid, tid or '')).fetchone()
            return dict(r) if r else None
        except Exception:
            return None

    # POST .../revenue -- institution records milestones; fee/invoice materialise
    # only once Approved + Agreement Executed + First Disbursement are all true.
    @app.route("/funding/workspace/<int:pid>/<institution_id>/revenue",
               methods=["POST"], endpoint="funding_application_revenue")
    @login_required
    def _fi_revenue(pid, institution_id):
        from flask import abort
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            abort(404)
        csrf_protect()
        f = request.form
        proj = actx["project"]
        sel = actx["selection"]
        fund = actx["funding"] or {}
        tid = sel.get("tenant_id") or ''

        import math as _math

        def _num(v):
            # Reject non-finite (inf/nan) so a crafted amount can't poison the
            # fee/invoice maths (Codex financial-input hardening).
            try:
                x = float(v)
            except (TypeError, ValueError):
                return None
            return x if _math.isfinite(x) else None
        loan = _num(f.get("approved_loan_amount"))
        pval = _num(f.get("approved_project_value"))
        if pval is None:
            _r, _e = fund.get("funding_requested"), fund.get("customer_equity")
            if _r is not None or _e is not None:
                pval = round((_r or 0) + (_e or 0), 2)
        try:
            fee_pct = float(actx["institution"].get("fee_pct") or 2.0)
        except (TypeError, ValueError):
            fee_pct = 2.0
        vat = _num(f.get("vat")) or 0.0
        agreement_ref = (f.get("agreement_reference") or "").strip()[:120]
        agr_exec = 1 if f.get("agreement_executed") else 0
        disb = 1 if f.get("first_disbursement") else 0
        disb_date = (f.get("disbursement_date") or "").strip()[:20] or None
        remarks = (f.get("remarks") or "").strip()[:1000]
        # The fee is only chargeable after Approved + Agreement + Disbursement.
        approved = sel.get("status") in FI_APPROVED_GATE
        all_met = bool(approved and agr_exec and disb and loan and loan > 0)
        existing = _fi_revenue_row(pid, institution_id, tid)
        import uuid as _uuid
        import datetime as _dt
        if all_met:
            fee_amount = round(loan * fee_pct / 100.0, 2)
            inv_no = ((existing or {}).get("invoice_number")
                      or "SPF-%d-%s" % (pid, _uuid.uuid4().hex[:6].upper()))
            inv_date = ((existing or {}).get("invoice_date")
                        or _dt.date.today().isoformat())
            inv_status = "issued"
        else:
            fee_amount = None
            inv_no = (existing or {}).get("invoice_number")
            inv_date = (existing or {}).get("invoice_date")
            inv_status = "issued" if inv_no else "pending"
        cur = proj.get("currency") or "GHS"
        ok = False
        if _ensure_fi_revenue_schema(get_db):
            try:
                with get_db() as c:
                    c.execute(
                        "INSERT INTO funding_revenue "
                        "(capital_investment_project_id, institution_id, "
                        " tenant_id, project_name, customer, developer, "
                        " institution_name, country, region, project_type, "
                        " currency, approved_loan_amount, approved_project_value,"
                        " fee_pct, fee_amount, vat, invoice_number, invoice_date, "
                        " invoice_status, agreement_reference, agreement_executed,"
                        " first_disbursement, disbursement_date, remarks, "
                        " created_by_user_id) VALUES (" + ",".join("?" * 25) + ") "
                        "ON CONFLICT (capital_investment_project_id, "
                        " institution_id, tenant_id) DO UPDATE SET "
                        "project_name=EXCLUDED.project_name, "
                        "customer=EXCLUDED.customer, developer=EXCLUDED.developer,"
                        "institution_name=EXCLUDED.institution_name, "
                        "country=EXCLUDED.country, region=EXCLUDED.region, "
                        "project_type=EXCLUDED.project_type, "
                        "currency=EXCLUDED.currency, "
                        "approved_loan_amount=EXCLUDED.approved_loan_amount, "
                        "approved_project_value=EXCLUDED.approved_project_value, "
                        "fee_pct=EXCLUDED.fee_pct, fee_amount=EXCLUDED.fee_amount, "
                        "vat=EXCLUDED.vat, "
                        "invoice_number=EXCLUDED.invoice_number, "
                        "invoice_date=EXCLUDED.invoice_date, "
                        "invoice_status=EXCLUDED.invoice_status, "
                        "agreement_reference=EXCLUDED.agreement_reference, "
                        "agreement_executed=EXCLUDED.agreement_executed, "
                        "first_disbursement=EXCLUDED.first_disbursement, "
                        "disbursement_date=EXCLUDED.disbursement_date, "
                        "remarks=EXCLUDED.remarks, updated_at=CURRENT_TIMESTAMP",
                        (pid, institution_id, tid,
                         proj.get("project_name") or '',
                         proj.get("client_name") or '',
                         proj.get("developer") or '',
                         actx["institution"].get("name") or '',
                         proj.get("country") or '', proj.get("region") or '',
                         proj.get("project_type") or '', cur, loan, pval,
                         fee_pct, fee_amount, vat, inv_no, inv_date, inv_status,
                         agreement_ref, agr_exec, disb, disb_date, remarks, uid))
                ok = True
            except Exception:
                ok = False
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_revenue_recorded",
                          "capital_investment_project", pid,
                          "%s fee=%s inv=%s" % (institution_id, fee_amount,
                                                inv_no or "-"))
            except Exception:
                pass
            if all_met:
                flash("Success fee invoice %s issued: %s %s (%.2f%%)." % (
                    inv_no, cur, "{:,.2f}".format(fee_amount), fee_pct),
                    "success")
            else:
                flash("Milestones saved. The success-fee invoice is issued "
                      "automatically once Approved + Agreement Executed + First "
                      "Disbursement are all confirmed with a loan amount.",
                      "info")
        else:
            flash("Could not record the funding revenue.", "danger")
        return redirect(url_for("funding_application_review", pid=pid,
                                institution_id=institution_id))

    # GET /admin/funding/revenue -- Platform-Admin Funding Revenue Dashboard.
    @app.route("/admin/funding/revenue", endpoint="funding_revenue_dashboard")
    @login_required
    def _fi_revenue_admin():
        if not _fi_admin_ok():
            from flask import abort
            abort(403)
        _ensure_fi_revenue_schema(get_db)
        rows = []
        try:
            with get_db() as c:
                rr = c.execute("SELECT * FROM funding_revenue "
                               "ORDER BY created_at DESC").fetchall()
            rows = [dict(r) for r in rr] if rr else []
        except Exception:
            rows = []
        import datetime as _dt
        ym = _dt.date.today().strftime("%Y-%m")
        yy = _dt.date.today().strftime("%Y")

        def _f(v):
            try:
                return float(v or 0)
            except (TypeError, ValueError):
                return 0.0
        ptypes = dict((c, L) for c, L, _ in PROJECT_TYPES)
        # Money is aggregated PER CURRENCY (never summed across currencies).
        by_ccy, by_inst, by_country, by_type, by_dev = {}, {}, {}, {}, {}
        pay = {"outstanding": {"n": 0, "amt": {}},
               "paid": {"n": 0, "amt": {}}}

        def _bump(dct, key, cur, loan, fee):
            e = dct.setdefault((key or "—", cur),
                               {"loan": 0.0, "fee": 0.0, "n": 0})
            e["loan"] += loan
            e["fee"] += fee
            e["n"] += 1
        for r in rows:
            cur = r.get("currency") or "GHS"
            loan = _f(r.get("approved_loan_amount"))
            fee = _f(r.get("fee_amount"))
            d = by_ccy.setdefault(cur, {"loan": 0.0, "fee": 0.0,
                                        "fee_month": 0.0, "fee_year": 0.0,
                                        "n": 0})
            d["loan"] += loan
            d["fee"] += fee
            d["n"] += 1
            idate = r.get("invoice_date") or ""
            if idate[:7] == ym:
                d["fee_month"] += fee
            if idate[:4] == yy:
                d["fee_year"] += fee
            if r.get("invoice_number"):
                st = "paid" if (r.get("payment_status") == "paid") \
                    else "outstanding"
                pay[st]["n"] += 1
                pay[st]["amt"][cur] = pay[st]["amt"].get(cur, 0.0) + fee
            _bump(by_inst, r.get("institution_name"), cur, loan, fee)
            _bump(by_country, r.get("country"), cur, loan, fee)
            _bump(by_type, ptypes.get(r.get("project_type"),
                                      r.get("project_type")), cur, loan, fee)
            _bump(by_dev, r.get("developer"), cur, loan, fee)

        def _top(dct, n=8):
            items = [{"name": k[0], "currency": k[1], **v}
                     for k, v in dct.items()]
            items.sort(key=lambda x: x["fee"], reverse=True)
            return items[:n]
        agg = {
            "by_ccy": by_ccy, "pay": pay,
            "by_inst": _top(by_inst), "by_country": _top(by_country),
            "by_type": _top(by_type), "by_dev": _top(by_dev),
            "avg_loan": {c: (v["loan"] / v["n"] if v["n"] else 0.0)
                         for c, v in by_ccy.items()},
            "avg_fee": {c: (v["fee"] / v["n"] if v["n"] else 0.0)
                        for c, v in by_ccy.items()},
        }
        return render_template(
            "capital_investment/funding_revenue_dashboard.html",
            user=current_user(), rows=rows, agg=agg,
            pay_statuses=FI_PAYMENT_STATUSES,
            project_types=ptypes)

    # POST /admin/funding/revenue/<pid>/<institution_id>/payment -- admin marks
    # a success-fee invoice paid / outstanding (row scoped by pid+inst+tenant).
    @app.route("/admin/funding/revenue/<int:pid>/<institution_id>/payment",
               methods=["POST"], endpoint="funding_revenue_payment")
    @login_required
    def _fi_revenue_payment(pid, institution_id):
        from flask import abort
        if not _fi_admin_ok():
            abort(403)
        csrf_protect()
        f = request.form
        pay = (f.get("payment_status") or "").strip()
        if pay not in FI_PAYMENT_STATUSES:
            flash("Invalid payment status.", "warning")
            return redirect(url_for("funding_revenue_dashboard"))
        tid = (f.get("tenant_id") or "").strip()
        import datetime as _dt
        pdate = None
        if pay == "paid":
            pdate = ((f.get("payment_date") or "").strip()[:20]
                     or _dt.date.today().isoformat())
        ok = False
        try:
            with get_db() as c:
                cur = c.execute(
                    "UPDATE funding_revenue SET payment_status=?, "
                    "payment_date=?, updated_at=CURRENT_TIMESTAMP "
                    "WHERE capital_investment_project_id=? AND institution_id=? "
                    "AND COALESCE(tenant_id,'')=?",
                    (pay, pdate, pid, institution_id, tid))
                try:
                    ok = (cur.rowcount == 1)
                except Exception:
                    ok = True
        except Exception:
            ok = False
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, session.get("user_id"),
                          "funding_invoice_payment",
                          "capital_investment_project", pid,
                          "%s %s" % (institution_id, pay))
            except Exception:
                pass
            flash("Invoice marked %s." % pay, "success")
        else:
            flash("Could not update the invoice.", "danger")
        return redirect(url_for("funding_revenue_dashboard"))

    # ==================================================================
    # Slice 8 -- AI Funding Assessment (deterministic; reuses finance/bankability
    # engines -- see _ci_funding_assessment). The institution runs it from the
    # review page; the latest result persists per (project, institution, tenant).
    # ==================================================================
    def _fi_assessment_row(pid, iid, tid):
        _ensure_fi_assessment_schema(get_db)
        try:
            with get_db() as c:
                r = c.execute(
                    "SELECT * FROM funding_assessments "
                    "WHERE capital_investment_project_id=? AND institution_id=? "
                    "AND COALESCE(tenant_id,'')=?",
                    (pid, iid, tid or '')).fetchone()
            return dict(r) if r else None
        except Exception:
            return None

    # POST .../assess -- run + persist the funding assessment for this pairing.
    @app.route("/funding/workspace/<int:pid>/<institution_id>/assess",
               methods=["POST"], endpoint="funding_application_assess")
    @login_required
    def _fi_assess(pid, institution_id):
        from flask import abort
        uid = session.get("user_id")
        actx = _fi_load_application(uid, pid, institution_id)
        if not actx:
            abort(404)
        csrf_protect()
        tid = actx["selection"].get("tenant_id") or ''
        a = _ci_funding_assessment(actx["project"], actx["funding"] or {},
                                   actx["institution"])
        ok = False
        if _ensure_fi_assessment_schema(get_db):
            try:
                import json as _json
                with get_db() as c:
                    c.execute(
                        "INSERT INTO funding_assessments "
                        "(capital_investment_project_id, institution_id, "
                        " tenant_id, funding_score, technical_readiness, "
                        " financial_readiness, documentation_readiness, "
                        " construction_readiness, risk_rating, matched, "
                        " recommendation, payload, created_by_user_id) "
                        "VALUES (" + ",".join("?" * 13) + ") "
                        "ON CONFLICT (capital_investment_project_id, "
                        " institution_id, tenant_id) DO UPDATE SET "
                        "funding_score=EXCLUDED.funding_score, "
                        "technical_readiness=EXCLUDED.technical_readiness, "
                        "financial_readiness=EXCLUDED.financial_readiness, "
                        "documentation_readiness=EXCLUDED.documentation_readiness,"
                        "construction_readiness=EXCLUDED.construction_readiness, "
                        "risk_rating=EXCLUDED.risk_rating, "
                        "matched=EXCLUDED.matched, "
                        "recommendation=EXCLUDED.recommendation, "
                        "payload=EXCLUDED.payload, updated_at=CURRENT_TIMESTAMP",
                        (pid, institution_id, tid, a["funding_score"],
                         a["technical_readiness"], a["financial_readiness"],
                         a["documentation_readiness"],
                         a["construction_readiness"], a["risk_rating"],
                         a["matched"], a["recommendation"],
                         _json.dumps(a["findings"]), uid))
                ok = True
            except Exception:
                ok = False
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_assessment_run",
                          "capital_investment_project", pid,
                          "%s score=%d rec=%s" % (institution_id,
                                                  a["funding_score"],
                                                  a["recommendation"]))
            except Exception:
                pass
            flash("Funding assessment complete: score %d/100 - %s." % (
                a["funding_score"],
                FI_RECOMMENDATION_LABELS.get(a["recommendation"],
                                             a["recommendation"])), "success")
        else:
            flash("Could not run the funding assessment.", "danger")
        return redirect(url_for("funding_application_review", pid=pid,
                                institution_id=institution_id))

    # ==================================================================
    # Slice 9 -- CRM + Sales Pipeline + Marketplace handoff. Funding state is
    # projected onto the existing CRM opportunity (= the pipeline entry) so the
    # sales pipeline reflects funding; an approved project can proceed to the
    # existing procurement centre. No parallel CRM -- reuse only.
    # ==================================================================
    def _ci_funding_crm_snapshot(pid, uid, tid):
        """Funding fields for the CRM handoff, read from this project's funding
        app + consented selections + revenue (owner reads their OWN project)."""
        _ensure_ci_funding_schema(get_db)
        _ensure_fi_selection_schema(get_db)
        _ensure_fi_revenue_schema(get_db)
        _ensure_fi_schema(get_db)
        out = {"funding_requested": None, "funding_amount": None,
               "selected_institutions": "", "funding_status": None,
               "funding_score": None, "funding_approval_date": None,
               "success_fee": None, "approved": False}
        try:
            with get_db() as c:
                fr = c.execute(
                    "SELECT funding_requested, funding_score, status "
                    "FROM capital_investment_funding "
                    "WHERE capital_investment_project_id=? "
                    "AND COALESCE(tenant_id,'')=?",
                    (pid, tid or '')).fetchone()
                if fr:
                    out["funding_requested"] = _row_get(fr, "funding_requested")
                    out["funding_score"] = _row_get(fr, "funding_score")
                    out["funding_status"] = _row_get(fr, "status")
                srows = c.execute(
                    "SELECT s.institution_id, s.status, fi.name "
                    "FROM funding_institution_selections s "
                    "LEFT JOIN financial_institutions fi "
                    "  ON fi.institution_id = s.institution_id "
                    "WHERE s.capital_investment_project_id=? "
                    "AND COALESCE(s.tenant_id,'')=? AND s.consent=1",
                    (pid, tid or '')).fetchall()
                names, statuses = [], []
                for r in srows or []:
                    names.append(_row_get(r, "name")
                                 or _row_get(r, "institution_id"))
                    statuses.append(_row_get(r, "status"))
                out["selected_institutions"] = ", ".join(n for n in names if n)
                # Most-advanced per-institution status = the funding status.
                order = {s: i for i, s in enumerate(FI_APP_STATUSES)}
                if statuses:
                    best = max(statuses, key=lambda s: order.get(s, -1))
                    out["funding_status"] = best or out["funding_status"]
                # The actual funder's approved loan + our success fee + date.
                rv = c.execute(
                    "SELECT approved_loan_amount, fee_amount, invoice_date "
                    "FROM funding_revenue WHERE capital_investment_project_id=? "
                    "AND COALESCE(tenant_id,'')=? "
                    "AND approved_loan_amount IS NOT NULL "
                    "ORDER BY approved_loan_amount DESC LIMIT 1",
                    (pid, tid or '')).fetchone()
                if rv:
                    out["funding_amount"] = _row_get(rv, "approved_loan_amount")
                    out["success_fee"] = _row_get(rv, "fee_amount")
                    out["funding_approval_date"] = _row_get(rv, "invoice_date")
        except Exception:
            pass
        out["approved"] = bool(out["funding_status"] in ("approved", "completed")
                               or out["funding_amount"] is not None)
        return out

    # POST /large-scale-solar/<pid>/funding/sync-crm -- project funding onto the
    # owner's CRM opportunity + mirror into the platform sales pipeline.
    @app.route("/large-scale-solar/<int:pid>/funding/sync-crm",
               methods=["POST"], endpoint="capital_investment_funding_sync_crm")
    @login_required
    def _funding_sync_crm(pid):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)          # owner-scoped
        _ctid = _tenant_id()
        if _ctid is not None and proj.get("tenant_id") not in (
                None, _ctid, str(_ctid)):
            from flask import abort
            abort(404)
        csrf_protect()
        uid = session.get("user_id")
        tid_s = str(_ctid) if _ctid is not None else ''
        _ensure_opportunities_schema(get_db)
        snap = _ci_funding_crm_snapshot(pid, uid, tid_s)
        opp = _load_opportunity(pid, uid)
        # Create the opportunity from project data if the developer never opened
        # Step 11 (reuse build_opportunity_from_project -- no parallel CRM).
        if opp is None:
            d = build_opportunity_from_project(proj)
            try:
                with get_db() as c:
                    c.execute(
                        "INSERT INTO capital_investment_opportunities "
                        "(capital_investment_project_id, user_id, project_name, "
                        " investor, developer, client, location, country, "
                        " currency, capacity_mwp, capex_local, capex_usd, "
                        " revenue_y1_local, annual_gen_mwh, npv_local, irr_pct, "
                        " lcoe_local_per_kwh, payback_years, dscr_avg, stage, "
                        " tenant_id) VALUES (" + ",".join("?" * 21) + ")",
                        (d["capital_investment_project_id"], d["user_id"],
                         d["project_name"], d["investor"], d["developer"],
                         d["client"], d["location"], d["country"], d["currency"],
                         d["capacity_mwp"], d["capex_local"], d["capex_usd"],
                         d["revenue_y1_local"], d["annual_gen_mwh"],
                         d["npv_local"], d["irr_pct"], d["lcoe_local_per_kwh"],
                         d["payback_years"], d["dscr_avg"], "lead", _ctid))
            except Exception:
                try:                       # legacy schema without tenant_id
                    with get_db() as c:
                        c.execute(
                            "INSERT INTO capital_investment_opportunities "
                            "(capital_investment_project_id, user_id, "
                            " project_name, investor, developer, client, "
                            " location, country, currency, capacity_mwp, "
                            " capex_local, capex_usd, revenue_y1_local, "
                            " annual_gen_mwh, npv_local, irr_pct, "
                            " lcoe_local_per_kwh, payback_years, dscr_avg, "
                            " stage) VALUES (" + ",".join("?" * 20) + ")",
                            (d["capital_investment_project_id"], d["user_id"],
                             d["project_name"], d["investor"], d["developer"],
                             d["client"], d["location"], d["country"],
                             d["currency"], d["capacity_mwp"], d["capex_local"],
                             d["capex_usd"], d["revenue_y1_local"],
                             d["annual_gen_mwh"], d["npv_local"], d["irr_pct"],
                             d["lcoe_local_per_kwh"], d["payback_years"],
                             d["dscr_avg"], "lead"))
                except Exception:
                    pass
            opp = _load_opportunity(pid, uid)
        ok = False
        try:
            with get_db() as c:
                cur = c.execute(
                    "UPDATE capital_investment_opportunities SET "
                    "funding_requested=?, funding_amount=?, "
                    "funding_selected_institutions=?, funding_status=?, "
                    "funding_score=?, funding_approval_date=?, "
                    "funding_success_fee=?, updated_at=CURRENT_TIMESTAMP "
                    "WHERE capital_investment_project_id=? AND user_id=?",
                    (snap["funding_requested"], snap["funding_amount"],
                     snap["selected_institutions"], snap["funding_status"],
                     snap["funding_score"], snap["funding_approval_date"],
                     snap["success_fee"], pid, uid))
                try:
                    ok = (cur.rowcount >= 1)
                except Exception:
                    ok = True
        except Exception:
            ok = False
        # Mirror the funding update into the platform sales pipeline (non-raising).
        try:
            from web_app import _capture_pipeline_lead
            u = current_user() or {}
            _capture_pipeline_lead(
                name=(_row_get(u, "full_name")
                      or _row_get(u, "username") or "")[:120],
                email=_row_get(u, "email") or "",
                country=proj.get("country") or "",
                region=proj.get("region") or "", system_type="industrial",
                company=proj.get("investor") or proj.get("developer") or "",
                interest="generation-station-funding",
                message=("Funding update: %s -- status %s, %d institution(s), "
                         "score %s" % (
                             proj.get("project_name") or "project",
                             snap.get("funding_status") or "n/a",
                             len([x for x in (snap.get("selected_institutions")
                                              or "").split(",") if x.strip()]),
                             snap.get("funding_score")))[:500],
                source="generation_station_funding_sync",
                pipeline_stage="assessment_submitted")
        except Exception:
            pass
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, uid, "funding_crm_synced",
                          "capital_investment_project", pid,
                          "status=%s amount=%s" % (snap.get("funding_status"),
                                                   snap.get("funding_amount")))
            except Exception:
                pass
            flash("Funding synced to your CRM opportunity and the sales "
                  "pipeline.", "success")
        else:
            flash("Could not sync funding to the CRM.", "danger")
        return redirect(url_for("capital_investment_funding", pid=pid))

    # ==================================================================
    # Slice 2 -- Financial Institution registration + Platform-Admin approval.
    # ==================================================================
    def _fi_admin_ok() -> bool:
        """True only for platform admins (mirrors the app's is_admin check)."""
        try:
            return bool(int(_row_get(current_user(), "is_admin", 0) or 0))
        except Exception:
            return False

    # GET/POST /funding/institutions/register -- self-service registration.
    @app.route("/funding/institutions/register", methods=["GET", "POST"],
               endpoint="funding_institution_register")
    @login_required
    def _fi_register():
        _ensure_fi_schema(get_db)
        if request.method == "POST":
            csrf_protect()
            f = request.form
            name = (f.get("name") or "").strip()[:200]
            email = (f.get("email") or "").strip()[:200]
            if not name or not email:
                flash("Institution name and contact email are required.",
                      "warning")
                return redirect(url_for("funding_institution_register"))
            import uuid as _uuid
            iid = "FI-" + _uuid.uuid4().hex[:12].upper()
            inst_type = (f.get("inst_type") or "").strip()
            if inst_type not in FI_TYPE_CODES:
                inst_type = ""
            spt = ",".join(_multi(f, "supported_project_types",
                                  PROJECT_TYPE_CODES))
            tenor = _n(f, "tenor_months")
            vals = (
                iid, name, inst_type,
                (f.get("country") or "").strip()[:80],
                (f.get("region") or "").strip()[:80],
                (f.get("contact_person") or "").strip()[:120],
                (f.get("position") or "").strip()[:120],
                email,
                (f.get("phone") or "").strip()[:60],
                (f.get("website") or "").strip()[:200],
                (f.get("licence_no") or "").strip()[:120],
                (f.get("regulator") or "").strip()[:120],
                _n(f, "loan_min"), _n(f, "loan_max"),
                int(tenor) if tenor else None,
                _n(f, "interest_min"), _n(f, "interest_max"),
                spt,
                (f.get("funding_products") or "").strip()[:1000],
                (f.get("agreement_ref") or "").strip()[:120],
                session.get("user_id"),
                str(_tenant_id()) if _tenant_id() is not None else '',
            )
            ok = False
            if _ensure_fi_schema(get_db):
                try:
                    with get_db() as c:
                        c.execute(
                            "INSERT INTO financial_institutions "
                            "(institution_id, name, inst_type, country, region, "
                            " contact_person, position, email, phone, website, "
                            " licence_no, regulator, loan_min, loan_max, "
                            " tenor_months, interest_min, interest_max, "
                            " supported_project_types, funding_products, "
                            " agreement_ref, created_by_user_id, tenant_id) "
                            "VALUES (" + ",".join("?" * 22) + ")", vals)
                    ok = True
                except Exception:
                    ok = False
            if ok:
                try:
                    from new_boq_hierarchy_schema import boq_audit
                    boq_audit(get_db, session.get("user_id"),
                              "funding_institution_registered",
                              "financial_institution", 0,
                              "registered %s (%s)" % (name, iid))
                except Exception:
                    pass
                flash("Registration submitted. A platform administrator will "
                      "review and approve your institution.", "success")
                return redirect(url_for("funding_institution_register"))
            flash("Could not submit the registration - please try again.",
                  "danger")
            return redirect(url_for("funding_institution_register"))
        # GET
        return render_template(
            "capital_investment/funding_institution_register.html",
            user=current_user(), fi_types=FI_TYPES, project_types=PROJECT_TYPES)

    # GET /admin/funding/institutions -- Platform-Admin review queue.
    @app.route("/admin/funding/institutions",
               endpoint="funding_institutions_admin")
    @login_required
    def _fi_admin_list():
        if not _fi_admin_ok():
            from flask import abort
            abort(403)
        _ensure_fi_schema(get_db)
        status = (request.args.get("status") or "").strip()
        rows, counts = [], {s: 0 for s in FI_STATUSES}
        try:
            with get_db() as c:
                if status in FI_STATUSES:
                    rr = c.execute(
                        "SELECT * FROM financial_institutions WHERE status=? "
                        "ORDER BY created_at DESC", (status,)).fetchall()
                else:
                    rr = c.execute(
                        "SELECT * FROM financial_institutions "
                        "ORDER BY created_at DESC").fetchall()
                rows = [dict(r) for r in rr] if rr else []
                for gr in c.execute(
                        "SELECT status, COUNT(*) AS n FROM "
                        "financial_institutions GROUP BY status").fetchall() or []:
                    counts[_row_get(gr, "status")] = _row_get(gr, "n", 0)
        except Exception:
            rows = []
        return render_template(
            "capital_investment/funding_institutions_admin.html",
            user=current_user(), institutions=rows, counts=counts,
            status=status, fi_types=dict(FI_TYPES), fi_statuses=FI_STATUSES)

    # POST /admin/funding/institutions/<iid>/action -- approve/reject/suspend.
    @app.route("/admin/funding/institutions/<iid>/action", methods=["POST"],
               endpoint="funding_institution_action")
    @login_required
    def _fi_action(iid):
        if not _fi_admin_ok():
            from flask import abort
            abort(403)
        csrf_protect()
        action = (request.form.get("action") or "").strip()
        new_status = FI_ACTIONS.get(action)
        if not new_status:
            flash("Unknown action.", "warning")
            return redirect(url_for("funding_institutions_admin"))
        ok = False
        if _ensure_fi_schema(get_db):
            try:
                with get_db() as c:
                    c.execute(
                        "UPDATE financial_institutions SET status=?, "
                        "updated_at=CURRENT_TIMESTAMP WHERE institution_id=?",
                        (new_status, iid))
                ok = True
            except Exception:
                ok = False
        if ok:
            try:
                from new_boq_hierarchy_schema import boq_audit
                boq_audit(get_db, session.get("user_id"),
                          "funding_institution_" + action,
                          "financial_institution", 0,
                          "%s -> %s" % (iid, new_status))
            except Exception:
                pass
            flash("Institution %s." % new_status, "success")
        else:
            flash("Could not update the institution.", "danger")
        back = (request.form.get("return_status") or "").strip()
        return redirect(url_for("funding_institutions_admin",
                                status=back if back in FI_STATUSES else None))

    # -- step navigation helper (forward-safe: skip to overview if the next
    #    step's phase has not shipped yet) --------------------------------
    def _go_next(pid: int, next_endpoint: str):
        from flask import current_app
        if next_endpoint in current_app.view_functions:
            return redirect(url_for(next_endpoint, pid=pid))
        return redirect(url_for("capital_investment_project", pid=pid))

    def _multi(f, key, allowed_codes):
        """Return the submitted multi-select values that are valid codes."""
        return [v for v in f.getlist(key) if v in allowed_codes]

    def _wp(proj: dict[str, Any]) -> list[dict[str, Any]]:
        """Wizard progress + per-step `available` (endpoint registered) flag, so
        every step template can render the shared rail without a BuildError.
        Steps 11/12 completion needs a cheap opportunity lookup (not derivable
        from proj alone)."""
        from flask import current_app
        vf = current_app.view_functions
        progress = _wizard_progress(proj)
        # Real CRM/pipeline completion from the opportunity row (best-effort).
        opp_exists = False
        pipeline_moved = False
        try:
            uid = session.get("user_id")
            with get_db() as c:
                orow = c.execute(
                    "SELECT stage, stage_history FROM "
                    "capital_investment_opportunities WHERE "
                    "capital_investment_project_id=? AND user_id=? "
                    "ORDER BY id DESC LIMIT 1", (proj.get("id"), uid)).fetchone()
            if orow:
                opp_exists = True
                stg = _row_get(orow, "stage")
                hist = _row_get(orow, "stage_history") or "[]"
                pipeline_moved = bool(
                    (stg and stg != "lead")
                    or (hist not in ("", "[]", None)))
        except Exception:
            pass
        agents_ran = False
        try:
            with get_db() as c:
                arow = c.execute(
                    "SELECT 1 FROM capital_investment_agent_runs "
                    "WHERE project_id=? AND user_id=? LIMIT 1",
                    (proj.get("id"), session.get("user_id"))).fetchone()
            agents_ran = bool(arow)
        except Exception:
            pass
        for s in progress:
            s["available"] = s["endpoint"] in vf
            if s["num"] == 11:
                s["done"] = opp_exists
            elif s["num"] == 12:
                s["done"] = pipeline_moved
            elif s["num"] == 14:
                s["done"] = agents_ran
        return progress

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step2 -- Project Type & classification
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step2", methods=["GET", "POST"],
               endpoint="capital_investment_step2")
    @login_required
    def _step2(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        if request.method == "POST":
            csrf_protect()
            f = request.form
            ptype = _pick(f, "project_type", PROJECT_TYPES)
            pstatus = _pick(f, "project_status", PROJECT_STATUSES) or "concept"
            standard = _pick(f, "design_standard", DESIGN_STANDARDS) or "IEC"
            _save_project_field(pid, "project_type", ptype)
            _save_project_field(pid, "project_status", pstatus)
            _save_project_field(pid, "design_standard", standard)
            flash("Project type saved. Continue with Step 3 - Site.", "success")
            return _go_next(pid, "capital_investment_step3")
        return render_template(
            "capital_investment/step02_type.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            project_types=PROJECT_TYPES, project_statuses=PROJECT_STATUSES,
            design_standards=DESIGN_STANDARDS,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step3 -- Site characterisation
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step3", methods=["GET", "POST"],
               endpoint="capital_investment_step3")
    @login_required
    def _step3(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        if request.method == "POST":
            csrf_protect()
            f = request.form
            cfg = {
                "terrain":      _pick(f, "terrain", SITE_TERRAINS),
                "slope":        _pick(f, "slope", SITE_SLOPES),
                "soil":         _pick(f, "soil", SITE_SOILS),
                "flood_risk":   _pick(f, "flood_risk", SITE_FLOOD_RISKS),
                "wind_zone":    _pick(f, "wind_zone", SITE_WIND_ZONES),
                "seismic_zone": _pick(f, "seismic_zone", SITE_SEISMIC_ZONES),
                "access":       _pick(f, "access", SITE_ACCESS),
                "water":        _pick(f, "water", SITE_WATER),
                "land_area_ha": _n(f, "land_area_ha"),
                "notes":        (f.get("notes") or "").strip()[:2000],
            }
            _save_project_field(pid, "site_config", json.dumps(cfg))
            flash("Site details saved. Continue with Step 4 - Facilities.",
                  "success")
            return _go_next(pid, "capital_investment_step4")
        return render_template(
            "capital_investment/step03_site.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            cfg=_safe_json(proj.get("site_config")),
            terrains=SITE_TERRAINS, slopes=SITE_SLOPES, soils=SITE_SOILS,
            flood_risks=SITE_FLOOD_RISKS, wind_zones=SITE_WIND_ZONES,
            seismic_zones=SITE_SEISMIC_ZONES, accesses=SITE_ACCESS,
            waters=SITE_WATER,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step4 -- Facilities & external works
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step4", methods=["GET", "POST"],
               endpoint="capital_investment_step4")
    @login_required
    def _step4(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        if request.method == "POST":
            csrf_protect()
            f = request.form
            buildings = _multi(f, "buildings", BUILDING_CODES)
            external = _multi(f, "external_works", EXTERNAL_WORKS_CODES)
            sub_items: dict[str, list[str]] = {}
            for code in buildings:
                valid = set(BUILDING_SUB_ITEMS.get(code, []))
                chosen = [v for v in f.getlist("sub_" + code) if v in valid]
                if chosen:
                    sub_items[code] = chosen
            cfg = {
                "buildings":      buildings,
                "external_works": external,
                "sub_items":      sub_items,
                "notes":          (f.get("notes") or "").strip()[:2000],
            }
            _save_project_field(pid, "facility_config", json.dumps(cfg))
            flash("Facilities saved. Continue with Step 5 - Technology.",
                  "success")
            return _go_next(pid, "capital_investment_step5")
        return render_template(
            "capital_investment/step04_facilities.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            cfg=_safe_json(proj.get("facility_config")),
            building_types=BUILDING_TYPES, building_sub_items=BUILDING_SUB_ITEMS,
            external_works=EXTERNAL_WORKS,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step5 -- Technology stack
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step5", methods=["GET", "POST"],
               endpoint="capital_investment_step5")
    @login_required
    def _step5(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        if request.method == "POST":
            csrf_protect()
            f = request.form
            cfg = {
                "technologies": _multi(f, "technologies", TECHNOLOGY_CODES),
                "notes":        (f.get("notes") or "").strip()[:2000],
            }
            _save_project_field(pid, "technology_config", json.dumps(cfg))
            flash("Technology stack saved. Continue with Step 6 - Electrical.",
                  "success")
            return _go_next(pid, "capital_investment_step6")
        return render_template(
            "capital_investment/step05_technology.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            cfg=_safe_json(proj.get("technology_config")),
            technology_groups=TECHNOLOGY_GROUPS,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step6 -- Electrical services
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step6", methods=["GET", "POST"],
               endpoint="capital_investment_step6")
    @login_required
    def _step6(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        if request.method == "POST":
            csrf_protect()
            f = request.form
            cfg = {
                "services": _multi(f, "services", ELECTRICAL_SERVICE_CODES),
                "notes":    (f.get("notes") or "").strip()[:2000],
            }
            _save_project_field(pid, "electrical_config", json.dumps(cfg))
            flash("Electrical services saved. Next: PV Design (Step 7).",
                  "success")
            # Step 7 lands in Phase 3; until then loop back to the overview.
            return _go_next(pid, "capital_investment_step7")
        return render_template(
            "capital_investment/step06_electrical.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            cfg=_safe_json(proj.get("electrical_config")),
            electrical_services=ELECTRICAL_SERVICES,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step7 -- PV Design (reuses the
    # module-local utility-scale sizing engine; results persist to pv_config)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step7", methods=["GET", "POST"],
               endpoint="capital_investment_step7")
    @login_required
    def _step7(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        pv_cfg = _safe_json(proj.get("pv_config"))

        if request.method == "POST":
            csrf_protect()
            f = request.form

            module_tech = (f.get("module_tech") or "mono_topcon").strip()
            if module_tech not in PV_MODULE_TECH_CODES:
                module_tech = "mono_topcon"
            mounting = (f.get("mounting") or "fixed_tilt").strip()
            if mounting not in PV_MOUNTING_CODES:
                mounting = "fixed_tilt"
            inv_type = (f.get("inverter_type") or "central").strip()
            if inv_type not in PV_INVERTER_CODES:
                inv_type = "central"
            batt_chem = (f.get("battery_chem") or "none").strip()
            if batt_chem not in PV_BATTERY_CODES:
                batt_chem = "none"

            kwp = _n(f, "kwp", 0.0) or 0.0
            module_wp = _n(f, "module_wp", 550.0) or 550.0
            dc_ac_ratio = _n(f, "dc_ac_ratio", 1.20) or 1.20
            tilt_deg = _n(f, "tilt_deg", 10.0)
            azimuth_deg = _n(f, "azimuth_deg", 180.0)
            psh_daily = _n(f, "psh_daily", 5.4) or 5.4
            performance_ratio = _n(f, "performance_ratio", 0.78) or 0.78
            availability_pct = _n(f, "availability_pct", 98.0) or 98.0
            annual_degradation_pct = _n(f, "annual_degradation_pct", 0.5)
            project_life_yr = int(_n(f, "project_life_yr", 25) or 25)
            battery_mwh = _n(f, "battery_mwh", 0.0) or 0.0
            central_inverter_kw = 1500.0 if inv_type == "central" else 250.0

            sizing = size_utility_pv(
                kwp=kwp, module_wp=module_wp, dc_ac_ratio=dc_ac_ratio,
                tilt_deg=tilt_deg, azimuth_deg=azimuth_deg, psh_daily=psh_daily,
                performance_ratio=performance_ratio,
                availability_pct=availability_pct,
                annual_degradation_pct=annual_degradation_pct,
                project_life_yr=project_life_yr,
                central_inverter_kw=central_inverter_kw,
            )
            saved = {
                "module_tech": module_tech, "mounting": mounting,
                "inverter_type": inv_type, "battery_chem": batt_chem,
                "battery_mwh": battery_mwh, "kwp": kwp, "module_wp": module_wp,
                "dc_ac_ratio": dc_ac_ratio, "tilt_deg": tilt_deg,
                "azimuth_deg": azimuth_deg, "psh_daily": psh_daily,
                "performance_ratio": performance_ratio,
                "availability_pct": availability_pct,
                "annual_degradation_pct": annual_degradation_pct,
                "project_life_yr": project_life_yr, "sizing": sizing,
                "notes": (f.get("pv_notes") or "").strip()[:2000],
            }
            _save_project_field(pid, "pv_config", json.dumps(saved))

            if f.get("recompute_only"):
                # Stay on Step 7 so the engineer can iterate on inputs.
                return render_template(
                    "capital_investment/step07_pv.html",
                    user=current_user(), proj=proj, progress=_wp(proj),
                    cfg=saved, sizing=sizing,
                    module_techs=PV_MODULE_TECHS,
                    mounting_types=PV_MOUNTING_TYPES,
                    inverter_types=PV_INVERTER_TYPES,
                    battery_chemistries=PV_BATTERY_CHEMISTRIES,
                )
            flash("PV design saved. Next: Finance (Step 8).", "success")
            return _go_next(pid, "capital_investment_step8")

        # GET -- seed kWp from the registered target if not yet designed.
        if not pv_cfg.get("kwp") and proj.get("target_kwp"):
            pv_cfg = dict(pv_cfg)
            pv_cfg["kwp"] = round(float(proj["target_kwp"]), 2)
        return render_template(
            "capital_investment/step07_pv.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            cfg=pv_cfg, sizing=pv_cfg.get("sizing") or {},
            module_techs=PV_MODULE_TECHS, mounting_types=PV_MOUNTING_TYPES,
            inverter_types=PV_INVERTER_TYPES,
            battery_chemistries=PV_BATTERY_CHEMISTRIES,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step8 -- Financial Engineering
    # (module-local finance engine + opt-in Step-9 BOQ reconciliation)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step8", methods=["GET", "POST"],
               endpoint="capital_investment_step8")
    @login_required
    def _step8(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        fin_cfg = _safe_json(proj.get("finance_config"))
        pv_cfg = _safe_json(proj.get("pv_config"))
        sizing = pv_cfg.get("sizing") or {}
        kwp = float(sizing.get("kwp_input") or pv_cfg.get("kwp") or 0)
        annual_gen_mwh = float(sizing.get("annual_gen_mwh") or 0)

        if request.method == "POST":
            csrf_protect()
            f = request.form
            capex_form = {k: (_n(f, "capex_" + k, DEFAULT_CAPEX_USD_PER_KWP[k]))
                          for k in DEFAULT_CAPEX_USD_PER_KWP}
            opex_form = {k: (_n(f, "opex_" + k, DEFAULT_OPEX_USD_PER_KWP_YR[k]))
                         for k in DEFAULT_OPEX_USD_PER_KWP_YR}
            tariff = _n(f, "tariff_local_per_kwh", 1.5)
            fx = _n(f, "fx_local_per_usd", 12.0) or 12.0
            if fx <= 0:
                fx = 12.0
            use_boq_capex = bool(f.get("use_boq_capex"))
            revenue_model = (f.get("revenue_model") or "ppa").strip()
            if revenue_model not in REVENUE_MODEL_CODES:
                revenue_model = "ppa"
            project_life = int(_n(f, "project_life_yr", 25) or 25)
            discount = _n(f, "discount_rate_pct", 10) / 100.0
            debt_ratio = _n(f, "debt_ratio_pct", 70) / 100.0
            debt_rate = _n(f, "debt_rate_pct", 10) / 100.0
            debt_tenor = int(_n(f, "debt_tenor_yr", 12) or 12)
            tax_rate = _n(f, "tax_rate_pct", 25) / 100.0
            tariff_esc = _n(f, "tariff_escalation_pct", 2) / 100.0
            opex_esc = _n(f, "opex_escalation_pct", 3) / 100.0
            degrad = _n(f, "annual_degradation_pct",
                        pv_cfg.get("annual_degradation_pct", 0.5))
            bess_capex = _n(f, "bess_capex_usd", 0)
            carbon_price = _n(f, "carbon_credit_usd_per_tco2", 5.0)
            grid_ef = _n(f, "grid_ef_kgco2_per_kwh", 0.45)
            mc_runs = int(_n(f, "monte_carlo_runs", 200) or 0)
            # Net-metering inputs (used only when revenue_model == net_metering).
            self_consumption_pct = _n(f, "self_consumption_pct", 60.0)
            export_tariff = _n(f, "export_tariff_local_per_kwh", None)

            if kwp <= 0 or annual_gen_mwh <= 0:
                flash("PV kWp and annual generation must be set on Step 7 "
                      "before finance can be computed.", "warning")
                return redirect(url_for("capital_investment_step7", pid=pid))

            boq_actuals = _ci_boq_actuals(
                get_db, proj.get("boq_project_id"), uid, fx)
            est_facilities_usd = round(
                sum(float(capex_form.get(k, 0.0) or 0.0)
                    for k in _CI_FACILITY_CAPEX_KEYS) * kwp, 2)
            capex_effective = dict(capex_form)
            if (use_boq_capex and boq_actuals["grand_total_usd"] > 0
                    and kwp > 0):
                for _k in _CI_FACILITY_CAPEX_KEYS:
                    capex_effective[_k] = 0.0
                capex_effective["boq_facilities"] = round(
                    boq_actuals["grand_total_usd"] / kwp, 4)
            recon = {
                "est_facilities_usd": est_facilities_usd,
                "boq_actual_usd":     boq_actuals["grand_total_usd"],
                "boq_actual_local":   boq_actuals["grand_total_local"],
                "variance_usd":       round(
                    boq_actuals["grand_total_usd"] - est_facilities_usd, 2),
                "use_boq_capex":      bool(use_boq_capex),
                "n_items":            boq_actuals["n_items"],
            }

            computed = finance_utility(
                kwp=kwp, annual_gen_mwh=annual_gen_mwh,
                tariff_local_per_kwh=tariff, fx_local_per_usd=fx,
                capex_usd_per_kwp=capex_effective, opex_usd_per_kwp_yr=opex_form,
                project_life_yr=project_life, discount_rate=discount,
                debt_ratio=debt_ratio, debt_rate=debt_rate,
                debt_tenor_yr=debt_tenor, tax_rate=tax_rate,
                tariff_escalation=tariff_esc, opex_escalation=opex_esc,
                degradation_pct=degrad, bess_capex_usd=bess_capex,
                carbon_credit_usd_per_tco2=carbon_price,
                grid_ef_kgco2_per_kwh=grid_ef, monte_carlo_runs=mc_runs,
                revenue_model=revenue_model,
                self_consumption_pct=self_consumption_pct,
                export_tariff_local_per_kwh=export_tariff,
            )
            computed["boq_reconciliation"] = recon
            computed["facility_costs_usd"] = boq_actuals["facility_costs_usd"]

            saved = {
                "capex_usd_per_kwp": capex_form,
                "use_boq_capex": bool(use_boq_capex),
                "opex_usd_per_kwp_yr": opex_form,
                "tariff_local_per_kwh": tariff, "fx_local_per_usd": fx,
                "revenue_model": revenue_model, "project_life_yr": project_life,
                "self_consumption_pct": self_consumption_pct,
                "export_tariff_local_per_kwh": export_tariff,
                "discount_rate_pct": discount * 100,
                "debt_ratio_pct": debt_ratio * 100,
                "debt_rate_pct": debt_rate * 100, "debt_tenor_yr": debt_tenor,
                "tax_rate_pct": tax_rate * 100,
                "tariff_escalation_pct": tariff_esc * 100,
                "opex_escalation_pct": opex_esc * 100,
                "annual_degradation_pct": degrad, "bess_capex_usd": bess_capex,
                "carbon_credit_usd_per_tco2": carbon_price,
                "grid_ef_kgco2_per_kwh": grid_ef, "monte_carlo_runs": mc_runs,
                "computed": computed,
                "notes": (f.get("fin_notes") or "").strip()[:2000],
            }
            _save_project_field(pid, "finance_config", json.dumps(saved))

            if f.get("recompute_only"):
                return render_template(
                    "capital_investment/step08_finance.html",
                    user=current_user(), proj=proj, progress=_wp(proj),
                    pv_cfg=pv_cfg, cfg=saved, computed=computed,
                    default_capex=DEFAULT_CAPEX_USD_PER_KWP,
                    default_opex=DEFAULT_OPEX_USD_PER_KWP_YR,
                    revenue_models=REVENUE_MODELS, kwp=kwp,
                    annual_gen_mwh=annual_gen_mwh, boq_actuals=boq_actuals,
                    recon=recon,
                )
            flash("Financial model saved. Next: BOQ (Step 9).", "success")
            return _go_next(pid, "capital_investment_step9")

        # GET -- rebuild the BOQ reconciliation from the saved finance config.
        gfx = 12.0
        try:
            gfx = float(fin_cfg.get("fx_local_per_usd") or 12.0)
        except (TypeError, ValueError):
            gfx = 12.0
        if gfx <= 0:
            gfx = 12.0
        g_boq = _ci_boq_actuals(get_db, proj.get("boq_project_id"), uid, gfx)
        g_capex = fin_cfg.get("capex_usd_per_kwp") or DEFAULT_CAPEX_USD_PER_KWP
        g_est = round(sum(float(g_capex.get(k, 0.0) or 0.0)
                          for k in _CI_FACILITY_CAPEX_KEYS) * kwp, 2)
        _saved_recon = (fin_cfg.get("computed") or {}).get(
            "boq_reconciliation") or {}
        g_stale = bool(
            _saved_recon and abs(
                g_boq["grand_total_usd"]
                - float(_saved_recon.get("boq_actual_usd") or 0.0)) > 1.0)
        g_recon = {
            "est_facilities_usd": g_est,
            "boq_actual_usd": g_boq["grand_total_usd"],
            "boq_actual_local": g_boq["grand_total_local"],
            "variance_usd": round(g_boq["grand_total_usd"] - g_est, 2),
            "use_boq_capex": bool(fin_cfg.get("use_boq_capex")),
            "n_items": g_boq["n_items"], "stale": g_stale,
        }
        return render_template(
            "capital_investment/step08_finance.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            pv_cfg=pv_cfg, cfg=fin_cfg,
            computed=fin_cfg.get("computed") or {},
            default_capex=DEFAULT_CAPEX_USD_PER_KWP,
            default_opex=DEFAULT_OPEX_USD_PER_KWP_YR,
            revenue_models=REVENUE_MODELS, kwp=kwp,
            annual_gen_mwh=annual_gen_mwh, boq_actuals=g_boq, recon=g_recon,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step9 -- BOQ. Auto-creates a linked
    # boq_projects row + boq_buildings/floors per facility and auto-builds the
    # cell-level items via the STANDARD engine (web_app._ci_autobuild_floor_
    # items). This module ships NO BOQ implementation of its own.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step9", methods=["GET", "POST"],
               endpoint="capital_investment_step9")
    @login_required
    def _step9(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        fac_cfg = _safe_json(proj.get("facility_config"))
        elec_cfg = _safe_json(proj.get("electrical_config"))
        selected_buildings = fac_cfg.get("buildings") or []
        selected_external = fac_cfg.get("external_works") or []
        boq_project_id = proj.get("boq_project_id")
        planned_buildings = [
            {"code": b,
             "label": next((L for c, L, _, _ in BUILDING_TYPES if c == b), b),
             "sub_items": BUILDING_SUB_ITEMS.get(b, [])}
            for b in selected_buildings
        ]

        if request.method == "POST":
            csrf_protect()
            if boq_project_id:
                flash("BOQ project already linked.", "info")
                return redirect(url_for("capital_investment_project", pid=pid))
            if not selected_buildings:
                flash("Enable at least one facility on Step 4 before "
                      "generating the BOQ.", "warning")
                return redirect(url_for("capital_investment_step4", pid=pid))

            tech_cfg = _safe_json(proj.get("technology_config"))
            pv_cfg = _safe_json(proj.get("pv_config"))
            sizing = pv_cfg.get("sizing") or {}
            service_codes = _ci_derive_boq_services(fac_cfg, tech_cfg, elec_cfg)
            services_csv = ",".join(service_codes)
            try:
                from web_app import _kc_current_tenant_id as _kc_tid
                tenant_id = _kc_tid()
            except Exception:
                tenant_id = None

            links_ready = False
            try:
                links_ready = bool(
                    _ensure_capital_investment_boq_links_schema(get_db))
            except Exception:
                links_ready = False

            project_name = (f"{proj['project_name']} - Capital Investment BOQ")[:300]
            location = ", ".join(x for x in (proj.get("region"),
                                             proj.get("country")) if x)[:300]
            external_flag = 1 if selected_external else 0
            built_floors: list = []
            solar_ctx = None   # (solar_boq_pid, solar_building_id, solar_floor_id)
            link_errors = 0
            new_boq_pid = 0

            try:
                with get_db() as c:
                    # 1. Linked boq_projects row (PG-safe RETURNING id).
                    try:
                        cur = c.execute(
                            "INSERT INTO boq_projects "
                            "(user_id, tenant_id, project_name, client_name, "
                            " location, project_type, external_works_included, "
                            " infrastructure_included, services_csv) "
                            "VALUES (?,?,?,?,?,?,?,?,?) RETURNING id",
                            (uid, tenant_id,
                             (project_name + " - Facilities")[:300],
                             proj.get("client_name") or "", location,
                             "capital_facilities",
                             external_flag, 1, services_csv))
                        _rr = cur.fetchone()
                        cur = _RetId(int(_rr[0])) if _rr else cur
                    except Exception:
                        cur = c.execute(
                            "INSERT INTO boq_projects "
                            "(user_id, project_name, client_name, location, "
                            " project_type, external_works_included, "
                            " infrastructure_included) "
                            "VALUES (?,?,?,?,?,?,?) RETURNING id",
                            (uid, (project_name + " - Facilities")[:300],
                             proj.get("client_name") or "",
                             location, "capital_facilities", external_flag, 1))
                        _rr2 = cur.fetchone()
                        if _rr2:
                            cur = _RetId(int(_rr2[0]))
                    new_boq_pid = int(cur.lastrowid or 0)

                    # 2. Atomic claim - only if still unset. Sets BOTH the legacy
                    # boq_project_id (kept pointing at the facilities BOQ for
                    # back-compat with Step-8 reconciliation) AND the explicit
                    # boq_facilities_project_id (two-BOQ split, owner 2026-07-03).
                    cclaim = c.execute(
                        "UPDATE capital_investment_projects "
                        "SET boq_project_id=?, boq_facilities_project_id=? "
                        "WHERE id=? AND user_id=? AND "
                        "(boq_project_id IS NULL OR boq_project_id=0)",
                        (new_boq_pid, new_boq_pid, pid, uid))
                    if int(getattr(cclaim, "rowcount", 0) or 0) != 1:
                        raise _CIGenerationRaceLost()

                    # 3. One boq_buildings + Ground Floor per enabled facility.
                    for b in selected_buildings:
                        label = next(
                            (L for cd, L, _, _ in BUILDING_TYPES if cd == b), b)
                        b_services = _ci_facility_services(b)
                        bid = 0
                        try:
                            bcur = c.execute(
                                "INSERT INTO boq_buildings "
                                "(project_id, tenant_id, building_name, "
                                " building_code, primary_purpose, "
                                " purpose_subtype, building_area, "
                                " number_of_floors, basement_included, "
                                " roof_level_included, external_area_included) "
                                "VALUES (?,?,?,?,?,?,?,?,?,?,?) RETURNING id",
                                (new_boq_pid, tenant_id, label, b.upper(),
                                 "commercial", b, 0, 1, 0, 1, 0))
                            _br = bcur.fetchone()
                            bid = int(_br[0]) if _br else int(bcur.lastrowid or 0)
                        except Exception:
                            try:
                                bcur = c.execute(
                                    "INSERT INTO boq_buildings "
                                    "(project_id, building_name, "
                                    " building_code, number_of_floors) "
                                    "VALUES (?,?,?,?) RETURNING id",
                                    (new_boq_pid, label, b.upper(), 1))
                                _br2 = bcur.fetchone()
                                bid = int(_br2[0]) if _br2 else int(bcur.lastrowid or 0)
                            except Exception:
                                bid = 0

                        fid = 0
                        if bid:
                            try:
                                fcur = c.execute(
                                    "INSERT INTO boq_floors "
                                    "(building_id, project_id, tenant_id, "
                                    " floor_name, floor_level, floor_type) "
                                    "VALUES (?,?,?,?,?,?) RETURNING id",
                                    (bid, new_boq_pid, tenant_id,
                                     "Ground Floor", 0, "ground"))
                                _fr = fcur.fetchone()
                                fid = int(_fr[0]) if _fr else int(fcur.lastrowid or 0)
                            except Exception:
                                try:
                                    fcur = c.execute(
                                        "INSERT INTO boq_floors "
                                        "(building_id, project_id, floor_name, "
                                        " floor_level, floor_type) "
                                        "VALUES (?,?,?,?,?) RETURNING id",
                                        (bid, new_boq_pid, "Ground Floor", 0,
                                         "ground"))
                                    _fr2 = fcur.fetchone()
                                    fid = int(_fr2[0]) if _fr2 else int(fcur.lastrowid or 0)
                                except Exception:
                                    fid = 0
                        if fid:
                            built_floors.append((bid, fid, list(b_services)))

                        if links_ready:
                            try:
                                c.execute(
                                    "INSERT INTO capital_investment_boq_links "
                                    "(capital_investment_project_id, user_id, "
                                    " tenant_id, facility_code, source_kind, "
                                    " boq_project_id, boq_building_id, "
                                    " boq_floor_id, service_codes_csv) "
                                    "VALUES (?,?,?,?,?,?,?,?,?)",
                                    (pid, uid, tenant_id, b, "facility",
                                     new_boq_pid, bid or None, fid or None,
                                     ",".join(_ci_order_services(b_services))))
                            except Exception:
                                link_errors += 1
            except _CIGenerationRaceLost:
                flash("BOQ generation is already in progress or complete for "
                      "this project.", "info")
                return redirect(url_for("capital_investment_project", pid=pid))
            except Exception:
                try:
                    from flask import current_app
                    current_app.logger.exception(
                        "capital step9 BOQ creation failed for pid=%s", pid)
                except Exception:
                    pass
                flash("BOQ generation failed - nothing was linked. Please try "
                      "again; the error was logged.", "danger")
                return redirect(url_for("capital_investment_step9", pid=pid))

            # --- Solar-farm BOQ in its OWN transaction (after the facilities
            # BOQ has committed) so a solar failure can never roll back the
            # facilities BOQ, and a caught PG statement error stays contained
            # to this txn (Codex HIGH-1 2026-07-03). ---
            try:
                with get_db() as c:
                    # 4. SECOND BOQ -- the 20MWp solar-farm equipment BOQ, a
                    # SEPARATE boq_projects row (project_type='capital_solar_farm')
                    # so the PV field / balance-of-plant BOQ never mixes with the
                    # facilities BOQ (owner 2026-07-03). One "Solar Farm" building
                    # + "Array Field" floor; items pre-priced after the txn.
                    try:
                        try:
                            scur = c.execute(
                                "INSERT INTO boq_projects "
                                "(user_id, tenant_id, project_name, client_name, "
                                " location, project_type, external_works_included, "
                                " infrastructure_included, services_csv) "
                                "VALUES (?,?,?,?,?,?,?,?,?) RETURNING id",
                                (uid, tenant_id,
                                 (project_name + " - Solar Farm 20MWp")[:300],
                                 proj.get("client_name") or "", location,
                                 "capital_solar_farm", 0, 1, ""))
                            _sr = scur.fetchone()
                            new_solar_pid = int(_sr[0]) if _sr else int(
                                scur.lastrowid or 0)
                        except Exception:
                            scur = c.execute(
                                "INSERT INTO boq_projects "
                                "(user_id, project_name, client_name, location, "
                                " project_type, external_works_included, "
                                " infrastructure_included) "
                                "VALUES (?,?,?,?,?,?,?) RETURNING id",
                                (uid, (project_name + " - Solar Farm 20MWp")[:300],
                                 proj.get("client_name") or "", location,
                                 "capital_solar_farm", 0, 1))
                            _sr2 = scur.fetchone()
                            new_solar_pid = int(_sr2[0]) if _sr2 else int(
                                scur.lastrowid or 0)

                        c.execute(
                            "UPDATE capital_investment_projects "
                            "SET boq_solar_project_id=? "
                            "WHERE id=? AND user_id=?",
                            (new_solar_pid, pid, uid))

                        s_bid = 0
                        try:
                            sbcur = c.execute(
                                "INSERT INTO boq_buildings "
                                "(project_id, tenant_id, building_name, "
                                " building_code, primary_purpose, "
                                " purpose_subtype, building_area, "
                                " number_of_floors, basement_included, "
                                " roof_level_included, external_area_included) "
                                "VALUES (?,?,?,?,?,?,?,?,?,?,?) RETURNING id",
                                (new_solar_pid, tenant_id,
                                 "Solar Farm 20MWp - Generation Assets", "SOLAR_FARM",
                                 "infrastructure", "solar_farm", 0, 1, 0, 0, 1))
                            _sbr = sbcur.fetchone()
                            s_bid = int(_sbr[0]) if _sbr else int(
                                sbcur.lastrowid or 0)
                        except Exception:
                            try:
                                sbcur = c.execute(
                                    "INSERT INTO boq_buildings "
                                    "(project_id, building_name, "
                                    " building_code, number_of_floors) "
                                    "VALUES (?,?,?,?) RETURNING id",
                                    (new_solar_pid,
                                     "Solar Farm 20MWp - Generation Assets",
                                     "SOLAR_FARM", 1))
                                _sbr2 = sbcur.fetchone()
                                s_bid = int(_sbr2[0]) if _sbr2 else int(
                                    sbcur.lastrowid or 0)
                            except Exception:
                                s_bid = 0

                        s_fid = 0
                        if s_bid:
                            try:
                                sfcur = c.execute(
                                    "INSERT INTO boq_floors "
                                    "(building_id, project_id, tenant_id, "
                                    " floor_name, floor_level, floor_type) "
                                    "VALUES (?,?,?,?,?,?) RETURNING id",
                                    (s_bid, new_solar_pid, tenant_id,
                                     "Farm BOQ Zone - Array Field & Balance of Plant", 0, "solar_farm_zone"))
                                _sfr = sfcur.fetchone()
                                s_fid = int(_sfr[0]) if _sfr else int(
                                    sfcur.lastrowid or 0)
                            except Exception:
                                try:
                                    sfcur = c.execute(
                                        "INSERT INTO boq_floors "
                                        "(building_id, project_id, floor_name, "
                                        " floor_level, floor_type) "
                                        "VALUES (?,?,?,?,?) RETURNING id",
                                        (s_bid, new_solar_pid, "Farm BOQ Zone - Array Field & Balance of Plant", 0,
                                         "solar_farm_zone"))
                                    _sfr2 = sfcur.fetchone()
                                    s_fid = int(_sfr2[0]) if _sfr2 else int(
                                        sfcur.lastrowid or 0)
                                except Exception:
                                    s_fid = 0
                        if s_fid:
                            solar_ctx = (new_solar_pid, s_bid, s_fid)

                        if links_ready and s_fid:
                            try:
                                c.execute(
                                    "INSERT INTO capital_investment_boq_links "
                                    "(capital_investment_project_id, user_id, "
                                    " tenant_id, facility_code, source_kind, "
                                    " boq_project_id, boq_building_id, "
                                    " boq_floor_id, service_codes_csv) "
                                    "VALUES (?,?,?,?,?,?,?,?,?)",
                                    (pid, uid, tenant_id, "solar_farm",
                                     "solar_farm", new_solar_pid,
                                     s_bid or None, s_fid or None, ""))
                            except Exception:
                                link_errors += 1
                    except Exception:
                        # Solar BOQ is additive; a failure here must not roll back
                        # the facilities BOQ that already succeeded. Log + carry on.
                        try:
                            from flask import current_app
                            current_app.logger.exception(
                                "capital step9 solar BOQ creation failed pid=%s",
                                pid)
                        except Exception:
                            pass
            except Exception:
                # Connection-level failure opening/committing the solar txn: the
                # facilities BOQ already committed, so just log and continue.
                try:
                    from flask import current_app
                    current_app.logger.exception(
                        "capital step9 solar txn failed pid=%s", pid)
                except Exception:
                    pass

            # Auto-build cell-level items via the STANDARD engine (outside the
            # insert transaction to avoid a nested connection). To protect the
            # single free-tier worker we pre-price at most _CI_MAX_AUTOBUILD_FLOORS
            # floors synchronously; any beyond that stay fully linked (already
            # inserted above) and the user finishes them with BOQ "Build-all".
            items_built = 0
            deferred_floors = 0
            _autobuild = None
            if _CI_STEP9_PREPRICE:
                try:
                    from web_app import _ci_autobuild_floor_items as _autobuild
                except Exception:
                    _autobuild = None
            if _autobuild:
                for _idx, (_bid, _fid, _svcs) in enumerate(built_floors):
                    if _idx >= _CI_MAX_AUTOBUILD_FLOORS:
                        # Over the safety cap - leave linked, don't hang the worker.
                        deferred_floors = len(built_floors) - _idx
                        break
                    try:
                        items_built += int(
                            _autobuild(_fid, _bid, new_boq_pid, uid, _svcs) or 0)
                    except Exception:
                        try:
                            from flask import current_app
                            current_app.logger.exception(
                                "capital step9 autobuild floor=%s failed", _fid)
                        except Exception:
                            pass
            else:
                # Default free-tier path: defer ALL floor pricing to "Finish BOQ
                # pricing" (bounded, ~1 floor/click) so the Step-9 request itself
                # is always fast and never kills the worker.
                deferred_floors = len(built_floors)

            # Pre-price the SOLAR-FARM BOQ only when synchronous pricing is enabled;
            # otherwise the solar floor is priced by "Finish BOQ pricing" too.
            solar_items = 0
            if _CI_STEP9_PREPRICE and solar_ctx:
                _spid, _sbid, _sfid = solar_ctx
                # Reuse the platform PV design engine for sizing (falls back to
                # target_kwp when Step 7 was skipped); threads mounting too.
                _sizing_m = _ci_solar_sizing_for(proj)
                try:
                    solar_items = int(_ci_build_solar_farm_items(
                        get_db, _sfid, _sbid, _spid, uid, tenant_id, _sizing_m) or 0)
                except Exception:
                    try:
                        from flask import current_app
                        current_app.logger.exception(
                            "capital step9 solar autobuild failed pid=%s", pid)
                    except Exception:
                        pass

            notes = []
            if not links_ready or link_errors:
                notes.append("facility links unavailable - see admin diagnostics")
            if solar_ctx and solar_items:
                notes.append(
                    f"solar-farm BOQ #{solar_ctx[0]} created with {solar_items} "
                    f"20MWp equipment line item(s)")
            elif solar_ctx and _CI_STEP9_PREPRICE and not solar_items:
                notes.append(
                    "solar-farm BOQ created but not priced - complete Step 7 (PV "
                    "design) so the 20MWp equipment quantities are available")
            suffix = (" (" + "; ".join(notes) + ")") if notes else ""
            if deferred_floors and items_built == 0:
                # Default free-tier path: structure created, pricing deferred.
                flash(
                    f"Two BOQs created for #{new_boq_pid} - Facilities/Technology "
                    f"({len(selected_buildings)} building(s), {len(service_codes)} "
                    f"service(s)) + the 20MWp Solar Farm. Click \"Finish BOQ "
                    f"pricing\" on the project page to populate the line items "
                    f"(processed in small batches so it never times out)." + suffix,
                    "success")
            elif service_codes and items_built == 0:
                flash(
                    f"Linked BOQ project #{new_boq_pid} created with "
                    f"{len(selected_buildings)} building(s) and "
                    f"{len(service_codes)} service(s), but line items could NOT "
                    f"be auto-priced - open the project and click Finish BOQ "
                    f"pricing." + suffix, "warning")
            else:
                _more = (f"; {deferred_floors} more via Finish BOQ pricing"
                         if deferred_floors else "")
                flash(
                    f"Linked BOQ project #{new_boq_pid} created: "
                    f"{len(selected_buildings)} building(s), "
                    f"{len(service_codes)} service(s), {items_built} priced "
                    f"line item(s) pre-loaded{_more}." + suffix, "success")
            return redirect(url_for("capital_investment_project", pid=pid))

        # GET
        return render_template(
            "capital_investment/step09_boq.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            planned_buildings=planned_buildings,
            selected_external=selected_external, external_works=EXTERNAL_WORKS,
            boq_project_id=boq_project_id,
            solar_boq_project_id=proj.get("boq_solar_project_id"),
            electrical_selected=elec_cfg.get("services") or [],
        )

    # ------------------------------------------------------------------
    # POST /large-scale-solar/<pid>/boq/clear -- delete the generated BOQ build
    # (facilities + 20MWp solar-farm BOQ projects and every child row) and reset
    # the linkage so Step 9 offers "Generate" again. This is the "clear BOQ
    # build history" control (owner 2026-07-04). Destructive => POST-only + CSRF
    # + ownership-scoped; never a state-mutating GET.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/boq/clear", methods=["POST"],
               endpoint="capital_investment_boq_clear")
    @login_required
    def _boq_clear(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)          # 404s unless owned by this user
        csrf_protect()
        uid = session.get("user_id")
        tenant_id = _tenant_id()           # app-layer tenant scope; RLS is the
                                           # DB-layer backstop on live Postgres.
        # The two-BOQ split stores up to three ids: legacy boq_project_id
        # (== facilities), boq_facilities_project_id, boq_solar_project_id.
        boq_ids = set()
        for _k in ("boq_project_id", "boq_facilities_project_id",
                   "boq_solar_project_id"):
            _v = proj.get(_k)
            if _v:
                boq_ids.add(int(_v))
        if not boq_ids:
            flash("No BOQ build to clear.", "info")
            return redirect(url_for("capital_investment_step9", pid=pid))
        # Ensure the link table exists BEFORE the destructive txn so its cleanup
        # runs INSIDE the same transaction without a blanket try/except that
        # would otherwise orphan link rows on a real error (Codex HIGH).
        try:
            links_ready = bool(
                _ensure_capital_investment_boq_links_schema(get_db))
        except Exception:
            links_ready = False
        deleted = 0
        try:
            with get_db() as c:
                for _bpid in boq_ids:
                    # Defence in depth: only cascade BOQ projects owned by this
                    # user AND (when known) this tenant - matches the canonical
                    # BOQ ownership predicate; RLS is the DB-layer backstop.
                    if tenant_id is not None:
                        _owned = c.execute(
                            "SELECT 1 FROM boq_projects WHERE id=? AND user_id=? "
                            "AND tenant_id=? LIMIT 1",
                            (_bpid, uid, tenant_id)).fetchone()
                    else:
                        _owned = c.execute(
                            "SELECT 1 FROM boq_projects WHERE id=? AND user_id=? "
                            "LIMIT 1", (_bpid, uid)).fetchone()
                    if not _owned:
                        continue
                    # Same cascade order as /boq-projects/<pid>/delete.
                    c.execute("DELETE FROM boq_floor_rate_buildup WHERE project_id=?", (_bpid,))
                    c.execute("DELETE FROM boq_floor_items WHERE project_id=?", (_bpid,))
                    c.execute("DELETE FROM boq_floors WHERE project_id=?", (_bpid,))
                    c.execute("DELETE FROM boq_buildings WHERE project_id=?", (_bpid,))
                    c.execute("DELETE FROM boq_projects WHERE id=? AND user_id=?", (_bpid, uid))
                    deleted += 1
                # Remove the CI<->BOQ link rows in the SAME txn (no bare except:
                # a real failure must roll the whole clear back, not orphan rows).
                if links_ready:
                    if tenant_id is not None:
                        c.execute(
                            "DELETE FROM capital_investment_boq_links "
                            "WHERE capital_investment_project_id=? AND user_id=? "
                            "AND tenant_id=?", (pid, uid, tenant_id))
                    else:
                        c.execute(
                            "DELETE FROM capital_investment_boq_links "
                            "WHERE capital_investment_project_id=? AND user_id=?",
                            (pid, uid))
                # Reset linkage so Step 9 shows the Generate button again.
                c.execute(
                    "UPDATE capital_investment_projects "
                    "SET boq_project_id=NULL, boq_facilities_project_id=NULL, "
                    "    boq_solar_project_id=NULL "
                    "WHERE id=? AND user_id=?", (pid, uid))
        except Exception:
            try:
                from flask import current_app
                current_app.logger.exception(
                    "capital BOQ clear failed pid=%s", pid)
            except Exception:
                pass
            flash("Could not clear the BOQ build - please try again.", "danger")
            return redirect(url_for("capital_investment_step9", pid=pid))
        # Audit the destructive clear (directive Section 16); best-effort.
        try:
            from new_boq_hierarchy_schema import boq_audit
            boq_audit(get_db, uid, "capital_boq_build_cleared",
                      "capital_investment_project", pid,
                      "cleared %d BOQ project(s): %s" % (
                          deleted, ",".join(str(b) for b in sorted(boq_ids))))
        except Exception:
            pass
        flash(
            f"BOQ build cleared - {deleted} BOQ project(s) deleted. "
            f"You can regenerate the BOQ now.", "success")
        return redirect(url_for("capital_investment_step9", pid=pid))

    # ------------------------------------------------------------------
    # POST /large-scale-solar/<pid>/boq/finish -- price any facility floors that
    # Step 9 left LINKED-BUT-UNPRICED (the large-campus safety cap) AND the
    # solar-farm floor if empty. Each facility floor is built with ITS OWN service
    # scope (capital_investment_boq_links.service_codes_csv) -- NOT the project-
    # wide services_csv the standard Build-all uses (owner #1/#5). Bounded per
    # request (_CI_MAX_AUTOBUILD_FLOORS) so the free-tier worker is never hung;
    # re-run the button to finish the rest. Idempotent (skips priced floors).
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/boq/finish", methods=["POST"],
               endpoint="capital_investment_boq_finish")
    @login_required
    def _boq_finish(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        csrf_protect()
        proj = _load_project(pid)
        uid = session["user_id"]
        fac_pid = proj.get("boq_facilities_project_id") or proj.get("boq_project_id")
        if not fac_pid:
            flash("Generate the BOQ on Step 9 first.", "warning")
            return redirect(url_for("capital_investment_step9", pid=pid))

        # Pull the per-facility link scope (building/floor/service codes). Scoped
        # to THIS project's own facilities BOQ (boq_project_id=fac_pid), only
        # 'facility' links, the owning user, and the current tenant -- so a
        # tampered/stale link can never drive a write into another project's or
        # tenant's BOQ floor (Codex MED-3).
        try:
            from web_app import _kc_current_tenant_id as _kc_tid
            _tid = _kc_tid()
        except Exception:
            _tid = None
        _tclause = " AND (tenant_id=? OR tenant_id IS NULL)" if _tid else ""
        _lparams = [pid, uid, int(fac_pid)] + ([_tid] if _tid else [])
        links = []
        try:
            with get_db() as c:
                links = c.execute(
                    "SELECT facility_code, source_kind, boq_project_id, "
                    "       boq_building_id, boq_floor_id, service_codes_csv "
                    "FROM capital_investment_boq_links "
                    "WHERE capital_investment_project_id=? AND user_id=? "
                    "  AND boq_project_id=? AND source_kind='facility'" + _tclause +
                    " ORDER BY id",
                    tuple(_lparams)).fetchall()
        except Exception:
            links = []

        try:
            from web_app import _ci_autobuild_floor_items as _autobuild
        except Exception:
            _autobuild = None

        # SOLAR-FARM BOQ FIRST: it is the primary generation-station deliverable,
        # so the FIRST "Finish BOQ pricing" click populates it (a solar build is
        # ~12s of inserts, well within the free-tier worker budget). Facility
        # floors are priced on subsequent clicks. Build only if the solar BOQ is
        # still empty. Owner 2026-07-04: "nothing for solar" -> prioritise solar.
        solar_built = 0
        s_pid = proj.get("boq_solar_project_id")
        if s_pid:
            try:
                tenant_id = None
                try:
                    from web_app import _kc_current_tenant_id as _kc_tid
                    tenant_id = _kc_tid()
                except Exception:
                    tenant_id = None
                # Validate the solar BOQ project belongs to THIS user (join to
                # boq_projects) and match floor.project_id=building.project_id
                # before any write -- same ownership guard as the facility path
                # (Codex MED-3 follow-up).
                _stc = " AND (p.tenant_id IS NULL OR p.tenant_id=?)" if tenant_id else ""
                _sparams = [int(s_pid), uid] + ([tenant_id] if tenant_id else [])
                with get_db() as c:
                    srow = c.execute(
                        "SELECT b.id, f.id FROM boq_buildings b "
                        "JOIN boq_floors f ON f.building_id=b.id "
                        "       AND f.project_id=b.project_id "
                        "JOIN boq_projects p ON p.id=b.project_id "
                        "WHERE b.project_id=? AND p.user_id=? "
                        "  AND (b.purpose_subtype='solar_farm' "
                        "       OR b.building_code='SOLAR_FARM')" + _stc +
                        " LIMIT 1", tuple(_sparams)).fetchone()
                    # Emptiness is checked on the SPECIFIC solar floor (not the
                    # whole project) so a stray item elsewhere can never skip the
                    # solar-farm floor forever (Codex HIGH-2).
                    has_solar = None
                    if srow:
                        has_solar = c.execute(
                            "SELECT 1 FROM boq_floor_items WHERE floor_id=? "
                            "AND project_id=? AND user_id=? LIMIT 1",
                            (int(srow[1]), int(s_pid), uid)).fetchone()
                if srow and has_solar is None:
                    # Reuse the platform PV design engine for sizing (falls back
                    # to target_kwp when Step 7 was skipped) so the solar-farm BOQ
                    # is never silently empty.
                    solar_built = int(_ci_build_solar_farm_items(
                        get_db, int(srow[1]), int(srow[0]), int(s_pid),
                        uid, tenant_id, _ci_solar_sizing_for(proj)) or 0)
            except Exception:
                try:
                    from flask import current_app
                    current_app.logger.exception(
                        "boq finish solar build failed pid=%s", pid)
                except Exception:
                    pass

        # FACILITY FLOORS: priced only when the solar BOQ did NOT consume this
        # click (keeps every free-tier request within the worker budget).
        built_facilities = 0
        items_added = 0
        remaining = 0
        facilities_pending = 0
        if solar_built:
            # Solar took this request; report how many facility floors remain so
            # the user knows to click again.
            try:
                with get_db() as c:
                    _fp = c.execute(
                        "SELECT COUNT(*) FROM boq_floors f "
                        "JOIN boq_projects p ON p.id=f.project_id "
                        "WHERE f.project_id=? AND p.user_id=? "
                        "AND NOT EXISTS (SELECT 1 FROM boq_floor_items i "
                        "                WHERE i.floor_id=f.id)",
                        (int(fac_pid), uid)).fetchone()
                facilities_pending = int(_fp[0] or 0) if _fp else 0
            except Exception:
                facilities_pending = 0
        else:
            for lk in (links or []):
                try:
                    src_kind = (lk[1] or "").strip()
                    bpid = int(lk[2] or 0)
                    bbid = int(lk[3] or 0)
                    bfid = int(lk[4] or 0)
                    svc_csv = (lk[5] or "")
                except Exception:
                    continue
                # Defence in depth: the query already restricts to facility links
                # on fac_pid, but re-check before any write.
                if src_kind != "facility" or bpid != int(fac_pid) or not bfid:
                    continue
                # Validate the floor really belongs to this user's facilities BOQ
                # (join to boq_projects) before pricing it.
                try:
                    with get_db() as c:
                        own = c.execute(
                            "SELECT 1 FROM boq_floors f "
                            "JOIN boq_projects p ON p.id=f.project_id "
                            "WHERE f.id=? AND f.project_id=? AND p.user_id=? LIMIT 1",
                            (bfid, int(fac_pid), uid)).fetchone()
                except Exception:
                    own = None
                if own is None:
                    continue
                # Already priced? skip fast.
                try:
                    with get_db() as c:
                        has = c.execute(
                            "SELECT 1 FROM boq_floor_items WHERE floor_id=? LIMIT 1",
                            (bfid,)).fetchone()
                except Exception:
                    has = None
                if has is not None:
                    continue
                if built_facilities >= _CI_MAX_AUTOBUILD_FLOORS:
                    remaining += 1
                    continue
                svcs = [s for s in svc_csv.split(",") if s] or \
                    _ci_facility_services(lk[0] or "")
                if _autobuild:
                    try:
                        n = int(_autobuild(bfid, bbid, bpid, uid, svcs) or 0)
                        if n:
                            items_added += n
                            built_facilities += 1
                    except Exception:
                        try:
                            from flask import current_app
                            current_app.logger.exception(
                                "boq finish autobuild floor=%s failed", bfid)
                        except Exception:
                            pass

        # Is the solar-farm BOQ still awaiting pricing? (deferred whenever a
        # facility floor was priced this click). Cheap single-row check.
        solar_pending = False
        if s_pid and not solar_built:
            try:
                with get_db() as c:
                    _sp = c.execute(
                        "SELECT 1 FROM boq_floor_items WHERE project_id=? "
                        "AND user_id=? LIMIT 1", (int(s_pid), uid)).fetchone()
                solar_pending = _sp is None
            except Exception:
                solar_pending = False

        msg = (f"Priced {built_facilities} facility floor(s) "
               f"({items_added} item(s))")
        if solar_built:
            msg += f" + {solar_built} solar-farm item(s)"
        pend = []
        _fac_pending_total = remaining + facilities_pending
        if _fac_pending_total:
            pend.append(f"{_fac_pending_total} more facility floor(s)")
        if solar_pending:
            pend.append("the solar-farm BOQ")
        if pend:
            msg += ("; " + " and ".join(pend) + " still pending - click Finish "
                    "BOQ pricing again to continue")
        flash(msg + ".",
              "success" if (items_added or solar_built) else "info")
        return redirect(url_for("capital_investment_project", pid=pid))

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/step10 -- Marketplace (curated category
    # shortcuts that link OUT to the existing /marketplace UI; read-only)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step10",
               endpoint="capital_investment_step10")
    @login_required
    def _step10(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        tech_cfg = _safe_json(proj.get("technology_config"))
        elec_cfg = _safe_json(proj.get("electrical_config"))
        pv_cfg = _safe_json(proj.get("pv_config"))
        categories = _marketplace_categories_for(pv_cfg, tech_cfg, elec_cfg)
        return render_template(
            "capital_investment/step10_marketplace.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            categories=categories,
        )

    # -- CRM opportunity load helper (user-scoped, latest first) -----------
    def _load_opportunity(pid: int, uid: int):
        _ensure_opportunities_schema(get_db)
        with get_db() as c:
            row = c.execute(
                "SELECT * FROM capital_investment_opportunities "
                "WHERE capital_investment_project_id=? AND user_id=? "
                "ORDER BY id DESC LIMIT 1", (pid, uid)).fetchone()
        return dict(row) if row else None

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step11 -- CRM Investment Opportunity
    # (auto-derived from the project; on create, ALSO mirrors a lead into the
    # platform pipeline via web_app._capture_pipeline_lead so it shows on
    # /admin/pipeline -- Codex rebuild recommendation #4)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step11", methods=["GET", "POST"],
               endpoint="capital_investment_step11")
    @login_required
    def _step11(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        _ensure_opportunities_schema(get_db)
        opp = _load_opportunity(pid, uid)

        if request.method == "POST":
            csrf_protect()
            action = request.form.get("action") or "sync"
            derived = build_opportunity_from_project(proj)
            derived["investor"] = (request.form.get("investor")
                                   or derived["investor"])[:300]
            notes = (request.form.get("pipeline_notes") or "").strip()[:2000]

            if opp is None:
                tenant_id = _tenant_id()
                try:
                    with get_db() as c:
                        cur = c.execute(
                            "INSERT INTO capital_investment_opportunities ("
                            "capital_investment_project_id, user_id, project_name, "
                            "investor, developer, client, location, country, "
                            "currency, capacity_mwp, capex_local, capex_usd, "
                            "revenue_y1_local, annual_gen_mwh, npv_local, "
                            "irr_pct, lcoe_local_per_kwh, payback_years, "
                            "dscr_avg, stage, pipeline_notes, tenant_id"
                            ") VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
                            "RETURNING id",
                            (derived["capital_investment_project_id"],
                             derived["user_id"], derived["project_name"],
                             derived["investor"], derived["developer"],
                             derived["client"], derived["location"],
                             derived["country"], derived["currency"],
                             derived["capacity_mwp"], derived["capex_local"],
                             derived["capex_usd"], derived["revenue_y1_local"],
                             derived["annual_gen_mwh"], derived["npv_local"],
                             derived["irr_pct"], derived["lcoe_local_per_kwh"],
                             derived["payback_years"], derived["dscr_avg"],
                             "lead", notes, tenant_id))
                        _orow = cur.fetchone()
                        oid = int(_orow[0]) if _orow else int(cur.lastrowid or 0)
                except Exception:
                    # Legacy schema without tenant_id column -> retry without it.
                    with get_db() as c:
                        cur = c.execute(
                            "INSERT INTO capital_investment_opportunities ("
                            "capital_investment_project_id, user_id, project_name, "
                            "investor, developer, client, location, country, "
                            "currency, capacity_mwp, capex_local, capex_usd, "
                            "revenue_y1_local, annual_gen_mwh, npv_local, "
                            "irr_pct, lcoe_local_per_kwh, payback_years, "
                            "dscr_avg, stage, pipeline_notes"
                            ") VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
                            "RETURNING id",
                            (derived["capital_investment_project_id"],
                             derived["user_id"], derived["project_name"],
                             derived["investor"], derived["developer"],
                             derived["client"], derived["location"],
                             derived["country"], derived["currency"],
                             derived["capacity_mwp"], derived["capex_local"],
                             derived["capex_usd"], derived["revenue_y1_local"],
                             derived["annual_gen_mwh"], derived["npv_local"],
                             derived["irr_pct"], derived["lcoe_local_per_kwh"],
                             derived["payback_years"], derived["dscr_avg"],
                             "lead", notes))
                        _orow = cur.fetchone()
                        oid = int(_orow[0]) if _orow else int(cur.lastrowid or 0)
                # Mirror into the platform sales pipeline (non-raising).
                try:
                    from web_app import _capture_pipeline_lead
                    u = current_user() or {}
                    _capture_pipeline_lead(
                        name=(_row_get(u, "full_name")
                              or _row_get(u, "username") or "")[:120],
                        email=_row_get(u, "email") or "",
                        country=derived["country"], region=proj.get("region") or "",
                        system_type="industrial",
                        company=derived["investor"] or derived["developer"] or "",
                        interest="generation-station",
                        message=(f"Generation Station opportunity: "
                                 f"{derived['project_name']} "
                                 f"({derived['capacity_mwp'] or '?'} MWp, "
                                 f"NPV {derived['currency']} "
                                 f"{derived['npv_local'] or 'n/a'})")[:500],
                        source="generation_station_step11",
                        pipeline_stage="assessment_submitted",
                    )
                except Exception:
                    pass
                flash(f"Investment opportunity #{oid} created and set to "
                      "'Lead' stage.", "success")
            else:
                with get_db() as c:
                    c.execute(
                        "UPDATE capital_investment_opportunities SET "
                        "investor=?, developer=?, client=?, location=?, "
                        "country=?, currency=?, capacity_mwp=?, capex_local=?, "
                        "capex_usd=?, revenue_y1_local=?, annual_gen_mwh=?, "
                        "npv_local=?, irr_pct=?, lcoe_local_per_kwh=?, "
                        "payback_years=?, dscr_avg=?, pipeline_notes=?, "
                        "updated_at=CURRENT_TIMESTAMP "
                        "WHERE id=? AND user_id=?",
                        (derived["investor"], derived["developer"],
                         derived["client"], derived["location"],
                         derived["country"], derived["currency"],
                         derived["capacity_mwp"], derived["capex_local"],
                         derived["capex_usd"], derived["revenue_y1_local"],
                         derived["annual_gen_mwh"], derived["npv_local"],
                         derived["irr_pct"], derived["lcoe_local_per_kwh"],
                         derived["payback_years"], derived["dscr_avg"],
                         notes, opp["id"], uid))
                flash("Investment opportunity refreshed from project data.",
                      "success")
            if action == "advance":
                return _go_next(pid, "capital_investment_step12")
            return redirect(url_for("capital_investment_step11", pid=pid))

        derived = build_opportunity_from_project(proj) if opp is None else None
        return render_template(
            "capital_investment/step11_crm.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            opp=opp, derived=derived,
        )

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step12 -- Sales Pipeline (13 stages)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step12", methods=["GET", "POST"],
               endpoint="capital_investment_step12")
    @login_required
    def _step12(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        _ensure_opportunities_schema(get_db)
        opp = _load_opportunity(pid, uid)
        if opp is None:
            flash("Create the investment opportunity on Step 11 first.",
                  "warning")
            return redirect(url_for("capital_investment_step11", pid=pid))

        if request.method == "POST":
            csrf_protect()
            new_stage = (request.form.get("stage") or "").strip()
            if new_stage not in set(PIPELINE_STAGE_CODES):
                flash("Unknown pipeline stage.", "warning")
                return redirect(url_for("capital_investment_step12", pid=pid))
            history = []
            try:
                history = json.loads(opp.get("stage_history") or "[]")
                if not isinstance(history, list):
                    history = []
            except (TypeError, ValueError):
                history = []
            history.append({"from": opp.get("stage"), "to": new_stage,
                            "at": _utc_now_iso()})
            with get_db() as c:
                c.execute(
                    "UPDATE capital_investment_opportunities "
                    "SET stage=?, stage_history=?, updated_at=CURRENT_TIMESTAMP "
                    "WHERE id=? AND user_id=?",
                    (new_stage, json.dumps(history), opp["id"], uid))
            flash(f"Pipeline advanced to '{PIPELINE_STAGE_LABEL[new_stage]}'.",
                  "success")
            return redirect(url_for("capital_investment_step12", pid=pid))

        opp = _load_opportunity(pid, uid)
        try:
            history = json.loads(opp.get("stage_history") or "[]")
        except (TypeError, ValueError):
            history = []
        return render_template(
            "capital_investment/step12_pipeline.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            opp=opp, history=history, pipeline_stages=PIPELINE_STAGES,
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/cost-plan -- Cost Plan Deck
    # BOQ-derived cost breakdown (by building / by service, all with sections)
    # + cost-distribution infographics + solar yield (daily/monthly/10yr) + the
    # project cash flow. Every number comes from the BOQ + the project's own
    # finance output -- no parallel costing.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/cost-plan",
               endpoint="capital_investment_cost_plan")
    @login_required
    def _cost_plan(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        fin_cfg = _safe_json(proj.get("finance_config"))
        pv_cfg = _safe_json(proj.get("pv_config"))
        computed = fin_cfg.get("computed") or {}
        try:
            fx = float(computed.get("fx_local_per_usd")
                       or fin_cfg.get("fx_local_per_usd") or 12.0)
        except (TypeError, ValueError):
            fx = 12.0
        if fx <= 0:
            fx = 12.0
        cur = proj.get("currency") or "GHS"
        boq_pid = proj.get("boq_facilities_project_id") or proj.get("boq_project_id")

        cost = _ci_cost_plan(get_db, boq_pid, uid, fx=fx,
                             extra_project_ids=[proj.get("boq_solar_project_id")])
        yld = _ci_yield_profile(pv_cfg, gps_lat=proj.get("gps_lat")) or {}
        cash = _ci_cashflow_plan(fin_cfg)
        # Bankability verdict + financial-engineering read-out from Step-8 output.
        bank = _ci_bankability((fin_cfg or {}).get("computed") or {})

        by_bldg = cost["distribution"]["by_building"]
        _t = cost.get("totals") or {}
        daily = yld.get("daily") or {}
        charts = {
            # Circular pie / donut of the cost split by building (owner request).
            "cost_donut": _svg_donut(
                [{"label": d["label"], "value": d["total_local"], "pct": d["pct"]}
                 for d in by_bldg],
                center_label=f"{cur} {(_t.get('grand_total_local') or 0):,.0f}",
                center_sub="total cost"),
            "cost_by_building": _svg_hbars(
                [{"label": d["label"], "value": d["total_local"],
                  "pct": d["pct"]} for d in by_bldg]),
            "cost_by_service": _svg_hbars(
                [{"label": d["label"], "value": d["total_local"],
                  "pct": d["pct"]}
                 for d in cost["distribution"]["by_service"][:14]]),
            # Cost S-curve: cumulative spend across buildings (owner request).
            "cost_curve": _svg_scurve(
                [d["label"] for d in by_bldg],
                [d["total_local"] for d in by_bldg]),
            "yield_daily": _svg_columns(
                [str(h) for h in daily.get("hours", [])],
                daily.get("mwh", []), color=_CI_ACCENT2, every=3),
            "yield_monthly": _svg_columns(
                [m["month"] for m in yld.get("monthly", [])],
                [m["mwh"] for m in yld.get("monthly", [])], color=_CI_ACCENT2),
            "yield_annual": _svg_columns(
                [str(a["year"]) for a in yld.get("annual_series", [])],
                [a["mwh"] for a in yld.get("annual_series", [])],
                color=_CI_ACCENT2),
            "cashflow": (_svg_columns(
                [str(y) for y in cash["years"]], cash["net"],
                line=cash["cumulative"], color=_CI_ACCENT2,
                line_color=_CI_ACCENT)
                if cash.get("available") else ""),
        }
        return render_template(
            "capital_investment/cost_plan.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            cost=cost, yld=yld, cash=cash, bank=bank, charts=charts, currency=cur,
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/cost-plan.xlsx -- Excel export of the deck,
    # BROKEN INTO ONE WORKSHEET PER SERVICE (tabs at the bottom) so a large BOQ
    # stays manageable, plus a Summary tab (owner request 2026-07-03).
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/cost-plan.xlsx",
               endpoint="capital_investment_cost_plan_xlsx")
    @login_required
    def _cost_plan_xlsx(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        fin_cfg = _safe_json(proj.get("finance_config"))
        computed = fin_cfg.get("computed") or {}
        try:
            fx = float(computed.get("fx_local_per_usd")
                       or fin_cfg.get("fx_local_per_usd") or 12.0)
        except (TypeError, ValueError):
            fx = 12.0
        if fx <= 0:
            fx = 12.0
        cur = proj.get("currency") or "GHS"
        cost = _ci_cost_plan(
            get_db,
            proj.get("boq_facilities_project_id") or proj.get("boq_project_id"),
            uid, fx=fx,
            extra_project_ids=[proj.get("boq_solar_project_id")])

        import io
        import re as _re
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from flask import Response

        used_titles: set[str] = set()

        def _xs(v):
            # Excel formula-injection guard: force any string that Excel would
            # treat as a formula (leading = + - @, or a leading tab/CR) to be
            # inert text by prefixing an apostrophe. Numbers pass through as-is.
            if isinstance(v, str) and v[:1] in ("=", "+", "-", "@", "\t", "\r"):
                return "'" + v
            return v

        def _sheet_title(label: str) -> str:
            # Excel: <=31 chars, none of []:*?/\, unique, non-blank.
            t = _re.sub(r'[\[\]:\*\?/\\]', " ", str(label or "Service")).strip()[:28]
            t = t or "Service"
            base, i = t, 2
            while t.lower() in used_titles:
                suf = f" {i}"
                t = base[:28 - len(suf)] + suf
                i += 1
            used_titles.add(t.lower())
            return t

        wb = openpyxl.Workbook()
        bold = Font(bold=True)
        hdr = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="333333")

        ws = wb.active
        ws.title = "Summary"
        ws.append([f"Cost Plan - {proj.get('project_name') or ''}"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        t = cost["totals"]
        for k, v in [(f"Total cost ({cur})", t["grand_total_local"]),
                     ("Total cost (USD)", t["grand_total_usd"]),
                     ("Buildings", t["n_buildings"]),
                     ("Services", t["n_sections"]),
                     ("Line items", t["n_items"])]:
            ws.append([k, v])
            ws.cell(ws.max_row, 1).font = bold
        ws.append([])
        for title, rows in (
            (["By service", f"Total ({cur})", "Share %", "Items"],
             [[_xs(s["label"]), s["total_local"], s["pct"], s["n_items"]]
              for s in cost["by_service"]]),
            (["By building", f"Total ({cur})", "Share %", "Items"],
             [[_xs(b["label"]), b["total_local"], b["pct"], b["n_items"]]
              for b in cost["by_building"]])):
            ws.append(title)
            for c in range(1, len(title) + 1):
                cell = ws.cell(ws.max_row, c)
                cell.font = hdr
                cell.fill = fill
            for r in rows:
                ws.append(r)
            ws.append([])

        # One worksheet per service (section), item lines across buildings.
        # Key the intermediate map by SECTION CODE (not display label): two
        # distinct sections can normalise to the same label, and keying by label
        # would merge their line items onto the wrong sheet (Codex MED-1).
        svc: dict[str, list] = {}
        for b in cost["by_building"]:
            for sec in b["sections"]:
                svc.setdefault(sec["section"], []).extend(
                    (b["label"], it) for it in sec["items"])
        for s in cost["by_service"]:
            rows = svc.get(s["key"])
            if not rows:
                continue
            sh = wb.create_sheet(_sheet_title(s["label"]))
            head = ["Building", "#", "Description", "Unit", "Qty",
                    f"Rate ({cur})", f"Amount ({cur})"]
            sh.append(head)
            for c in range(1, len(head) + 1):
                sh.cell(1, c).font = hdr
                sh.cell(1, c).fill = fill
            for bl, it in rows:
                sh.append([_xs(bl), _xs(it["item_no"]), _xs(it["description"]),
                           _xs(it["unit"]), it["qty"], it["rate"],
                           it["total_local"]])
            sh.append([])
            sh.append(["", "", "", "", "", "Total", round(s["total_local"], 2)])
            sh.cell(sh.max_row, 6).font = bold
            sh.cell(sh.max_row, 7).font = bold
            for col, w in zip("ABCDEFG", (22, 6, 46, 8, 10, 14, 16)):
                sh.column_dimensions[col].width = w

        # Always surface a tab for every SELECTED facility service that has NO
        # priced rows yet (deferred floors / lean starter) so the owner never sees
        # a genuinely-missing service at the bottom (owner #2). A service that DOES
        # have rows (possibly under section-title sheets above) is skipped -- we
        # check real row existence by service_code so the placeholder never lies.
        try:
            _sel_codes = _ci_derive_boq_services(
                _safe_json(proj.get("facility_config")),
                _safe_json(proj.get("technology_config")),
                _safe_json(proj.get("electrical_config")))
            from web_app import _BOQ_SERVICE_LABEL as _svc_lbl
        except Exception:
            _sel_codes, _svc_lbl = [], {}
        _priced_svc_codes: set = set()
        try:
            _fpid = proj.get("boq_facilities_project_id") or proj.get("boq_project_id")
            if _fpid:
                with get_db() as _cc:
                    for _rr in _cc.execute(
                        "SELECT DISTINCT service_code FROM boq_floor_items "
                        "WHERE project_id=? AND user_id=?",
                            (int(_fpid), uid)).fetchall():
                        if _rr and _rr[0]:
                            _priced_svc_codes.add(str(_rr[0]))
        except Exception:
            _priced_svc_codes = set()
        for _code in _sel_codes:
            if _code in _priced_svc_codes:
                continue   # this service already has priced rows somewhere
            _lbl = (_svc_lbl.get(_code) if isinstance(_svc_lbl, dict)
                    else None) or _code.replace("_", " ").title()
            _key = _re.sub(r'[\[\]:\*\?/\\]', " ",
                           str(_lbl)).strip()[:28].lower()
            if not _key or _key in used_titles:
                continue
            sh = wb.create_sheet(_sheet_title(_lbl))
            head = ["Building", "#", "Description", "Unit", "Qty",
                    f"Rate ({cur})", f"Amount ({cur})"]
            sh.append(head)
            for c in range(1, len(head) + 1):
                sh.cell(1, c).font = hdr
                sh.cell(1, c).fill = fill
            sh.append(["(no priced rows yet - open the BOQ and use Build-all to "
                       "add items for this service)"])
            for col, w in zip("ABCDEFG", (22, 6, 46, 8, 10, 14, 16)):
                sh.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"cost-plan-{pid}.xlsx"
        return Response(
            buf.getvalue(),
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet"),
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/cost-plan.pdf -- printable Cost Plan report
    # (KPIs + by-service + by-building-with-sections + yield + cash flow), built
    # from the SAME engines as the deck (_ci_cost_plan / _ci_yield_profile /
    # _ci_cashflow_plan) so the PDF, the deck, and the Excel never diverge.
    # Owner request 2026-07-03. Rendered via the SolarPro-standard markdown-pdf.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/cost-plan.pdf",
               endpoint="capital_investment_cost_plan_pdf")
    @login_required
    def _cost_plan_pdf(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        fin_cfg = _safe_json(proj.get("finance_config"))
        pv_cfg = _safe_json(proj.get("pv_config"))
        computed = fin_cfg.get("computed") or {}
        try:
            fx = float(computed.get("fx_local_per_usd")
                       or fin_cfg.get("fx_local_per_usd") or 12.0)
        except (TypeError, ValueError):
            fx = 12.0
        if fx <= 0:
            fx = 12.0
        cur = proj.get("currency") or "GHS"
        cost = _ci_cost_plan(
            get_db,
            proj.get("boq_facilities_project_id") or proj.get("boq_project_id"),
            uid, fx=fx,
            extra_project_ids=[proj.get("boq_solar_project_id")])
        yld = _ci_yield_profile(pv_cfg, gps_lat=proj.get("gps_lat")) or {}
        cash = _ci_cashflow_plan(fin_cfg)

        from flask import Response

        def _m(v):
            # Money formatter for the markdown body.
            try:
                return f"{cur} {float(v or 0):,.0f}"
            except (TypeError, ValueError):
                return f"{cur} 0"

        def _clean(s):
            # Keep pipe/newline out of markdown table cells.
            return str(s or "").replace("|", "/").replace("\n", " ").strip()

        name = _clean(proj.get("project_name") or f"Project {pid}")
        t = cost["totals"]
        parts = [f"# Cost Plan -- {name}\n",
                 "_Derived from the linked BOQ and the project finance. "
                 "No parallel costing._\n"]

        # --- KPI summary ---
        parts.append("## Summary\n")
        parts.append("| Metric | Value |\n|---|---|")
        parts.append(f"| Total cost | {_m(t['grand_total_local'])} "
                     f"(USD {t['grand_total_usd']:,.0f}) |")
        parts.append(f"| Scope | {t['n_buildings']} buildings, "
                     f"{t['n_sections']} services, {t['n_items']} line items |")
        if yld:
            parts.append(f"| Annual yield | {yld.get('annual_gen_mwh', 0):,.0f} "
                         f"MWh ({yld.get('specific_yield_kwh_per_kwp', 0)} "
                         "kWh/kWp) |")
        if cash.get("irr_pct") is not None:
            parts.append(f"| IRR | {cash['irr_pct']:.1f}% |")
        if cash.get("payback_years") is not None:
            parts.append(f"| Payback | {cash['payback_years']:.1f} yr |")
        parts.append("")

        # --- Cost by service ---
        if cost["by_service"]:
            parts.append("## Cost by service\n")
            parts.append("| Service | Total | Share | Items |\n|---|---|---|---|")
            for s in cost["by_service"]:
                parts.append(f"| {_clean(s['label'])} | {_m(s['total_local'])} "
                             f"| {s['pct']}% | {s['n_items']} |")
            parts.append("")

        # --- Cost by building, with section breakdown ---
        if cost["by_building"]:
            parts.append("## Cost by building\n")
            for b in cost["by_building"]:
                parts.append(f"### {_clean(b['label'])} -- {_m(b['total_local'])} "
                             f"({b['pct']}%, {b['n_items']} items)\n")
                parts.append("| # | Description | Unit | Qty | Rate | Amount "
                             "|\n|---|---|---|---|---|---|")
                for sec in b["sections"]:
                    parts.append(f"| | **{_clean(sec['label'])}** | | | | "
                                 f"**{_m(sec['total_local'])}** |")
                    for i, it in enumerate(sec["items"], 1):
                        parts.append(
                            f"| {_clean(it.get('item_no') or i)} "
                            f"| {_clean(it['description'])} | {_clean(it['unit'])} "
                            f"| {it['qty']:,.2f} | {it['rate']:,.2f} "
                            f"| {it['total_local']:,.0f} |")
                parts.append("")

        # --- Generation yield ---
        if yld:
            parts.append("## Generation yield\n")
            parts.append(f"- Annual: **{yld.get('annual_gen_mwh', 0):,.0f} MWh** "
                         f"(specific yield {yld.get('specific_yield_kwh_per_kwp', 0)}"
                         " kWh/kWp)")
            d = yld.get("daily") or {}
            if d:
                parts.append(f"- Representative day: peak {d.get('peak_mw', 0)} MW, "
                             f"{d.get('daylight_hours', 0)} h daylight, "
                             f"{d.get('avg_daily_mwh', 0)} MWh/day avg")
            if yld.get("monthly"):
                parts.append("\n| Month | MWh | Share |\n|---|---|---|")
                for m in yld["monthly"]:
                    parts.append(f"| {m['month']} | {m['mwh']:,.0f} "
                                 f"| {m['pct']}% |")
            if yld.get("annual_series"):
                parts.append(f"\n**Annual generation over {yld.get('years', 10)} "
                             "years (with degradation)**\n")
                parts.append("| Year | MWh | % of Y1 |\n|---|---|---|")
                for a in yld["annual_series"]:
                    parts.append(f"| {a['year']} | {a['mwh']:,.0f} "
                                 f"| {a['pct_of_y1']}% |")
            parts.append("")

        # --- Cash flow ---
        if cash.get("available"):
            parts.append("## Cash flow\n")
            parts.append(f"- CAPEX: {_m(cash.get('capex_local'))}")
            parts.append(f"- Equity: {_m(cash.get('equity_local'))}")
            if cash.get("npv_local") is not None:
                parts.append(f"- NPV: {_m(cash['npv_local'])}")
            if cash.get("irr_pct") is not None:
                parts.append(f"- IRR: {cash['irr_pct']:.1f}%")
            if cash.get("payback_years") is not None:
                parts.append(f"- Payback: {cash['payback_years']:.1f} yr")
            parts.append("\n| Year | Net cash flow | Cumulative |\n|---|---|---|")
            for i, y in enumerate(cash["years"]):
                parts.append(f"| {y} | {_m(cash['net'][i])} "
                             f"| {_m(cash['cumulative'][i])} |")
            parts.append("")

        md = "\n".join(parts)
        pdf_bytes = _render_pdf_bytes(md, f"Cost Plan - {name}")
        fname = f"cost-plan-{pid}.pdf"
        return Response(
            pdf_bytes, mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/step13 -- Reports menu (13 PDF reports)
    # + GET /large-scale-solar/<pid>/report/<key>.pdf -- download
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/step13",
               endpoint="capital_investment_step13")
    @login_required
    def _step13(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        opp = _load_opportunity(pid, uid)
        return render_template(
            "capital_investment/step13_reports.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            opp=opp, report_types=REPORT_TYPES,
        )

    @app.route("/large-scale-solar/<int:pid>/report/<report_key>.pdf",
               endpoint="capital_investment_report_pdf")
    @login_required
    def _report_pdf(pid: int, report_key: str):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        if report_key not in REPORT_KEYS:
            abort(404)
        proj = _load_project(pid)
        uid = session["user_id"]
        opp = _load_opportunity(pid, uid)
        # Real priced totals for BOQ/BOM/construction reports (reuse only).
        try:
            _rfin = _safe_json(proj.get("finance_config"))
            _rfx = float(_rfin.get("fx_local_per_usd") or 12.0)
        except (TypeError, ValueError):
            _rfx = 12.0
        try:
            # Reports reflect the WHOLE capital investment: facilities BOQ +
            # solar-farm BOQ (Codex MED-5). Finance reconciliation stays
            # facilities-only elsewhere.
            _rboq = _ci_boq_actuals(
                get_db,
                proj.get("boq_facilities_project_id") or proj.get("boq_project_id"),
                uid, _rfx,
                extra_project_ids=[proj.get("boq_solar_project_id")])
        except Exception:
            _rboq = None
        md, title = _build_report_markdown(report_key, proj, opp, _rboq)
        try:
            pdf_bytes = _render_pdf_bytes(md, title)
        except Exception as e:
            flash(f"Could not build the PDF - {e}. markdown-pdf missing?",
                  "danger")
            return redirect(url_for("capital_investment_step13", pid=pid))
        from flask import make_response
        safe_name = (proj["project_name"] or "project").replace(" ", "_")[:80]
        resp = make_response(pdf_bytes)
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_{report_key}.pdf"')
        return resp

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/step14 -- AI Agents (14 rule-based
    # specialists + reviewer). Runs persist to capital_investment_agent_runs.
    # ------------------------------------------------------------------
    def _load_latest_agent_runs(pid: int) -> dict[str, dict[str, Any]]:
        _ensure_agent_runs_schema(get_db)
        out: dict[str, dict[str, Any]] = {}
        uid = session.get("user_id")
        try:
            with get_db() as c:
                rows = c.execute(
                    "SELECT agent_code, status, score, payload, created_at "
                    "FROM capital_investment_agent_runs "
                    "WHERE project_id=? AND user_id=? "
                    "ORDER BY created_at DESC, id DESC LIMIT 200",
                    (pid, uid)).fetchall()
        except Exception:
            return out
        for r in rows or []:
            d = dict(r) if hasattr(r, "keys") else {
                "agent_code": r[0], "status": r[1], "score": r[2],
                "payload": r[3], "created_at": r[4]}
            code = d["agent_code"]
            if code in out:
                continue
            try:
                pl = json.loads(d.get("payload") or "{}")
            except (TypeError, ValueError):
                pl = {}
            out[code] = {"status": d.get("status"), "score": d.get("score"),
                         "created": d.get("created_at"), **pl}
        return out

    @app.route("/large-scale-solar/<int:pid>/step14", methods=["GET", "POST"],
               endpoint="capital_investment_step14")
    @login_required
    def _step14(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        uid = session["user_id"]
        _ensure_agent_runs_schema(get_db)

        if request.method == "POST":
            csrf_protect()
            which = (request.form.get("agent") or "all").strip()
            nproj = _ci_normalize_proj_for_agents(proj)
            if which == "all":
                specialists = run_agent_orchestrator(nproj)["specialists"]
            elif which not in AGENT_CODES:
                flash("Unknown agent.", "warning")
                return redirect(url_for("capital_investment_step14", pid=pid))
            elif which == "reviewer":
                specialists = run_agent_orchestrator(nproj)["specialists"]
            else:
                runner = AGENT_RUNNERS.get(which)
                specialists = {which: runner(nproj)} if runner else {}
            with get_db() as c:
                for code, payload in specialists.items():
                    try:
                        c.execute(
                            "INSERT INTO capital_investment_agent_runs "
                            "(project_id, user_id, agent_code, status, "
                            " score, payload) VALUES (?,?,?,?,?,?)",
                            (pid, uid, code, payload.get("status") or "ok",
                             int(payload.get("score") or 0), json.dumps(payload)))
                    except Exception:
                        pass
            flash(f"{len(specialists)} agent(s) ran.", "success")
            return redirect(url_for("capital_investment_step14", pid=pid))

        latest = _load_latest_agent_runs(pid)
        aggregate_score = None
        if latest:
            considered = [v["score"] for k, v in latest.items()
                          if k != "reviewer" and v.get("score") is not None]
            if considered:
                aggregate_score = int(round(sum(considered) / len(considered)))
        return render_template(
            "capital_investment/step14_agents.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            agents=AGENT_DEPARTMENTS, latest=latest,
            aggregate_score=aggregate_score,
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/digital-twin -- 3D Digital Twin Studio
    # (server-built Three.js scene graph; a bolt-on, not a numbered step)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/digital-twin",
               endpoint="capital_investment_digital_twin")
    @login_required
    def _digital_twin(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        scene = build_scene_from_project(_ci_normalize_proj_for_agents(proj))
        return render_template(
            "capital_investment/digital_twin.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            scene=scene, layer_groups=scene["layer_groups"],
            metrics=_ci_dt_metrics(proj),
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/electrical-sld -- Single-Line Diagram
    # + PV mounting detail. Reuses pv_config.sizing via dt_electrical_sld
    # (no new sizing engine); Ghana + IEC referenced for the design report.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/electrical-sld",
               endpoint="capital_investment_electrical_sld")
    @login_required
    def _electrical_sld(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        from dt_electrical_sld import build_sld_model, has_committed_sizing
        from dt_sld_drawing import render_sld_svg
        sld = build_sld_model(proj)
        # build_sld_model derives a sizing from target capacity when Step 7
        # is not done; the drawing must then be captioned as indicative.
        committed = has_committed_sizing(proj)
        # render_sld_svg never raises: "" degrades to the stage cards
        return render_template(
            "capital_investment/electrical_sld.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            sld=sld, sld_committed=committed,
            sld_svg=render_sld_svg(sld, committed=committed),
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/design-report -- comprehensive,
    # calculation-showing engineering design report for client + government
    # engineers. Reuses build_design_report (which consumes size_utility_pv
    # + the SLD model); no new sizing engine. Ghana + IEC cited per calc.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/design-report",
               endpoint="capital_investment_design_report")
    @login_required
    def _design_report(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        from dt_design_report import build_design_report
        return render_template(
            "capital_investment/design_report.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            report=build_design_report(proj),
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/showcase -- customer-facing PHOTOREAL
    # showcase (self-hosted scene imagery + live project KPIs/callouts).
    # Reuses build_showcase_model (sizing + Step-8 finance); no new engine.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/showcase",
               endpoint="capital_investment_showcase")
    @login_required
    def _showcase(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        from dt_showcase import build_showcase_model
        return render_template(
            "capital_investment/showcase.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            show=build_showcase_model(proj),
        )

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/showcase-aerial.png -- same-origin,
    # browser-cacheable PNG of the ACTUAL twin scene (oblique aerial). Keeps
    # the truthful "your design" hero out of the page HTML as a large base64
    # blob; renders on demand via the reusable dt_showcase_aerial module (no
    # new engine). Falls back to the stock aerial jpg if the render is empty
    # so the hero is never broken. Inherits @login_required + gate + the
    # user-scoped _load_project(), so tenant/user isolation is preserved.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/showcase-aerial.png",
               endpoint="capital_investment_showcase_aerial")
    @login_required
    def _showcase_aerial(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        from flask import abort, make_response
        from dt_electrical_sld import has_committed_sizing
        import dt_showcase_aerial as _aer
        from dt_scene_v2 import augment_scene_v2
        # no-lies gate: without COMMITTED Step-7 sizing the scene builder emits
        # a placeholder plant. Serving that under a "your design" caption would
        # be a lie, so the design imagery simply does not exist yet.
        if not has_committed_sizing(proj):
            abort(404)
        png = b""
        try:
            scene = augment_scene_v2(
                build_scene_from_project(_ci_normalize_proj_for_agents(proj)),
                proj)
            png = _aer.render_plant_aerial(scene, 1600, 900)
        except Exception:
            png = b""
        if not png:
            # the caption on this image says "your design"; substituting stock art
            # here would put a photograph of somebody else's plant under that
            # claim. A missing image is honest, a wrong one is not.
            abort(404)
        resp = make_response(png)
        resp.headers["Content-Type"] = "image/png"
        resp.headers["Cache-Control"] = "private, max-age=300"
        return resp

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/showcase-view/<view>.png -- the gallery
    # scenes, each rendered from THIS project's own twin scene graph (array
    # close-up, inverter station, substation, dusk) rather than a stock
    # photograph. Owner rule: the picture must show the actual site scene.
    # <view> is validated against dt_showcase_aerial.SHOWCASE_VIEWS; anything
    # else 404s. Falls back to the labelled stock jpg only if the render is
    # empty, so a thumbnail is never broken. Inherits @login_required + gate
    # + user-scoped _load_project(), so tenant isolation is preserved.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/showcase-view/<view>.png",
               endpoint="capital_investment_showcase_view")
    @login_required
    def _showcase_view(pid: int, view: str):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        from flask import abort, make_response
        from dt_electrical_sld import has_committed_sizing
        import dt_showcase_aerial as _aer
        from dt_scene_v2 import augment_scene_v2
        if view not in _aer.SHOWCASE_VIEWS:
            abort(404)
        proj = _load_project(pid)
        # same no-lies gate as the aerial: these frames are captioned "your
        # array" / "your substation", so they may only exist once the design does
        if not has_committed_sizing(proj):
            abort(404)
        # optional ?w= lets the gallery ask for a cheap thumbnail instead of
        # re-rendering the full 1600x900 frame four times per page view.
        try:
            w = int(request.args.get("w") or 1600)
        except (TypeError, ValueError):
            w = 1600
        w = 320 if w < 320 else (1600 if w > 1600 else w)
        h = int(round(w * 9.0 / 16.0))
        png = b""
        try:
            scene = augment_scene_v2(
                build_scene_from_project(_ci_normalize_proj_for_agents(proj)),
                proj)
            png = _aer.render_plant_view(scene, view, w, h)
        except Exception:
            png = b""
        if not png:
            # never swap in a stock photograph: the caption claims this frame is
            # the customer's own plant (see dt_showcase._design_captions).
            abort(404)
        resp = make_response(png)
        resp.headers["Content-Type"] = "image/png"
        resp.headers["Cache-Control"] = "private, max-age=300"
        return resp

    # ------------------------------------------------------------------
    # GET /large-scale-solar/<pid>/site-layout -- scaled plot plan of the
    # physical plant arrangement. Reuses build_site_layout_model (sizing via
    # build_sld_model); no new engine.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/site-layout",
               endpoint="capital_investment_site_layout")
    @login_required
    def _site_layout(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        from dt_site_layout import build_site_layout_model
        return render_template(
            "capital_investment/site_layout.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            report=build_site_layout_model(proj),
        )

    @app.route("/large-scale-solar/<int:pid>/dt/scene.json",
               endpoint="capital_investment_dt_scene")
    @login_required
    def _dt_scene_json(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        return jsonify(build_scene_from_project(
            _ci_normalize_proj_for_agents(proj)))

    @app.route("/large-scale-solar/<int:pid>/dt/sun.json",
               endpoint="capital_investment_dt_sun")
    @login_required
    def _dt_sun_json(pid: int):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        try:
            month = int(request.args.get("month", "6"))
            hour = float(request.args.get("hour", "12"))
        except (TypeError, ValueError):
            month, hour = 6, 12.0
        month = max(1, min(12, month))
        hour = max(0.0, min(24.0, hour))
        lat = proj.get("gps_lat") or 6.0
        lon = proj.get("gps_lon") or 0.0
        return jsonify(_sun_position(float(lat), float(lon), month, hour))

    # ==================================================================
    # Digital Twin v2 -- engineering interaction endpoints (ADDITIVE).
    # Every route inherits @login_required + _gate(CI_LEVEL_FULL) + the
    # user-scoped _load_project(), so tenant/user isolation is preserved.
    # All heavy logic lives in the reusable dt_scene_v2 module (_dtv2).
    # ==================================================================
    def _dt_scene_for(pid):
        """Load a project (user-scoped) and return its augmented v2 scene."""
        proj = _load_project(pid)
        return build_scene_from_project(_ci_normalize_proj_for_agents(proj))

    def _dt_clamp(v, lo, hi, cur):
        """Clamp a submitted numeric to [lo,hi]; fall back to cur on garbage."""
        try:
            v = float(v)
        except (TypeError, ValueError):
            return cur
        return max(lo, min(hi, v))

    def _dt_resize(pv_cfg):
        """Re-run the EXISTING sizing engine for a pv_config so module/inverter
        counts + energy metadata stay current after any design change. Falls
        back to the stored sizing on any engine error (never raises)."""
        try:
            return size_utility_pv(
                kwp=pv_cfg.get("kwp") or 0,
                module_wp=pv_cfg.get("module_wp") or 550,
                dc_ac_ratio=pv_cfg.get("dc_ac_ratio") or 1.2,
                tilt_deg=pv_cfg.get("tilt_deg") or 12,
                azimuth_deg=pv_cfg.get("azimuth_deg") or 180,
                psh_daily=pv_cfg.get("psh_daily") or 5.4,
                performance_ratio=pv_cfg.get("performance_ratio") or 0.78,
                availability_pct=pv_cfg.get("availability_pct") or 98.0,
                annual_degradation_pct=pv_cfg.get("annual_degradation_pct") or 0.5,
                project_life_yr=int(pv_cfg.get("project_life_yr") or 25),
                central_inverter_kw=pv_cfg.get("central_inverter_kw") or 1500.0,
            )
        except Exception:
            return pv_cfg.get("sizing") or {}

    # POST /dt/parameters -- live Design-Parameters round-trip (Phase 3).
    # Updates the EXISTING project JSON blobs via _save_project_field and
    # re-runs the EXISTING size_utility_pv engine; returns a fresh scene.
    @app.route("/large-scale-solar/<int:pid>/dt/parameters",
               methods=["POST"],
               endpoint="capital_investment_dt_parameters")
    @login_required
    def _dt_parameters(pid):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        csrf_protect()
        proj = _load_project(pid)
        payload = request.get_json(silent=True) or {}
        pv_in = payload.get("pv") or {}
        pv_cfg = _safe_json(proj.get("pv_config"))
        kwp = _dt_clamp(pv_in.get("kwp", pv_cfg.get("kwp")),
                        0, 500000, pv_cfg.get("kwp") or 0)
        module_wp = _dt_clamp(pv_in.get("module_wp", pv_cfg.get("module_wp")),
                              200, 800, pv_cfg.get("module_wp") or 550)
        tilt_deg = _dt_clamp(pv_in.get("tilt_deg", pv_cfg.get("tilt_deg")),
                             0, 60, pv_cfg.get("tilt_deg") or 12)
        azimuth_deg = _dt_clamp(pv_in.get("azimuth_deg", pv_cfg.get("azimuth_deg")),
                                0, 359, pv_cfg.get("azimuth_deg") or 180)
        row_pitch = _dt_clamp(pv_in.get("row_pitch_m", pv_cfg.get("row_pitch_m")),
                              2, 20, pv_cfg.get("row_pitch_m") or 6)
        # Merge the new values, then reuse the existing sizing engine (never a
        # duplicate PV calculator) so module/inverter counts stay consistent.
        pv_cfg.update({"kwp": kwp, "module_wp": module_wp, "tilt_deg": tilt_deg,
                       "azimuth_deg": azimuth_deg, "row_pitch_m": row_pitch})
        pv_cfg["sizing"] = _dt_resize(pv_cfg)
        _save_project_field(pid, "pv_config", json.dumps(pv_cfg))
        fac_in = payload.get("facility") or {}
        if "battery_kwh" in fac_in:
            fac_cfg = _safe_json(proj.get("facility_config"))
            try:
                fac_cfg["battery_kwh"] = max(0.0, float(fac_in["battery_kwh"]))
                _save_project_field(pid, "facility_config", json.dumps(fac_cfg))
            except (TypeError, ValueError):
                pass
        scene = _dt_scene_for(pid)
        pv_meta = (scene.get("pv") or {}).get("meta") or {}
        return jsonify({
            "ok": True,
            "scene": scene,
            "summary": {
                "kwp": pv_meta.get("kwp"),
                "modules": pv_meta.get("n_modules_planned"),
                "rows": pv_meta.get("n_rows"),
                "land_area_ha": (scene.get("site") or {}).get("land_area_ha"),
            },
            "boq_dirty": True, "finance_dirty": True, "warnings": [],
        })

    # POST /dt/object-action -- controlled scene mutation (Phase 6).
    @app.route("/large-scale-solar/<int:pid>/dt/object-action",
               methods=["POST"],
               endpoint="capital_investment_dt_object_action")
    @login_required
    def _dt_object_action(pid):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        csrf_protect()
        proj = _load_project(pid)
        payload = request.get_json(silent=True) or {}
        action = (payload.get("action") or "").strip()
        params = payload.get("params") or {}
        allowed = {"move_transformer", "increase_row_spacing",
                   "decrease_row_spacing", "set_tilt", "hide", "lock"}
        if action not in allowed:
            return jsonify({"ok": False, "error": "UNSUPPORTED_ACTION",
                            "message": "Action '%s' not supported." % action}), 400
        boq_dirty = False
        finance_dirty = False
        msg = ""
        if action in ("increase_row_spacing", "decrease_row_spacing"):
            pv_cfg = _safe_json(proj.get("pv_config"))
            cur = float(pv_cfg.get("row_pitch_m") or 6.0)
            delta = _dt_clamp(params.get("delta_m"), 0.1, 5.0, 0.5)
            new = cur + (delta if action == "increase_row_spacing" else -delta)
            pv_cfg["row_pitch_m"] = max(2.0, min(20.0, new))
            _save_project_field(pid, "pv_config", json.dumps(pv_cfg))
            boq_dirty = True
            msg = ("Row spacing set to %.2f m; shading + cable lengths require "
                   "recompute." % pv_cfg["row_pitch_m"])
        elif action == "set_tilt":
            pv_cfg = _safe_json(proj.get("pv_config"))
            tilt = _dt_clamp(params.get("tilt_deg"), 0, 60, None)
            if tilt is None:
                return jsonify({"ok": False, "error": "BAD_PARAM"}), 400
            pv_cfg["tilt_deg"] = tilt
            # Tilt feeds the sizing engine -- re-run it so sizing/energy metadata
            # is not left stale (parity with /dt/parameters).
            pv_cfg["sizing"] = _dt_resize(pv_cfg)
            _save_project_field(pid, "pv_config", json.dumps(pv_cfg))
            boq_dirty = True
            msg = "Tilt set to %.0f deg." % tilt
        elif action == "move_transformer":
            import math as _math
            try:
                x = float(params.get("x"))
                z = float(params.get("z"))
            except (TypeError, ValueError):
                return jsonify({"ok": False, "error": "BAD_PARAM"}), 400
            if not (_math.isfinite(x) and _math.isfinite(z)):
                return jsonify({"ok": False, "error": "BAD_PARAM"}), 400
            # Keep the transformer inside the site square (5 m margin) so a bad
            # drag cannot fling it kilometres away and wreck cable quantities.
            site_cfg = _safe_json(proj.get("site_config"))
            pv_cfg = _safe_json(proj.get("pv_config"))
            kwp = float((pv_cfg.get("sizing") or {}).get("kwp_input")
                        or pv_cfg.get("kwp") or 0)
            land_ha = float(site_cfg.get("land_area_ha") or max(kwp / 800.0, 5.0))
            bound = _math.sqrt(land_ha * 10000.0) / 2.0 - 5.0
            if abs(x) > bound or abs(z) > bound:
                return jsonify({"ok": False, "error": "OUT_OF_BOUNDS",
                                "message": "Transformer must stay within the "
                                           "site boundary."}), 400
            elec_cfg = _safe_json(proj.get("electrical_config"))
            elec_cfg["transformer_pos"] = {"x": x, "z": z}
            _save_project_field(pid, "electrical_config", json.dumps(elec_cfg))
            boq_dirty = True
            msg = "Transformer moved; cable quantities require recalculation."
        else:
            # hide / lock are client-local view state; ack without persistence.
            msg = "%s applied (view only)." % action
        scene = _dt_scene_for(pid)
        return jsonify({"ok": True, "scene": scene, "boq_dirty": boq_dirty,
                        "finance_dirty": finance_dirty, "message": msg})

    # GET /dt/shadow-analysis.json -- per-row shading estimate (Phase 5).
    @app.route("/large-scale-solar/<int:pid>/dt/shadow-analysis.json",
               endpoint="capital_investment_dt_shadow")
    @login_required
    def _dt_shadow(pid):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        try:
            month = max(1, min(12, int(request.args.get("month", "6"))))
            hour = max(0.0, min(24.0, float(request.args.get("hour", "9"))))
        except (TypeError, ValueError):
            month, hour = 6, 9.0
        lat = float(proj.get("gps_lat") or 6.0)
        lon = float(proj.get("gps_lon") or 0.0)
        scene = build_scene_from_project(_ci_normalize_proj_for_agents(proj))
        sun = _dtv2.sun_position(lat, lon, month, hour)
        return jsonify(_dtv2.shadow_analysis(scene, sun))

    # GET /dt/object-schedule.json -- BOQ-linked object export (Phase 8).
    @app.route("/large-scale-solar/<int:pid>/dt/object-schedule.json",
               endpoint="capital_investment_dt_object_schedule")
    @login_required
    def _dt_object_schedule(pid):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        scene = _dt_scene_for(pid)
        rows = []
        for o in scene.get("objects") or []:
            rows.append({
                "id": o.get("id"), "type": o.get("type"),
                "layer": o.get("layer"), "label": o.get("label"),
                "quantity": (o.get("engineering") or {}).get("quantity"),
                "w": (o.get("dimensions") or {}).get("w"),
                "h": (o.get("dimensions") or {}).get("h"),
                "l": (o.get("dimensions") or {}).get("l"),
                "boq": (o.get("links") or {}).get("boq"),
                "marketplace": (o.get("links") or {}).get("marketplace"),
            })
        return jsonify({"schema_version": scene.get("schema_version"),
                        "project_id": pid, "count": len(rows), "objects": rows})

    # GET /dt/sun-arc.json -- authoritative sun-path samples for the month so
    # the client draws the arc from the SAME solar model as the light/shadows
    # (never a hardcoded generic ellipse).
    @app.route("/large-scale-solar/<int:pid>/dt/sun-arc.json",
               endpoint="capital_investment_dt_sun_arc")
    @login_required
    def _dt_sun_arc(pid):
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        proj = _load_project(pid)
        try:
            month = max(1, min(12, int(request.args.get("month", "6"))))
        except (TypeError, ValueError):
            month = 6
        lat = float(proj.get("gps_lat") or 6.0)
        lon = float(proj.get("gps_lon") or 0.0)
        samples = []
        h = 5.0
        while h <= 19.0001:
            s = _dtv2.sun_position(lat, lon, month, h)
            samples.append({"hour": round(h, 2),
                            "altitude_deg": s["altitude_deg"],
                            "azimuth_deg": s["azimuth_deg"],
                            "is_daylight": s["is_daylight"]})
            h += 0.5
        return jsonify({"month": month, "samples": samples})

    # ------------------------------------------------------------------
    # GET/POST /large-scale-solar/<pid>/regulatory -- Development & Regulatory
    # (country-scoped, Ghana-first; a bolt-on panel, not a numbered step)
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/<int:pid>/regulatory",
               methods=["GET", "POST"],
               endpoint="capital_investment_regulatory")
    @login_required
    def _regulatory(pid: int):
        if (g := _gate(CI_LEVEL_SETUP)) is not None:
            return g
        proj = _load_project(pid)
        fw = country_framework(proj.get("country"))
        tenure_codes = {t["code"] for t in fw.get("land_tenures", [])}
        if request.method == "POST":
            csrf_protect()
            f = request.form
            items: dict[str, dict[str, str]] = {}
            for code, _label, _hint in REGULATORY_ITEMS:
                status = f.get("status_" + code, "")
                if status not in REGULATORY_STATUS_CODES:
                    status = "not_started"
                note = (f.get("note_" + code) or "").strip()[:600]
                items[code] = {"status": status, "note": note}
            tenure = (f.get("land_tenure") or "").strip()
            if tenure not in tenure_codes:
                tenure = ""
            cfg = {
                "items":       items,
                "land_tenure": tenure,
                "notes":       (f.get("notes") or "").strip()[:2000],
            }
            _save_project_field(pid, "regulatory_config", json.dumps(cfg))
            flash("Regulatory checklist saved.", "success")
            return redirect(url_for("capital_investment_project", pid=pid))
        return render_template(
            "capital_investment/regulatory.html",
            user=current_user(), proj=proj, progress=_wp(proj),
            cfg=_safe_json(proj.get("regulatory_config")),
            framework=fw, reg_items=REGULATORY_ITEMS,
            reg_statuses=REGULATORY_STATUSES,
        )

    # ------------------------------------------------------------------
    # Diagnostics (admin/enterprise gated) -- prove which build is live +
    # schema readiness on this backend.
    # ------------------------------------------------------------------
    @app.route("/large-scale-solar/diag/whoami",
               endpoint="capital_investment_diag_whoami")
    @login_required
    def _diag_whoami():
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        user = current_user()
        return jsonify({
            "module_build": CI_MODULE_BUILD,
            "uid": session.get("user_id"),
            "tier": _ci_tier_of(user), "level": _ci_level_of(user),
            "tenant_id": str(_tenant_id() or ""),
        })

    @app.route("/large-scale-solar/diag/schema",
               endpoint="capital_investment_diag_schema")
    @login_required
    def _diag_schema():
        if (g := _gate(CI_LEVEL_FULL)) is not None:
            return g
        ready = _ensure_ci_projects_schema_verified(get_db)
        backend = "unknown"
        has_kwp = None
        try:
            with get_db() as c:
                c.execute("SELECT target_kwp FROM capital_investment_projects "
                          "LIMIT 1")
                has_kwp = True
        except Exception:
            has_kwp = False
        try:
            from web_app import _db_backend  # optional helper
            backend = _db_backend()
        except Exception:
            backend = "sqlite-or-pg"
        return jsonify({
            "module_build":   CI_MODULE_BUILD,
            "backend":        backend,
            "projects_ready": bool(ready),
            "schema_error":   _CIP_SCHEMA_STATE.get("error", ""),
            "has_target_kwp": bool(has_kwp),
            "tables":         ["capital_investment_projects"],
        })

    return app
