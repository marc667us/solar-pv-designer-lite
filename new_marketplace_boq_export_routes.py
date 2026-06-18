# ─── Routes — BOM Rates Panel + Excel + PDF export ────────────────────────────
# Slice 8: convert a marketplace BOM (basic supplier prices) into a full BOQ
# by adding labour + overhead + profit + VAT markups, then export as Excel
# (.xlsx via openpyxl) or PDF (via solar's existing MarkdownPdf chain).

_BOM_DEFAULT_RATES = {
    "labour_pct":   15.0,   # % of basic supply rate added as install labour
    "overhead_pct":  8.0,   # % of (supply + labour) added as overhead
    "profit_pct":   12.0,   # % of (supply + labour + overhead) added as profit
    "vat_pct":       0.0,   # % VAT applied AFTER profit
}


def _ensure_bom_rates_table():
    """Idempotent — runs on both SQLite (dev) and Postgres (Render)."""
    if bool(os.environ.get("DATABASE_URL")):
        _ensure_marketplace_schema_postgres()
        # Postgres path: separately create the rates table since the
        # marketplace bootstrap above is one-shot at first marketplace hit.
        for ddl in [
            """CREATE TABLE IF NOT EXISTS marketplace_bom_rates (
                bom_id       INTEGER PRIMARY KEY,
                labour_pct   REAL DEFAULT 15,
                overhead_pct REAL DEFAULT 8,
                profit_pct   REAL DEFAULT 12,
                vat_pct      REAL DEFAULT 0,
                updated_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
        ]:
            try:
                with get_db() as c:
                    c.execute(ddl)
            except Exception:
                pass
        return
    with get_db() as c:
        c.executescript(
            """
            CREATE TABLE IF NOT EXISTS marketplace_bom_rates (
                bom_id       INTEGER PRIMARY KEY,
                labour_pct   REAL DEFAULT 15,
                overhead_pct REAL DEFAULT 8,
                profit_pct   REAL DEFAULT 12,
                vat_pct      REAL DEFAULT 0,
                updated_at   TEXT DEFAULT CURRENT_TIMESTAMP
            );
            """
        )


def _bom_rates_for(bom_id: int) -> dict:
    """Return the active rates for a BOM. Falls back to defaults if no
    row exists yet — never raises."""
    _ensure_bom_rates_table()
    try:
        with get_db() as c:
            row = c.execute(
                "SELECT labour_pct, overhead_pct, profit_pct, vat_pct "
                "FROM marketplace_bom_rates WHERE bom_id=?", (bom_id,),
            ).fetchone()
    except Exception:
        row = None
    if row:
        return {
            "labour_pct":   float(row["labour_pct"]   or 0),
            "overhead_pct": float(row["overhead_pct"] or 0),
            "profit_pct":   float(row["profit_pct"]   or 0),
            "vat_pct":      float(row["vat_pct"]      or 0),
        }
    return dict(_BOM_DEFAULT_RATES)


def _bom_totals_with_rates(items, rates: dict) -> dict:
    """Compute per-line basic / install / overhead / profit / VAT / total
    rate / amount columns + grand total + per-category subtotals.

    Returns the same shape as the original _bom_totals() so existing
    templates that only use {lines, category_totals, grand_total} keep
    working — but each line dict now also carries the rate breakdown."""
    lab_pct  = max(0.0, float(rates.get("labour_pct",   0)))
    ovh_pct  = max(0.0, float(rates.get("overhead_pct", 0)))
    prf_pct  = max(0.0, float(rates.get("profit_pct",   0)))
    vat_pct  = max(0.0, float(rates.get("vat_pct",      0)))

    lines = []
    cat_totals: dict = {}
    grand = 0.0
    for it in items:
        basic_rate = float(
            (it["unit_price_override"] if it["unit_price_override"] is not None
             else (it["catalog_price"] or 0)) or 0
        )
        install_labour = basic_rate * lab_pct / 100.0
        supply_install = basic_rate + install_labour
        overhead       = supply_install * ovh_pct / 100.0
        with_overhead  = supply_install + overhead
        profit         = with_overhead * prf_pct / 100.0
        before_vat     = with_overhead + profit
        vat            = before_vat * vat_pct / 100.0
        total_rate     = before_vat + vat
        qty            = float(it["qty"] or 0)
        line_total     = total_rate * qty
        cat = it["category_name"] or "Uncategorised"
        cat_totals[cat] = cat_totals.get(cat, 0) + line_total
        grand += line_total
        lines.append({
            "item": it,
            "basic_rate": basic_rate,
            "install_labour": install_labour,
            "overhead": overhead,
            "profit": profit,
            "vat": vat,
            "total_rate": total_rate,
            "line_total": line_total,
            # Backward-compat with the old template:
            "unit_price": basic_rate,
        })
    return {
        "lines": lines,
        "category_totals": cat_totals,
        "grand_total": grand,
        "rates": rates,
        "totals_basic": sum(l["basic_rate"] * (l["item"]["qty"] or 0) for l in lines),
        "totals_labour": sum(l["install_labour"] * (l["item"]["qty"] or 0) for l in lines),
        "totals_overhead": sum(l["overhead"] * (l["item"]["qty"] or 0) for l in lines),
        "totals_profit": sum(l["profit"] * (l["item"]["qty"] or 0) for l in lines),
        "totals_vat": sum(l["vat"] * (l["item"]["qty"] or 0) for l in lines),
    }


# ──────────────────────── POST /boms/<id>/rates ─────────────────────────


@app.route("/boms/<int:bom_id>/rates", methods=["POST"])
@login_required
def boms_save_rates(bom_id):
    uid = session["user_id"]
    _bom_owned_or_404(bom_id, uid)
    csrf_protect()
    _ensure_bom_rates_table()
    f = request.form
    def _pct(name: str) -> float:
        try:
            v = float(f.get(name, _BOM_DEFAULT_RATES[name]))
        except (TypeError, ValueError):
            v = _BOM_DEFAULT_RATES[name]
        # Clamp to a sane window — protects the totals from a runaway 1e9 value.
        return max(0.0, min(100.0, v))
    lab, ovh, prf, vat = _pct("labour_pct"), _pct("overhead_pct"), _pct("profit_pct"), _pct("vat_pct")
    with get_db() as c:
        # UPSERT — INSERT OR REPLACE on SQLite; ON CONFLICT on Postgres
        # (db_adapter translates INSERT OR REPLACE → ON CONFLICT DO UPDATE).
        c.execute(
            "INSERT OR REPLACE INTO marketplace_bom_rates "
            "(bom_id, labour_pct, overhead_pct, profit_pct, vat_pct) "
            "VALUES (?, ?, ?, ?, ?)",
            (bom_id, lab, ovh, prf, vat),
        )
        c.execute(
            "UPDATE marketplace_boms SET updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (bom_id,),
        )
    flash(f"Rates updated — labour {lab}% / overhead {ovh}% / profit {prf}% / VAT {vat}%.", "success")
    return redirect(url_for("boms_view", bom_id=bom_id))


# ──────────────────────── GET /boms/<id>/boq.xlsx ────────────────────────


@app.route("/boms/<int:bom_id>/boq.xlsx")
@login_required
def boms_boq_xlsx(bom_id):
    uid = session["user_id"]
    bom = _bom_owned_or_404(bom_id, uid)
    items = _bom_items_with_prices(bom_id)
    rates = _bom_rates_for(bom_id)
    totals = _bom_totals_with_rates(items, rates)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="B45309")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    cat_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(border_style="thin", color="D1D5DB")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title + meta
    ws["A1"] = f"Bill of Quantities — {bom['title']}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Project: {bom['project_name'] or '-'}"
    ws["A3"] = f"Client : {bom['client_name'] or '-'}"
    ws["A4"] = f"Date   : {bom['updated_at']}"
    ws["A6"] = (
        f"Rates applied: labour {rates['labour_pct']}% · overhead {rates['overhead_pct']}% "
        f"· profit {rates['profit_pct']}% · VAT {rates['vat_pct']}%"
    )
    ws.merge_cells("A6:I6")

    # Header row
    headers = ["#", "Description", "Category", "Qty", "Unit", "Basic (USD)",
               "Install (USD)", "Total Rate (USD)", "Amount (USD)"]
    HROW = 8
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=HROW, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = box
        cell.alignment = Alignment(horizontal="center")

    row = HROW + 1
    prev_cat = None
    for idx, line in enumerate(totals["lines"], 1):
        it = line["item"]
        cat = it["category_name"] or "Uncategorised"
        if cat != prev_cat:
            ws.cell(row=row, column=1, value=cat).font = bold
            ws.cell(row=row, column=1).fill = cat_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            row += 1
            prev_cat = cat
        # Excel cell-content safety: prefix any leading '=' with apostrophe
        # so Excel doesn't try to evaluate user-supplied text as a formula
        # (CSV/XLSX formula-injection defence).
        def _sanitize(v: str) -> str:
            s = str(v or "")
            return "'" + s if s and s[0] in ("=", "+", "-", "@") else s
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=_sanitize(it["custom_name"]))
        ws.cell(row=row, column=3, value=_sanitize(cat))
        ws.cell(row=row, column=4, value=float(it["qty"] or 0))
        ws.cell(row=row, column=5, value=_sanitize(it["unit"]))
        ws.cell(row=row, column=6, value=round(line["basic_rate"], 2))
        ws.cell(row=row, column=7, value=round(line["install_labour"], 2))
        ws.cell(row=row, column=8, value=round(line["total_rate"], 2))
        ws.cell(row=row, column=9, value=round(line["line_total"], 2))
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = box
        row += 1

    # Subtotals + grand total
    row += 1
    for cat, sub in totals["category_totals"].items():
        ws.cell(row=row, column=2, value=f"{cat} subtotal").font = bold
        ws.cell(row=row, column=9, value=round(sub, 2)).font = bold
        row += 1
    row += 1
    ws.cell(row=row, column=2, value="GRAND TOTAL").font = title_font
    ws.cell(row=row, column=9, value=round(totals["grand_total"], 2)).font = title_font

    # Column widths
    for col, w in enumerate([5, 40, 24, 8, 8, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Category"; ws2["B1"] = "Subtotal (USD)"
    ws2["A1"].font = bold; ws2["B1"].font = bold
    r2 = 2
    for cat, sub in totals["category_totals"].items():
        ws2.cell(row=r2, column=1, value=cat)
        ws2.cell(row=r2, column=2, value=round(sub, 2))
        r2 += 1
    ws2.cell(row=r2 + 1, column=1, value="GRAND TOTAL").font = bold
    ws2.cell(row=r2 + 1, column=2, value=round(totals["grand_total"], 2)).font = bold

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = re.sub(r"[^A-Za-z0-9_.-]+", "_", bom["title"])[:60]
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"BOQ_{safe_title}.xlsx",
    )


# ──────────────────────── GET /boms/<id>/boq.pdf ─────────────────────────


@app.route("/boms/<int:bom_id>/boq.pdf")
@login_required
def boms_boq_pdf(bom_id):
    uid = session["user_id"]
    bom = _bom_owned_or_404(bom_id, uid)
    items = _bom_items_with_prices(bom_id)
    rates = _bom_rates_for(bom_id)
    totals = _bom_totals_with_rates(items, rates)

    md = []
    md.append(f"# Bill of Quantities — {bom['title']}")
    if bom["project_name"]:
        md.append(f"**Project:** {bom['project_name']}  ")
    if bom["client_name"]:
        md.append(f"**Client:** {bom['client_name']}  ")
    md.append(f"**Generated:** {bom['updated_at']}")
    md.append("")
    md.append("")
    md.append("## Rates applied")
    md.append("")
    md.append(
        f"- Install labour: **{rates['labour_pct']}%** of basic supply\n"
        f"- Overhead: **{rates['overhead_pct']}%**\n"
        f"- Profit: **{rates['profit_pct']}%**\n"
        f"- VAT: **{rates['vat_pct']}%**\n"
    )
    md.append("")
    md.append("## Line items")
    md.append("")
    md.append("| # | Description | Category | Qty | Unit | Basic Rate (USD) | Total Rate (USD) | Amount (USD) |")
    md.append("|---|---|---|---|---|---|---|---|")
    prev_cat = None
    for idx, line in enumerate(totals["lines"], 1):
        it = line["item"]
        cat = it["category_name"] or "Uncategorised"
        if cat != prev_cat:
            md.append(f"| | **{cat}** | | | | | | |")
            prev_cat = cat
        md.append(
            f"| {idx} | {it['custom_name']} | {cat} | "
            f"{it['qty']:.2f} | {it['unit']} | "
            f"{line['basic_rate']:.2f} | "
            f"{line['total_rate']:.2f} | "
            f"{line['line_total']:.2f} |"
        )
    md.append("")
    md.append("## Category subtotals")
    md.append("")
    md.append("| Category | Subtotal (USD) |")
    md.append("|---|---|")
    for cat, sub in totals["category_totals"].items():
        md.append(f"| {cat} | {sub:.2f} |")
    md.append("")
    md.append(f"## Grand total\n\n**USD {totals['grand_total']:.2f}**\n")

    safe_title = re.sub(r"[^A-Za-z0-9_.-]+", "_", bom["title"])[:60]
    return _render_pdf(
        f"BOQ — {bom['title']}",
        "\n".join(md),
        f"BOQ_{safe_title}.pdf",
    )
