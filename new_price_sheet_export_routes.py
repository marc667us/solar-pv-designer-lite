# new_price_sheet_export_routes.py
# Excel + PDF + Email export for the Basic Price Sheet -- closes the
# 4-route 404 gap surfaced by Diag Export Audit.
#
# Columns mirror the price_sheet_view.html grid:
#   # | Product / item description | Qty | Unit | Price (currency)
#     | Supplier | Brand | Phone | Email | Address


@app.route("/price-sheets/<int:sheet_id>/export.xlsx")
@app.route("/price-sheets/<int:sheet_id>.xlsx")
@login_required
def price_sheet_xlsx(sheet_id):
    _ensure_price_sheet_tables()
    uid = session["user_id"]
    sheet = _price_sheet_owned_or_404(sheet_id, uid)
    with get_db() as c:
        items = c.execute(
            "SELECT * FROM marketplace_price_sheet_items "
            "WHERE sheet_id=? ORDER BY id",
            (sheet_id,),
        ).fetchall()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Sheet"

    title_font  = Font(bold=True, size=14, color="B45309")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    bold        = Font(bold=True)
    thin        = Side(border_style="thin", color="D1D5DB")
    box         = Border(left=thin, right=thin, top=thin, bottom=thin)

    cur = sheet["currency"] or "GHS"
    ws["A1"] = f"Basic Price Sheet -- {sheet['title']}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:J1")
    ws["A2"] = f"Currency: {cur}"
    ws["A3"] = f"Generated: {sheet['created_at']}"

    headers = ["#", "Product / item description", "Qty", "Unit",
               f"Price ({cur})", "Supplier", "Brand",
               "Phone", "Email", "Address"]
    HROW = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=HROW, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = box
        cell.alignment = Alignment(horizontal="center")

    def _san(v):
        s = str(v or "")
        return "'" + s if s and s[0] in ("=", "+", "-", "@") else s

    row = HROW + 1
    for idx, it in enumerate(items, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=_san(it["custom_name"]))
        ws.cell(row=row, column=3, value=1)
        ws.cell(row=row, column=4, value=_san(it["unit"]))
        ws.cell(row=row, column=5, value=round(float(it["price_at_add"] or 0), 2))
        ws.cell(row=row, column=6, value=_san(it["supplier_name"]))
        ws.cell(row=row, column=7, value=_san(it["supplier_brand"]))
        ws.cell(row=row, column=8, value=_san(it["supplier_phone"]))
        ws.cell(row=row, column=9, value=_san(it["supplier_email"]))
        ws.cell(row=row, column=10, value=_san(it["supplier_address"]))
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = box
        row += 1

    for col, w in enumerate([5, 38, 6, 8, 14, 22, 16, 16, 24, 32], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", sheet["title"])[:60]
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"PriceSheet_{safe_name}.xlsx",
    )


def _price_sheet_markdown(sheet_id: int) -> str:
    """Markdown render for the PDF + email body."""
    with get_db() as c:
        sheet = c.execute(
            "SELECT * FROM marketplace_price_sheets WHERE id=?", (sheet_id,)
        ).fetchone()
        items = c.execute(
            "SELECT * FROM marketplace_price_sheet_items "
            "WHERE sheet_id=? ORDER BY id",
            (sheet_id,),
        ).fetchall()
    cur = sheet["currency"] or "GHS"
    md = [
        f"# Basic Price Sheet -- {sheet['title']}",
        "",
        f"**Currency:** {cur}  ",
        f"**Generated:** {sheet['created_at']}",
        "",
        f"| # | Product / item description | Qty | Unit | Price ({cur}) | Supplier | Brand | Phone | Email |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for idx, it in enumerate(items, 1):
        md.append(
            f"| {idx} | {it['custom_name']} | 1 | {it['unit']} | "
            f"{float(it['price_at_add'] or 0):.2f} | "
            f"{it['supplier_name'] or '-'} | "
            f"{it['supplier_brand'] or '-'} | "
            f"{it['supplier_phone'] or '-'} | "
            f"{it['supplier_email'] or '-'} |"
        )
    md.append("")
    md.append("---")
    md.append("")
    md.append("**Notes.** Qty = 1 per line because this is a reference price sheet, "
              "not a purchase order. Prices use indicative FX rates -- verify with "
              "the supplier before ordering.")
    return "\n".join(md)


@app.route("/price-sheets/<int:sheet_id>/export.pdf")
@app.route("/price-sheets/<int:sheet_id>.pdf")
@login_required
def price_sheet_pdf(sheet_id):
    _ensure_price_sheet_tables()
    uid = session["user_id"]
    sheet = _price_sheet_owned_or_404(sheet_id, uid)
    md = _price_sheet_markdown(sheet_id)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", sheet["title"])[:60]
    return _render_pdf(
        f"Basic Price Sheet -- {sheet['title']}",
        md,
        f"PriceSheet_{safe_name}.pdf",
    )


@app.route("/price-sheets/<int:sheet_id>/email", methods=["POST"])
@login_required
def price_sheet_email(sheet_id):
    _ensure_price_sheet_tables()
    uid = session["user_id"]
    sheet = _price_sheet_owned_or_404(sheet_id, uid)
    csrf_protect()
    to_email = (request.form.get("to_email") or "").strip().lower()[:200]
    if "@" not in to_email or "." not in to_email:
        flash("Invalid recipient email address.", "warning")
        return redirect(url_for("price_sheet_view", sheet_id=sheet_id))
    subject = (request.form.get("subject")
               or f"Basic Price Sheet -- {sheet['title']}")[:200]
    body_text = (request.form.get("body")
                 or f"Please find attached the price sheet '{sheet['title']}'.\n"
                    f"Generated {sheet['created_at']}.")

    md = _price_sheet_markdown(sheet_id)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", sheet["title"])[:60]

    attachment_bytes = b""
    try:
        from markdown_pdf import MarkdownPdf, Section
        pdf = MarkdownPdf(toc_level=2)
        pdf.add_section(Section(md, toc=False))
        import tempfile
        tf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        pdf.save(tf.name); tf.close()
        with open(tf.name, "rb") as fh:
            attachment_bytes = fh.read()
    except Exception as e:
        try: app.logger.warning("price-sheet email PDF build failed: %s", e)
        except Exception: pass

    sent = False
    try:
        from api_manager import _send_email  # type: ignore
        sent = bool(_send_email(
            to_email, subject, body_text,
            attachment_name=f"PriceSheet_{safe_name}.pdf",
            attachment_bytes=attachment_bytes,
        ))
    except Exception:
        try:
            from api_manager import send_email_with_attachment  # type: ignore
            sent = bool(send_email_with_attachment(
                to_email, subject, body_text,
                f"PriceSheet_{safe_name}.pdf", attachment_bytes,
            ))
        except Exception:
            sent = False

    flash(
        f"Email {'sent' if sent else 'queued (no email backend available)'} to {to_email}.",
        "success" if sent else "warning",
    )
    return redirect(url_for("price_sheet_view", sheet_id=sheet_id))
