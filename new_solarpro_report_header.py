# new_solarpro_report_header.py
# Shared "SolarPro Marketplace Services" branded report header
# modelled on Apinto / Agenda Commercial Limited price-list layout.
#
# Used by: boms_boq_pdf / boms_boq_xlsx (Quick Cost Estimate),
# price_sheets_pdf / price_sheets_xlsx (Basic Price Sheet),
# boq_project_pdf (formal BOQ).
#
# Layout (top of every report):
#
#   ============================================================
#   SOLARPRO MARKETPLACE SERVICES
#   Procurement | BOM | BOQ | Cost Estimate | Price Sheet
#   ============================================================
#   <REPORT SUBTITLE — e.g. BILL OF QUANTITIES>
#   ------------------------------------------------------------
#   Project    : <project name>
#   Client     : <client name>
#   Document   : <e.g. Quick Cost Estimate #42>
#   Currency   : <ISO code + symbol>
#   Generated  : <ISO timestamp>
#   ------------------------------------------------------------
#
# Apinto-style table columns (9):
#   #  Description  Qty  Unit  Basic Rate (LOCAL)  Basic Rate (US$)
#   Brand  Supplier  Phone
#

from datetime import datetime, timezone


def _solarpro_report_header_md(subtitle, project_name=None, client_name=None,
                               doc_label=None, currency="GHS", generated_at=None):
    """Return a list of markdown lines for the SolarPro Marketplace Services
    branded header. Mirrors the Apinto/Agenda price-list block."""
    lines = []
    lines.append("# SolarPro Marketplace Services")
    lines.append("")
    lines.append("*Procurement &middot; BOM &middot; BOQ &middot; Cost Estimate &middot; Price Sheet*")
    lines.append("")
    if subtitle:
        lines.append(f"## {subtitle}")
        lines.append("")
    lines.append("")
    lines.append("| | |")
    lines.append("|---|---|")
    if project_name:
        lines.append(f"| **Project** | {project_name} |")
    if client_name:
        lines.append(f"| **Client** | {client_name} |")
    if doc_label:
        lines.append(f"| **Document** | {doc_label} |")
    lines.append(f"| **Currency** | {currency} (per row: local + US$ reference) |")
    if generated_at:
        lines.append(f"| **Generated** | {generated_at} |")
    else:
        lines.append(f"| **Generated** | {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} |")
    lines.append("")
    return lines


def _solarpro_report_footer_md():
    """Branded footer block matching supplier price-list style."""
    return [
        "",
        "---",
        "",
        "**SolarPro Marketplace Services** &mdash; Africa's procurement marketplace for solar, "
        "electrical and ICT contractors. Verified supplier pricing, brand schedules, "
        "and bill-of-quantities tooling at https://solarpro.aiappinvent.com",
        "",
        "*This document is an estimate generated against the SolarPro catalogue at the "
        "stated timestamp. Final pricing is subject to supplier confirmation, currency rate "
        "at order date, freight, duties and statutory taxes.*",
        "",
    ]


def _solarpro_report_header_xlsx(ws, subtitle, project_name=None, client_name=None,
                                 doc_label=None, currency="GHS", generated_at=None):
    """Write the SolarPro branded header into an openpyxl worksheet ``ws``.

    Returns the next free row (1-based) that callers should use for the table.
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        Font = Alignment = PatternFill = None

    row = 1
    ws.cell(row=row, column=1, value="SolarPro Marketplace Services")
    if Font is not None:
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=16, bold=True, color="C49A2D")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    ws.cell(row=row, column=1, value="Procurement  -  BOM  -  BOQ  -  Cost Estimate  -  Price Sheet")
    if Font is not None:
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, italic=True, color="606060")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 2
    if subtitle:
        ws.cell(row=row, column=1, value=subtitle.upper())
        if Font is not None:
            ws.cell(row=row, column=1).font = Font(name="Calibri", size=13, bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
    row += 1
    meta_rows = []
    if project_name:
        meta_rows.append(("Project", project_name))
    if client_name:
        meta_rows.append(("Client", client_name))
    if doc_label:
        meta_rows.append(("Document", doc_label))
    meta_rows.append(("Currency", f"{currency} (each row: local + US$ reference)"))
    if generated_at:
        meta_rows.append(("Generated", str(generated_at)))
    else:
        meta_rows.append(("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")))

    for label, value in meta_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if Font is not None:
            ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
            ws.cell(row=row, column=2).font = Font(name="Calibri", size=10)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        row += 1
    row += 1
    return row


def _solarpro_fx_to_usd(local_amount, fx_rate_local_per_usd):
    """Convert a local-currency amount back to a US$ reference value.
    fx_rate_local_per_usd is the multiplier used to convert USD -> local
    (so amount_usd = amount_local / fx_rate)."""
    try:
        r = float(fx_rate_local_per_usd or 1.0)
        if r <= 0:
            return 0.0
        return float(local_amount or 0.0) / r
    except Exception:
        return 0.0


def _solarpro_xlsx_apply_borders_and_a4(ws):
    """Owner directive 2026-06-21: all reports must print A4 with solid
    dark column borders. Applies thin black borders to every used cell and
    sets A4 portrait page setup with fit-to-width."""
    try:
        from openpyxl.styles import Border, Side
        from openpyxl.worksheet.page import PageMargins
    except Exception:
        return
    thin = Side(border_style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=1, max_col=max_col):
        for cell in row:
            cell.border = box
    try:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.gridLines = False
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                      header=0.3, footer=0.3)
    except Exception:
        pass
