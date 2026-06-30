"""Generate Floor BOQ + 3 exports (2026-06-30 owner directive).

Adds:
  * "Generate Floor BOQ" button on the floor view (boq_floor_view.html).
  * GET  /boq-projects/<pid>/buildings/<bid>/floors/<fid>/boq
         -> renders boq_floor_generate.html (review page with 3 export buttons).
  * GET  /boq-projects/<pid>/buildings/<bid>/floors/<fid>/boq.pdf
         -> markdown-pdf, downloadable as application/pdf.
  * GET  /boq-projects/<pid>/buildings/<bid>/floors/<fid>/boq.xlsx
         -> openpyxl workbook with Summary sheet + one sheet per service.
  * POST /boq-projects/<pid>/buildings/<bid>/floors/<fid>/boq/email
         -> sends the PDF as an attachment. Recipient defaults to
            project.client_email, user can override in the modal.

The shared helper _floor_boq_rows(pid, fid) builds the row list +
service breakdown + bill subtotals. Each export route calls it.
"""
from __future__ import annotations
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent
CRLF = b"\r\n"

def crlf(s: bytes) -> bytes:
    return s.replace(b"\r\n", b"\n").replace(b"\n", CRLF)

def replace_once(d, old, new, label, *, crlf_target):
    if crlf_target:
        old_c, new_c = crlf(old), crlf(new)
    else:
        old_c, new_c = old, new
    if new_c in d:
        print(f"  {label}: already patched, skipping"); return d
    n = d.count(old_c)
    if n != 1:
        sys.exit(f"  {label}: expected 1 OLD match, found {n}")
    print(f"  {label}: patched")
    return d.replace(old_c, new_c, 1)


# ============================================================
# 1. web_app.py -- inject helper + 4 routes
# ============================================================
WEB = REPO / "web_app.py"
data = WEB.read_bytes()

# Anchor: just before the existing building-summary route. Include the
# decorators in OLD so the injected block lands BEFORE them.
ANCHOR_OLD = b'@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/summary")\r\n@login_required\r\ndef boq_building_summary(pid, bid):\r\n'

INJECT = b'''def _floor_boq_rows(pid, fid):
    """Fetch boq_floor_items rows for ONE floor, joined with the rate
    buildup, ordered by bill -> section -> item. Returns:
      rows           list of Row objects ready for template/spreadsheet
      bills          list of {bill_no, bill_name, subtotal}
      services_breakdown list of {code, label, sections, service_total}
      floor_total    sum of all row totals
    """
    t_clause, t_params = _boq_tenant_clause(alias="i")
    with get_db() as c:
        rows = c.execute(
            "SELECT i.*, "
            "       rb.basic_price, rb.supply_rate, rb.install_rate, "
            "       rb.overhead_pct, rb.profit_pct, rb.contingency_pct, rb.vat_pct, "
            "       rb.final_built_up_rate AS bu_final "
            "FROM boq_floor_items i "
            "LEFT JOIN boq_floor_rate_buildup rb ON rb.floor_item_id=i.id "
            "WHERE i.floor_id=? AND i.project_id=?" + t_clause + " "
            "ORDER BY COALESCE(i.bill_no,0), "
            "         COALESCE(i.section_letter,''), "
            "         COALESCE(i.display_order,0), "
            "         COALESCE(NULLIF(i.item_no_display,''),'0'), i.id",
            (fid, pid) + t_params,
        ).fetchall()
    floor_total = sum(float(r["total_amount"] or 0) for r in rows)
    # Per-bill subtotals
    bill_map = {}
    for r in rows:
        bn = int(r["bill_no"] or 0)
        bill_map.setdefault(bn, {
            "bill_no":  bn,
            "bill_name": (r["bill_name"] or _boq_lookup_bill_name(bn) or "OTHER"),
            "subtotal": 0.0,
        })
        bill_map[bn]["subtotal"] += float(r["total_amount"] or 0)
    bills = sorted(bill_map.values(), key=lambda b: b["bill_no"])
    # Service breakdown
    svc_buckets = {}
    for r in rows:
        code = (r["service_code"] or "").strip().lower()
        label = _BOQ_SERVICE_LABEL.get(code, code.replace("_", " ").title() if code else "Uncategorised")
        bucket = svc_buckets.setdefault(code or "_uncategorised", {
            "code": code or "_uncategorised",
            "label": label,
            "sections_by_letter": {},
            "rows": [],
            "service_total": 0.0,
        })
        bucket["rows"].append(r)
        bucket["service_total"] += float(r["total_amount"] or 0)
        sec_key = (int(r["bill_no"] or 0), (r["section_letter"] or "").upper())
        sec = bucket["sections_by_letter"].setdefault(sec_key, {
            "bill_no":  int(r["bill_no"] or 0),
            "bill_name": (r["bill_name"] or _boq_lookup_bill_name(int(r["bill_no"] or 0)) or "OTHER"),
            "section_letter": (r["section_letter"] or "").upper(),
            "section_title":  (r["section"] or ""),
            "subtotal": 0.0,
            "rows": [],
        })
        sec["subtotal"] += float(r["total_amount"] or 0)
        sec["rows"].append(r)
    services_breakdown = []
    for sb in svc_buckets.values():
        sb["sections"] = sorted(sb["sections_by_letter"].values(),
                                key=lambda s: (s["bill_no"], s["section_letter"]))
        del sb["sections_by_letter"]
        services_breakdown.append(sb)
    services_breakdown.sort(key=lambda b: (b["code"] == "_uncategorised", -b["service_total"]))
    return rows, bills, services_breakdown, floor_total


@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/floors/<int:fid>/boq")
@login_required
def boq_floor_boq_review(pid, bid, fid):
    """Floor BOQ review page -- list every saved item on this floor +
    per-service + per-bill breakdowns + 3 export buttons (PDF, Excel,
    Email)."""
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    building = _boq_building_owned_or_404(bid, pid)
    floor = _boq_floor_owned_or_404(fid, bid)
    rows, bills, services_breakdown, floor_total = _floor_boq_rows(pid, fid)
    return render_template(
        "boq_floor_generate.html",
        user=current_user(),
        project=project, building=building, floor=floor,
        rows=rows, bills=bills,
        services_breakdown=services_breakdown,
        floor_total=floor_total,
    )


def _floor_boq_build_pdf_bytes(project, building, floor, rows, bills, services_breakdown, floor_total):
    """Return PDF bytes for a floor BOQ using markdown-pdf."""
    from markdown_pdf import MarkdownPdf, Section
    sym = "GHS"
    title = f"Floor BOQ -- {project['project_name']} / {building['building_name']} / {floor['floor_name']}"
    md = f"# {title}\\n\\n"
    if project["client_name"]:
        md += f"**Client:** {project['client_name']}\\n\\n"
    md += "---\\n\\n## Items\\n\\n"
    md += "| Item | Description | Qty | Unit | Basic | Supply | Install | Total Rate | Amount |\\n"
    md += "|---|---|---|---|---|---|---|---|---|\\n"
    prev_bill = prev_sec = None
    for r in rows:
        bn = int(r["bill_no"] or 0)
        sl = (r["section_letter"] or "").upper()
        if bn != prev_bill:
            md += f"\\n**BILL No. {bn} -- {r['bill_name'] or 'OTHER'}**\\n\\n"
            md += "| Item | Description | Qty | Unit | Basic | Supply | Install | Total Rate | Amount |\\n"
            md += "|---|---|---|---|---|---|---|---|---|\\n"
            prev_bill = bn; prev_sec = None
        if sl != prev_sec:
            md += f"_{sl}. {(r['section'] or '').upper()}_\\n\\n"
            prev_sec = sl
        desc = str(r["description"] or "").replace("|", " / ")
        md += (f"| {r['item_no_display'] or r['item_no'] or ''} | {desc} | "
               f"{int(r['qty'] or 0)} | {r['unit'] or ''} | "
               f"{float(r['basic_price'] or 0):,.2f} | "
               f"{float(r['supply_rate'] or 0):,.2f} | "
               f"{float(r['install_rate'] or 0):,.2f} | "
               f"{float(r['final_built_up_rate'] or 0):,.2f} | "
               f"**{float(r['total_amount'] or 0):,.2f}** |\\n")
    md += f"\\n\\n**FLOOR TOTAL CARRIED TO BUILDING SUMMARY: {sym} {floor_total:,.2f}**\\n\\n"
    md += "---\\n\\n## Per-service totals\\n\\n| Service | Amount |\\n|---|---|\\n"
    for s in services_breakdown:
        md += f"| {s['label']} | {sym} {s['service_total']:,.2f} |\\n"
    md += "\\n## Per-bill totals\\n\\n| Bill | Amount |\\n|---|---|\\n"
    for b in bills:
        md += f"| BILL No. {b['bill_no']} -- {b['bill_name']} | {sym} {b['subtotal']:,.2f} |\\n"

    pdf = MarkdownPdf()
    pdf.meta["title"] = title
    pdf.add_section(Section(md, toc=False))
    import io
    buf = io.BytesIO()
    pdf.save(buf)
    return buf.getvalue()


@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/floors/<int:fid>/boq.pdf")
@login_required
@limiter.limit("10 per minute")
def boq_floor_boq_pdf(pid, bid, fid):
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    building = _boq_building_owned_or_404(bid, pid)
    floor = _boq_floor_owned_or_404(fid, bid)
    rows, bills, services_breakdown, floor_total = _floor_boq_rows(pid, fid)
    pdf_bytes = _floor_boq_build_pdf_bytes(project, building, floor, rows, bills, services_breakdown, floor_total)
    fname = f"floor_boq_{building['building_name']}_{floor['floor_name']}.pdf"
    fname = "".join(ch if (ch.isalnum() or ch in "._-") else "_" for ch in fname)
    from flask import send_file
    import io
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name=fname)


def _floor_boq_build_xlsx_bytes(project, building, floor, rows, bills, services_breakdown, floor_total):
    """Excel workbook: Summary sheet + one sheet per service."""
    import openpyxl, io, re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="B45309")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    bill_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(border_style="thin", color="D1D5DB")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _san(v):
        s = str(v or "")
        return "'" + s if s and s[0] in ("=", "+", "-", "@") else s

    # ----- Summary sheet -----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Floor BOQ Summary -- {project['project_name']}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Building : {building['building_name']}"
    ws["A3"] = f"Floor    : {floor['floor_name']}"
    ws["A4"] = f"Client   : {project['client_name'] or '-'}"

    # Per-service totals
    r2 = 6
    ws.cell(row=r2, column=1, value="Per-service totals").font = bold
    r2 += 1
    for col, h in enumerate(["Service", "Amount", "% of Floor"], 1):
        c_ = ws.cell(row=r2, column=col, value=h)
        c_.font = header_font; c_.fill = header_fill; c_.border = box
    r2 += 1
    for s in services_breakdown:
        ws.cell(row=r2, column=1, value=_san(s["label"]))
        ws.cell(row=r2, column=2, value=round(s["service_total"], 2))
        ws.cell(row=r2, column=3, value=(round(s["service_total"] * 100 / floor_total, 1) if floor_total > 0 else 0))
        for col in range(1, 4):
            ws.cell(row=r2, column=col).border = box
        r2 += 1

    r2 += 1
    ws.cell(row=r2, column=1, value="Per-bill totals").font = bold
    r2 += 1
    for col, h in enumerate(["Bill", "Amount", "% of Floor"], 1):
        c_ = ws.cell(row=r2, column=col, value=h)
        c_.font = header_font; c_.fill = header_fill; c_.border = box
    r2 += 1
    for b in bills:
        ws.cell(row=r2, column=1, value=f"BILL No. {b['bill_no']} -- {_san(b['bill_name'])}")
        ws.cell(row=r2, column=2, value=round(b["subtotal"], 2))
        ws.cell(row=r2, column=3, value=(round(b["subtotal"] * 100 / floor_total, 1) if floor_total > 0 else 0))
        for col in range(1, 4):
            ws.cell(row=r2, column=col).border = box
        r2 += 1

    r2 += 2
    ws.cell(row=r2, column=1, value="FLOOR TOTAL CARRIED TO BUILDING SUMMARY").font = title_font
    ws.cell(row=r2, column=2, value=round(floor_total, 2)).font = title_font
    ws.cell(row=r2, column=2).fill = bill_fill
    for col, w in enumerate([42, 16, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ----- Per-service sheets -----
    def _safe_sheet_name(name):
        # Excel sheet names: max 31 chars, no []:*?/\\
        name = re.sub(r"[\\[\\]:*?/\\\\]", "_", name or "Service")[:31] or "Service"
        return name

    used_names = {"Summary"}
    for s in services_breakdown:
        base = _safe_sheet_name(s["label"])
        name, n = base, 1
        while name in used_names:
            suffix = f" ({n})"; name = base[:31 - len(suffix)] + suffix; n += 1
        used_names.add(name)
        ws_s = wb.create_sheet(name)
        ws_s["A1"] = f"{s['label']} -- {floor['floor_name']}"
        ws_s["A1"].font = title_font
        ws_s.merge_cells("A1:I1")
        ws_s["A2"] = f"Service Total: {s['service_total']:,.2f}"
        ws_s["A2"].font = bold

        headers = ["Item", "Description", "Qty", "Unit", "Basic",
                   "Supply Rate", "Install Rate", "Total Rate", "Amount"]
        HROW = 4
        for col, h in enumerate(headers, 1):
            c_ = ws_s.cell(row=HROW, column=col, value=h)
            c_.font = header_font; c_.fill = header_fill; c_.border = box
            c_.alignment = Alignment(horizontal="center")

        r3 = HROW + 1
        # Group by bill -> section within the service
        for sec in s["sections"]:
            ws_s.cell(row=r3, column=1, value=f"BILL No. {sec['bill_no']} -- {sec['bill_name']}").font = bold
            ws_s.cell(row=r3, column=1).fill = bill_fill
            ws_s.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=9)
            r3 += 1
            ws_s.cell(row=r3, column=1, value=f"  {sec['section_letter']}. {(sec['section_title'] or '').upper()}").font = bold
            ws_s.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=9)
            r3 += 1
            for rr in sec["rows"]:
                ws_s.cell(row=r3, column=1, value=_san(rr["item_no_display"] or rr["item_no"] or ""))
                ws_s.cell(row=r3, column=2, value=_san(rr["description"]))
                ws_s.cell(row=r3, column=3, value=float(rr["qty"] or 0))
                ws_s.cell(row=r3, column=4, value=_san(rr["unit"]))
                ws_s.cell(row=r3, column=5, value=round(float(rr["basic_price"] or 0), 2))
                ws_s.cell(row=r3, column=6, value=round(float(rr["supply_rate"] or 0), 2))
                ws_s.cell(row=r3, column=7, value=round(float(rr["install_rate"] or 0), 2))
                ws_s.cell(row=r3, column=8, value=round(float(rr["final_built_up_rate"] or 0), 2))
                ws_s.cell(row=r3, column=9, value=round(float(rr["total_amount"] or 0), 2))
                for col in range(1, 10):
                    ws_s.cell(row=r3, column=col).border = box
                r3 += 1
            r3 += 1
        ws_s.cell(row=r3, column=8, value="SERVICE TOTAL").font = title_font
        ws_s.cell(row=r3, column=9, value=round(s["service_total"], 2)).font = title_font
        ws_s.cell(row=r3, column=9).fill = bill_fill
        for col, w in enumerate([8, 50, 8, 8, 14, 14, 14, 14, 16], 1):
            ws_s.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/floors/<int:fid>/boq.xlsx")
@login_required
@limiter.limit("10 per minute")
def boq_floor_boq_xlsx(pid, bid, fid):
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    building = _boq_building_owned_or_404(bid, pid)
    floor = _boq_floor_owned_or_404(fid, bid)
    rows, bills, services_breakdown, floor_total = _floor_boq_rows(pid, fid)
    xlsx_bytes = _floor_boq_build_xlsx_bytes(project, building, floor, rows, bills, services_breakdown, floor_total)
    fname = f"floor_boq_{building['building_name']}_{floor['floor_name']}.xlsx"
    fname = "".join(ch if (ch.isalnum() or ch in "._-") else "_" for ch in fname)
    from flask import send_file
    import io
    return send_file(io.BytesIO(xlsx_bytes),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/floors/<int:fid>/boq/email", methods=["POST"])
@login_required
@limiter.limit("5 per hour")
def boq_floor_boq_email(pid, bid, fid):
    uid = session["user_id"]
    project = _boq_project_owned_or_404(pid, uid)
    building = _boq_building_owned_or_404(bid, pid)
    floor = _boq_floor_owned_or_404(fid, bid)
    csrf_protect()
    to = (request.form.get("to", "") or "").strip()
    if not to or "@" not in to:
        flash("Please provide a valid email address.", "danger")
        return redirect(url_for("boq_floor_boq_review", pid=pid, bid=bid, fid=fid))
    subject = (request.form.get("subject", "") or
               f"Floor BOQ for review - {project['project_name']} / {floor['floor_name']}").strip()
    body_text = (request.form.get("body", "") or
                 f"Hi,\\n\\nPlease find attached the floor BOQ for "
                 f"{building['building_name']} / {floor['floor_name']} "
                 f"({project['project_name']}) for your review.\\n\\nRegards,\\nSolarPro Global").strip()
    rows, bills, services_breakdown, floor_total = _floor_boq_rows(pid, fid)
    try:
        pdf_bytes = _floor_boq_build_pdf_bytes(project, building, floor, rows, bills, services_breakdown, floor_total)
    except Exception as e:
        app.logger.exception("boq_floor_boq_email PDF build failed pid=%s fid=%s", pid, fid)
        flash(f"PDF build failed: {e}", "danger")
        return redirect(url_for("boq_floor_boq_review", pid=pid, bid=bid, fid=fid))
    fname = f"floor_boq_{building['building_name']}_{floor['floor_name']}.pdf"
    fname = "".join(ch if (ch.isalnum() or ch in "._-") else "_" for ch in fname)
    html = ("<div style='font-family:sans-serif;padding:18px;color:#1a1a2e'>"
            "<pre style='white-space:pre-wrap'>" + body_text + "</pre></div>")
    try:
        ok, err = _send_email(to, subject, html, text_body=body_text,
                              attachments=[(fname, pdf_bytes, "application/pdf")])
        if ok:
            flash(f"Floor BOQ emailed to {to}.", "success")
        else:
            flash(f"Email send failed: {err}", "danger")
    except Exception as e:
        app.logger.exception("boq_floor_boq_email send failed pid=%s fid=%s", pid, fid)
        flash(f"Email send raised: {e}", "danger")
    return redirect(url_for("boq_floor_boq_review", pid=pid, bid=bid, fid=fid))


@app.route("/boq-projects/<int:pid>/buildings/<int:bid>/summary")
@login_required
def boq_building_summary(pid, bid):
'''

INJECT_CRLF = INJECT.replace(b"\n", b"\r\n")
if INJECT_CRLF in data:
    print("  1: web_app.py floor-boq routes already injected, skipping")
else:
    if ANCHOR_OLD not in data:
        sys.exit("  1: anchor (building-summary route) not found")
    data = data.replace(ANCHOR_OLD, INJECT_CRLF, 1)
    print("  1: web_app.py floor-boq routes injected")
WEB.write_bytes(data)


# ============================================================
# 2. templates/boq_floor_view.html -- add "Generate Floor BOQ" button
# ============================================================
FV = REPO / "templates" / "boq_floor_view.html"
fv = FV.read_bytes()

FV_OLD = b'''    <a href="{{ url_for('boq_section_setup', pid=project.id, bid=building.id, fid=floor.id) }}" class="btn btn-warning fw-bold btn-sm">
      <i class="bi bi-list-ol me-1"></i>Section-by-Section
    </a>'''

FV_NEW = b'''    <a href="{{ url_for('boq_section_setup', pid=project.id, bid=building.id, fid=floor.id) }}" class="btn btn-warning fw-bold btn-sm">
      <i class="bi bi-list-ol me-1"></i>Section-by-Section
    </a>
    <a href="{{ url_for('boq_floor_boq_review', pid=project.id, bid=building.id, fid=floor.id) }}" class="btn btn-outline-warning fw-bold btn-sm"
       title="Generate this floor's BOQ for review + export as PDF / Excel / Email">
      <i class="bi bi-file-earmark-text me-1"></i>Generate Floor BOQ
    </a>'''

fv = replace_once(fv, FV_OLD, FV_NEW, "2: floor-view 'Generate Floor BOQ' button", crlf_target=True)
FV.write_bytes(fv)

print("done.")
