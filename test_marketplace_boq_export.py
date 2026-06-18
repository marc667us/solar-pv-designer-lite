"""Slice 8 — BOM rates panel + Excel/PDF export tests."""
from __future__ import annotations

import importlib.util
import re
import time
import uuid
from pathlib import Path

import pytest


_RUN_TAG = f"{int(time.time())}_{uuid.uuid4().hex[:6]}"
_SEQ = [0]


def _u(label: str) -> str:
    _SEQ[0] += 1
    return f"bq_{label}{_SEQ[0]}_{_RUN_TAG}"[:32]


@pytest.fixture(scope="module")
def app():
    spec = importlib.util.spec_from_file_location(
        "web_app", Path(__file__).resolve().parent / "web_app.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    try:
        mod.limiter.enabled = False
    except Exception:
        pass
    return mod


@pytest.fixture
def client(app):
    return app.app.test_client()


def _csrf(html: str) -> str:
    m = re.search(r'name="_csrf"\s+value="([^"]+)"', html)
    assert m, "CSRF token not found"
    return m.group(1)


def _db_row(app, sql, params):
    with app.app.app_context():
        from web_app import get_db
        with get_db() as c:
            return c.execute(sql, params).fetchone()


def _register_user(client, app, label) -> int:
    csrf = _csrf(client.get("/register").get_data(as_text=True))
    uname = _u(label)
    r = client.post(
        "/register",
        data={
            "_csrf": csrf, "username": uname,
            "email": f"{uname}@example.com",
            "password": "longenoughpw9",
            "name": "T", "company": "", "country": "Ghana",
            "terms_agreed": "1",
        },
        follow_redirects=False,
    )
    assert r.status_code in (302, 303)
    row = _db_row(app, "SELECT id FROM users WHERE username=?", (uname,))
    uid = row["id"]
    with client.session_transaction() as sess:
        sess["user_id"] = uid
    return uid


def _create_bom_with_item(client, app, title, qty=2, override_price=100):
    csrf = _csrf(client.get("/boms/new").get_data(as_text=True))
    r = client.post(
        "/boms/new", data={"_csrf": csrf, "title": title},
        follow_redirects=False,
    )
    bom_id = int(r.headers["Location"].rsplit("/", 1)[-1])
    page = client.get(f"/boms/{bom_id}").get_data(as_text=True)
    csrf = _csrf(page)
    client.post(
        f"/boms/{bom_id}/items/add",
        data={
            "_csrf": csrf, "name": "Test Cable", "qty": str(qty),
            "unit": "m", "unit_price_override": str(override_price),
        },
        follow_redirects=False,
    )
    return bom_id


# ───────────────────────── rates ─────────────────────────────────────────


def test_default_rates_apply_when_unset(client, app):
    _register_user(client, app, "A")
    bom_id = _create_bom_with_item(client, app, f"Default {_RUN_TAG}", qty=1, override_price=100)
    # No rates row → defaults: 15% labour, 8% overhead, 12% profit, 0% VAT
    # 100 * 1.15 * 1.08 * 1.12 * 1.0 = 139.10
    page = client.get(f"/boms/{bom_id}").get_data(as_text=True)
    assert "139.10" in page or "139.10" in page.replace(",", "")


def test_rates_can_be_saved(client, app):
    _register_user(client, app, "B")
    bom_id = _create_bom_with_item(client, app, f"Save {_RUN_TAG}")
    page = client.get(f"/boms/{bom_id}").get_data(as_text=True)
    csrf = _csrf(page)
    r = client.post(
        f"/boms/{bom_id}/rates",
        data={
            "_csrf": csrf, "labour_pct": "20", "overhead_pct": "10",
            "profit_pct": "15", "vat_pct": "12.5",
        },
        follow_redirects=False,
    )
    assert r.status_code in (302, 303)
    row = _db_row(
        app,
        "SELECT labour_pct, overhead_pct, profit_pct, vat_pct "
        "FROM marketplace_bom_rates WHERE bom_id=?", (bom_id,),
    )
    assert row["labour_pct"] == 20
    assert row["overhead_pct"] == 10
    assert row["profit_pct"] == 15
    assert row["vat_pct"] == 12.5


def test_rates_clamp_out_of_range_input(client, app):
    """Out-of-range percent (negative or > 100) gets clamped, not stored raw."""
    _register_user(client, app, "C")
    bom_id = _create_bom_with_item(client, app, f"Clamp {_RUN_TAG}")
    csrf = _csrf(client.get(f"/boms/{bom_id}").get_data(as_text=True))
    client.post(
        f"/boms/{bom_id}/rates",
        data={
            "_csrf": csrf, "labour_pct": "-50", "overhead_pct": "1000",
            "profit_pct": "0", "vat_pct": "0",
        },
        follow_redirects=False,
    )
    row = _db_row(
        app, "SELECT labour_pct, overhead_pct FROM marketplace_bom_rates WHERE bom_id=?",
        (bom_id,),
    )
    assert row["labour_pct"] == 0      # clamped from -50
    assert row["overhead_pct"] == 100  # clamped from 1000


def test_other_user_cannot_save_rates(client, app):
    a = _register_user(client, app, "ownerA")
    bom_id = _create_bom_with_item(client, app, f"OwnerA {_RUN_TAG}")
    _register_user(client, app, "intruder")
    # Even with a valid CSRF (from another page), IDOR blocks via _bom_owned_or_404
    csrf = _csrf(client.get("/boms/new").get_data(as_text=True))
    r = client.post(
        f"/boms/{bom_id}/rates",
        data={"_csrf": csrf, "labour_pct": "99",
              "overhead_pct": "0", "profit_pct": "0", "vat_pct": "0"},
        follow_redirects=False,
    )
    assert r.status_code == 404


def test_total_rate_math_matches_formula(client, app):
    """100 × (1 + 20/100) × (1 + 10/100) × (1 + 15/100) × (1 + 12.5/100)
    = 100 × 1.2 × 1.1 × 1.15 × 1.125 = 170.775"""
    _register_user(client, app, "math")
    bom_id = _create_bom_with_item(client, app, f"Math {_RUN_TAG}", qty=1, override_price=100)
    csrf = _csrf(client.get(f"/boms/{bom_id}").get_data(as_text=True))
    client.post(
        f"/boms/{bom_id}/rates",
        data={"_csrf": csrf, "labour_pct": "20", "overhead_pct": "10",
              "profit_pct": "15", "vat_pct": "12.5"},
        follow_redirects=False,
    )
    page = client.get(f"/boms/{bom_id}").get_data(as_text=True)
    # The grand total should match 170.78 to 2 d.p.
    assert "170.78" in page or "170.77" in page  # rounding flexibility


# ───────────────────────── Excel + PDF export ────────────────────────────


def test_excel_export_returns_xlsx(client, app):
    _register_user(client, app, "xl")
    bom_id = _create_bom_with_item(client, app, f"XL {_RUN_TAG}")
    r = client.get(f"/boms/{bom_id}/boq.xlsx")
    assert r.status_code == 200
    assert r.headers["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    # XLSX is a ZIP — magic bytes 'PK\x03\x04'
    assert r.data[:4] == b"PK\x03\x04"
    assert len(r.data) > 2000  # non-trivial workbook


def test_excel_export_sanitises_formula_injection(client, app):
    """A product name starting with '=' would be evaluated as a formula by
    Excel — defence is to prefix with an apostrophe."""
    uid = _register_user(client, app, "form")
    bom_id = _create_bom_with_item(client, app, f"Form {_RUN_TAG}")
    csrf = _csrf(client.get(f"/boms/{bom_id}").get_data(as_text=True))
    client.post(
        f"/boms/{bom_id}/items/add",
        data={
            "_csrf": csrf,
            "name": "=SUM(A1:A99)",   # hostile leading '='
            "qty": "1", "unit": "No.", "unit_price_override": "10",
        },
        follow_redirects=False,
    )
    r = client.get(f"/boms/{bom_id}/boq.xlsx")
    assert r.status_code == 200
    # Open the workbook via openpyxl to inspect the actual cell content
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.data), data_only=False)
    ws = wb.active
    raw_descriptions = [
        ws.cell(row=row, column=2).value
        for row in range(9, 20)
        if ws.cell(row=row, column=2).value
    ]
    hostile = [v for v in raw_descriptions if v and v.startswith("=SUM")]
    assert not hostile, f"unescaped formula leaked into XLSX: {hostile}"
    # The escaped (apostrophe-prefixed) form must be present.
    safe = [v for v in raw_descriptions if v == "'=SUM(A1:A99)"]
    assert safe, "expected apostrophe-prefixed sanitised value"


def test_pdf_export_returns_pdf(client, app):
    _register_user(client, app, "pdf")
    bom_id = _create_bom_with_item(client, app, f"PDF {_RUN_TAG}")
    r = client.get(f"/boms/{bom_id}/boq.pdf")
    assert r.status_code == 200
    assert r.headers["Content-Type"] == "application/pdf"
    # PDF magic bytes: '%PDF'
    assert r.data[:4] == b"%PDF"
    assert len(r.data) > 1000


def test_export_routes_require_login(client):
    r1 = client.get("/boms/1/boq.xlsx", follow_redirects=False)
    r2 = client.get("/boms/1/boq.pdf", follow_redirects=False)
    assert r1.status_code == 302
    assert r2.status_code == 302
    assert "/login" in r1.headers.get("Location", "")
    assert "/login" in r2.headers.get("Location", "")


def test_export_idor_blocks_other_users(client, app):
    _register_user(client, app, "ownerB")
    bom_id = _create_bom_with_item(client, app, f"OwnerB {_RUN_TAG}")
    _register_user(client, app, "thief")
    r1 = client.get(f"/boms/{bom_id}/boq.xlsx", follow_redirects=False)
    r2 = client.get(f"/boms/{bom_id}/boq.pdf", follow_redirects=False)
    assert r1.status_code == 404
    assert r2.status_code == 404
